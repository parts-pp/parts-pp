import os
import json
import re
import asyncio
import uuid
import html
import logging
from datetime import datetime, timezone, timedelta
import threading
import time
from http.server import HTTPServer, BaseHTTPRequestHandler
from dotenv import load_dotenv

# ===== Atomic Excel Save (SAFE PATCH) =====
# مكانه: بعد import load_dotenv مباشرة (قبل أي استخدام لـ openpyxl)
import tempfile
from openpyxl.workbook.workbook import Workbook

if not hasattr(Workbook, "_atomic_save_patched"):
    _orig_save = Workbook.save

    def _atomic_save(self, filename):
        # إذا لم يكن مسارًا نصيًا، استخدم الحفظ الأصلي
        if not isinstance(filename, (str, bytes, os.PathLike)):
            return _orig_save(self, filename)

        folder = os.path.dirname(os.path.abspath(filename)) or "."
        base = os.path.basename(filename)

        # إنشاء ملف مؤقت بنفس المجلد (ضروري لـ atomic replace)
        fd, tmp_path = tempfile.mkstemp(prefix=base + ".", suffix=".tmp", dir=folder)
        os.close(fd)

        try:
            # حفظ فعلي إلى الملف المؤقت
            _orig_save(self, tmp_path)

            # استبدال ذري: إما القديم كامل أو الجديد كامل
            os.replace(tmp_path, filename)
        finally:
            # تنظيف في حال بقي ملف مؤقت
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except Exception:
                pass

    Workbook.save = _atomic_save
    Workbook._atomic_save_patched = True
# ===== End Atomic Excel Save =====

from aiohttp import web

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Image

from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    InputMediaPhoto,
    InputMediaVideo,
    InputMediaDocument,
    InputFile,
)
from telegram.error import Forbidden, BadRequest, TimedOut
from telegram.constants import ChatType
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    ContextTypes,
    ChatMemberHandler,   # ✅ أضف هذا السطر
    ChatJoinRequestHandler,
    filters,
)
from telegram.request import HTTPXRequest

from pp_states import *

from pp_excel import (
    ensure_workbook,
    add_order,
    add_items,
    generate_order_id,
    update_order_fields,
    update_order_payment,
    update_order_status,
    update_delivery,
    get_order_user_id,
    get_order_assignment,
    get_order_bundle,
    mark_order_forwarded,
    get_trader_profile,
    upsert_trader_profile,
    list_orders,
    list_orders_for_trader,
    compute_admin_financials,
    compute_revenue_breakdown,
    get_setting,
    set_setting,
    append_legal_log,
    list_traders,
    set_trader_enabled,
    is_trader_enabled,
    list_legal_log,
    month_key_utc,
    upsert_trader_subscription,
    get_trader_subscription,
    list_trader_subscriptions,
)


# ===== Excel write lock + short-lived bundle cache (SAFE PATCH) =====
# الهدف: تقليل الأعطال (Race/Corruption) بدون تغيير منطق الدوال في pp_excel
# - قفل واحد لكل عمليات الإكسل (write + read الحساسة)
# - Cache قصير جدًا لقراءة نفس الطلب عدة مرات خلال نفس الثواني
# ملاحظة: pp_excel دوالها Sync، لذلك نستخدم threading.RLock كـ "قفل فعلي".
# (asyncio.Lock وحده لا يصلح داخل دوال sync بدون await/executor)

_EXCEL_WRITE_LOCK = threading.RLock()

# Cache: order_id -> (ts_monotonic, bundle)
_ORDER_BUNDLE_CACHE: dict[str, tuple[float, dict]] = {}
_ORDER_BUNDLE_TTL_SECONDS = 1.5

# احتفظ بالأصول قبل إعادة التعريف
_pp_get_order_bundle = get_order_bundle
_pp_update_order_fields = update_order_fields
_pp_update_order_payment = update_order_payment
_pp_update_order_status = update_order_status
_pp_update_delivery = update_delivery
_pp_add_order = add_order
_pp_add_items = add_items
_pp_mark_order_forwarded = mark_order_forwarded
_pp_set_setting = set_setting
_pp_append_legal_log = append_legal_log
_pp_upsert_trader_profile = upsert_trader_profile
_pp_set_trader_enabled = set_trader_enabled
_pp_upsert_trader_subscription = upsert_trader_subscription

def _bundle_cache_drop(order_id: str | None = None) -> None:
    try:
        if order_id:
            _ORDER_BUNDLE_CACHE.pop(str(order_id).strip(), None)
        else:
            _ORDER_BUNDLE_CACHE.clear()
    except Exception:
        pass

def get_order_bundle(order_id: str):
    oid = str(order_id or "").strip()
    if not oid:
        return {"order": {}, "items": []}

    now = time.monotonic()
    try:
        hit = _ORDER_BUNDLE_CACHE.get(oid)
        if hit and (now - float(hit[0])) <= _ORDER_BUNDLE_TTL_SECONDS:
            return hit[1]
    except Exception:
        pass

    # قفل قراءة/فتح الملف (يقلل 400/Timeout من تزامن I/O)
    with _EXCEL_WRITE_LOCK:
        b = _pp_get_order_bundle(oid)

    try:
        if isinstance(b, dict):
            _ORDER_BUNDLE_CACHE[oid] = (now, b)
    except Exception:
        pass
    return b

def update_order_fields(order_id: str, fields: dict):
    oid = str(order_id or "").strip()
    with _EXCEL_WRITE_LOCK:
        r = _pp_update_order_fields(oid, fields)
    _bundle_cache_drop(oid)
    return r

def update_order_payment(order_id: str, **kwargs):
    oid = str(order_id or "").strip()
    with _EXCEL_WRITE_LOCK:
        r = _pp_update_order_payment(oid, **kwargs)
    _bundle_cache_drop(oid)
    return r

def update_order_status(order_id: str, status: str, **kwargs):
    oid = str(order_id or "").strip()
    with _EXCEL_WRITE_LOCK:
        r = _pp_update_order_status(oid, status, **kwargs)
    _bundle_cache_drop(oid)
    return r

def update_delivery(order_id: str, *args, **kwargs):
    oid = str(order_id or "").strip()
    with _EXCEL_WRITE_LOCK:
        r = _pp_update_delivery(oid, *args, **kwargs)
    _bundle_cache_drop(oid)
    return r

def add_order(*args, **kwargs):
    with _EXCEL_WRITE_LOCK:
        return _pp_add_order(*args, **kwargs)

def add_items(*args, **kwargs):
    with _EXCEL_WRITE_LOCK:
        return _pp_add_items(*args, **kwargs)

def mark_order_forwarded(order_id: str, *args, **kwargs):
    oid = str(order_id or "").strip()
    with _EXCEL_WRITE_LOCK:
        r = _pp_mark_order_forwarded(oid, *args, **kwargs)
    _bundle_cache_drop(oid)
    return r

def set_setting(key: str, value: str):
    with _EXCEL_WRITE_LOCK:
        return _pp_set_setting(key, value)

def append_legal_log(*args, **kwargs):
    with _EXCEL_WRITE_LOCK:
        return _pp_append_legal_log(*args, **kwargs)

def upsert_trader_profile(*args, **kwargs):
    with _EXCEL_WRITE_LOCK:
        return _pp_upsert_trader_profile(*args, **kwargs)

def set_trader_enabled(*args, **kwargs):
    with _EXCEL_WRITE_LOCK:
        return _pp_set_trader_enabled(*args, **kwargs)

def upsert_trader_subscription(*args, **kwargs):
    with _EXCEL_WRITE_LOCK:
        return _pp_upsert_trader_subscription(*args, **kwargs)

# ===== End Excel write lock + bundle cache =====

from pp_security import parse_admin_ids


load_dotenv()

BOT_TOKEN = (os.getenv("PP_BOT_TOKEN") or "").strip()

TEAM_CHAT_ID_RAW = (os.getenv("PARTS_TEAM_CHAT_ID") or "").strip()
TEAM_CHAT_ID = int(TEAM_CHAT_ID_RAW) if TEAM_CHAT_ID_RAW.lstrip("-").isdigit() else None

# ✅ مجموعة التجار (لازم البوت يكون عضو فيها)
TRADERS_GROUP_ID_RAW = (os.getenv("PP_TRADERS_GROUP_ID") or "").strip()
TRADERS_GROUP_ID = int(TRADERS_GROUP_ID_RAW) if TRADERS_GROUP_ID_RAW.lstrip("-").isdigit() else None

ADMIN_IDS = parse_admin_ids()

# ===== Backup (scenario الجديد: Manual Backup + Restore Last Pinned) =====
BACKUP_CHAT_ID_RAW = (os.getenv("PP_BACKUP_CHAT_ID") or "").strip()
PP_BACKUP_CHAT_ID = int(BACKUP_CHAT_ID_RAW) if BACKUP_CHAT_ID_RAW.lstrip("-").isdigit() else None

# ✅ حد أدنى بين نسختين (لمنع التكرار/السبام)
PP_BACKUP_MIN_SECONDS = int((os.getenv("PP_BACKUP_MIN_SECONDS") or "600").strip() or "600")

PP_BOT_USERNAME = (os.getenv("PP_BOT_USERNAME") or "ppartsbot").strip().lstrip("@")
PP_BOT_DEEPLINK = f"https://t.me/{PP_BOT_USERNAME}?start=1"

# =========================
# ✅ بوابة التجار (طلبات الانضمام + رفع المستندات) — تدفق مراحل احترافي
# - لا يتم إرسال أي رسالة داخل مجموعة التجار إطلاقاً
# - إشعارات الإدارة تذهب للخاص (ADMIN_IDS) فقط
# - المتقدم يرفع: (1) السجل التجاري → (2) رخصة/إثبات المتجر → (3) الرقم الضريبي (نص)
# - الإدارة تراسل المتقدم "داخل البوت" وباسم الإدارة
# =========================
STAGE_JOIN_CR = "pp_join_cr"
STAGE_JOIN_LICENSE = "pp_join_license"
STAGE_JOIN_VAT = "pp_join_vat"
STAGE_JOIN_WAIT = "pp_join_wait"
STAGE_APPLICANT_CHAT_ADMIN = "pp_applicant_chat_admin"

def _join_portal_brand() -> str:
    return "🚗 P Parts | بوابة التجار"

def _join_portal_open_url() -> str:
    u = (PP_BOT_USERNAME or "ppartsbot").strip().lstrip("@")
    return f"https://t.me/{u}?start=join"

def _join_portal_applicant_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[InlineKeyboardButton("🔗 فتح البوت والتقديم", url=_join_portal_open_url())]])

def _join_portal_admin_kb(applicant_id: int) -> InlineKeyboardMarkup:
    aid = int(applicant_id or 0)
    # ✅ 3 أزرار فقط (مراسلة/قبول/رفض) — بدون زر "فتح البوت" عند الإدارة
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("💬 مراسلة المتقدم", callback_data=f"pp_join_chat|{aid}")],
        [InlineKeyboardButton("✅ قبول الطلب", callback_data=f"pp_join_appr|{aid}"),
         InlineKeyboardButton("⛔ رفض الطلب", callback_data=f"pp_join_decl|{aid}")],
    ])

def _join_portal_applicant_reply_kb(admin_id: int) -> InlineKeyboardMarkup:
    aid = int(admin_id or 0)
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✍️ رد للإدارة", callback_data=f"pp_applicant_chat_admin|{aid}")],
        [InlineKeyboardButton("✖️ إنهاء", callback_data="pp_applicant_chat_admin_done")],
    ])

async def _notify_admins_private(context: ContextTypes.DEFAULT_TYPE, text: str, kb=None):
    for aid in (ADMIN_IDS or []):
        try:
            await context.bot.send_message(
                chat_id=int(aid),
                text=text,
                parse_mode="HTML",
                reply_markup=kb,
                disable_web_page_preview=True,
            )
        except Exception:
            pass

def _join_ud(context: ContextTypes.DEFAULT_TYPE, user_id: int) -> dict:
    ud = get_ud(context, int(user_id))
    d = ud.get("join_portal")
    if not isinstance(d, dict):
        d = {}
        ud["join_portal"] = d
    return d

def _join_reset(context: ContextTypes.DEFAULT_TYPE, user_id: int):
    ud = get_ud(context, int(user_id))
    ud.pop("join_portal", None)
    set_stage(context, int(user_id), STAGE_NONE)

def _who_html(u) -> tuple[str, str]:
    full_name = (" ".join([getattr(u, "first_name", "") or "", getattr(u, "last_name", "") or ""])).strip()
    username = (getattr(u, "username", None) or "").strip()
    who = html.escape(full_name or (f"@{username}" if username else "—"), quote=False)
    return who, username

async def traders_join_request_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    jr = getattr(update, "chat_join_request", None)
    if not jr:
        return
    try:
        if TRADERS_GROUP_ID and int(jr.chat.id) != int(TRADERS_GROUP_ID):
            return
    except Exception:
        return

    u = jr.from_user
    uid = int(getattr(u, "id", 0) or 0)
    if not uid:
        return

    brand = _join_portal_brand()
    who, username = _who_html(u)

    applicant_text = (
        f"{brand}\n\n"
        "هذه مجموعة خاصة بالتجار وموردي قطع الغيار فقط.\n\n"
        "لإكمال طلب الانضمام: افتح البوت وابدأ التقديم، ثم أرسل المستندات بالترتيب التالي:\n"
        "1) السجل التجاري\n"
        "2) رخصة/إثبات المتجر\n"
        "3) الرقم الضريبي (نص)\n\n"
        "بعد الإرسال سيتم مراجعة الطلب من الإدارة.\n\n"
        f"✅ التقديم والطلبات عبر @{PP_BOT_USERNAME} فقط"
    )

    dm_sent = False
    try:
        await context.bot.send_message(
            chat_id=uid,
            text=applicant_text,
            reply_markup=_join_portal_applicant_kb(),
            disable_web_page_preview=True,
        )
        dm_sent = True
    except Exception:
        dm_sent = False

    admin_text = (
        f"{brand}\n\n"
        "🛡️ <b>طلب انضمام جديد لمجموعة التجار</b>\n\n"
        f"👤 الاسم: <b>{who}</b>\n"
        f"🆔 ID: <code>{uid}</code>\n"
        + (f"🔗 المستخدم: @{html.escape(username, quote=False)}\n" if username else "")
        + "\n"
        "📌 المطلوب قبل القبول:\n"
        "1) السجل التجاري\n"
        "2) رخصة/إثبات المتجر\n"
        "3) الرقم الضريبي\n\n"
        + ("✅ تم إرسال التعليمات له بالخاص." if dm_sent else "⚠️ لم نستطع مراسلته بالخاص (غالباً لم يفتح البوت).")
        + "\n\n"
    )
    await _notify_admins_private(context, admin_text, kb=_join_portal_admin_kb(uid))

async def join_portal_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message:
        return
    user_id = int(update.effective_user.id)
    _join_reset(context, user_id)
    set_stage(context, user_id, STAGE_JOIN_CR)

    txt = (
        f"{_join_portal_brand()}\n\n"
        "📥 <b>تقديم طلب الانضمام (للتجار وموردي قطع الغيار فقط)</b>\n\n"
        "الخطوة 1/3\n"
        "📎 أرسل الآن <b>السجل التجاري</b> (صورة أو PDF).\n\n"
    )
    kb = InlineKeyboardMarkup([[InlineKeyboardButton("✖️ إنهاء", callback_data="pp_join_done")]])
    await update.message.reply_text(txt, parse_mode="HTML", reply_markup=kb, disable_web_page_preview=True)

async def pp_join_done_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    try:
        await q.answer()
    except Exception:
        pass
    uid = int(getattr(q.from_user, "id", 0) or 0)
    if not uid:
        return
    _join_reset(context, uid)
    try:
        await q.message.reply_text("✅ تم إنهاء التقديم.")
    except Exception:
        pass

async def pp_join_chat_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    try:
        await q.answer()
    except Exception:
        pass
    actor_id = int(getattr(q.from_user, "id", 0) or 0)
    if actor_id not in (ADMIN_IDS or []):
        try:
            await q.answer("غير مصرح", show_alert=True)
        except Exception:
            pass
        return
    data = (q.data or "").strip()
    try:
        _, uid_str = data.split("|", 1)
        applicant_id = int(uid_str)
    except Exception:
        return

    ud = get_ud(context, actor_id)
    ud["admin_chat_order_id"] = "JOIN"
    ud["admin_chat_peer_id"] = applicant_id
    ud["admin_chat_role"] = "applicant"
    set_stage(context, actor_id, STAGE_ADMIN_CHAT)

    await q.message.reply_text(
        f"👤 الإدارة\n🧾 بوابة التجار\n🆔 المتقدم: <code>{applicant_id}</code>\nاكتب رسالتك الآن.",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("✖️ إنهاء", callback_data="pp_admin_chat_done")]]),
    )

async def applicant_chat_admin_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    try:
        await q.answer()
    except Exception:
        pass
    actor_id = int(getattr(q.from_user, "id", 0) or 0)
    data = (q.data or "").strip()
    try:
        _, admin_id = data.split("|", 1)
        admin_id = int(admin_id)
    except Exception:
        return
    ud = get_ud(context, actor_id)
    ud["applicant_chat_admin_id"] = admin_id
    set_stage(context, actor_id, STAGE_APPLICANT_CHAT_ADMIN)
    await q.message.reply_text("🟨 مراسلة الإدارة\nاكتب رسالتك الآن.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("✖️ إنهاء", callback_data="pp_applicant_chat_admin_done")]]))

async def applicant_chat_admin_done_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    try:
        await q.answer()
    except Exception:
        pass
    actor_id = int(getattr(q.from_user, "id", 0) or 0)
    ud = get_ud(context, actor_id)
    ud.pop("applicant_chat_admin_id", None)
    set_stage(context, actor_id, STAGE_NONE)
    try:
        await q.message.reply_text("تم إنهاء وضع المراسلة.")
    except Exception:
        pass

async def pp_join_admin_action_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    try:
        await q.answer()
    except Exception:
        pass
    actor = int(getattr(q.from_user, "id", 0) or 0)
    if actor not in (ADMIN_IDS or []):
        try:
            await q.answer("غير مصرح", show_alert=True)
        except Exception:
            pass
        return

    data = (q.data or "").strip()
    try:
        action, uid_str = data.split("|", 1)
        applicant_id = int(uid_str)
    except Exception:
        return

    brand = _join_portal_brand()
    ok = False
    approved = False
    err = ""
    try:
        if action == "pp_join_appr":
            approved = True
            await context.bot.approve_chat_join_request(chat_id=int(TRADERS_GROUP_ID), user_id=applicant_id)
            ok = True
        else:
            approved = False
            await context.bot.decline_chat_join_request(chat_id=int(TRADERS_GROUP_ID), user_id=applicant_id)
            ok = True
    except Exception as e:
        ok = False
        err = str(e)
    # إشعار المتقدم
    try:
        if ok and approved:
            await context.bot.send_message(
                chat_id=applicant_id,
                text=(f"{brand}\n\n✅ تم قبول طلب انضمامك."),
                disable_web_page_preview=True,
            )
        elif ok and (not approved):
            await context.bot.send_message(
                chat_id=applicant_id,
                text=(f"{brand}\n\n⛔ تم رفض طلب الانضمام. يمكنك إعادة التقديم.\n\n✅ التقديم والطلبات عبر @{PP_BOT_USERNAME} فقط"),
                disable_web_page_preview=True,
            )
    except Exception:
        pass

    # إشعار الإدارة بالنتيجة + زر مراسلة داخلي
    who = "—"
    uname = ""
    try:
        ch = await context.bot.get_chat(applicant_id)
        full_name = (" ".join([getattr(ch, "first_name", "") or "", getattr(ch, "last_name", "") or ""])).strip()
        uname = (getattr(ch, "username", None) or "").strip()
        who = html.escape(full_name or (f"@{uname}" if uname else "—"), quote=False)
    except Exception:
        pass

    status_line = "✅ تم قبول الطلب" if ok and approved else "⛔ تم رفض الطلب" if ok else "⚠️ تعذر تنفيذ الإجراء"
    admin_txt = (
        f"{brand}\n\n"
        f"{status_line}\n\n"
        f"👤 الاسم: <b>{who}</b>\n"
        f"🆔 ID: <code>{applicant_id}</code>\n"
        + (f"🔗 المستخدم: @{html.escape(uname, quote=False)}\n" if uname else "")
        + (f"🛑 السبب: <code>{html.escape(err, quote=False)[:250]}</code>\n" if (not ok and err) else "")
    )
    await _notify_admins_private(context, admin_txt, kb=_join_portal_admin_kb(applicant_id))



async def _is_trader_group_member(context: ContextTypes.DEFAULT_TYPE, user_id: int) -> bool:
    """
    True فقط إذا كان المستخدم عضو/ادمن/منشئ داخل مجموعة التجار.
    لازم البوت يكون عضو (ويفضل Admin) في مجموعة التجار.
    """
    if not TRADERS_GROUP_ID:
        return False
    try:
        m = await context.bot.get_chat_member(chat_id=TRADERS_GROUP_ID, user_id=int(user_id))
        st = (getattr(m, "status", None) or "").lower()
        return st in ("member", "administrator", "creator")
    except Exception:
        return False

# ==============================
# Quote Sessions (Order-scoped)
# ==============================

def _qses(td: dict) -> dict:
    qs = td.get("quote_sessions")
    if not isinstance(qs, dict):
        qs = {}
        td["quote_sessions"] = qs
    return qs

def _qget(td: dict, order_id: str, create: bool = True) -> dict:
    qs = _qses(td)
    oid = str(order_id or "").strip()
    if not oid:
        return {}
    s = qs.get(oid)
    if not isinstance(s, dict):
        if not create:
            return {}
        s = {}
        qs[oid] = s
    # ضمان مفاتيح أساسية
    if not isinstance(s.get("item_prices"), dict):
        s["item_prices"] = {}
    return s

def _qreset(td: dict, order_id: str):
    qs = _qses(td)
    oid = str(order_id or "").strip()
    if oid:
        qs.pop(oid, None)

def _q_is_sent(s: dict) -> bool:
    try:
        return str((s or {}).get("sent") or "").strip().lower() in ("1", "yes", "true", "sent", "done")
    except Exception:
        return False

def _q_mark_sent(s: dict):
    try:
        s["sent"] = "1"
        s["sent_at_utc"] = utc_now_iso()
    except Exception:
        pass

def _q_get_version(s: dict) -> int:
    try:
        v = int((s or {}).get("version") or 1)
        return v if v >= 1 else 1
    except Exception:
        return 1

def _q_bump_version(s: dict) -> int:
    try:
        v = _q_get_version(s) + 1
        s["version"] = v
        # إعادة فتح التحرير للإصدار الجديد
        s.pop("sent", None)
        s.pop("sent_at_utc", None)
        return v
    except Exception:
        return 1

def _qgc(td: dict, keep_last: int = 50):
    """تنظيف جلسات قديمة: يبقي آخر N جلسات فقط."""
    qs = _qses(td)
    if len(qs) <= keep_last:
        return
    items = []
    for oid, s in qs.items():
        lt = 0
        try:
            lt = int((s or {}).get("last_touch") or 0)
        except Exception:
            lt = 0
        items.append((lt, oid))
    items.sort(reverse=True)
    keep = set([oid for _, oid in items[:keep_last]])
    for oid in list(qs.keys()):
        if oid not in keep:
            qs.pop(oid, None)

async def _notify_invoice_error(context, order_id: str, stage: str, err: Exception):
    # ✅ لوق عربي واضح (بدون لمس الاكسل)
    try:
        log.exception("🧾 فشل الفاتورة | الطلب=%s | المرحلة=%s | الخطأ=%s", order_id, stage, err)
    except Exception:
        pass

    try:
        log_event("فشل نظام الفواتير", order_id=order_id, stage=stage, error=str(err))
    except Exception:
        pass

    msg = (
        "⚠️ فشل نظام الفواتير الداخلية\n\n"
        f"🧾 رقم الطلب: {order_id}\n"
        f"📍 المرحلة: {stage}\n"
        f"🛑 الخطأ:\n{err}"
    )
    for aid in ADMIN_IDS:
        try:
            await context.bot.send_message(chat_id=aid, text=msg)
        except Exception:
            pass

def _is_maintenance_mode() -> bool:
    try:
        v = (get_setting("maintenance_mode", "off") or "").strip().lower()
        return v in ("on", "1", "yes", "true")
    except Exception:
        return False

# ===== Platform Fee Free Mode (settings) =====
PLATFORM_FEE_FREE_KEY = "platform_fee_free"  # 1/0 in settings sheet

def _is_platform_fee_free_mode() -> bool:
    try:
        v = str(get_setting(PLATFORM_FEE_FREE_KEY) or "").strip().lower()
    except Exception:
        v = ""
    return v in ("1", "true", "yes", "on", "enable", "enabled")

def _set_platform_fee_free_mode(enable: bool) -> None:
    try:
        set_setting(PLATFORM_FEE_FREE_KEY, "1" if enable else "0")
    except Exception:
        pass

def _maintenance_block_text() -> str:
    return (
        "🟧 <b>تنبيه صيانة</b>\n"
        "المنصة حاليا في وضع الصيانة المؤقتة.\n"
        "⛔ تم ايقاف استقبال الطلبات الجديدة وتقديم عروض السعر مؤقتا.\n"
        "يرجى المحاولة لاحقا."
    )

def _trader_is_disabled(tid: int) -> bool:
    try:
        return not bool(is_trader_enabled(int(tid)))
    except Exception as e:
        # ✅ Fail-closed: إذا فشلنا نقرأ الحالة، نعتبره موقوف (أمان للمبيعات)
        try:
            log.exception("TRADER_ENABLE_CHECK_FAILED tid=%s", tid)
        except Exception:
            pass
        return True

async def _deny_disabled_trader_q(q, reason: str = "حساب التاجر موقوف"):
    # 1) تنبيه سريع (Alert)
    try:
        await _alert(q, f"⛔ {reason}")
    except Exception:
        try:
            await q.answer(f"⛔ {reason}", show_alert=True)
        except Exception:
            pass

    # 2) رسالة خاصة واضحة للتاجر + زر مراسلة الإدارة
    try:
        uid = int(getattr(q, "from_user", None).id or 0)
    except Exception:
        uid = 0

    if not uid:
        return

    try:
        bot = q.get_bot()
    except Exception:
        bot = None

    if not bot:
        return

    try:
        await bot.send_message(
            chat_id=uid,
            text=(
                f"{_user_name(q)}\n"
                "⛔ حسابك موقوف مؤقتًا.\n\n"
                "هذا الزر غير متاح لك الآن.\n"
                "راجع لوحة التاجر لمعرفة الحالة، أو تواصل مع المنصة عبر الزر بالأسفل."
            ),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("📩 اتصل بالمنصة", callback_data="pp_support_open")],
            ]),
            disable_web_page_preview=True,
        )
    except Exception:
        pass

def _trader_disabled_msg() -> str:
    return "⛔ حسابك موقوف مؤقتًا. راجع لوحة التاجر لمعرفة الحالة، أو تواصل مع الإدارة عبر زر (مراسلة الإدارة)."


def _bot_username(context: ContextTypes.DEFAULT_TYPE = None) -> str:
    # اسم المنصة للروابط (deep-link). يعتمد على PP_BOT_USERNAME من env
    try:
        return (PP_BOT_USERNAME or '').strip().lstrip('@') or 'ppartsbot'
    except Exception:
        return 'ppartsbot'



# ===== Order ID display helper (NO LINKS / NO CODE) =====
# الهدف: إبقاء رقم الطلب كنص عادي حتى يتعامل معه تيليجرام تلقائياً
# (يظهر الجزء الرقمي أزرق ويعطي خيارات النسخ/إضافة جهات الاتصال… كما في الصورة)
def _order_id_link_html(order_id: str, context: "ContextTypes.DEFAULT_TYPE" = None) -> str:
    oid = ("" if order_id is None else str(order_id)).strip()
    if not oid:
        return "—"
    # نُبقيه كنص فقط (بدون <a> وبدون <code>)
    # ونستخدم escape لأن أغلب الرسائل تعمل بـ parse_mode="HTML"
    return html.escape(oid, quote=False)
# ===== End Order ID display helper =====




def _money(v) -> str:
    try:
        s = "" if v is None else str(v).strip()   # ✅ بدل (v or "")
        s = re.sub(r"[^0-9.]+", "", s)
        if not s:
            return ""
        f = float(s)
        if f.is_integer():
            return f"{int(f):,} ر.س"
        return f"{f:,.2f} ر.س"
    except Exception:
        return str(v or "").strip()

def _pay_status_ar(x: object) -> str:
    """Helper موحد لترجمة حالات الدفع/الطلب إلى عربي (مستخدم داخل رسائل التاجر/الفواتير/الإشعارات)."""
    v = str(x or "").strip().lower()
    if not v:
        return "—"

    # حالات الطلب (Order Status)
    order_map = {
        "preparing": "جاري تجهيز الطلب",
        "prep": "جاري تجهيز الطلب",
        "ready_to_ship": "الطلب جاهز للشحن",
        "ready": "الطلب جاهز للشحن",
        "shipped": "تم شحن الطلب",
        "delivered": "تم تسليم الطلب",
        "closed": "مغلق",
        "cancelled": "ملغي",
        "canceled": "ملغي",
    }
    if v in order_map:
        return order_map[v]

    # حالات الدفع (Payment Status)
    pay_map = {
        "awaiting": "بانتظار الدفع",
        "awaiting_confirm": "بانتظار التأكيد",
        "pending": "قيد المعالجة",
        "processing": "قيد المعالجة",
        "confirmed": "مؤكد",
        "paid": "مدفوع",
        "rejected": "مرفوض",
        "failed": "فشل",
        "expired": "منتهي",
    }
    if v in pay_map:
        return pay_map[v]

    return str(x).strip()

def _order_status_display(o: dict) -> str:
    """
    عرض حالة موحد + يوضح جهة الإلغاء عند الإلغاء.
    - ملغي من قبل العميل / ملغي من قبل الإدارة / ملغي
    """
    try:
        ost = str((o or {}).get("order_status") or "").strip().lower()
    except Exception:
        ost = ""

    # افتراضي
    base = _pay_status_ar(ost or "—")

    # تخصيص الإلغاء
    if ost in ("cancelled", "canceled", "ملغي"):
        try:
            by_client = str((o or {}).get("cancelled_by_client_id") or "").strip()
        except Exception:
            by_client = ""
        try:
            by_admin = str((o or {}).get("cancelled_by_admin_id") or "").strip()
        except Exception:
            by_admin = ""

        if by_client:
            return "ملغي من قبل العميل"
        if by_admin:
            return "ملغي من قبل الإدارة"
        return "ملغي"

    return base or "—"
def _trader_label(uid: int, fallback_name: str = "") -> str:
    try:
        tp = get_trader_profile(int(uid or 0)) or {}
    except Exception:
        tp = {}
    dn = (tp.get("display_name") or "").strip()
    cn = (tp.get("company_name") or "").strip()
    if not dn:
        dn = (fallback_name or "").strip() or "التاجر"
    if cn:
        return f"{dn} ({cn})"
    return dn

def _trade_payment_block(tp: dict) -> str:
    bank = (tp.get("bank_name") or "").strip()
    iban = (tp.get("iban") or "").strip()
    stc = (tp.get("stc_pay") or "").strip()

    if not bank and not iban and not stc:
        return "غير مضافة بعد"

    parts = []
    if bank:
        parts.append(f"🏦 البنك: {bank}")
    if iban:
        parts.append(f"💳 IBAN: {iban}")
    if stc:
        parts.append(f"📱 STC Pay: {stc}")
    return "\n".join(parts)

async def ui_close_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    try:
        await q.message.delete()
    except Exception:
        # fallback: لو ما قدر يحذف، نخليه يعدّل
        try:
            await q.message.edit_text("✅ تم الإغلاق")
        except Exception:
            pass

async def ui_locked_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    # ✅ Alert فقط بدون حذف/تعديل أزرار
    try:
        await _alert(
            q,
            "✅ تم إرسال العرض مسبقًا.\nلإصدار عرض جديد استخدم زر (إنشاء عرض جديد).",
            force=True
        )
    except Exception:
        try:
            await q.answer(
                "✅ تم إرسال العرض مسبقًا.\nلإصدار عرض جديد استخدم زر (إنشاء عرض جديد).",
                show_alert=True
            )
        except Exception:
            pass

# Manual payment info (required)
PP_BANK_NAME = (os.getenv("PP_BANK_NAME") or "").strip()
PP_BENEFICIARY = (os.getenv("PP_BENEFICIARY") or "").strip()
PP_IBAN = (os.getenv("PP_IBAN") or "").strip()
PP_STC_PAY = (os.getenv("PP_STC_PAY") or "").strip()
# optional
PP_PAY_LINK_URL = (os.getenv("PP_PAY_LINK_URL") or "").strip()

PP_SUPPORT_LABEL = (os.getenv("PP_SUPPORT_LABEL") or "الإدارة").strip()
PP_TRADER_LABEL  = (os.getenv("PP_TRADER_LABEL")  or "التاجر").strip()

MAX_ITEMS = 30

# ===== منصة الدعم المباشر (أمر سلاش فقط) =====
# خمول: 10 دقائق / حد أقصى: 60 دقيقة
SUPPORT_IDLE_SECONDS = 10 * 60
SUPPORT_MAX_SECONDS  = 60 * 60
STAGE_SUPPORT_ADMIN_REPLY = "pp_support_admin_reply"
STAGE_CONFIRM_CLIENT_PREVIEW = "confirm_client_preview"

STAGE_ADMIN_TRADER_MSG = "pp_admin_trader_msg"

VIN_RE = re.compile(r"^[A-HJ-NPR-Z0-9]{17}$")  # 17 chars, excludes I O Q

logging.basicConfig(
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    level=logging.INFO
)
log = logging.getLogger("PP")


def _swallow(err: Exception | None = None, tag: str = "") -> None:
    """بديل آمن لـ except: pass — لوق مختصر بدون كسر تدفق البوت."""
    try:
        if tag:
            log.debug("SWALLOW|%s|%s", tag, err, exc_info=True)
        else:
            log.debug("SWALLOW|%s", err, exc_info=True)
    except Exception:
        pass


def log_event(event: str, **kwargs):
    """
    لوق عربي واضح في Render بدون لمس الاكسل
    """
    try:
        ts = datetime.now(timezone.utc).isoformat()
        # نختصر القيم الطويلة عشان ما يتفجر اللوق
        clean = {}
        for k, v in (kwargs or {}).items():
            try:
                s = str(v)
                if len(s) > 500:
                    s = s[:500] + "…"
                clean[k] = s
            except Exception:
                clean[k] = "?"
        log.info("🧾 [%s] %s | %s", ts, event, clean)
    except Exception as e:
        _swallow(e)

def utc_now_iso() -> str:
    # UTC aware دائمًا
    return datetime.now(timezone.utc).isoformat()

def _dt_utc_now() -> datetime:
    # UTC aware دائمًا
    return datetime.now(timezone.utc)

def _as_utc_aware(dt: datetime | None) -> datetime | None:
    """
    يحوّل أي datetime إلى UTC-aware (بدون تغيير الوقت إذا كان UTC أصلاً)
    """
    if not dt:
        return None
    try:
        if dt.tzinfo is None:
            return dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except Exception:
        try:
            return dt.replace(tzinfo=timezone.utc)
        except Exception:
            return None

def _parse_utc_iso(s: str) -> datetime | None:
    """
    يقبل:
    - 2026-02-04T19:55:53Z
    - 2026-02-04T19:55:53+00:00
    - 2026-02-04T19:55:53   (بدون timezone) => نعتبره UTC
    ويرجع datetime UTC aware دائمًا
    """
    try:
        v = (s or "").strip()
        if not v:
            return None
        if v.endswith("Z"):
            v = v[:-1] + "+00:00"
        dt = datetime.fromisoformat(v)
        return _as_utc_aware(dt)
    except Exception:
        return None

def set_stage(context: ContextTypes.DEFAULT_TYPE, user_id: int, stage: str):
    ud = context.user_data.setdefault(user_id, {})
    ud[ACTION_KEY] = ACTION_PAID_PARTS
    ud[STAGE_KEY] = stage

def get_ud(context: ContextTypes.DEFAULT_TYPE, user_id: int) -> dict:
    return context.user_data.setdefault(user_id, {})

def reset_flow(context: ContextTypes.DEFAULT_TYPE, user_id: int):
    context.user_data.setdefault(user_id, {}).clear()

def _support_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[InlineKeyboardButton("🔒 إنهاء", callback_data="pp_support_close")]])

def _support_is_open(ud: dict) -> bool:
    return bool(ud.get("support_open"))

def _support_touch(ud: dict):
    now = utc_now_iso()
    ud["support_last_at_utc"] = now
    ud.setdefault("support_started_at_utc", now)

def _support_should_close_by_time(ud: dict) -> bool:
    """True if support chat should auto-close (idle or max duration)."""
    if not _support_is_open(ud):
        return False
    try:
        now = datetime.now(timezone.utc)
        started = datetime.fromisoformat((ud.get("support_started_at_utc") or "").replace("Z", "+00:00"))
        last = datetime.fromisoformat((ud.get("support_last_at_utc") or ud.get("support_started_at_utc") or "").replace("Z", "+00:00"))
        if (now - last).total_seconds() >= SUPPORT_IDLE_SECONDS:
            return True
        if (now - started).total_seconds() >= SUPPORT_MAX_SECONDS:
            return True
        return False
    except Exception:
        # إذا تعذر قراءة التوقيتات نقفل احتياطياً
        return True

async def _support_close(update_or_q, context: ContextTypes.DEFAULT_TYPE, user_id: int, reason: str = ""):
    ud = get_ud(context, user_id)
    ud.pop("support_open", None)
    ud.pop("support_started_at_utc", None)
    ud.pop("support_last_at_utc", None)
    try:
        txt = "✅ تم إغلاق قناة التواصل مع الإدارة"
        if reason:
            txt += f"\n{reason}"
        if hasattr(update_or_q, "callback_query") and update_or_q.callback_query:
            q = update_or_q.callback_query
            try:
                await q.answer("تم الإغلاق")
            except Exception as e:
                _swallow(e)
            try:
                await q.message.reply_text(txt)
            except Exception as e:
                _swallow(e)
        elif hasattr(update_or_q, "message") and update_or_q.message:
            await update_or_q.message.reply_text(txt)
        else:
            await context.bot.send_message(chat_id=user_id, text=txt)
    except Exception as e:
        _swallow(e)

def price_for_count(c: int) -> int:
    """رسوم المنصة حسب عدد القطع (غير الاستهلاكية)."""

    # ✅ عرض مجاني للمنصة: رسوم المنصة = 0
    if _is_platform_fee_free_mode():
        return 0

    if c <= 0:
        return 0

    return 25 if c <= 5 else 39

# ===== مساعدات تنسيق رسائل الإدارة (بدون تشوه بصري) =====
STAGE_ADMIN_SEND_PAYLINK = "admin_send_paylink"

def _trim_caption(s: str, limit: int = 950) -> str:
    s = (s or "").strip()
    if len(s) <= limit:
        return s
    return s[: max(0, limit-1)].rstrip() + "…"

def _build_admin_order_caption(order_id: str, ud: dict, order: dict, title: str, extra_lines=None) -> str:
    extra_lines = extra_lines or []
    user_name = (ud.get("user_name") or order.get("user_name") or "").strip()
    car = (ud.get("car_name") or order.get("car_name") or "").strip()
    model = (ud.get("car_model") or order.get("car_model") or "").strip()
    vin = (ud.get("vin") or order.get("vin") or "").strip()
    fee = ud.get("price_sar", order.get("price_sar", ""))
    ship_method = (ud.get("ship_method") or order.get("ship_method") or "").strip()
    delivery_details = (ud.get("delivery_details") or order.get("delivery_details") or "").strip()

    # ✅ إضافة: قراءة ملاحظات العميل (fallback: ud -> order)
    notes = (ud.get("notes") or order.get("notes") or "").strip()

    parts = []
    try:
        b = get_order_bundle(order_id)
        items = b.get("items", []) or []
        for i, it in enumerate(items, start=1):
            nm = (it.get("name") or "").strip()
            pn = (it.get("part_no") or it.get("item_part_no") or "").strip()
            if not nm:
                continue
            parts.append(f"{i}- {nm}" + (f" ({pn})" if pn else ""))
            if len(parts) >= 6:
                break
    except Exception:
        parts = []
    parts_txt = "\n".join(parts) if parts else "—"

    lines = [title, f"🧾 رقم الطلب: {_order_id_link_html(order_id)}"]
    if user_name:
        lines.append(f"👤 العميل: {user_name}")
    if car or model:
        lines.append(f"🚗 السيارة: {(car + ' ' + model).strip()}")
    if vin:
        lines.append(f"🔎 VIN: {vin}")
    if str(fee).strip() not in ("", "0", "0.0"):
        lines.append(f"💰 رسوم المنصة: {fee} ريال")

    # ✅ إضافة: إظهار الملاحظات في رسالة المجموعة
    if notes:
        lines += ["", "📝 ملاحظات العميل:", notes]

    lines.extend(extra_lines)
    lines += ["", "🧩 القطع:", parts_txt]

    if ship_method or delivery_details:
        lines += ["", "📦 طريقة التسليم:"]
        if ship_method:
            lines.append(ship_method)
        if delivery_details:
            lines += ["", "📍 تفاصيل التسليم:", delivery_details]

    return _trim_caption("\n".join(lines))

# ✅ MUST be defined BEFORE _is_consumable_part()
_CONSUMABLE_KEYWORDS = [
    # Arabic
    "زيت", "زيوت", "فلتر", "فلاتر", "سيفون",
    "بوجي", "بواجي", "شمعة اشعال", "شمعات اشعال",
    "سير", "سيور",
    "سائل", "سوائل",
    "فحمات", "فحمات اشعال", "فحمات إشعال",
    "صرة", "صره", "صوفة", "صوفه", "جاسكيت",
    # English
    "oil", "filter", "filters", "spark plug", "spark plugs", "plug", "plugs",
    "belt", "belts",
    "fluid", "fluids", "coolant",
    "gasket",
    "brake pad", "brake pads", "pads",
]

def _is_consumable_part(name: str) -> bool:
    s = (name or "").strip().lower()
    if not s:
        return False
    # توحيد بسيط
    s = re.sub(r"\s+", " ", s)
    # بحث احتوائي (يشمل مفرد/جمع وتنوعات بسيطة)
    return any(k in s for k in _CONSUMABLE_KEYWORDS)

def _platform_fee_for_items(items: list[dict]) -> tuple[int, int, int]:
    """Returns (fee_sar, non_consumable_count, consumable_count)."""
    if not items:
        return 0, 0, 0
    c_cons = 0
    c_non = 0
    for it in items:
        nm = (it.get("name") or "").strip()
        if _is_consumable_part(nm):
            c_cons += 1
        else:
            c_non += 1
    fee = 0 if (c_non == 0 and c_cons > 0) else price_for_count(c_non)
    return fee, c_non, c_cons

def main_menu_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ بدء طلب جديد", callback_data="pp_start_new")]
    ])

def more_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ انهاء وارسال الطلب", callback_data="pp_more_no")],
        [InlineKeyboardButton("✖️ الغاء الطلب", callback_data="pp_cancel")],
    ])

def photo_prompt_kb():
    # زر انهاء يظهر دائما حتى لو العميل ما رفع صورة
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ انهاء وارسال الطلب", callback_data="pp_more_no")],
        [InlineKeyboardButton("✖️ الغاء الطلب", callback_data="pp_cancel")],
    ])

def partno_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("⏭️ تخطي", callback_data="pp_partno_skip")],
        [InlineKeyboardButton("✖️ الغاء الطلب", callback_data="pp_cancel")],
    ])

def prepay_notes_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("⏭️ تخطي والانتقال للشحن", callback_data="pp_prepay_notes_skip")],
        [InlineKeyboardButton("✖️ الغاء الطلب", callback_data="pp_cancel")],
    ])

def build_order_preview(ud: dict) -> str:
    # مهم: لازم يكون عندك import html أعلى الملف
    # import html

    def esc(x) -> str:
        # يمنع اختفاء الرقم 0
        return html.escape("" if x is None else str(x), quote=False)

    order_id = esc(ud.get("order_id", ""))
    uname = esc(ud.get("user_name", ""))
    car = esc(ud.get("car_name", ""))
    model = esc(ud.get("car_model", ""))
    vin = esc(ud.get("vin", ""))
    notes_raw = _norm(ud.get("notes", ""))
    notes = esc(notes_raw)

    items = ud.get("items", []) or []

    # رسوم المنصة (تظهر 0 دائمًا لو ما فيه رسوم)
    price = ud.get("price_sar", 0)
    if price is None or str(price).strip() == "":
        price = 0

    lines = []

    lines.append(f"🧾 <b>معاينة الطلب</b> <i>#{_order_id_link_html(order_id)}</i>")
    lines.append(f"👤 <b>العميل</b>: <i>{uname}</i>")
    lines.append(f"🚗 <b>السيارة</b>: <i>{car}</i>")
    lines.append(f"📌 <b>الموديل</b>: <i>{model}</i>")
    lines.append(f"🔎 <b>VIN</b>: <i>{vin}</i>")
    lines.append(f"📝 <b>الملاحظات</b>: <i>{notes if notes else 'لا يوجد'}</i>")
    lines.append("")

    lines.append(f"🧩 <b>القطع المطلوبة</b> <i>({len(items)})</i>:")
    for i, it in enumerate(items, start=1):
        nm = esc((it.get("name") or "").strip())
        pn = esc((it.get("part_no") or "").strip())
        if nm:
            if pn:
                lines.append(f"  🔹 <b>{i}</b>- <i>{nm}</i> <b>رقم</b>: <code>{pn}</code>")
            else:
                lines.append(f"  🔹 <b>{i}</b>- <i>{nm}</i>")

    lines.append("")
    lines.append(f"💰 <b>رسوم المنصة</b>: <i>{esc(price)} ريال</i>")

    if str(price) == "0":
        lines.append("✅ <i>لا توجد رسوم منصة على هذا الطلب</i>")

    return "\n".join(lines)

def pay_method_kb():
    rows = [
        [InlineKeyboardButton("🏦 تحويل بنكي", callback_data="pp_pay_bank")],
        [InlineKeyboardButton("📱 STC Pay", callback_data="pp_pay_stc")],
        [InlineKeyboardButton("🔗 رابط دفع سريع", callback_data="pp_pay_link")],
        [InlineKeyboardButton("✖️  الغاء الطلب", callback_data="pp_cancel")],
    ]
    return InlineKeyboardMarkup(rows)

# === Structured Quote Engine (Trader Private Wizard) ===

def trader_quote_start_kb(order_id: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(_wide_btn_label("🚀 بدء بناء عرض السعر"), callback_data=f"ppq_begin|{order_id}")],
    ])

# --- UI helper: make single-column buttons look consistently wide ---
_WIDE_FILL = "\u2800"  # braille blank (renders as a visible width placeholder)

def _wide_btn_label(s: str, target: int = 22) -> str:
    s = "" if s is None else str(s)
    # Pad with braille blanks to make rows feel equally wide in Telegram.
    pad = max(0, int(target) - len(s))
    return s + (_WIDE_FILL * pad)

def trader_quote_type_kb(order_id: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(_wide_btn_label("✔️ وكالة"), callback_data=f"ppq_type|{order_id}|agency")],
        [InlineKeyboardButton(_wide_btn_label("✔️ وكلاء محليين"), callback_data=f"ppq_type|{order_id}|local_dealers")],
        [InlineKeyboardButton(_wide_btn_label("✔️ تجاري"), callback_data=f"ppq_type|{order_id}|aftermarket")],
        [InlineKeyboardButton(_wide_btn_label("✔️ مختلط"), callback_data=f"ppq_type|{order_id}|mixed")],
    ])

def trader_quote_shipping_method_kb(order_id: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(_wide_btn_label("🚚 محلي"), callback_data=f"ppq_ship|{order_id}|local")],
        [InlineKeyboardButton(_wide_btn_label("✈️ دولي"), callback_data=f"ppq_ship|{order_id}|intl")],
    ])

def trader_quote_shipping_included_kb(order_id: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(_wide_btn_label("✅ السعر يشمل الشحن"), callback_data=f"ppq_shipinc|{order_id}|yes")],
        [InlineKeyboardButton(_wide_btn_label("❌ الشحن غير مشمول"), callback_data=f"ppq_shipinc|{order_id}|no")],
    ])

def trader_quote_eta_kb(order_id: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(_wide_btn_label("⏱ 1-2 يوم"), callback_data=f"ppq_eta|{order_id}|1-2")],
        [InlineKeyboardButton(_wide_btn_label("⏱ 3-5 ايام"), callback_data=f"ppq_eta|{order_id}|3-5")],
        [InlineKeyboardButton(_wide_btn_label("⏱ 7-14 يوم"), callback_data=f"ppq_eta|{order_id}|7-14")],
        [InlineKeyboardButton(_wide_btn_label("✍️ مدة اخرى"), callback_data=f"ppq_eta|{order_id}|custom")],
    ])

def trader_quote_availability_kb(order_id: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(_wide_btn_label("⏱ 1-2 يوم"), callback_data=f"ppq_avail|{order_id}|1-2")],
        [InlineKeyboardButton(_wide_btn_label("⏱ 3-5 ايام"), callback_data=f"ppq_avail|{order_id}|3-5")],
        [InlineKeyboardButton(_wide_btn_label("⏱ 7-14 يوم"), callback_data=f"ppq_avail|{order_id}|7-14")],
        [InlineKeyboardButton(_wide_btn_label("✍️ مدة اخرى"), callback_data=f"ppq_avail|{order_id}|custom")],
    ])

def _ppq_type_label(v: str) -> str:
    return {
        "agency": "وكالة",
        "local_dealers": "وكلاء محليين",
        "aftermarket": "تجاري",
        "mixed": "مختلط",
    }.get(v, "غير محدد")

def trader_quote_preview_kb(order_id: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(_wide_btn_label("✅ إرسال العرض للعميل"), callback_data=f"ppq_preview_send|{order_id}")],
        [InlineKeyboardButton(_wide_btn_label("🔁 إعادة بناء العرض"), callback_data=f"ppq_preview_restart|{order_id}")],
        [InlineKeyboardButton(_wide_btn_label("✖️ إلغاء"), callback_data=f"ppq_preview_cancel|{order_id}")],
    ])

def trader_quote_preview_kb_locked(order_id: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(_wide_btn_label("✅ تم إرسال العرض للعميل"), callback_data="pp_ui_locked")],
        [InlineKeyboardButton(_wide_btn_label("🆕 إنشاء عرض جديد (إصدار جديد)"), callback_data=f"ppq_new_version|{order_id}")],
        [InlineKeyboardButton(_wide_btn_label("✖️ إغلاق"), callback_data="pp_ui_close")],
    ])

def _ppq_ship_label(v: str) -> str:
    return {"local": "محلي", "intl": "دولي"}.get(v, "غير محدد")

def build_legal_shipping_block(method: str, fee_sar: str, eta: str, included: str) -> str:
    # صيغة موحدة قانونيا يعاد استخدامها (بدون اسم شركة الشحن)

    included_norm = str(included or "").strip().lower()
    inc_txt = "مشمولة" if included_norm in ("yes", "true", "1", "included", "مشمولة", "مشمول") else "غير مشمولة"

    fee_txt = str(fee_sar or "").strip()

    # القاعدة الجديدة:
    # - مشمول => 0
    # - غير مشمول + لا قيمة => يحددها التاجر
    if not fee_txt:
        fee_txt = "0" if inc_txt == "مشمولة" else "يحددها التاجر"

    return (
        "🚚 الشحن:\n"
        f"طريقة الشحن: {_ppq_ship_label(method)}\n"
        f"مدة الشحن: {eta}\n"
        f"تكلفة الشحن: {inc_txt}\n"
        f"قيمة الشحن: {fee_txt}"
    )

def build_official_quote_text(
    order_id: str,
    client_name: str,
    goods_amount_sar: str,
    parts_type: str,
    ship_block: str,
    availability: str,
    shipping_fee_sar: str = "",
    ship_included: bool = False,
) -> str:
    """
    يبني نص عرض سعر رسمي (نص عادي بدون HTML):
    - قيمة القطع
    - قيمة الشحن (أو مشمول)
    - إجمالي شامل (القطع + الشحن) بسطر واحد واضح
    ❌ بدون أي ذكر لرسوم المنصة
    ✅ بدون تكرار قيمة الشحن داخل ship_block
    """

    def _s(x: object) -> str:
        return ("" if x is None else str(x)).strip()

    def _to_float(x: object) -> float:
        try:
            return float(str(x).replace(",", "").strip())
        except Exception:
            return 0.0

    def _money(x: object) -> str:
        try:
            v = _to_float(x)
            if abs(v - int(v)) < 1e-9:
                return f"{int(v)} ريال"
            return f"{v:.2f} ريال"
        except Exception:
            return "—"

    def _strip_parens(s: str) -> str:
        return (s or "").replace("(", "").replace(")", "").replace("（", "").replace("）", "").strip()

    def _clean_ship_block(block: str) -> str:
        """
        يمنع تكرار قيمة الشحن:
        نحذف أي سطر داخل ship_block يذكر (قيمة/رسوم/تكلفة الشحن) لأننا سنعرضها مرة واحدة فقط في قسم المبالغ.
        """
        b = _s(block)
        if not b:
            return ""
        drop_keys = (
            "قيمة الشحن",
            "رسوم الشحن",
            "تكلفة الشحن",
            "مبلغ الشحن",
            "سعر الشحن",
            "shipping fee",
        )
        lines = []
        for ln in b.splitlines():
            lns = ln.strip()
            if not lns:
                continue
            low = lns.lower()
            if any(k in lns for k in drop_keys) or any(k in low for k in drop_keys):
                continue
            lines.append(ln)
        return "\n".join(lines).strip()

    order_id = _strip_parens(order_id)
    client_name = _strip_parens(client_name)
    availability = _strip_parens(availability)

    goods_val = _to_float(goods_amount_sar)
    goods_txt = _money(goods_val)

    if ship_included:
        shipping_txt = "مشمول (ضمن الإجمالي)"
        shipping_val = 0.0
    else:
        shipping_val = _to_float(shipping_fee_sar)
        shipping_txt = _money(shipping_val) if _s(shipping_fee_sar) else "غير محدد"

    total_val = goods_val + shipping_val
    total_txt = _money(total_val) if (ship_included or _s(shipping_fee_sar)) else "غير محدد"

    ship_block_clean = _clean_ship_block(ship_block)

    head = (
        "💰 عرض سعر رسمي\n"
        f"👤 العميل: {client_name}\n"
        f"🧾 رقم الطلب: {_order_id_link_html(order_id)}\n\n"
        "📦 تفاصيل المبالغ\n"
        f"🧩 قيمة القطع: {goods_txt}\n"
        f"🚚 قيمة الشحن: {shipping_txt}\n"
        f"✅ الإجمالي شامل الشحن: {total_txt}\n\n"
        "🔧 نوع القطع:\n"
        f"✔️ {_ppq_type_label(parts_type)}\n"
    )

    if ship_block_clean:
        head += "\n" + ship_block_clean + "\n"

    head += (
        "\n"
        f"⏳ مدة التجهيز: {availability}\n\n"
        "يرجى مراجعة العرض ثم اختيار القرار من الأزرار بالأسفل.\n"
        "في حال قبول العرض سيتم فتح قناة تواصل مباشرة بين التاجر والعميل."
    )
    return head


def quote_client_kb(order_id: str, trader_id: int) -> InlineKeyboardMarkup:
    tid = int(trader_id or 0)
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton(
                "✅ أقبل العرض",
                callback_data=f"pp_quote_ok|{order_id}|{tid}"
            ),
        ],
        [
            InlineKeyboardButton(
                "❌ أرفض العرض",
                callback_data=f"pp_quote_no|{order_id}|{tid}"
            ),
        ],
    ])

def trader_status_kb(order_id: str) -> InlineKeyboardMarkup:
    """
    ✅ كيبورد ديناميكي حسب السيناريو الجديد (بدون Alerts وبدون رسائل تحذير إضافية)
    🔒 مهم: ممنوع إظهار أي زر (مراسلة/اتصال) للعميل للتاجر من هنا.
    يظهر للتاجر زر المراسلة/الاتصال فقط عبر إشعار إيصال الدفع (trader_goods_receipt_kb)
    أو إشعار الاستلام (trader_received_notice_kb) حسب التدفق.
    """
    try:
        ob = get_order_bundle(order_id) or {}
        order = (ob.get("order") or {}) if isinstance(ob, dict) else {}
        ost = str(order.get("order_status") or "").strip().lower()
        inv_file = (str(order.get("seller_invoice_file_id") or order.get("shop_invoice_file_id") or "")).strip()
        pay_method = str(order.get("goods_payment_method") or "").strip().lower()
        pay_link = (str(order.get("goods_payment_link") or "")).strip()
        if (not inv_file) and pay_link and pay_method in ("pay_link", "link", "payment_link"):
            inv_file = "__PAYLINK__"
        gps = str(order.get("goods_payment_status") or "").strip().lower()
    except Exception:
        ost = ""
        inv_file = ""
        gps = ""

    goods_paid = gps in ("confirmed", "paid", "success", "successful", "done", "ok")

    # 1) بداية بعد الموافقة
    if ost in ("", "new", "accepted", "quoted"):
        return InlineKeyboardMarkup([
            [InlineKeyboardButton("🟡 جاري تجهيز الطلب", callback_data=f"pp_trader_status|prep|{order_id}")],
        ])

    # 2) بعد تجهيز
    if ost in ("preparing", "prep"):
        return InlineKeyboardMarkup([
            [InlineKeyboardButton("🟢 الطلب جاهز للشحن", callback_data=f"pp_trader_status|ready|{order_id}")],
        ])

    # 3) جاهز للشحن
    if ost in ("ready_to_ship", "ready"):
        # قبل رفع فاتورة المتجر: يبقى نفس الزر فقط
        if not inv_file:
            return InlineKeyboardMarkup([
                [InlineKeyboardButton("🟢 الطلب جاهز للشحن", callback_data=f"pp_trader_status|ready|{order_id}")],
            ])

        # بعد رفع الفاتورة وقبل الدفع: لا نعرض أي اتصال/شات للتاجر
        if not goods_paid:
            return InlineKeyboardMarkup([
                [InlineKeyboardButton("⏳ بانتظار سداد العميل لقيمة البضاعة", callback_data="pp_ui_close")],
            ])

        # بعد الدفع: يسمح بالشحن (بدون زر مراسلة هنا)
        return InlineKeyboardMarkup([
            [InlineKeyboardButton("🚚 تم شحن الطلب", callback_data=f"pp_trader_status|shipped|{order_id}")],
        ])

    # 4) بعد الشحن
    if ost in ("shipped",):
        return InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ تأكيد التسليم بنجاح", callback_data=f"pp_trader_status|delivered|{order_id}")],
            [InlineKeyboardButton("🔒 إنهاء / إقفال الطلب (منجز)", callback_data=f"pp_order_finish|{order_id}")],
        ])

    # 5) بعد التسليم / الإغلاق
    if ost in ("delivered", "closed"):
        return InlineKeyboardMarkup([
            [InlineKeyboardButton("🔒 إنهاء / إقفال الطلب (منجز)", callback_data=f"pp_order_finish|{order_id}")],
        ])

    # 🛡️ fallback ذكي — حالة غير متوقعة
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📌 مراجعة حالة الطلب", callback_data=f"pp_order_review|{order_id}")],
    ])

def trader_received_notice_kb(order_id: str) -> InlineKeyboardMarkup:
    """
    🔒 هذا الإشعار يصل للتاجر بعد الاستلام/التسليم حسب تدفقك.
    ✅ هنا مسموح إظهار مراسلة العميل (لأنه ليس قبل الدفع).
    """
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("💬 مراسلة العميل داخل المنصة", callback_data=f"pp_chat_open|{order_id}")],
        [InlineKeyboardButton("✅ إشعار الإدارة: تم التسليم بنجاح", callback_data=f"pp_trader_status|delivered|{order_id}")],
        [InlineKeyboardButton("🔒 إنهاء / إقفال الطلب (منجز)", callback_data=f"pp_order_finish|{order_id}")],
    ])


def pay_goods_method_kb(order_id: str, pay_mode: str = "manual", has_link: bool = False) -> InlineKeyboardMarkup:
    """
    ✅ للعميل: طرق دفع قيمة القطع.
    - في وضع manual: نعرض (تحويل بنكي + STC) (وممكن رابط إذا متوفر).
    - في وضع link: لا نعرض مسارات التحويل إطلاقاً (يبقى فقط مراسلة التاجر).
    """
    pm = (str(pay_mode or "").strip().lower() or "manual")

    rows = []
    if pm in ("link", "pay_link", "payment_link"):
        rows = [
            [InlineKeyboardButton("💬 مراسلة التاجر", callback_data=f"pp_chat_trader|{order_id}")],
        ]
        return InlineKeyboardMarkup(rows)

    # manual (افتراضي)
    rows.append([InlineKeyboardButton("🏦 تحويل بنكي", callback_data=f"pp_goods_pay_bank|{order_id}")])
    rows.append([InlineKeyboardButton("📱 STC Pay", callback_data=f"pp_goods_pay_stc|{order_id}")])
    if has_link:
        rows.append([InlineKeyboardButton("🔗 رابط الدفع", callback_data=f"pp_goods_pay_link|{order_id}")])
    rows.append([InlineKeyboardButton("💬 مراسلة التاجر", callback_data=f"pp_chat_trader|{order_id}")])
    return InlineKeyboardMarkup(rows)


def team_goods_confirm_kb(order_id: str):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ تأكيد استلام قيمة القطع", callback_data=f"pp_team_goods_confirm|{order_id}")]
    ])


def trader_goods_receipt_kb(order_id: str, user_id: int) -> InlineKeyboardMarkup:
    """
    ✅ للتاجر: هذه هي النقطة الأساسية التي يُسمح فيها بإظهار مراسلة العميل (بعد الدفع)
    🔒 مهم: لا نعرض رقم/اتصال هنا إلا إذا كان عندك زر اتصال منفصل ترغب به
    (حالياً نكتفي بالمراسلة داخل المنصة حسب شرطك).
    """
    uid = int(user_id or 0)
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ تأكيد استلام قيمة القطع", callback_data=f"pp_team_goods_confirm|{order_id}")],
        [InlineKeyboardButton("💬 مراسلة العميل داخل المنصة", callback_data=f"pp_chat_open|{order_id}")],
        [InlineKeyboardButton("🔒 إنهاء / إقفال الطلب (منجز)", callback_data=f"pp_order_finish|{order_id}")],
        [InlineKeyboardButton("✖️ إغلاق", callback_data="pp_ui_close")],
    ])

def admin_free_order_kb(order_id: str, client_id: int) -> InlineKeyboardMarkup:
    oid = (order_id or "").strip()
    uid = int(client_id or 0)

    rows = []
    if oid and uid:
        rows.append([InlineKeyboardButton("💬 مراسلة العميل", callback_data=f"pp_admin_reply|{oid}|{uid}")])

    if oid:
        rows.append([InlineKeyboardButton("⛔ إلغاء الطلب", callback_data=f"pp_admin_cancel|{oid}")])
        rows.append([InlineKeyboardButton("🔁 إعادة نشر الطلب", callback_data=f"pp_admin_republish|{oid}")])

    rows.append([InlineKeyboardButton("✖️ إغلاق", callback_data="pp_ui_close")])
    return InlineKeyboardMarkup(rows)


def bank_info_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📋 نسخ IBAN", callback_data="pp_copy_iban")],
        [InlineKeyboardButton("❌  الغاء الطلب", callback_data="pp_cancel")],
    ])

def stc_info_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📋 نسخ رقم STC Pay", callback_data="pp_copy_stc")],
        [InlineKeyboardButton("❌  الغاء الطلب", callback_data="pp_cancel")],
    ])

def delivery_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🚚 شحن", callback_data="pp_delivery_ship")],
        [InlineKeyboardButton("📍 استلام من الموقع", callback_data="pp_delivery_pickup")],
    ])



def cancel_only_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✖️ الغاء الطلب", callback_data="pp_cancel")],
    ])

def _flow_nav_kb(back_to: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("↩️ رجوع", callback_data=f"pp_back|{back_to}")],
        [InlineKeyboardButton("✖️ إلغاء", callback_data="pp_cancel")],
    ])

def track_kb(order_id: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🔎 مراسلة المنصة", callback_data=f"pp_track|{order_id}")],
        [InlineKeyboardButton("⛔ ايقاف الطلب لم اعد بحاجة له", callback_data=f"pp_rb_stop|{order_id}")],
    ])

def admin_reply_kb(order_id: str, user_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✍️ رد كالإدارة", callback_data=f"pp_admin_reply|{order_id}|{user_id}")],
    ])

def admin_reply_done_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ انهاء الرد", callback_data="pp_admin_reply_done")],
    ])

from io import BytesIO
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfgen import canvas
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False

def client_preview_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ تأكيد ومتابعة", callback_data="pp_client_confirm_preview")],
        [InlineKeyboardButton("↩️ رجوع لتعديل العنوان", callback_data="pp_back|delivery")],
        [InlineKeyboardButton("✖️ إلغاء", callback_data="pp_cancel")],
    ])


async def client_confirm_preview_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    user_id = q.from_user.id
    ud = get_ud(context, user_id)

    # لازم يكون في مرحلة المعاينة
    if ud.get(STAGE_KEY) != STAGE_CONFIRM_CLIENT_PREVIEW:
        return

    mode = str(ud.get("preview_mode") or "").strip()   # "free_ship" / "free_pickup" / "pay_ship" / "pay_pickup"
    details = str(ud.get("preview_details") or "").strip()
    order_id = str(ud.get("order_id") or "").strip()

    if not order_id:
        reset_flow(context, user_id)
        try:
            await q.message.reply_text("تعذر تحديد الطلب، ابدأ من جديد بكتابة pp")
        except Exception as e:
            _swallow(e)
        return

    # امسح أزرار المعاينة
    try:
        await q.message.edit_reply_markup(reply_markup=None)
    except Exception as e:
        _swallow(e)

    # ✅ لو مدفوع: افتح شاشة اختيار الدفع
    if mode.startswith("pay_"):
        set_stage(context, user_id, STAGE_AWAIT_PAY_METHOD)
        try:
            safe_details = html.escape(details)
            await q.message.reply_text(
                build_order_preview(ud)
                + "\n\n<b>📍 تفاصيل العنوان</b>:\n<pre>"
                + safe_details
                + "</pre>\n\n"
                "<b>⬇️ اختر طريقة دفع رسوم المنصة</b>",
                parse_mode="HTML",
                reply_markup=pay_method_kb(),
                disable_web_page_preview=True,
            )
        except Exception:
            await q.message.reply_text("⬇️ اختر طريقة دفع رسوم المنصة", reply_markup=pay_method_kb())
        return

    # ✅ لو مجاني: نفّذ الإرسال للمجموعة الآن (نفس منطقك السابق)
    # (ننسخ نفس بلوكات “free mode” الموجودة عند ship/pickup ونحطها هنا)
    try:
        _save_order_once(ud)
    except Exception as e:
        _swallow(e)

    try:
        update_order_fields(order_id, {
            "price_sar": 0,
            "payment_method": "free",
            "payment_status": "confirmed",
            "payment_confirmed_at_utc": utc_now_iso(),
        })
    except Exception as e:
        _swallow(e)

    try:
        await send_platform_invoice_pdf(context, order_id, kind="preliminary", admin_only=False)
    except Exception as e:
        _swallow(e)

    try:
        await notify_team(context, ud)
    except Exception as e:
        _swallow(e)

    try:
        await notify_admins_free_order(context, ud, client_id=user_id)
    except Exception as e:
        _swallow(e)

    try:
        safe_details = html.escape(details)
        await q.message.reply_text(
            build_order_preview(ud)
            + "\n\n<b>📍 تفاصيل العنوان</b>:\n<pre>"
            + safe_details
            + "</pre>\n"
            "<b>✅ تم استلام طلبك وستصلك العروض قريباً</b>",
            parse_mode="HTML",
            reply_markup=track_kb(order_id),
            disable_web_page_preview=True,
        )
    except Exception as e:
        _swallow(e)

    set_stage(context, user_id, STAGE_DONE)
    return


async def send_trader_ledger_pdf(
    context: ContextTypes.DEFAULT_TYPE,
    trader_id: int,
    admin_chat_id: int,
):
    """
    ✅ PDF "سجل التاجر" (للأدمن فقط)
    - نفس ستايل الفواتير (Header/Badges/Sections/Tables) لكن ثيم برتقالي
    - بدون ختم (مدفوع) لأنه ليس فاتورة
    """
    # ✅ tempfile
    try:
        import tempfile
    except Exception as e:
        try:
            await context.bot.send_message(chat_id=admin_chat_id, text=f"⚠️ تعذر إنشاء PDF (tempfile): {e}")
        except Exception:
            pass
        return

    import os, re, uuid
    from datetime import datetime, timezone, timedelta

    tid = int(trader_id or 0)
    if tid <= 0:
        return

    # --- Arabic RTL + shaping (keep tags, shape text parts) ---
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display

        _tag_re = re.compile(r"(<[^>]+>)")

        def _ar(s: str) -> str:
            s = "" if s is None else str(s)
            if not s:
                return s
            try:
                parts = _tag_re.split(s)
                out = []
                for part in parts:
                    if not part:
                        continue
                    if part.startswith("<") and part.endswith(">"):
                        out.append(part)
                    else:
                        try:
                            out.append(get_display(arabic_reshaper.reshape(part)))
                        except Exception:
                            out.append(part)
                return "".join(out)
            except Exception:
                try:
                    return get_display(arabic_reshaper.reshape(s))
                except Exception:
                    return s

    except Exception:

        def _ar(s: str) -> str:
            return "" if s is None else str(s)

    # ✅ reportlab imports
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph, Image as RLImage
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except Exception as e:
        try:
            await context.bot.send_message(chat_id=admin_chat_id, text=f"⚠️ تعذر إنشاء PDF (reportlab): {e}")
        except Exception:
            pass
        return

    # ---------------- Data ----------------
    try:
        prof = get_trader_profile(tid) or {}
    except Exception:
        prof = {}

    def _pick(*keys: str) -> str:
        for k in keys:
            try:
                v = (prof.get(k) or "").strip()
                if v:
                    return v
            except Exception:
                continue
        return ""

    tname = _pick("display_name", "name", "full_name")
    tcompany = _pick("company_name", "shop_name", "store_name")

    # username
    tuser = _pick(
        "username",
        "tg_username",
        "user_name",
        "telegram_username",
        "telegram_user",
        "telegram",
        "tg_user",
        "telegramUser",
        "user_username",
    )
    tuser = (tuser or "").strip()
    if tuser and not tuser.startswith("@"):
        tuser = "@" + tuser

    # ✅ fallback: fetch username from Telegram if not stored
    if not tuser:
        try:
            ch = await context.bot.get_chat(int(tid))
            u = getattr(ch, "username", None) or ""
            u = str(u or "").strip()
            if u:
                tuser = u if u.startswith("@") else ("@" + u)
        except Exception:
            pass

    shop_phone = _pick("shop_phone", "phone", "mobile", "shop_mobile", "store_phone")
    cr_no = _pick("cr_no", "cr", "cr_number", "commercial_register", "commercial_registration")
    vat_no = _pick("vat_no", "vat", "vat_number", "tax_no", "tax_number")
    bank = _pick("bank_name", "bank")
    iban = _pick("iban")
    stc = _pick("stc_pay", "stc", "stcpay")
    joined = _pick("joined_at_utc", "joined_at", "created_at_utc", "created_at")
    upd = _pick("updated_at_utc", "updated_at")

    try:
        enabled = bool(is_trader_enabled(tid))
    except Exception:
        enabled = True
    enabled_txt = "مفعل" if enabled else "موقوف"

    month = month_key_utc()
    sub_status = "متأخر"
    try:
        subs = list_trader_subscriptions(month) or []
        for s in subs:
            try:
                if int(s.get("trader_id") or 0) != tid:
                    continue
            except Exception:
                continue
            stv = str(s.get("payment_status") or "").strip().lower()
            if stv == "confirmed":
                sub_status = "مدفوع"
            elif stv in ("pending", "awaiting"):
                sub_status = "قيد التحقق"
            else:
                sub_status = "متأخر"
            break
    except Exception:
        pass

    try:
        orders = list_orders_for_trader(tid) or []
    except Exception:
        orders = []

    def _parse_dt(s: str) -> datetime:
        v = (s or "").strip()
        if not v:
            return datetime.min.replace(tzinfo=timezone.utc)
        try:
            return datetime.fromisoformat(v.replace("Z", "+00:00"))
        except Exception:
            return datetime.min.replace(tzinfo=timezone.utc)

    def _num(x: object) -> float:
        try:
            s = str(x or "").strip()
            s = re.sub(r"[^0-9.]+", "", s)
            return float(s or 0)
        except Exception:
            return 0.0

    def _money_tail(x: object, fb: str = "0") -> str:
        try:
            s = _money(x)
        except Exception:
            s = ""
        s = (s or "").strip()
        s = re.sub(r"^\s*(ر\.?\s*س|ر\.س|SAR|SR|s\.r|s\.r\.?)\s*", "", s, flags=re.I)
        s = re.sub(r"\s*(ر\.?\s*س|ر\.س|SAR|SR|s\.r|s\.r\.?)\s*$", "", s, flags=re.I)
        s = s.strip() or fb
        return f"{s} ﷼"

    # ملخص الطلبات
    total_orders = len(orders)
    done_orders = 0
    pending_orders = 0
    sum_goods = 0.0
    sum_ship = 0.0

    for o in orders:
        ost = str(o.get("order_status") or "").strip().lower()
        if ost in ("cancelled", "canceled"):
            continue
        if ost in ("closed", "delivered", "canceled", "cancelled"):
            done_orders += 1
        else:
            pending_orders += 1
        sum_goods += _num(o.get("goods_amount_sar"))
        sum_ship += _num(o.get("shipping_fee_sar"))

    sum_total = sum_goods + sum_ship

    orders_sorted = sorted(orders, key=lambda x: _parse_dt(str(x.get("created_at_utc") or "")), reverse=True)
    last15 = orders_sorted[:15]

    # ---------------- PDF meta ----------------
    # ✅ توقيت السعودية
    try:
        from zoneinfo import ZoneInfo

        ksa_tz = ZoneInfo("Asia/Riyadh")
        now = datetime.now(ksa_tz)
    except Exception:
        ksa_tz = timezone(timedelta(hours=3))
        now = datetime.now(ksa_tz)

    inv_title = "سجل التاجر"
    inv_no = f"{tid}-{now.strftime('%y%m%d')}"
    platform_bar = "منصة قطع غيار PPARTS"

    # --------------- temp pdf ---------------
    tmpdir = tempfile.gettempdir()
    pdf_path = os.path.join(tmpdir, f"pp_trader_ledger_{tid}_{uuid.uuid4().hex[:6]}.pdf")

    # --------------- Arabic font ---------------
    font_name = "Helvetica"
    chosen = ""
    try:
        base_dir = os.path.dirname(__file__)
        amiri_path = os.path.join(base_dir, "Amiri-Regular.ttf")
        noto_path = os.path.join(base_dir, "NotoNaskhArabic-Regular.ttf")

        if os.path.exists(amiri_path):
            chosen = amiri_path
        elif os.path.exists(noto_path):
            chosen = noto_path
        else:
            dejavu = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
            if os.path.exists(dejavu):
                chosen = dejavu

        if chosen:
            pdfmetrics.registerFont(TTFont("AR", chosen))
            font_name = "AR"
    except Exception:
        font_name = "Helvetica"

    # --------------- Styles ---------------
    ss = getSampleStyleSheet()
    base = ParagraphStyle(
        "base",
        parent=ss["Normal"],
        fontName=font_name,
        fontSize=9.6,
        leading=12.0,
        textColor=colors.HexColor("#0B1220"),
    )
    right = ParagraphStyle("right", parent=base, alignment=TA_RIGHT)
    left = ParagraphStyle("left", parent=base, alignment=TA_LEFT)
    center = ParagraphStyle("center", parent=base, alignment=TA_CENTER)
    tiny_r = ParagraphStyle("tiny_r", parent=right, fontSize=8.6, leading=10.2)
    tiny_c = ParagraphStyle("tiny_c", parent=center, fontSize=8.6, leading=10.2)
    tiny_l = ParagraphStyle("tiny_l", parent=left, fontSize=8.6, leading=10.2)

    # ✅ BADGES 6-COLS styles (السطر الأول فقط)
    # الكلمة يسار + القيمة يمين
    badge_lbl = ParagraphStyle("badge_lbl", parent=base, alignment=TA_LEFT, fontSize=8.9, leading=10.6)
    badge_val = ParagraphStyle("badge_val", parent=base, alignment=TA_RIGHT, fontSize=8.9, leading=10.6)

    # --------------- Colors (Orange theme) ---------------
    def _with_alpha(c, a=1.0):
        try:
            r, g, b = c.red, c.green, c.blue
            return colors.Color(r, g, b, alpha=a)
        except Exception:
            return c

    C_BORDER = colors.HexColor("#CBD5E1")
    C_DARK = colors.HexColor("#9A3412")
    C_DARK_2 = colors.HexColor("#C2410C")
    BADGE_BG = colors.HexColor("#FFF7ED")
    SEC_HDR = colors.HexColor("#7C2D12")
    SEC_HDR_2 = colors.HexColor("#B45309")
    ROW_TINT1 = "#FFF7ED"
    ROW_TINT2 = "#FFEDD5"

    # ✅ شبكة خفيفة (للطابع العام)
    GRID = _with_alpha(C_BORDER, 0.45)
    # ✅ شبكة أوضح/أغمق للفواصل بين السطور (حل “تداخل بصري”)
    GRID_BOLD = _with_alpha(C_BORDER, 0.88)

    ROW_BG1 = _with_alpha(colors.HexColor(ROW_TINT1), 0.38)
    ROW_BG2 = _with_alpha(colors.HexColor(ROW_TINT2), 0.38)

    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=A4,
        rightMargin=0.85 * cm,
        leftMargin=0.85 * cm,
        topMargin=0.65 * cm,
        bottomMargin=0.75 * cm,
        title=inv_title,
        author="PP Platform",
    )

    def P(txt: str, st):
        return Paragraph(_ar(txt), st)

    full_w = A4[0] - doc.leftMargin - doc.rightMargin
    story = []

    # -------- Logo path --------
    logo_path = ""
    try:
        p1 = os.path.join(os.path.dirname(__file__), "pparts.jpg")
        if os.path.exists(p1):
            logo_path = p1
        elif os.path.exists("pparts.jpg"):
            logo_path = "pparts.jpg"
    except Exception:
        logo_path = ""

    # ===== Header: Bigger Logo centered =====
    logo_cell = ""
    try:
        if logo_path and os.path.exists(logo_path):
            img = RLImage(logo_path)
            img.drawHeight = 3.00 * cm
            img.drawWidth = 3.00 * cm
            logo_cell = img
    except Exception:
        logo_cell = ""

    header_tbl = Table([[logo_cell if logo_cell else P("PPARTS", center)]], colWidths=[full_w])
    header_tbl.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
            ]
        )
    )
    story.append(header_tbl)
    story.append(Spacer(1, 2))

    title_bar = Table(
        [
            [
                P(
                    f"<b>{platform_bar}</b>    |    <b>{inv_title}</b>",
                    ParagraphStyle(
                        "tbar",
                        parent=center,
                        textColor=colors.white,
                        fontSize=10.6,
                        leading=12.0,
                        fontName=font_name,
                    ),
                )
            ]
        ],
        colWidths=[full_w],
    )
    title_bar.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), C_DARK),
                ("BOX", (0, 0), (-1, -1), 0.0, colors.white),
                ("LINEBELOW", (0, 0), (-1, 0), 1.6, _with_alpha(C_DARK_2, 0.95)),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(title_bar)
    story.append(Spacer(1, 3))

    # =========================
    # ✅ (01) BADGES FIX (نهائي فعلي):
    # - حذف FSI/PDI نهائيًا (سبب المربعات)
    # - استخدام LRE/PDF للأرقام/اليوزر فقط (عادة غير مرئية)
    # - عدم لف الاسم العربي بأي محارف اتجاه (حتى لا يتفكك)
    # - الكلمة يسار + القيمة يمين (فقط بالسطر الأول)
    # =========================
    LRE = "\u202A"
    PDF = "\u202C"

    def _ltr(x: str) -> str:
        s = "" if x is None else str(x)
        s = s.strip()
        return f"{LRE}{s}{PDF}" if s else "—"

    trader_name = (tname or str(tid)).strip() or "—"
    trader_value = trader_name
    if tuser:
        trader_value = f"{trader_name}  {_ltr(tuser)}"

    time_value = _ltr(now.strftime("%H:%M"))
    date_value = _ltr(now.strftime("%Y-%m-%d"))

    # ✅ widths: (قيمة عريضة + كلمة ضيقة) × 3  -> تقارب قوي
    VAL_W = full_w * 0.255
    LBL_W = full_w * 0.078

    badges = Table(
        [
            [
                P(time_value, badge_val),
                P("<b>الوقت</b>", badge_lbl),
                P(date_value, badge_val),
                P("<b>التاريخ</b>", badge_lbl),
                P(trader_value, badge_val),
                P("<b>التاجر</b>", badge_lbl),
            ]
        ],
        colWidths=[VAL_W, LBL_W, VAL_W, LBL_W, VAL_W, LBL_W],
    )

    PAIR1 = _with_alpha(colors.HexColor("#FFEFD8"), 1.0)  # وقت
    PAIR2 = _with_alpha(colors.HexColor("#FFF2E5"), 1.0)  # تاريخ
    PAIR3 = _with_alpha(colors.HexColor("#FFE7CC"), 1.0)  # تاجر

    badges.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 1.05, GRID_BOLD),
                # ❌ احذف INNERGRID بالكامل (هو اللي يطلع خطوط داخلية واضحة)
                # ("INNERGRID", (0, 0), (-1, -1), 0.85, GRID_BOLD),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BACKGROUND", (0, 0), (-1, -1), BADGE_BG),
                ("BACKGROUND", (0, 0), (1, 0), PAIR1),
                ("BACKGROUND", (2, 0), (3, 0), PAIR2),
                ("BACKGROUND", (4, 0), (5, 0), PAIR3),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                # القيم يمين
                ("ALIGN", (0, 0), (0, 0), "RIGHT"),
                ("ALIGN", (2, 0), (2, 0), "RIGHT"),
                ("ALIGN", (4, 0), (4, 0), "RIGHT"),
                # الكلمات يسار
                ("ALIGN", (1, 0), (1, 0), "LEFT"),
                ("ALIGN", (3, 0), (3, 0), "LEFT"),
                ("ALIGN", (5, 0), (5, 0), "LEFT"),
                # ✅ فواصل الأزواج فقط (بين وقت/تاريخ/تاجر)
                ("LINEBEFORE", (2, 0), (2, 0), 0.85, GRID_BOLD),
                ("LINEBEFORE", (4, 0), (4, 0), 0.85, GRID_BOLD),
                # ✅ لا يوجد أي خط بين القيمة والكلمة داخل الزوج
            ]
        )
    )
    story.append(badges)
    story.append(Spacer(1, 5))

    def section_header(title: str):
        t = Table(
            [
                [
                    P(
                        f"<b>{title}</b>",
                        ParagraphStyle(
                            "sec",
                            parent=right,
                            textColor=colors.white,
                            fontSize=9.8,
                            leading=12.0,
                            fontName=font_name,
                        ),
                    )
                ]
            ],
            colWidths=[full_w],
        )
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), SEC_HDR),
                    ("LINEBELOW", (0, 0), (-1, 0), 1.2, _with_alpha(SEC_HDR_2, 0.95)),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(t)
        story.append(Spacer(1, 3))

    def kv_table(rows: list[tuple[str, str]]):
        data = []
        for k, v in rows:
            vtxt = v if (v is not None and str(v).strip() != "") else "—"
            data.append([P(str(vtxt), right), P(f"<b>{k}</b>", right)])

        t = Table(data, colWidths=[full_w * 0.56, full_w * 0.44])
        ts = TableStyle(
            [
                # ✅ حدود أوضح
                ("BOX", (0, 0), (-1, -1), 1.05, GRID_BOLD),
                ("INNERGRID", (0, 0), (-1, -1), 0.85, GRID_BOLD),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )

        for r in range(0, len(data)):
            ts.add("BACKGROUND", (0, r), (-1, r), ROW_BG1 if r % 2 == 0 else ROW_BG2)

        t.setStyle(ts)
        story.append(t)
        story.append(Spacer(1, 6))

    # ===== Sections =====
    section_header("بيانات التاجر")
    kv_table(
        [
            ("رقم السجل الداخلي", inv_no),
            ("معرّف التاجر", str(tid)),
            ("اسم التاجر", tname or "—"),
            ("اسم المتجر", tcompany or "—"),
            ("يوزر تيليجرام", tuser or "—"),
            ("رقم اتصال المتجر", shop_phone or "—"),
            ("رقم السجل التجاري", cr_no or "—"),
            ("الرقم الضريبي", vat_no or "—"),
            ("تاريخ الانضمام", joined or "—"),
            ("آخر تحديث", upd or "—"),
        ]
    )

    section_header("حالة الحساب")
    kv_table(
        [
            ("الحالة", enabled_txt),
            (f"اشتراك الشهر ({month})", sub_status),
        ]
    )

    section_header("بيانات الدفع")
    kv_table(
        [
            ("البنك", bank or "—"),
            ("IBAN", iban or "—"),
            ("STC Pay", stc or "—"),
        ]
    )

    section_header("ملخص الطلبات")
    kv_table(
        [
            ("عدد الطلبات (إجمالي)", str(total_orders)),
            ("طلبات منجزة", str(done_orders)),
            ("طلبات معلقة", str(pending_orders)),
            ("إجمالي القطع", _money_tail(sum_goods, fb="0")),
            ("إجمالي الشحن", _money_tail(sum_ship, fb="0")),
            ("الإجمالي", _money_tail(sum_total, fb="0")),
        ]
    )

    section_header("آخر 15 طلب")
    tbl = [
        [
            P("<b>رقم الطلب</b>", center),
            P("<b>التاريخ</b>", center),
            P("<b>الحالة</b>", center),
            P("<b>قيمة القطع</b>", center),
            P("<b>الشحن</b>", center),
        ]
    ]
    for o in last15:
        oid = str(o.get("order_id") or "").strip() or "—"
        dt = _parse_dt(str(o.get("created_at_utc") or ""))
        dt_s = dt.strftime("%Y-%m-%d") if dt.year > 1900 else "—"
        ost = str(o.get("order_status") or o.get("status") or "").strip()
        goods = _money_tail(_num(o.get("goods_amount_sar")), fb="0")
        ship = _money_tail(_num(o.get("shipping_fee_sar")), fb="0")
        tbl.append([P(oid, center), P(dt_s, center), P(ost or "—", center), P(goods, center), P(ship, center)])

    t = Table(tbl, colWidths=[full_w * 0.24, full_w * 0.16, full_w * 0.22, full_w * 0.19, full_w * 0.19])
    t.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 1.05, GRID_BOLD),
                ("INNERGRID", (0, 0), (-1, -1), 0.85, GRID_BOLD),
                ("BACKGROUND", (0, 0), (-1, 0), SEC_HDR),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )

    try:
        for r in range(1, len(tbl)):
            t.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, r), (-1, r), ROW_BG1 if r % 2 == 1 else ROW_BG2),
                    ]
                )
            )
    except Exception:
        pass

    story.append(t)
    story.append(Spacer(1, 2))

    def _draw_footer(canvas, docx):
        try:
            page_w, page_h = A4
            y = 0.55 * cm
            left_txt = "P202126P@HOTMAIL.CPM"
            right_txt = "منصة PPARTS احد الخدمات المساندة لنظام GO"

            canvas.saveState()
            try:
                canvas.setFillAlpha(0.85)
            except Exception:
                pass
            canvas.setFont(font_name, 8.6)
            canvas.setFillColor(colors.HexColor("#0B1220"))
            canvas.drawRightString(page_w - doc.rightMargin, y, _ar(right_txt))
            canvas.drawString(doc.leftMargin, y, left_txt)
            canvas.restoreState()
        except Exception:
            pass

    def _wm(canvas, docx):
        try:
            if logo_path and os.path.exists(logo_path):
                page_w, page_h = A4
                wm_w = page_w * 0.86
                wm_h = wm_w
                x = (page_w - wm_w) / 2.0
                y = (page_h - wm_h) / 2.0 + (0.9 * cm)
                try:
                    canvas.setFillAlpha(0.10)
                except Exception:
                    pass
                canvas.drawImage(
                    logo_path,
                    x,
                    y,
                    width=wm_w,
                    height=wm_h,
                    mask="auto",
                    preserveAspectRatio=True,
                    anchor="c",
                )
                try:
                    canvas.setFillAlpha(1.0)
                except Exception:
                    pass

            _draw_footer(canvas, docx)
        except Exception:
            pass

    try:
        doc.build(story, onFirstPage=_wm, onLaterPages=_wm)
    except Exception as e:
        try:
            await context.bot.send_message(chat_id=admin_chat_id, text=f"⚠️ فشل بناء PDF: {e}")
        except Exception:
            pass
        try:
            if os.path.exists(pdf_path):
                os.remove(pdf_path)
        except Exception:
            pass
        return

    try:
        with open(pdf_path, "rb") as f:
            await context.bot.send_document(
                chat_id=int(admin_chat_id),
                document=InputFile(f, filename=f"سجل-التاجر-{tid}.pdf"),
                caption=f"🧾 سجل التاجر: {tname or tid}" + (f" {tuser}" if tuser else ""),
            )
    except Exception as e:
        try:
            await context.bot.send_message(chat_id=admin_chat_id, text=f"⚠️ تعذر إرسال PDF: {e}")
        except Exception:
            pass
    finally:
        try:
            if os.path.exists(pdf_path):
                os.remove(pdf_path)
        except Exception:
            pass
            
async def send_platform_invoice_pdf(
    context: ContextTypes.DEFAULT_TYPE,
    order_id: str,
    kind: str = "preliminary",
    tracking_number: str = "",
    admin_only: bool = False,
):
    # فاتورة المنصة: رسوم المنصة فقط + كل بيانات العميل/الطلب
    return await send_invoice_pdf(
        context=context,
        order_id=order_id,
        kind=kind,
        tracking_number=tracking_number,
        admin_only=admin_only,
        invoice_for="platform",
    )


async def send_trader_invoice_pdf(
    context: ContextTypes.DEFAULT_TYPE,
    order_id: str,
    kind: str = "preliminary",
    tracking_number: str = "",
    admin_only: bool = False,
):
    # فاتورة التاجر: قيمة القطع + الشحن فقط (بدون رسوم المنصة)
    return await send_invoice_pdf(
        context=context,
        order_id=order_id,
        kind=kind,
        tracking_number=tracking_number,
        admin_only=admin_only,
        invoice_for="trader",
    )
    
    
async def send_invoice_pdf(
    context: ContextTypes.DEFAULT_TYPE,
    order_id: str,
    kind: str = "preliminary",
    tracking_number: str = "",
    admin_only: bool = False,
    invoice_for: str = "platform",   # "platform" or "trader"
    include_admins: bool = True,  # True: يرسل للعميل + الإدارة (الافتراضي) | False: يرسل للعميل فقط
    debug: bool = False,
):
    """
    Compact one-page Arabic invoice (Platform/Trader) with:
    ✅ اسم المنصة داخل الصف الملون: منصة قطع غيار PARTS / فاتورة داخلية (عند admin_only)
    ✅ تقسيم البيانات: (معلومات العميل) / (معلومات السيارة) / (تفاصيل الشحن)
    ✅ KV عربي مثل تفاصيل القطع: المعرّف يمين والمعلومة يساره (عمودين واضحين)
    ✅ جدول القطع RTL: # أقصى اليمين + أعمدة منفصلة (اسم القطعة / رقم القطعة)
    ✅ ختم مدفوع احترافي ثابت أسفل الصفحة (مرة واحدة) + "الخدمات المساندة GO" تحته
    ✅ العلامة المائية خلف المحتوى ومرفوعة للأعلى وتظهر (بدون ما تغطيها خلفيات بيضاء)
    ✅ ألوان مختلفة (المنصة أزرق / التاجر أخضر)
    ✅ توحيد وقت الفاتورة على KSA + تحسين عرض رقم الطلب (عرض فقط) + تقصير عرض رقم الفاتورة (عرض فقط)
    """

    # ✅ tempfile
    try:
        import tempfile
    except Exception as e:
        await _notify_invoice_error(context, order_id, "تهيئة (tempfile)", e)
        return

    import os, html, uuid, re, json
    from datetime import datetime, timezone, timedelta

    # --- Arabic RTL + shaping ---
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display

        def _ar(s: str) -> str:
            s = "" if s is None else str(s)
            if not s:
                return s
            try:
                return get_display(arabic_reshaper.reshape(s))
            except Exception:
                return s
    except Exception:
        def _ar(s: str) -> str:
            return "" if s is None else str(s)

    # ✅ reportlab imports
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph, Image as RLImage
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except Exception as e:
        await _notify_invoice_error(context, order_id, "استيراد مكتبات PDF (reportlab)", e)
        return

    # 1) اقرأ الطلب
    try:
        b = get_order_bundle(order_id)
        order = b.get("order", {}) or {}
        items = b.get("items", []) or []
    except Exception as e:
        await _notify_invoice_error(context, order_id, "قراءة بيانات الطلب من الإكسل", e)
        return

    invoice_for_norm = (str(invoice_for or "platform").strip().lower())
    if invoice_for_norm not in ("platform", "trader"):
        invoice_for_norm = "platform"

    # ---------------- Helpers ----------------
    def _s(x: object) -> str:
        return ("" if x is None else str(x)).strip()

    def _to_float(x: object) -> float:
        try:
            return float(str(x or 0).replace(",", "").strip() or 0)
        except Exception:
            return 0.0

    def _money_safe(x: object, fb: str = "0") -> str:
        try:
            s = _money(x)
            s = (s or "").strip()
            return s if s else fb
        except Exception:
            return fb

    def _extract_phone(txt: str) -> str:
        t = _s(txt)
        if not t:
            return ""
        m = re.search(r'(\+?9665\d{8}|9665\d{8}|05\d{8})', t)
        return m.group(1) if m else ""

    def _cell_clip(s: str, max_chars: int = 120) -> str:
        s = _s(s)
        s = re.sub(r"\s+", " ", s).strip()
        if len(s) <= max_chars:
            return s
        return s[: max(0, max_chars - 1)].rstrip() + "…"

    # ✅ (02) تنسيق المبلغ: رقم + ﷼ (بدل ر.س/س.ر)
    def _money_tail(x: object, fb: str = "0") -> str:
        s = _money_safe(x, fb=fb)
        s = _s(s)
        s = re.sub(r'^\s*(ر\.?\s*س|ر\.س|SAR|SR|s\.r|s\.r\.?)\s*', '', s, flags=re.I)
        s = re.sub(r'\s*(ر\.?\s*س|ر\.س|SAR|SR|s\.r|s\.r\.?)\s*$', '', s, flags=re.I)
        s = s.strip() or fb
        return f"{s} ﷼"

    # ✅ استخراج خريطة أسعار القطع (لفواتير التاجر فقط)
    def _load_item_prices_map() -> dict:
        pm = {}
        try:
            import ast
        except Exception:
            ast = None

        for k in ("quote_item_prices", "item_prices", "goods_item_prices", "quote_items_prices"):
            raw = order.get(k)
            if raw is None or raw == "":
                continue

            src = None
            try:
                if isinstance(raw, dict):
                    src = raw
                else:
                    sraw = str(raw).strip()
                    if sraw.startswith("{") and sraw.endswith("}"):
                        try:
                            src = json.loads(sraw)
                        except Exception:
                            src = None
                        if src is None and ast:
                            try:
                                src = ast.literal_eval(sraw)
                            except Exception:
                                src = None
            except Exception:
                src = None

            if isinstance(src, dict):
                for kk, vv in src.items():
                    ks = str(kk).strip()
                    vs = str(vv).strip()
                    if ks.isdigit() and vs not in ("", "0", "0.0", "0.00"):
                        pm[ks] = vs

        return pm

    def _pick_item_price(i: int, it: dict, pm: dict) -> str:
        cand_keys = ("price_sar", "item_price", "price", "unit_price", "amount_sar", "cost_sar", "cost", "sar")
        for ck in cand_keys:
            v = it.get(ck) if isinstance(it, dict) else None
            vs = _s(v)
            if vs and vs not in ("0", "0.0", "0.00"):
                return vs

        vs = _s(pm.get(str(i)))
        if vs and vs not in ("0", "0.0", "0.00"):
            return vs

        return ""

    # ---------------- IDs / dates ----------------
    client_id = int(order.get("user_id") or 0) if _s(order.get("user_id")).isdigit() else 0
    trader_id = int(order.get("accepted_trader_id") or 0) if _s(order.get("accepted_trader_id")).isdigit() else 0

    # ✅ KSA timezone بشكل مباشر (بدون UTC)
    KSA_TZ = timezone(timedelta(hours=3))
    now_dt = datetime.now(KSA_TZ)
    inv_date = now_dt.strftime("%Y-%m-%d")
    inv_time = now_dt.strftime("%H:%M")

    # ✅ تحسين عرض رقم الطلب داخل الفاتورة فقط:
    # - لو order_id يحتوي تاريخ مدمج + تسلسل
    # - نعيد بناء "عرض" بتاريخ السعودية + نفس التسلسل
    # - بدون تغيير order_id الحقيقي المستخدم بالنظام
    def _order_id_display(oid: str) -> str:
        s = _s(oid)
        if not s:
            return "—"

        # استخرج التسلسل من آخر جزء إذا كان رقم
        seq = ""
        try:
            parts = [p for p in re.split(r"[-_]", s) if p]
            last = parts[-1] if parts else ""
            if last.isdigit():
                seq = last.zfill(4) if len(last) <= 4 else last
        except Exception:
            seq = ""

        # إن لم نجد تسلسل، نرجع كما هو
        if not seq:
            return s

        # تاريخ السعودية الحالي بصيغة ddmmyy لتطابق رقم الطلب
        d6 = now_dt.strftime("%d%m%y")

        # حافظ على بادئة PP إذا كانت موجودة، وإلا خله PP
        prefix = "PP"
        try:
            if s.lower().startswith("pp"):
                prefix = "PP"
        except Exception:
            prefix = "PP"

        # الشكل المحسن المختصر
        return f"{prefix}-{d6}-{seq}"

    order_id_disp = _order_id_display(order_id)

    kind_norm = (kind or "preliminary").strip().lower()
    if kind_norm not in ("preliminary", "shipping"):
        kind_norm = "preliminary"

    def _get_existing_inv():
        if kind_norm == "preliminary":
            return _s(order.get("invoice_pre_no"))
        if kind_norm == "shipping":
            return _s(order.get("invoice_ship_no"))
        return ""

    inv_no = _get_existing_inv()
    if not inv_no:
        inv_no = f"PP-{order_id}-{kind_norm.upper()}-{now_dt.strftime('%Y%m%d')}-{uuid.uuid4().hex[:4].upper()}"

    # خزّن رقم الفاتورة (مرة واحدة)
    try:
        if kind_norm == "preliminary" and not _s(order.get("invoice_pre_no")):
            update_order_fields(order_id, {"invoice_pre_no": inv_no})
        if kind_norm == "shipping" and not _s(order.get("invoice_ship_no")):
            update_order_fields(order_id, {"invoice_ship_no": inv_no})
    except Exception as e:
        _swallow(e)

    # ✅ تقصير عرض رقم الفاتورة داخل PDF فقط (بدون تغيير المخزن)
    def _inv_no_display(full: str) -> str:
        s = _s(full)
        if not s:
            return "—"
        tail4 = s[-4:] if len(s) >= 4 else s
        kn = "PRE" if kind_norm == "preliminary" else "SHP"
        return f"{order_id_disp}-{kn}-{tail4}"

    inv_no_disp = _inv_no_display(inv_no)

    # ---------------- Anti-duplicate send (Excel flags) ----------------
    if invoice_for_norm == "trader":
        sent_flag_field = "invoice_trader_pre_pdf_sent" if kind_norm == "preliminary" else "invoice_trader_ship_pdf_sent"
        legacy_flag = "invoice_trader_pdf_sent"
    else:
        sent_flag_field = "invoice_platform_pre_pdf_sent" if kind_norm == "preliminary" else "invoice_platform_ship_pdf_sent"
        legacy_flag = "invoice_platform_pdf_sent"

    def _is_yes(v) -> bool:
        return _s(v).strip().lower() in {"1", "yes", "true", "sent", "done"}

    if (_is_yes(order.get(sent_flag_field)) or _is_yes(order.get(legacy_flag))) and (not debug):
        return

    # ---------------- Data ----------------
    client_name = _s(order.get("user_name")) or "—"

    client_phone = _s(order.get("ship_phone") or order.get("pickup_phone"))
    if not client_phone:
        client_phone = _extract_phone(_s(order.get("delivery_details")))
    if not client_phone:
        client_phone = _extract_phone(_s(order.get("address_text")))
    if not client_phone:
        client_phone = _extract_phone(_s(order.get("full_address")))
    if not client_phone:
        client_phone = _extract_phone(_s(order.get("address")))
    if not client_phone:
        client_phone = "—"

    car_name = _s(order.get("car_name")) or "—"
    car_model = _s(order.get("car_model")) or "—"
    vin = _s(order.get("vin")) or "—"

    # ---------------- Trader profile (for trader invoices) ----------------
    _tp = {}
    if trader_id:
        try:
            _tp = get_trader_profile(int(trader_id)) or {}
        except Exception:
            _tp = {}

    def _tp_pick(*keys: str) -> str:
        for k in keys:
            try:
                v = _s((_tp or {}).get(k))
                if v:
                    return v
            except Exception:
                continue
        return ""

    trader_name = _s(order.get("accepted_trader_name") or order.get("quoted_trader_name")) or _tp_pick("display_name", "name", "full_name")
    trader_company = _tp_pick("company_name", "shop_name", "store_name")
    trader_phone = _tp_pick("shop_phone", "phone", "mobile", "shop_mobile", "store_phone")
    trader_cr_no = _tp_pick("cr_no", "cr", "cr_number", "commercial_register", "commercial_registration")
    trader_vat_no = _tp_pick("vat_no", "vat", "vat_number", "tax_no", "tax_number")

    if not trader_name and trader_id:
        trader_name = trader_company
    trader_name = trader_name or "—"

    ship_method = _s(order.get("delivery_type") or order.get("ship_method") or order.get("delivery_choice")) or "—"
    ship_city = _s(order.get("ship_city") or order.get("pickup_city"))
    ship_district = _s(order.get("ship_district"))
    ship_short = _s(order.get("ship_short_address"))
    delivery_blob = _s(order.get("delivery_details") or order.get("address_text") or order.get("full_address") or order.get("address"))
    delivery_details = _s(delivery_blob)

    raw_platform_fee = order.get("price_sar")
    raw_goods_amount = order.get("goods_amount_sar")

    ship_included = str(order.get("ship_included") or "").strip().lower()
    raw_shipping_fee = order.get("shipping_fee_sar")

    if raw_shipping_fee is None or str(raw_shipping_fee).strip() == "":
        raw_shipping_fee = 0 if ship_included in ("yes", "true", "1", "included", "مشمولة", "مشمول") else ""

    try:
        if invoice_for_norm != "trader":
            pf = raw_platform_fee
            pf_f = _to_float(pf) if pf not in (None, "") else 0.0
            if pf_f <= 0:
                auto_fee = _platform_fee_for_items(items)
                if auto_fee and _to_float(auto_fee) > 0:
                    raw_platform_fee = auto_fee
                    try:
                        update_order_fields(order_id, {"price_sar": auto_fee})
                    except Exception as e:
                        _swallow(e)
    except Exception as e:
        _swallow(e)

    platform_fee = _money_safe(raw_platform_fee or 0, fb="0")
    goods_amount = _money_safe(raw_goods_amount or 0, fb="0")

    if invoice_for_norm == "trader":
        pay_method = _s(order.get("goods_payment_method")) or _s(order.get("payment_method"))
        pay_status_raw = _s(order.get("goods_payment_status")) or _s(order.get("payment_status"))
        pay_status = _pay_status_ar(pay_status_raw)

        gt_val = _to_float(raw_goods_amount) + _to_float(raw_shipping_fee)
        _ = _money_safe(gt_val, fb=goods_amount if goods_amount != "0" else "0")

        inv_title = "فاتورة تاجر - داخلية - قطع + شحن"
    else:
        pay_method = _s(order.get("payment_method")) or _s(order.get("goods_payment_method"))
        pay_status_raw = _s(order.get("payment_status")) or _s(order.get("goods_payment_status"))
        pay_status = _pay_status_ar(pay_status_raw)

        inv_title = "فاتورة داخلية"

    pay_status = "مؤكد"

    if kind_norm == "shipping":
        inv_title = "فاتورة شحن" if invoice_for_norm == "trader" else "فاتورة شحن - منصة"

    platform_bar = "منصة قطع غيار PPARTS"
    if admin_only:
        platform_bar = platform_bar + " / فاتورة داخلية"

    # --------------- temp pdf ---------------
    tmpdir = tempfile.gettempdir()
    pdf_path = os.path.join(tmpdir, f"pp_invoice_{order_id}_{kind_norm}_{uuid.uuid4().hex[:6]}.pdf")

    # --------------- Arabic font ---------------
    font_name = "Helvetica"
    chosen = ""
    try:
        base_dir = os.path.dirname(__file__)
        amiri_path = os.path.join(base_dir, "Amiri-Regular.ttf")
        noto_path = os.path.join(base_dir, "NotoNaskhArabic-Regular.ttf")

        if os.path.exists(amiri_path):
            chosen = amiri_path
        elif os.path.exists(noto_path):
            chosen = noto_path
        else:
            dejavu = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
            if os.path.exists(dejavu):
                chosen = dejavu

        if chosen:
            font_name = "PP_AR"
            try:
                pdfmetrics.getFont(font_name)
            except Exception:
                pdfmetrics.registerFont(TTFont(font_name, chosen))
    except Exception:
        font_name = "Helvetica"

    stamp_font = font_name
    try:
        if chosen:
            stamp_font = "PP_AR_STAMP"
            try:
                pdfmetrics.getFont(stamp_font)
            except Exception:
                pdfmetrics.registerFont(TTFont(stamp_font, chosen))
    except Exception:
        stamp_font = font_name

    # --------------- Colors (per invoice type) ---------------
    C_BORDER = colors.HexColor("#CBD5E1")
    C_TEXT = colors.HexColor("#0B1220")

    if invoice_for_norm == "trader":
        C_DARK    = colors.HexColor("#065F46")
        C_DARK_2  = colors.HexColor("#0B7A57")
        BADGE_BG  = colors.HexColor("#E9FFF6")
        SEC_HDR   = colors.HexColor("#0F3D2E")
        SEC_HDR_2 = colors.HexColor("#145A43")
        STAMP     = colors.HexColor("#DC2626")
        ROW_TINT1 = "#ECFDF5"
        ROW_TINT2 = "#E6FFFA"
    else:
        C_DARK    = colors.HexColor("#0B3A6E")
        C_DARK_2  = colors.HexColor("#145AA0")
        BADGE_BG  = colors.HexColor("#EAF2FF")
        SEC_HDR   = colors.HexColor("#0A2E57")
        SEC_HDR_2 = colors.HexColor("#123E6D")
        STAMP     = colors.HexColor("#DC2626")
        ROW_TINT1 = "#EFF6FF"
        ROW_TINT2 = "#E8F1FF"

    def _with_alpha(c, a: float):
        try:
            return colors.Color(c.red, c.green, c.blue, alpha=max(0.0, min(1.0, float(a))))
        except Exception:
            return c

    def _hexA(hx: str, a: float):
        try:
            c = colors.HexColor(hx)
            return colors.Color(c.red, c.green, c.blue, alpha=max(0.0, min(1.0, float(a))))
        except Exception:
            return colors.HexColor(hx)

    # --------------- Styles (tight to keep 1 page) ---------------
    styles = getSampleStyleSheet()

    kv_label = ParagraphStyle(
        "kv_label",
        parent=styles["Normal"],
        alignment=TA_RIGHT,
        fontSize=8.6,
        leading=10.2,
        fontName=font_name,
        textColor=C_TEXT
    )

    kv_value = ParagraphStyle(
        "kv_value",
        parent=styles["Normal"],
        alignment=TA_RIGHT,
        fontSize=8.6,
        leading=10.2,
        fontName=font_name,
        textColor=C_TEXT
    )

    center  = ParagraphStyle("center", parent=styles["Normal"], alignment=TA_CENTER, fontSize=11.4, leading=12.6, fontName=font_name)
    tiny_c  = ParagraphStyle("tiny_c", parent=styles["Normal"], alignment=TA_CENTER, fontSize=8.8, leading=10.2, fontName=font_name)

    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=A4,
        rightMargin=0.85 * cm,
        leftMargin=0.85 * cm,
        topMargin=0.65 * cm,
        bottomMargin=0.75 * cm,
        title=inv_title,
        author="PP Platform",
    )

    def P(txt: str, st):
        return Paragraph(_ar(txt), st)

    full_w = A4[0] - doc.leftMargin - doc.rightMargin
    story = []

    # -------- Logo path --------
    logo_path = ""
    try:
        p1 = os.path.join(os.path.dirname(__file__), "pparts.jpg")
        if os.path.exists(p1):
            logo_path = p1
        elif os.path.exists("pparts.jpg"):
            logo_path = "pparts.jpg"
    except Exception:
        logo_path = ""

    # ===== Header: Bigger Logo centered =====
    logo_cell = ""
    try:
        if logo_path and os.path.exists(logo_path):
            img = RLImage(logo_path)
            img.drawHeight = 3.00 * cm
            img.drawWidth = 3.00 * cm
            logo_cell = img
    except Exception:
        logo_cell = ""

    header_tbl = Table([[logo_cell if logo_cell else P("PPARTS", center)]], colWidths=[full_w])
    header_tbl.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
    ]))
    story.append(header_tbl)
    story.append(Spacer(1, 2))

    title_bar = Table([[
        P(f"<b>{platform_bar}</b>    |    <b>{inv_title}</b>",
          ParagraphStyle("tbar", parent=center, textColor=colors.white, fontSize=10.6, leading=12.0, fontName=font_name))
    ]], colWidths=[full_w])
    title_bar.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), C_DARK),
        ("BOX", (0, 0), (-1, -1), 0.0, colors.white),
        ("LINEBELOW", (0, 0), (-1, 0), 1.6, _with_alpha(C_DARK_2, 0.95)),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(title_bar)
    story.append(Spacer(1, 3))

    # ✅ badges: استخدم عرض رقم الطلب المحسن + عرض رقم الفاتورة المختصر + وقت KSA
    badges = Table([[
        P(f"رقم الفاتورة: <b>{inv_no_disp}</b>", tiny_c),
        P(f"رقم الطلب: <b>{order_id_disp}</b>", tiny_c),
        P(f"{inv_date}  {inv_time} (KSA)", tiny_c),
    ]], colWidths=[0.40 * full_w, 0.30 * full_w, 0.30 * full_w])
    badges.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), _with_alpha(BADGE_BG, 0.58)),
        ("BOX", (0, 0), (-1, -1), 0.6, C_BORDER),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(badges)
    story.append(Spacer(1, 3))

    def section_kv(title: str, rows: list):
        hdr = Table([[
            P(f"<b>{title}</b>",
              ParagraphStyle("sh", parent=kv_label, fontSize=9.1, leading=10.6,
                             textColor=colors.white, fontName=font_name))
        ]], colWidths=[full_w])
        hdr.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), _with_alpha(SEC_HDR, 0.92)),
            ("LINEBELOW", (0, 0), (-1, 0), 1.1, _with_alpha(SEC_HDR_2, 0.92)),
            ("BOX", (0, 0), (-1, -1), 0.6, C_BORDER),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(hdr)

        data = []
        for k, v in rows:
            data.append([
                P(html.escape(str(v)), kv_value),
                P("", kv_value),
                P(f"<b>{html.escape(str(k))}</b>", kv_label),
            ])

        t = Table(data, colWidths=[0.64 * full_w, 0.03 * full_w, 0.33 * full_w])
        t.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.6, C_BORDER),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, C_BORDER),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LINEBEFORE", (1, 0), (1, -1), 0, colors.white),
            ("LINEAFTER",  (1, 0), (1, -1), 0, colors.white),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [_hexA("#FFFFFF", 0.00), _hexA("#FFFFFF", 0.00)]),
        ]))
        story.append(t)
        story.append(Spacer(1, 3))

    rows_client = [("اسم العميل", client_name), ("رقم الجوال", client_phone)]
    pmethod = _s(order.get("goods_payment_method")) or _s(order.get("payment_method")) or ""
    if pmethod:
        rows_client.append(("طريقة الدفع", pmethod))
    rows_client.append(("حالة الدفع", "مؤكد"))
    section_kv("معلومات العميل", rows_client)

    rows_car = [
        ("اسم السيارة", car_name),
        ("الموديل", car_model),
        ("رقم الهيكل VIN", vin),
    ]
    section_kv("معلومات السيارة", rows_car)

    # ✅ إضافة بيانات التاجر/المتجر داخل فاتورة التاجر فقط (نفس تنسيق KV الحالي)
    if invoice_for_norm == "trader":
        rows_trader = [
            ("اسم التاجر", trader_name or "—"),
            ("اسم المتجر", trader_company or "—"),
            ("رقم اتصال المتجر", trader_phone or "—"),
            ("رقم السجل التجاري", trader_cr_no or "—"),
            ("الرقم الضريبي", trader_vat_no or "—"),
        ]
        section_kv("بيانات التاجر", rows_trader)

    rows_ship = [("نوع التسليم", ship_method)]
    if ship_city:
        rows_ship.append(("المدينة", ship_city))
    if ship_district:
        rows_ship.append(("الحي", ship_district))
    if ship_short:
        rows_ship.append(("العنوان المختصر", ship_short))
    if delivery_details:
        rows_ship.append(("تفاصيل العنوان", _cell_clip(delivery_details, 140)))

    # ✅ رقم التتبع: يظهر إذا موجود (من tracking_number أو من order.shipping_tracking)
    # ✅ وإذا غير موجود: يكتب توضيح بدل الشرط القديم المرتبط بـ kind_norm
    _trk = _s(tracking_number) or _s(order.get("shipping_tracking"))
    if _trk:
        rows_ship.append(("رقم التتبع", _trk))
    else:
        rows_ship.append(("رقم التتبع", "لا يوجد رقم تتبع"))

    section_kv("تفاصيل الشحن", rows_ship)

    sec_parts = Table([[P("<b>تفاصيل القطع</b>",
                          ParagraphStyle("sh2", parent=kv_label, fontSize=9.0, leading=10.5,
                                         textColor=colors.white, fontName=font_name))]],
                      colWidths=[full_w])
    sec_parts.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), _with_alpha(SEC_HDR, 0.92)),
        ("LINEBELOW", (0, 0), (-1, 0), 1.1, _with_alpha(SEC_HDR_2, 0.92)),
        ("BOX", (0, 0), (-1, -1), 0.6, C_BORDER),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(sec_parts)

    parts_cell_r = ParagraphStyle("parts_cell_r", parent=styles["Normal"], alignment=TA_RIGHT, fontSize=8.2, leading=9.6, fontName=font_name)
    parts_cell_num = ParagraphStyle("parts_cell_num", parent=styles["Normal"], alignment=TA_RIGHT, fontSize=8.2, leading=9.6, fontName=font_name)

    price_map = _load_item_prices_map() if invoice_for_norm == "trader" else {}

    if invoice_for_norm == "trader":
        parts_rows = [[
            P("<b>سعر القطعة</b>", ParagraphStyle("ph0", parent=parts_cell_r, textColor=colors.white)),
            P("<b>رقم القطعة</b>", ParagraphStyle("ph1", parent=parts_cell_r, textColor=colors.white)),
            P("<b>اسم القطعة</b>", ParagraphStyle("ph2", parent=parts_cell_r, textColor=colors.white)),
            P("<b>#</b>", ParagraphStyle("ph3", parent=parts_cell_r, textColor=colors.white)),
        ]]

        shown_any = False
        if items:
            for i, it in enumerate(items, start=1):
                nm = _cell_clip(it.get("name") or it.get("item_name") or "—", 60) or "—"
                pn = _cell_clip(it.get("part_no") or it.get("item_part_no") or it.get("number") or "—", 40) or "—"

                pr = _pick_item_price(i, it if isinstance(it, dict) else {}, price_map)
                if pr:
                    price_txt = _money_tail(pr, fb="0")
                else:
                    price_txt = "غير متوفرة"

                shown_any = True
                parts_rows.append([
                    Paragraph(_ar(html.escape(price_txt)), parts_cell_r),
                    Paragraph(_ar(html.escape(pn)), parts_cell_r),
                    Paragraph(_ar(html.escape(nm)), parts_cell_r),
                    Paragraph(_ar(str(i)), parts_cell_num),
                ])

        if not shown_any:
            parts_rows.append([
                Paragraph(_ar("—"), parts_cell_r),
                Paragraph(_ar("—"), parts_cell_r),
                Paragraph(_ar("—"), parts_cell_r),
                Paragraph(_ar("1"), parts_cell_num),
            ])

        col_w = [0.20 * full_w, 0.24 * full_w, 0.48 * full_w, 0.08 * full_w]
    else:
        parts_rows = [[
            P("<b>رقم القطعة</b>", ParagraphStyle("ph1", parent=parts_cell_r, textColor=colors.white)),
            P("<b>اسم القطعة</b>", ParagraphStyle("ph2", parent=parts_cell_r, textColor=colors.white)),
            P("<b>#</b>", ParagraphStyle("ph3", parent=parts_cell_r, textColor=colors.white)),
        ]]

        if items:
            for i, it in enumerate(items, start=1):
                nm = _cell_clip(it.get("name") or it.get("item_name") or "—", 60) or "—"
                pn = _cell_clip(it.get("part_no") or it.get("item_part_no") or it.get("number") or "—", 40) or "—"
                parts_rows.append([
                    Paragraph(_ar(html.escape(pn)), parts_cell_r),
                    Paragraph(_ar(html.escape(nm)), parts_cell_r),
                    Paragraph(_ar(str(i)), parts_cell_num),
                ])
        else:
            parts_rows.append([
                Paragraph(_ar("—"), parts_cell_r),
                Paragraph(_ar("—"), parts_cell_r),
                Paragraph(_ar("1"), parts_cell_num),
            ])

        col_w = [0.34 * full_w, 0.58 * full_w, 0.08 * full_w]

    row_h = 0.62 * cm
    parts_tbl = Table(parts_rows, colWidths=col_w, rowHeights=[row_h] * len(parts_rows), repeatRows=1)
    parts_tbl.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.7, C_BORDER),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, C_BORDER),
        ("BACKGROUND", (0, 0), (-1, 0), _with_alpha(C_DARK, 0.92)),
        ("LINEBELOW", (0, 0), (-1, 0), 1.2, _with_alpha(C_DARK_2, 0.95)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 1.2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 1.2),
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ("RIGHTPADDING", (len(col_w) - 1, 0), (len(col_w) - 1, -1), 0.8),
        ("LEFTPADDING", (len(col_w) - 1, 0), (len(col_w) - 1, -1), 0.8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [_hexA(ROW_TINT1, 0.18), _hexA(ROW_TINT2, 0.12)]),
    ]))
    story.append(parts_tbl)
    story.append(Spacer(1, 3))

    # ===== Financial Summary Header =====
    sec_fin = Table([[P("<b>الملخص المالي</b>",
                        ParagraphStyle("sh3", parent=kv_label, fontSize=9.0, leading=10.5,
                                       textColor=colors.white, fontName=font_name))]],
                    colWidths=[full_w])
    sec_fin.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), _with_alpha(SEC_HDR, 0.92)),
        ("LINEBELOW", (0, 0), (-1, 0), 1.1, _with_alpha(SEC_HDR_2, 0.92)),
        ("BOX", (0, 0), (-1, -1), 0.6, C_BORDER),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(sec_fin)

    fin_lbl_w = ParagraphStyle("fin_lbl_w", parent=tiny_c, alignment=TA_RIGHT,
                               fontSize=9.0, leading=10.2, fontName=font_name, textColor=colors.white)
    fin_lbl_d = ParagraphStyle("fin_lbl_d", parent=tiny_c, alignment=TA_RIGHT,
                               fontSize=9.0, leading=10.2, fontName=font_name, textColor=C_TEXT)

    fin_amt_w = ParagraphStyle("fin_amt_w", parent=tiny_c, alignment=TA_RIGHT,
                               fontSize=10.0, leading=11.0, fontName=font_name, textColor=colors.white)
    fin_amt_d = ParagraphStyle("fin_amt_d", parent=tiny_c, alignment=TA_RIGHT,
                               fontSize=10.0, leading=11.0, fontName=font_name, textColor=C_TEXT)

    if invoice_for_norm == "trader":
        gt_val = _to_float(raw_goods_amount) + _to_float(raw_shipping_fee)

        BG_TOTAL = _with_alpha(C_DARK, 0.82)
        BG_SHIP  = _hexA("#DFF7EA", 0.18)
        BG_PARTS = _hexA("#D8F0FF", 0.18)

        money_box = Table([
            [
                P("<b>الإجمالي</b>", fin_lbl_w),
                P("<b>رسوم الشحن</b>", fin_lbl_d),
                P("<b>قيمة القطع</b>", fin_lbl_d),
            ],
            [
                Paragraph(_money_tail(gt_val, fb="0"), fin_amt_w),
                Paragraph(_money_tail(raw_shipping_fee, fb="0"), fin_amt_d),
                Paragraph(_money_tail(raw_goods_amount, fb="0"), fin_amt_d),
            ],
        ], colWidths=[0.34 * full_w, 0.33 * full_w, 0.33 * full_w])

        money_box.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.8, C_BORDER),
            ("INNERGRID", (0, 0), (-1, -1), 0.35, C_BORDER),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
            ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("BACKGROUND", (0, 0), (0, 1), BG_TOTAL),
            ("BACKGROUND", (1, 0), (1, 1), BG_SHIP),
            ("BACKGROUND", (2, 0), (2, 1), BG_PARTS),
            ("LINEABOVE", (0, 1), (-1, 1), 0.6, C_BORDER),
        ]))
        story.append(money_box)

    else:
        BG_TOTAL = _with_alpha(C_DARK, 0.82)
        BG_FEE   = _hexA("#D7E7FF", 0.18)

        platform_total_val = _to_float(raw_platform_fee)

        one_box = Table([
            [P("<b>الإجمالي</b>", fin_lbl_w), P("<b>رسوم المنصة</b>", fin_lbl_d)],
            [
                Paragraph(_money_tail(platform_total_val, fb="0"), fin_amt_w),
                Paragraph(_money_tail(raw_platform_fee, fb="0"), fin_amt_d),
            ],
        ], colWidths=[0.45 * full_w, 0.55 * full_w])

        one_box.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.8, C_BORDER),
            ("INNERGRID", (0, 0), (-1, -1), 0.35, C_BORDER),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
            ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("BACKGROUND", (0, 0), (0, 1), BG_TOTAL),
            ("BACKGROUND", (1, 0), (1, 1), BG_FEE),
            ("LINEABOVE", (0, 1), (-1, 1), 0.6, C_BORDER),
        ]))
        story.append(one_box)

    story.append(Spacer(1, 2))

    footer_email = "p200126p@hotmail.com"
    rights_line = "/ الخدمات المساندة GO ومنصة PP"

    def _draw_extras(canvas, _doc, *, draw_stamp: bool):
        canvas.saveState()

        try:
            if logo_path and os.path.exists(logo_path):
                from reportlab.lib.utils import ImageReader
                img = ImageReader(logo_path)
                page_w, page_h = A4

                wm_w = 17.2 * cm
                wm_h = 17.2 * cm
                x = (page_w - wm_w) / 2.0
                y = (page_h - wm_h) / 2.0 + (3.2 * cm)

                try:
                    canvas.setFillAlpha(0.16)
                except Exception as e:
                    _swallow(e)

                canvas.drawImage(
                    img, x, y,
                    width=wm_w, height=wm_h,
                    mask='auto',
                    preserveAspectRatio=True,
                    anchor='c'
                )

                try:
                    canvas.setFillAlpha(1)
                except Exception as e:
                    _swallow(e)
        except Exception as e:
            _swallow(e)

        canvas.setStrokeColor(C_BORDER)
        canvas.setLineWidth(0.55)
        canvas.line(doc.leftMargin, 0.92 * cm, A4[0] - doc.rightMargin, 0.92 * cm)

        canvas.setFillColor(C_TEXT)
        try:
            canvas.setFont(font_name, 7.6)
        except Exception:
            canvas.setFont("Helvetica", 7.6)

        canvas.drawString(doc.leftMargin, 0.60 * cm, _ar(rights_line))
        canvas.drawRightString(A4[0] - doc.rightMargin, 0.60 * cm, _ar(footer_email))

        if draw_stamp:
            if invoice_for_norm == "trader":
                stamp_cx = doc.leftMargin + (0.34 * full_w) / 2.0
            else:
                stamp_cx = doc.leftMargin + (0.45 * full_w) / 2.0

            stamp_cy = 2.55 * cm
            r = 1.22 * cm

            try:
                canvas.setFillColor(STAMP)
                canvas.setStrokeColor(STAMP)
                canvas.setLineWidth(1.2)
                canvas.circle(stamp_cx, stamp_cy, r, stroke=1, fill=1)

                canvas.setStrokeColor(colors.white)
                canvas.setLineWidth(1.15)
                canvas.circle(stamp_cx, stamp_cy, r - (0.06 * cm), stroke=1, fill=0)

                canvas.setStrokeColor(_with_alpha(colors.white, 0.65))
                canvas.setLineWidth(0.9)
                canvas.circle(stamp_cx, stamp_cy, r - (0.18 * cm), stroke=1, fill=0)
            except Exception as e:
                _swallow(e)

            try:
                canvas.setFillColor(colors.white)

                try:
                    canvas.setFont(stamp_font, 13.2)
                except Exception:
                    canvas.setFont("Helvetica-Bold", 13.2)
                canvas.drawCentredString(stamp_cx, stamp_cy + 0.42 * cm, _ar("مدفوع"))

                try:
                    canvas.setFont(stamp_font, 6.5)
                except Exception:
                    canvas.setFont("Helvetica", 6.5)
                canvas.drawCentredString(stamp_cx, stamp_cy + 0.04 * cm, _ar("منصة قطع الغيار PP"))

                try:
                    canvas.setFont(stamp_font, 6.4)
                except Exception:
                    canvas.setFont("Helvetica", 6.4)
                canvas.drawCentredString(stamp_cx, stamp_cy - 0.34 * cm, _ar("الخدمات المساندة GO"))
            except Exception as e:
                _swallow(e)

        canvas.restoreState()

    def _on_first(canvas, _doc):
        _draw_extras(canvas, _doc, draw_stamp=True)

    def _on_later(canvas, _doc):
        _draw_extras(canvas, _doc, draw_stamp=False)

    try:
        doc.build(story, onFirstPage=_on_first, onLaterPages=_on_later)
    except Exception as e:
        await _notify_invoice_error(context, order_id, f"إنشاء PDF ({kind_norm})", e)
        try:
            os.remove(pdf_path)
        except Exception as e:
            _swallow(e)
        return

    # Send PDF
    caption = f"📄 {inv_title}\nرقم الطلب: {order_id_disp}\nرقم الفاتورة: {inv_no_disp}"
    filename = f"PP_Invoice_{inv_no}.pdf"

    targets = []
    if admin_only:
        for aid in ADMIN_IDS:
            try:
                targets.append(int(aid))
            except Exception as e:
                _swallow(e)
    else:
        if client_id:
            targets.append(int(client_id))
        if include_admins:
            for aid in ADMIN_IDS:
                try:
                    targets.append(int(aid))
                except Exception as e:
                    _swallow(e)

    targets = [x for i, x in enumerate(targets) if x and x not in targets[:i]]

    failed = []
    sent_any = False

    for cid in targets:
        try:
            try:
                log_event("محاولة إرسال فاتورة PDF", order_id=order_id, target_chat_id=cid, filename=filename)
            except Exception as e:
                _swallow(e)

            with open(pdf_path, "rb") as f:
                await context.bot.send_document(
                    chat_id=cid,
                    document=f,
                    filename=filename,
                    caption=caption,
                    disable_content_type_detection=False,
                )

            sent_any = True
            try:
                log_event("تم إرسال فاتورة PDF بنجاح", order_id=order_id, target_chat_id=cid)
            except Exception as e:
                _swallow(e)

        except Exception as e:
            emsg = getattr(e, "message", None) or str(e)
            failed.append((cid, emsg))
            try:
                log_event("فشل إرسال فاتورة PDF", order_id=order_id, target_chat_id=cid, error=emsg)
            except Exception as e:
                _swallow(e)

    if failed:
        lines = []
        for cid, err in failed[:8]:
            lines.append(f"- chat_id={cid}: {err}")
        more = f"\n(+{len(failed)-8} أخطاء أخرى)" if len(failed) > 8 else ""
        await _notify_invoice_error(
            context,
            order_id,
            f"إرسال PDF ({kind_norm}){' - لم يُرسل لأي جهة' if not sent_any else ''}",
            "\n".join(lines) + more
        )
    else:
        try:
            update_order_fields(order_id, {sent_flag_field: "yes", legacy_flag: "yes"})
        except Exception as e:
            _swallow(e)

    try:
        os.remove(pdf_path)
    except Exception as e:
        _swallow(e)

def client_trader_chat_kb(order_id: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("💬 مراسلة التاجر", callback_data=f"pp_chat_trader|{order_id}")],
    ])

def chat_nav_kb(order_id: str, end_cb: str) -> InlineKeyboardMarkup:
    """كيبورد موحد داخل وضع المراسلة: سجل + لوحة + إنهاء.
    إذا لم يوجد order_id صالح (مثلاً بوابة التجار JOIN) يظهر زر الإنهاء فقط.
    """
    oid = ("" if order_id is None else str(order_id)).strip()
    rows = []
    if oid and oid.upper() != "JOIN":
        rows.append([
            InlineKeyboardButton("📜 سجل الطلب", callback_data=f"pp_order_legal|{oid}"),
            InlineKeyboardButton("🧾 لوحة الطلب", callback_data=f"pp_open_order|{oid}"),
        ])
    rows.append([InlineKeyboardButton("✖️ إنهاء المراسلة", callback_data=str(end_cb or "").strip() or "ui_close")])
    return InlineKeyboardMarkup(rows)

def chat_nav_kb_for(context: ContextTypes.DEFAULT_TYPE, actor_id: int, order_id: str, fallback_end_cb: str) -> InlineKeyboardMarkup:
    """يرجع chat_nav_kb مع اختيار زر الإنهاء المناسب حسب نوع جلسة المراسلة."""
    end_cb = (str(fallback_end_cb or "").strip() or "ui_close")
    try:
        sessions = context.bot_data.get("pp_chat_sessions") or {}
        sess = sessions.get(str(int(actor_id or 0)))
        if isinstance(sess, dict):
            so = (sess.get("order_id") or "").strip()
            if so and (not str(order_id or "").strip() or str(order_id).strip() == so):
                end_cb = f"pp_chat_end|{so}"
    except Exception:
        pass
    return chat_nav_kb(order_id, end_cb)




def notice_kb_for(
    context: ContextTypes.DEFAULT_TYPE,
    actor_id: int,
    order_id: str,
    include_chat_trader: bool = False,
    include_support: bool = True,
    fallback_end_cb: str = "ui_close",
) -> InlineKeyboardMarkup:
    """
    كيبورد إشعارات موحد.
    - في وضع الإشعار (fallback_end_cb = ui_close / pp_ui_close): زرّين فقط (سجل + مراسلة) بدون إنهاء وبدون لوحة.
    - في وضع المراسلة الفعلية (غير ذلك): يرجع السلوك الكامل (سجل + لوحة + مراسلة + إنهاء).
    """
    oid = ("" if order_id is None else str(order_id)).strip()
    rows: list[list[InlineKeyboardButton]] = []

    # تحديد هل هذا "إشعار" فقط
    end_cb = (fallback_end_cb or "").strip()
    is_notice_only = end_cb in ("", "ui_close", "pp_ui_close")

    # هل المرسل أدمن؟
    aid = 0
    try:
        aid = int(actor_id or 0)
    except Exception:
        aid = 0
    is_admin = str(aid) in set([str(x) for x in (ADMIN_IDS or [])])

    if not oid or oid.upper() == "JOIN":
        return InlineKeyboardMarkup(rows)

    # ===== وضع الإشعار فقط: زرّين عمليين بدون خلط =====
    if is_notice_only:
        # 1) سجل الطلب دائمًا
        rows.append([InlineKeyboardButton("📜 سجل الطلب", callback_data=f"pp_order_legal|{oid}")])

        # 2) زر المراسلة يظهر فقط لغير الأدمن
        if not is_admin:
            if include_chat_trader:
                rows.append([InlineKeyboardButton("💬 مراسلة التاجر", callback_data=f"pp_chat_trader|{oid}")])
            elif include_support:
                rows.append([InlineKeyboardButton("🔎 مراسلة المنصة", callback_data=f"pp_track|{oid}")])

        # بدون إنهاء / بدون لوحة
        return InlineKeyboardMarkup(rows)

    # ===== وضع المراسلة الفعلية: السلوك السابق =====
    rows.append([
        InlineKeyboardButton("📜 سجل الطلب", callback_data=f"pp_order_legal|{oid}"),
        InlineKeyboardButton("🧾 لوحة الطلب", callback_data=f"pp_open_order|{oid}"),
    ])

    if include_chat_trader:
        rows.append([InlineKeyboardButton("💬 مراسلة التاجر", callback_data=f"pp_chat_trader|{oid}")])

    # الأدمن لا يحتاج "مراسلة المنصة"
    if include_support and (not is_admin):
        rows.append([InlineKeyboardButton("🔎 مراسلة المنصة", callback_data=f"pp_track|{oid}")])

    # زر الإنهاء فقط في وضع المراسلة
    try:
        nav = chat_nav_kb_for(context, int(aid or 0), oid, end_cb)
        end_row = (nav.inline_keyboard or [])[-1] if nav else None
        if end_row:
            rows.append(end_row)
        else:
            rows.append([InlineKeyboardButton("✖️ إنهاء", callback_data=end_cb or "ui_close")])
    except Exception:
        rows.append([InlineKeyboardButton("✖️ إنهاء", callback_data=end_cb or "ui_close")])

    return InlineKeyboardMarkup(rows)

def client_trader_chat_done_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ انهاء المراسلة", callback_data="pp_chat_trader_done")],
    ])

def trader_reply_kb(order_id: str, user_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("💬 رد على العميل", callback_data=f"pp_trader_reply|{order_id}|{user_id}")],
    ])

def trader_reply_done_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ انهاء الرد", callback_data="pp_trader_reply_done")],
    ])

def team_group_kb(order_id: str, bot_username: str | None = None) -> InlineKeyboardMarkup:
    """Keyboard used inside TEAM group for the initial order post.

    Requirement: only allow starting a quote from the group.
    All quote details are collected in private to avoid clutter and to keep finance/details private.
    """
    # افضل تجربة: زر URL يفتح الخاص مباشرة بدون ما يبحث التاجر عن البوت.
    if bot_username:
        deeplink = f"https://t.me/{bot_username}?start=ppq_{order_id}"
        return InlineKeyboardMarkup([
            [InlineKeyboardButton("💰 تقديم عرض سعر ➜", url=deeplink)],
        ])

    # fallback (لو لم يتوفر اسم البوت)
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("💰 تقديم عرض سعر ➜", callback_data=f"pp_team_quote|{order_id}")],
    ])

def _norm(s: str) -> str:
    return (s or "").strip()

def _user_name(update_or_q) -> str:
    try:
        u = update_or_q.effective_user if hasattr(update_or_q, "effective_user") else update_or_q.from_user
        return (u.full_name or "عميلنا").strip()
    except Exception:
        return "عميلنا"


def _admin_public_name() -> str:
    """اسم الإدارة الظاهر للطرف الآخر (بدون كشف اسم الأدمن الحقيقي)."""
    try:
        v = (PP_SUPPORT_LABEL or "الإدارة").strip()
        return v if v else "الإدارة"
    except Exception:
        return "الإدارة"

def _order_parties(order_id: str) -> tuple[str, str]:
    """يرجع (اسم العميل الحقيقي, اسم التاجر الحقيقي) من بيانات الطلب/الملف."""
    oid = (order_id or "").strip()
    if not oid:
        return "—", "—"

    try:
        b = get_order_bundle(oid) or {}
        o = b.get("order", {}) or {}
    except Exception:
        o = {}

    # client
    client_name = (str(o.get("user_name") or "").strip()) or "—"

    # trader
    trader_name = (str(o.get("accepted_trader_name") or o.get("quoted_trader_name") or "").strip())
    if not trader_name:
        try:
            tid = int(o.get("accepted_trader_id") or 0)
        except Exception:
            tid = 0
        if tid:
            try:
                tp = get_trader_profile(int(tid)) or {}
                trader_name = (tp.get("display_name") or "").strip() or (tp.get("company_name") or "").strip()
            except Exception:
                trader_name = ""
    trader_name = trader_name or "—"

    return client_name, trader_name

def _order_tag_plain(order_id: str) -> str:
    cn, tn = _order_parties(order_id)
    return f"🧾 رقم الطلب: {order_id} | 👤 العميل: {cn} | 🧑‍🔧 التاجر: {tn}"

def _order_tag_html(order_id: str) -> str:
    cn, tn = _order_parties(order_id)
    return (
        f"🧾 رقم الطلب: {html.escape(str(order_id))} | "
        f"👤 العميل: <b>{html.escape(str(cn))}</b> | "
        f"🧑‍🔧 التاجر: <b>{html.escape(str(tn))}</b>"
    )

def _looks_like_vin(s: str) -> bool:
    s = _norm(s).replace(" ", "").upper()
    return bool(VIN_RE.match(s))

def _sanitize_delivery_details(details: str, hide_phone: bool = True) -> str:
    d = (details or "").strip()
    if not hide_phone:
        return d
    # remove any line that contains phone/contact
    lines = []
    for ln in d.splitlines():
        if "رقم الاتصال" in ln or "الجوال" in ln or "الهاتف" in ln:
            continue
        lines.append(ln)
    return "\n".join(lines).strip()

def _save_order_once(ud: dict):
    if ud.get("order_saved"):
        return
    add_order({
        "order_id": ud.get("order_id",""),
        "user_id": ud.get("user_id",0),
        "user_name": ud.get("user_name",""),
        "car_name": ud.get("car_name",""),
        "car_model": ud.get("car_model",""),
        "vin": ud.get("vin",""),
        "notes": ud.get("notes",""),
        "items_count": len(ud.get("items",[])),
        "price_sar": ud.get("price_sar",0),
        "status": "payment_pending",
        "payment_method": ud.get("payment_method",""),
        "payment_status": "pending",
        "receipt_file_id": "",
        "payment_confirmed_at_utc": "",
        "delivery_choice": "",
        "delivery_details": "",
        "created_at_utc": ud.get("created_at_utc", utc_now_iso()),
    })
    add_items(ud.get("order_id",""), _items_for_excel(ud.get("items",[])))
    ud["order_saved"] = True

def _items_for_excel(items: list[dict]) -> list[dict]:
    out = []
    for it in items or []:
        name = it.get("name","")
        part_no = it.get("part_no","") or ""  # ✅ جديد
        photo = it.get("photo_file_id","") or it.get("file_id","") or ""
        out.append({
            "name": name,
            "part_no": part_no,  # ✅ جديد (يروح لعمود item_part_no)
            "photo_file_id": photo,
            "created_at_utc": it.get("created_at_utc", utc_now_iso()),
        })
    return out

def _pay_method_ar(method: str) -> str:
    m = (method or "").strip().lower()
    return {
        "bank_transfer": "🏦 تحويل بنكي",
        "stc_pay": "📱 STC Pay",
        "pay_link": "🔗 رابط دفع",
        "free": "🆓 مجاني",
    }.get(m, method or "—")

async def send_trader_subscription_invoice_pdf(
    context: ContextTypes.DEFAULT_TYPE,
    trader_id: int,
    month: str,
    amount_sar: int = 99,
):
    """Generate and send a simple 1-page PDF invoice for trader subscription."""
    try:
        from reportlab.pdfgen import canvas as rl_canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib import colors
    except Exception:
        return

    # Arabic shaping (best effort)
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
        def _ar(s: str) -> str:
            s = "" if s is None else str(s)
            if not s:
                return s
            try:
                return get_display(arabic_reshaper.reshape(s))
            except Exception:
                return s
    except Exception:
        def _ar(s: str) -> str:
            return "" if s is None else str(s)

    font_name = "Helvetica"
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        if os.path.exists("Amiri-Regular.ttf"):
            pdfmetrics.registerFont(TTFont("Amiri", "Amiri-Regular.ttf"))
            font_name = "Amiri"
    except Exception as e:
        _swallow(e)

    month = str(month or "").strip() or month_key_utc()
    amount_sar = int(float(amount_sar or 99))

    tmp = f"sub_invoice_{trader_id}_{month.replace('-', '')}.pdf"
    path = os.path.join("/tmp", tmp)

    c = rl_canvas.Canvas(path, pagesize=A4)
    w, h = A4

    c.setFillColor(colors.HexColor("#0B3D91"))
    c.rect(0, h-35*mm, w, 35*mm, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont(font_name, 16)
    c.drawRightString(w-15*mm, h-18*mm, _ar("فاتورة اشتراك منصة"))
    c.setFont(font_name, 10)
    c.drawRightString(w-15*mm, h-26*mm, _ar(f"الشهر: {month}"))

    c.setFillColor(colors.whitesmoke)
    c.rect(15*mm, h-140*mm, w-30*mm, 90*mm, fill=1, stroke=0)

    c.setFillColor(colors.black)
    c.setFont(font_name, 12)
    c.drawRightString(w-20*mm, h-70*mm, _ar("البند: رسوم اشتراك منصة"))
    c.drawRightString(w-20*mm, h-85*mm, _ar(f"المبلغ: {amount_sar} ريال"))
    c.drawRightString(w-20*mm, h-100*mm, _ar(f"المرجع: SUB-{trader_id}-{month}"))

    c.setFont(font_name, 9)
    c.setFillColor(colors.gray)
    c.drawString(15*mm, 15*mm, "PP / GO - Platform Subscription Invoice")

    c.showPage()
    c.save()

    caption = f"🧾 فاتورة اشتراك منصة — {month} — {amount_sar} ريال"
    try:
        with open(path, "rb") as f:
            await context.bot.send_document(chat_id=int(trader_id), document=f, caption=caption)
    except Exception as e:
        _swallow(e)

    for aid in ADMIN_IDS:
        try:
            with open(path, "rb") as f:
                await context.bot.send_document(chat_id=int(aid), document=f, caption=f"(نسخة) {caption} — trader_id {trader_id}")
        except Exception as e:
            _swallow(e)

    try:
        os.remove(path)
    except Exception as e:
        _swallow(e)

async def _send_client_payment_preview(
    context: ContextTypes.DEFAULT_TYPE,
    client_id: int,
    order_id: str,
    pay_scope: str = "platform",  # platform / goods
) -> None:
    if not client_id or not order_id:
        return

    try:
        b = get_order_bundle(order_id)
        order = b.get("order", {}) or {}
        items = b.get("items", []) or []
    except Exception:
        order = {}
        items = []

    # ---------- helpers ----------
    def _s(x: object) -> str:
        return ("" if x is None else str(x)).strip()

    def _money_num_str(x: object) -> str:
        s0 = _s(x)
        if not s0:
            return ""
        try:
            f = float(str(s0).replace(",", "").strip())
            if abs(f - int(f)) < 1e-9:
                return str(int(f))
            return f"{f:.2f}".rstrip("0").rstrip(".")
        except Exception:
            return str(s0).strip()

    def _to_float(x: object) -> float:
        s0 = _s(x)
        if not s0:
            return 0.0
        try:
            return float(str(s0).replace(",", "").strip())
        except Exception:
            return 0.0

    def _money(x: object) -> str:
        s = _money_num_str(x)
        if not s or s in ("0", "0.0", "0.00"):
            return "—"
        return f"{html.escape(s)} ﷼"

    def _load_item_prices_map_from_order_local(o: dict) -> dict:
        pm = {}
        try:
            import json as _json
        except Exception:
            _json = None
        try:
            import ast
        except Exception:
            ast = None

        for k in ("quote_item_prices", "item_prices", "goods_item_prices", "quote_items_prices"):
            raw = o.get(k)
            if raw is None or raw == "":
                continue

            src = None
            try:
                if isinstance(raw, dict):
                    src = raw
                else:
                    sraw = str(raw).strip()
                    if sraw.startswith("{") and sraw.endswith("}"):
                        if _json:
                            try:
                                src = _json.loads(sraw)
                            except Exception:
                                src = None
                        if src is None and ast:
                            try:
                                src = ast.literal_eval(sraw)
                            except Exception:
                                src = None
            except Exception:
                src = None

            if isinstance(src, dict):
                for kk, vv in src.items():
                    ks = str(kk).strip()
                    vs = str(vv).strip()
                    if ks.isdigit() and vs and vs not in ("0", "0.0", "0.00"):
                        pm[ks] = vs

        return pm

    # ---------- data ----------
    car = _s(order.get("car_name"))
    model = _s(order.get("car_model"))
    vin = _s(order.get("vin"))

    ship_method = _s(order.get("ship_method"))
    delivery_details = _s(order.get("delivery_details"))

    ship_inc_txt = _s(order.get("ship_included"))
    ship_eta_txt = _s(order.get("ship_eta")) or ""
    availability_txt = _s(order.get("availability_days")) or ""

    trader_name = _s(order.get("accepted_trader_name") or order.get("quoted_trader_name"))
    trader_company = ""
    try:
        trader_id0 = int(order.get("accepted_trader_id") or order.get("quoted_trader_id") or 0)
    except Exception:
        trader_id0 = 0

    if trader_id0:
        try:
            tp = get_trader_profile(int(trader_id0)) or {}
            if not trader_name:
                trader_name = _s(tp.get("display_name")) or _s(tp.get("company_name"))
            trader_company = _s(tp.get("company_name"))
        except Exception:
            trader_company = ""

    trader_name = trader_name or ""
    trader_company = trader_company or ""

    trader_header = ""
    if pay_scope == "goods" and (trader_name or trader_company):
        if trader_name and trader_company and trader_company != trader_name:
            trader_header = f"🧑‍💼 <b>{html.escape(trader_name)}</b> — 🏪 <b>{html.escape(trader_company)}</b>\n"
        elif trader_name:
            trader_header = f"🧑‍💼 <b>{html.escape(trader_name)}</b>\n"
        else:
            trader_header = f"🏪 <b>{html.escape(trader_company)}</b>\n"

    # ✅ استخراج رسوم الشحن من أكثر من حقل (حسب مساراتك المختلفة)
    ship_included_norm = str(order.get("ship_included") or "").strip().lower()
    raw_shipping_fee = (
        order.get("shipping_fee_sar")
        or order.get("quote_shipping_fee")
        or order.get("shipping_fee")
        or order.get("ship_fee")
        or ""
    )

    missing_ship = (raw_shipping_fee is None or str(raw_shipping_fee).strip() == "")

    # ✅ لو الشحن "مشمولة" نثبتها 0، ولو غير مشمولة وبدون قيمة نخليها غير محددة
    if missing_ship:
        raw_shipping_fee = 0 if ship_included_norm in ("yes", "true", "1", "included", "مشمولة") else ""

    ship_fee_val = _to_float(raw_shipping_fee)

    # ✅ نص الشحن (يظهر فقط إذا محدد فعلاً أو مشمول)
    if missing_ship and ship_included_norm not in ("yes", "true", "1", "included", "مشمولة"):
        ship_fee_total_txt = "—"
        ship_fee_txt = "—"
    else:
        ship_fee_total_txt = _money(raw_shipping_fee)
        ship_fee_txt = ship_fee_total_txt

    # ---------- scope ----------
    if pay_scope == "goods":
        amount = order.get("goods_amount_sar") or ""
        method = order.get("goods_payment_method") or ""
        title = "📦 تم استلام إيصال قيمة القطع"
        status_line = "⏳ جاري التحقق من الإيصال — ثم تجهيز الشحن"
        next_line = "بعد التحقق سيتم اعتماد الدفع والتجهيز لشحن القطع."
        icon_scope = "🧩"
    else:
        amount = order.get("price_sar") or ""
        method = order.get("payment_method") or ""
        title = "🧾 تم استلام إيصال رسوم المنصة"
        status_line = "⏳ جاري التحقق من الإيصال"
        next_line = "بعد التحقق سيتم متابعة طلبك واستكمال الإجراء."
        icon_scope = "🧾"

    amt_txt = _money(amount)
    method_txt = html.escape(_pay_method_ar(_s(method)))

    # ✅ الإجمالي الصحيح: (قيمة القطع + الشحن) عندما تكون قيمة الشحن محددة/مشمولة
    if pay_scope == "goods":
        inv_total_val = _to_float(amount) + ship_fee_val
        inv_total_txt = _money(inv_total_val)
    else:
        inv_total_txt = amt_txt

    # ---------- parts list ----------
    priced_count = 0
    unpriced_count = 0
    pm = _load_item_prices_map_from_order_local(order) if pay_scope == "goods" else {}

    parts_lines = []
    max_items = 14

    if isinstance(items, list) and items:
        for i, it in enumerate(items, start=1):
            if not isinstance(it, dict):
                continue

            nm = _s(it.get("name") or it.get("item_name"))
            pn = _s(it.get("part_no") or it.get("item_part_no"))

            if not nm:
                continue

            base = f"• <b>{html.escape(nm)}</b>"
            if pn:
                base += f"  <i>({html.escape(pn)})</i>"

            if pay_scope == "goods":
                pr = _s(pm.get(str(i)))
                if pr:
                    priced_count += 1
                    base += f"\n   ✅ السعر: 💰 <b>{_money(pr)}</b>"
                else:
                    unpriced_count += 1
                    base += "\n   ⚠️ <b>غير مسعّرة</b>"

            parts_lines.append(base)
            if len(parts_lines) >= max_items:
                break

    parts_txt = "\n".join(parts_lines) if parts_lines else "• —"

    car_line = "—"
    if car or model:
        car_line = html.escape((car + " " + model).strip())

    delivery_block = ""
    dd = _sanitize_delivery_details(delivery_details, hide_phone=True) if delivery_details else ""
    if ship_method or dd:
        delivery_block += "\n\n📦 <b>التسليم</b>\n"
        if ship_method:
            delivery_block += f"• 🛻 <b>{html.escape(ship_method)}</b>\n"
        if ship_inc_txt:
            delivery_block += f"• 📌 الشحن: <b>{html.escape(ship_inc_txt)}</b>\n"
        if ship_fee_txt and ship_fee_txt != "—":
            delivery_block += f"• 💵 قيمة الشحن: <b>{ship_fee_txt}</b>\n"
        if ship_eta_txt:
            delivery_block += f"• ⏱ مدة الشحن: <b>{html.escape(ship_eta_txt)}</b>\n"
        if availability_txt:
            delivery_block += f"• 🛠 مدة التجهيز: <b>{html.escape(availability_txt)}</b>\n"
        if dd:
            delivery_block += f"\n📍 <b>تفاصيل</b>:\n<pre>{html.escape(dd)}</pre>"

    parts_summary = ""
    if pay_scope == "goods" and isinstance(items, list) and items:
        total_items = len([x for x in items if isinstance(x, dict) and _s(x.get("name") or x.get("item_name"))])
        parts_summary = (
            "━━━━━━━━━━━━\n"
            + f"📌 <b>ملخص القطع</b>\n"
            + f"✅ مسعّرة: <b>{priced_count}</b>  |  ⚠️ غير مسعّرة: <b>{unpriced_count}</b>  |  📦 الإجمالي: <b>{total_items}</b>\n"
            + "━━━━━━━━━━━━\n"
        )

    # ✅ سطر الشحن يُعرض فقط عندما يكون محدد/مشمولة
    ship_line = ""
    if pay_scope == "goods" and ship_fee_total_txt and ship_fee_total_txt != "—":
        ship_line = f"🚚 <b>قيمة الشحن</b>: <b>{ship_fee_total_txt}</b>\n"

    msg = (
        f"✅ <b>{html.escape(title)}</b>\n"
        f"<i>{html.escape(status_line)}</i>\n\n"
        f"{trader_header}"
        f"🧾 <b>رقم الطلب</b>: {html.escape(order_id)}\n"
        f"💳 <b>طريقة الدفع</b>: <b>{method_txt}</b>\n"
        f"🧾 <b>الإجمالي</b>: <b>{inv_total_txt}</b>\n"
        + (f"💰 <b>قيمة القطع</b>: <b>{amt_txt}</b>\n" if pay_scope == "goods" else f"💰 <b>رسوم المنصة</b>: <b>{amt_txt}</b>\n")
        + ship_line
        + (f"\n{parts_summary}" if parts_summary else "\n━━━━━━━━━━━━\n")
        + (
            f"🚗 <b>بيانات السيارة</b>\n"
            f"• {car_line}\n"
            + (f"• 🔎 VIN: <code>{html.escape(vin)}</code>\n" if vin else "")
        )
        + "\n"
        f"{icon_scope} <b>تفاصيل القطع</b>\n"
        f"{parts_txt}\n"
        f"\n🟦 <b>ماذا الآن؟</b>\n"
        f"• {html.escape(next_line)}"
        + delivery_block
    )

    kb_client = track_kb(order_id) if pay_scope != "goods" else client_trader_chat_kb(order_id)

    # ✅ إرسال للعميل فقط (بدون تكرار للتاجر)
    try:
        cid_int = int(client_id)
    except Exception:
        cid_int = 0

    if cid_int:
        try:
            await context.bot.send_message(
                chat_id=cid_int,
                text=msg,
                parse_mode="HTML",
                reply_markup=kb_client,
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)

    # ❌ تم إلغاء إرسال نفس الرسالة للتاجر لأنها تخص العميل فقط
    return
def _delivery_brief(order: dict, ud: dict) -> str:
    ship_method = (str(order.get("ship_method") or "")).strip() or (str(ud.get("ship_method") or "")).strip()
    ship_city = (str(order.get("ship_city") or "")).strip() or (str(ud.get("ship_city") or "")).strip()
    pickup_city = (str(order.get("pickup_city") or "")).strip() or (str(ud.get("pickup_city") or "")).strip()
    pickup_loc = (str(order.get("pickup_location") or "")).strip() or (str(ud.get("pickup_location") or "")).strip()

    d_choice = (str(order.get("delivery_choice") or ud.get("delivery_choice") or "")).strip().lower()
    d_details = (str(order.get("delivery_details") or ud.get("delivery_details") or "")).strip()

    if not ship_method:
        if d_choice == "ship" or "شحن" in d_details:
            ship_method = "شحن"
        elif d_choice == "pickup" or "استلام" in d_details:
            ship_method = "استلام من الموقع"

    if ship_method and not ship_city and ("شحن" in ship_method or d_choice == "ship"):
        m = re.search(r"المدينة\s*:\s*([^\n\r]+)", d_details)
        if m:
            ship_city = (m.group(1) or "").strip()

    if ship_method and ("استلام" in ship_method or d_choice == "pickup"):
        if not pickup_city and d_details:
            m = re.search(r"مدينة\s*الاستلام\s*:\s*([^\n\r]+)", d_details)
            if m:
                pickup_city = (m.group(1) or "").strip()

    if not ship_method and not ship_city and not pickup_city:
        return "<i>غير محدد بعد</i>"

    parts = []
    if ship_method:
        parts.append(f"<b>طريقة التسليم</b>: <i>{html.escape(ship_method)}</i>")
    if ship_city:
        parts.append(f"<b>مدينة التسليم</b>: <i>{html.escape(ship_city)}</i>")
    if pickup_city:
        parts.append(f"<b>مدينة الاستلام</b>: <i>{html.escape(pickup_city)}</i>")
    if pickup_loc and ship_method and "استلام" in ship_method:
        parts.append(f"<b>موقع الاستلام</b>: <i>{html.escape(pickup_loc)}</i>")
    return "\n".join(parts)
    
async def _alert(q, text: str | None = None, force: bool = False):
    """
    🔒 تنبيه مقيّد:
    - لا يظهر أي Popup إلا إذا كان الزر مقفل/مجمّد فعلاً
    - النص الفارغ => فقط إيقاف التحميل (ACK)
    - Popup يظهر فقط عند force=True أو نص منع صريح
    """

    def _user_name(q) -> str:
        try:
            fu = getattr(q, "from_user", None)
            n = (getattr(fu, "first_name", "") or "").strip()
            if n:
                return n
            u = (getattr(fu, "username", "") or "").strip()
            if u:
                return u
        except Exception as e:
            _swallow(e)
        return "عزيزي المستخدم"

    try:
        # 1) لا نص + لا force => لا Popup (زر طبيعي)
        if (text is None or not str(text).strip()) and not force:
            await q.answer()  # فقط إيقاف التحميل
            return

        s = (str(text).strip() if text else "").strip()

        # 2) force=True بدون نص => نص افتراضي واضح
        if force and not s:
            name = _user_name(q)
            s = (
                f"{name} 👋\n\n"
                "⛔ هذا الإجراء غير متاح حاليًا.\n"
                "🔒 الزر مقفل حسب حالة الطلب الحالية.\n\n"
                "يرجى متابعة الخطوات المتاحة فقط."
            )
            await q.answer(text=s, show_alert=True)
            return

        # 3) نص صريح + force => Popup
        if force:
            await q.answer(text=s, show_alert=True)
            return

        # 4) نص عادي بدون force => Toast خفيف فقط
        await q.answer(text=s, show_alert=False)

    except Exception:
        # fallback آمن
        try:
            await q.answer()
        except Exception as e:
            _swallow(e)

def _nice_person_name_from_q(q, fallback: str = "عزيزي المستخدم") -> str:
    """
    يرجّع اسم لطيف للشخص من callback_query:
    - يحاول first_name ثم full_name
    - وإذا ما فيه يرجع fallback
    """
    try:
        u = getattr(q, "from_user", None)
        if not u:
            return fallback
        fn = (getattr(u, "first_name", "") or "").strip()
        full = (getattr(u, "full_name", "") or "").strip()
        return fn or full or fallback
    except Exception:
        return fallback

def _nice_greeting(role: str = "", name: str = "") -> str:
    """
    يبني سطر ترحيب لطيف حسب الدور:
    role: trader / client / admin / user
    """
    role = (role or "").strip().lower()
    if role == "trader":
        base = "عزيزي التاجر"
    elif role == "client":
        base = "عزيزي العميل"
    elif role == "admin":
        base = "عزيزي المسؤول"
    else:
        base = "عزيزي المستخدم"

    nm = (name or "").strip()
    return f"{base} {nm}".strip()

async def alert_nice(q, body: str, role: str = "user", force: bool = False, name: str = ""):
    """
    Wrapper: يضيف سطر لطيف + الاسم ثم يستدعي _alert
    """
    nm = name.strip() if isinstance(name, str) else ""
    if not nm:
        nm = _nice_person_name_from_q(q, fallback="")
    head = _nice_greeting(role=role, name=nm)

    msg = (body or "").strip()
    if not msg:
        # لو ما فيه نص، نخليها مجرد answer
        return await _alert(q, "", force=force)

    # إذا النص أصلاً يبدأ باسم/تحية لا نكرر
    if msg.startswith("عزيزي") or msg.startswith("عزيزتي"):
        return await _alert(q, msg, force=force)

    return await _alert(q, f"{head}\n{msg}", force=force)

async def _need_complete_trader_profile_notice(context: ContextTypes.DEFAULT_TYPE, user_id: int, name: str, order_id: str = ""):
    bot_username = getattr(context.bot, "username", "") or ""
    # ديب لينك آمن على نفس العضو (مقفل على user_id)
    panel_url = f"https://t.me/{bot_username}?start=trader_{int(user_id)}"

    extra = f"\n\n🧾 رقم الطلب: {order_id}" if order_id else ""

    try:
        tp = get_trader_profile(int(user_id or 0)) or {}
    except Exception:
        tp = {}

    # ✅ تعريف موحّد للحقول المطلوبة + أمثلة واضحة
    schema = [
        ("display_name", "اسم التاجر", "مثال: أبو ثامر"),
        ("company_name", "اسم المتجر", "مثال: قطع غيار الرياض"),
        ("shop_phone", "رقم اتصال المتجر", "مثال: 05xxxxxxxx (10 أرقام)"),
        ("cr_no", "رقم السجل التجاري", "مثال: 1010xxxxxx (أرقام فقط)"),
        ("vat_no", "الرقم الضريبي", "مثال: 15 رقم"),
        ("bank_name", "اسم البنك", "مثال: الأهلي / الراجحي"),
        ("iban", "رقم IBAN", "مثال: SAxxxxxxxxxxxxxxxxxxxxxx (24 خانة)"),
        ("stc_pay", "رقم STC Pay", "مثال: 05xxxxxxxx"),
    ]

    missing = []
    for k, label, ex in schema:
        v = (tp.get(k) or "").strip()
        if not v:
            missing.append((label, ex))

    if not missing:
        return

    miss_lines = []
    for i, (label, ex) in enumerate(missing, start=1):
        miss_lines.append(f"{i}) {label}\n   {ex}")

    msg = (
        f"{name}\n"
        "⛔ لا يمكنك تقديم عرض سعر قبل إكمال ملف التاجر بالكامل.\n\n"
        "🔻 الحقول الناقصة لديك الآن:\n"
        + "\n".join(miss_lines)
        + "\n\n"
        "افتح لوحة التاجر من الزر بالأسفل وأكمل البيانات، ثم ارجع وابدأ عرض السعر."
        + extra
    )

    kb = InlineKeyboardMarkup([[InlineKeyboardButton("🧑‍💼 فتح لوحة التاجر", url=panel_url)]])

    try:
        await context.bot.send_message(
            chat_id=int(user_id),
            text=msg,
            reply_markup=kb,
            disable_web_page_preview=True,
        )
    except Exception as e:
        _swallow(e)

async def _deny_disabled_trader_msg(update: Update, reason: str = "حساب التاجر موقوف"):
    try:
        if update and update.message:
            await update.message.reply_text(f"⛔ {reason}")
    except Exception as e:
        _swallow(e)

async def on_error(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    # 🟥 لوق عربي واضح مع تتبع كامل
    try:
        log.exception("❌ خطأ غير معالج داخل البوت: %s", context.error)
    except Exception as e:
        _swallow(e)

    # 🟧 تسجيل سياق الخطأ (من المستخدم / نوع الشات)
    try:
        if isinstance(update, Update):
            uid = getattr(update.effective_user, "id", None)
            uname = getattr(update.effective_user, "full_name", "") if update.effective_user else ""
            chat_id = getattr(update.effective_chat, "id", None)
            chat_type = getattr(update.effective_chat, "type", "") if update.effective_chat else ""

            log_event(
                "تفاصيل سياق الخطأ",
                user_id=uid,
                user_name=uname,
                chat_id=chat_id,
                chat_type=chat_type,
                error=str(context.error),
            )
    except Exception as e:
        _swallow(e)

    try:
        # إذا الخطأ جاء من CallbackQuery → تنبيه مربع فقط
        if isinstance(update, Update) and update.callback_query:
            await update.callback_query.answer(
                "⚠️ حدث خطأ غير متوقع\nيرجى المحاولة مرة أخرى",
                show_alert=True
            )
            return

        # ❌ لا نرسل أي رسالة نصية للشات
    except Exception as e:
        _swallow(e)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    ensure_workbook()

    # Deep-link args
    try:
        args = getattr(context, "args", []) or []
    except Exception:
        args = []


    if args and isinstance(args[0], str) and args[0].strip().lower() == "join":
        await join_portal_start(update, context)
        return

    # =========================
    # (1) ppq_ : لوحة عرض السعر للتاجر
    # =========================
    if args and isinstance(args[0], str) and args[0].startswith("ppq_"):
        order_id = args[0][4:].strip()

        td = context.user_data.setdefault(update.effective_user.id, {})
        td["quote_order_id"] = order_id
        td["quote_step"] = "start"
        set_stage(context, update.effective_user.id, STAGE_TRADER_SET_QUOTE)

        # ملخص سريع للطلب
        order_snapshot = ""
        try:
            b = get_order_bundle(order_id)
            order = b.get("order", {}) or {}
            items = b.get("items", []) or []

            # طريقة الشحن + المدينة
            ship_method = (order.get("quote_ship_method") or td.get("quote_ship_method") or order.get("ship_method") or order.get("shipping_method") or "").strip()
            delivery_details = (order.get("delivery_details") or order.get("address") or "").strip()

            city = (order.get("city") or "").strip()
            if not city and delivery_details:
                try:
                    for ln in delivery_details.splitlines():
                        ln2 = (ln or "").strip()
                        if ln2.startswith("المدينة"):
                            if ":" in ln2:
                                city = ln2.split(":", 1)[1].strip()
                            elif "-" in ln2:
                                city = ln2.split("-", 1)[1].strip()
                            break
                except Exception as e:
                    _swallow(e)

            lines = []
            for i, it in enumerate(items, start=1):
                nm = (it.get("name") or "").strip()
                if nm:
                    lines.append(f"{i}- {nm}")

            parts_txt = "\n".join(lines) if lines else "لا يوجد"

            order_snapshot = (
                "📌 ملخص الطلب\n"
                f"رقم الطلب: {order_id}\n"
                f"السيارة: {order.get('car_name','')}\n"
                f"الموديل: {order.get('car_model','')}\n"
                f"VIN: {order.get('vin','')}\n"
                f"طريقة الشحن: {ship_method or 'غير محدد'}\n"
                f"المدينة: {city or 'غير محددة'}\n"
                f"الملاحظات: {order.get('notes','') or 'لا يوجد'}\n\n"
                f"القطع:\n{parts_txt}"
            )
        except Exception:
            order_snapshot = f"رقم الطلب: {order_id}"

        await update.message.reply_text(
            f"{_user_name(update)}\n"
            "✨ اهلا بك في لوحة عرض السعر\n\n"
            "هذه الخطوات مصممة لتبني عرض منسق واحترافي\n\n"
            f"{order_snapshot}\n\n"
            "اضغط زر البدء بالاسفل ثم اتبع الخطوات خطوة بخطوة",
            reply_markup=trader_quote_start_kb(order_id),
            parse_mode="HTML",
        )
        return
    
    # =========================
    # (2) trader_ : فتح لوحة التاجر من زر الترحيب (deeplink)
    # =========================
    if args and isinstance(args[0], str) and args[0].startswith("trader_"):
        try:
            await start_trader_deeplink(update, context)
        except Exception:
            try:
                await update.message.reply_text(f"{_user_name(update)}\nتعذر فتح لوحة التاجر حاليا")
            except Exception as e:
                _swallow(e)
        return

    # =========================
    # (2) ppopen_ : فتح لوحة الطلب
    # =========================
    if args and isinstance(args[0], str) and args[0].startswith("ppopen_"):
        order_id = args[0][7:].strip()

        try:
            b = get_order_bundle(order_id)
            order = b.get("order", {}) or {}
        except Exception:
            order = {}

        try:
            acc = int(order.get("accepted_trader_id") or 0)
        except Exception:
            acc = 0

        qs = str(order.get("quote_status") or "").strip().lower()
        locked = str(order.get("quote_locked") or "").strip().lower() == "yes"

        if locked or qs == "accepted":
            tid = acc
        else:
            try:
                qid = int(order.get("quoted_trader_id") or 0)
            except Exception:
                qid = 0
            tid = acc or qid

        if not tid:
            await update.message.reply_text("🔒 لم يتم إسناد الطلب لتاجر بعد")
            return

        actor_id = update.effective_user.id

        accepted_name = (order.get("accepted_trader_name") or order.get("quoted_trader_name") or "").strip()
        if not accepted_name:
            try:
                tp = get_trader_profile(int(tid)) or {}
                accepted_name = (tp.get("display_name") or "").strip()
            except Exception:
                accepted_name = ""

        who = accepted_name or "التاجر المستلم"

        if tid != actor_id and actor_id not in ADMIN_IDS:
            await update.message.reply_text(
                "🔒 هذه اللوحة مخصصة لتاجر محدد\n"
                f"🧾 رقم الطلب: {order_id}\n"
                f"👤 التاجر: {who}\n\n"
                "✅ إذا كنت أنت التاجر المستلم افتح المنصة من نفس الحساب الذي استلم الطلب"
            )
            return

        try:
            if acc and actor_id == acc:
                notified = str(order.get("accepted_trader_notified") or "").strip().lower() == "yes"
                if not notified:
                    await context.bot.send_message(
                        chat_id=acc,
                        text=(
                            "✅ تم قبول عرض السعر من العميل\n"
                            f"🧾 رقم الطلب: {order_id}\n"
                            "🧰 ابدأ تجهيز الطلب ثم حدّث الحالة من لوحة التحكم"
                        ),
                        reply_markup=trader_status_kb(order_id),
                        disable_web_page_preview=True,
                    )
                    try:
                        update_order_fields(order_id, {"accepted_trader_notified": "yes"})
                    except Exception as e:
                        _swallow(e)
        except Exception as e:
            _swallow(e)

        await update.message.reply_text(
            f"🧰 لوحة التحكم للطلب\n"
            f"🧾 رقم الطلب: {order_id}\n"
            f"👤 التاجر: {who}",
            reply_markup=trader_status_kb(order_id),
            disable_web_page_preview=True,
        )
        return

    # =========================
    # (3) Start normal
    # =========================
    name = _user_name(update)
    await update.message.reply_text(
        f"<i>اهلا {name}</i>\n\n"
        "<b>✨ مرحبا بك في PP</b>\n\n"
        "<i>"
        "تجربة احترافية صممت بعناية للبحث الدقيق عن قطع سيارتك\n"
        "وتقديم تسعيرة واضحة وموثوقة قبل اتخاذ القرار\n"
        "</i>\n\n"
        "<b>🔍 ماذا يميز هذه الخدمة؟</b>\n"
        "<i>"
        "تحليل دقيق لبيانات سيارتك\n"
        "تحقق كامل من التوافق والتوفر\n"
        "وتسعيرة مبنية على واقع السوق بكل شفافية\n"
        "</i>\n\n"
        "<b>📋 للبدء نحتاج فقط:</b>\n"
        "• <i>اسم السيارة</i>\n"
        "• <i>الموديل (سنة من 4 ارقام)</i>\n"
        "• <i>رقم الهيكل VIN من 17 خانة</i>\n\n"
        "<b>🤝 هدفنا</b>\n"
        "<i>"
        "ان تصل الى القطعة والتسعيرة من خلال شركات السيارات او وكلاء محليين / عالميين باسرع وقت وتلقي عروض مختلفة\n"
        "</i>\n\n"
        "<b>⬇️ اضغط الزر بالاسفل لبدء طلب قطع غيار</b>",
        parse_mode="HTML",
        reply_markup=main_menu_kb(),
    )

async def start_new_order_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")

    # يبدأ في الخاص فقط
    try:
        if q.message and q.message.chat and q.message.chat.type != ChatType.PRIVATE:
            await _alert(q, "البدء من الخاص فقط", force=True)
            return
    except Exception:
        pass

    await begin_flow(update, context)

async def start_trader_deeplink(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    name = _user_name(update)

    # لازم يكون في الخاص فقط
    if update.effective_chat and update.effective_chat.type != ChatType.PRIVATE:
        try:
            await update.message.reply_text("ℹ️ هذا الخيار يعمل في الخاص فقط")
        except Exception as e:
            _swallow(e)
        return

    # ✅ قفل الرابط على نفس العضو (anti-abuse)
    try:
        args = getattr(context, "args", []) or []
    except Exception:
        args = []

    token = (args[0] if args else "") or ""
    allowed_id = 0
    try:
        allowed_id = int(token.split("_", 1)[1])
    except Exception:
        allowed_id = 0

    if allowed_id and int(allowed_id) != int(user_id):
        try:
            await update.message.reply_text(f"{name}\n⛔ هذا اللوحة لتاجر اخر  ليس لك")
        except Exception as e:
            _swallow(e)
        return

    # تحقق العضوية
    is_admin = user_id in ADMIN_IDS
    is_member = False
    try:
        is_member = await _is_trader_group_member(context, user_id)
    except Exception:
        is_member = False

    if not is_member and not is_admin:
        try:
            await update.message.reply_text(f"{name}\nغير مصرح")
        except Exception as e:
            _swallow(e)
        return

    # فحص ملف التاجر
    tp = {}
    try:
        tp = get_trader_profile(int(user_id or 0)) or {}
    except Exception:
        tp = {}

    # ✅ فحص كامل لملف التاجر (نفس حقول شرط تقديم العرض)
    schema = [
        ("display_name", "اسم التاجر"),
        ("company_name", "اسم المتجر"),
        ("shop_phone", "رقم اتصال المتجر"),
        ("cr_no", "رقم السجل التجاري"),
        ("vat_no", "الرقم الضريبي"),
        ("bank_name", "اسم البنك"),
        ("iban", "رقم IBAN"),
        ("stc_pay", "رقم STC Pay"),
    ]
    missing = []
    for k, label in schema:
        if not (str(tp.get(k) or "").strip()):
            missing.append(label)

    # ✅ رسالة دقيقة (بدون تضليل)
    try:
        if missing:
            await update.message.reply_text(
                f"{name}\n⚠️ ملف التاجر ناقص: " + "، ".join(missing) + "\nافتح لوحة التاجر وأكمل البيانات."
            )
        else:
            await update.message.reply_text(f"{name}\n✅ ملف التاجر مكتمل. يمكنك تقديم عروض السعر الآن.")
    except Exception as e:
        _swallow(e)

    set_stage(context, user_id, STAGE_NONE)
    try:
        await show_trader_panel(update, context, user_id)
    except Exception:
        try:
            await update.message.reply_text(f"{name}\nتعذر فتح لوحة التاجر حاليا")
        except Exception as e:
            _swallow(e)

async def chatid(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat = update.effective_chat
    await update.message.reply_text(f"Chat ID: {chat.id}\nType: {chat.type}")
    
async def support_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/منصة و /help : يفتح قناة تواصل مباشر مع الإدارة داخل الخاص فقط."""
    chat = update.effective_chat
    user_id = update.effective_user.id

    if chat.type != ChatType.PRIVATE:
        try:
            await update.message.reply_text("ℹ️ هذا الأمر يعمل في الخاص فقط")
        except Exception as e:
            _swallow(e)
        return

    ud = get_ud(context, user_id)

    # إذا كان المستخدم داخل مراحل طلب/عملية، لا نفتح منصة حتى لا تتداخل المدخلات
    try:
        cur_stage = ud.get(STAGE_KEY)
    except Exception:
        cur_stage = None
    if cur_stage and cur_stage != STAGE_NONE:
        try:
            await update.message.reply_text(
                "⚠️ أنت الآن داخل خطوة/عملية. أكملها أو الغِها ثم أعد كتابة /منصة\n"
                "(حتى لا تختلط رسائل الطلب برسائل الإدارة)",
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)
        return

    ud["support_open"] = True
    _support_touch(ud)

    try:
        await update.message.reply_text(
            "✅ تم فتح قناة تواصل مباشر مع الإدارة\n"
            "اكتب رسالتك الآن (استفسار/شكوى/ملاحظة)…",
            reply_markup=_support_kb(),
        )
    except Exception as e:
        _swallow(e)
    
async def support_open_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    user_id = q.from_user.id

    ud = get_ud(context, user_id)

    # ✅ استثناء مهم: التاجر الموقوف يسمح له بفتح منصة حتى لو داخل خطوة/عملية
    # الهدف: زر "مراسلة الإدارة" لا يُقفل بسبب STAGE حتى لا يُحرم الموقوف من التواصل
    try:
        if _trader_is_disabled(int(user_id or 0)):
            ud["support_open"] = True
            _support_touch(ud)

            # ✅ إشعار فوري للإدارة باسم التاجر
            try:
                tp = get_trader_profile(int(user_id or 0)) or {}
            except Exception:
                tp = {}
            tname = (tp.get("display_name") or "").strip() or (tp.get("company_name") or "").strip() or _user_name(q)
            tco = (tp.get("company_name") or "").strip()

            admin_ping = (
                "📩 <b>تواصل جديد مع الإدارة</b>\n"
                f"🆔 التاجر: <code>{user_id}</code>\n"
                f"👤 الاسم: <b>{html.escape(str(tname or '—'))}</b>\n"
                + (f"🏪 المتجر: <b>{html.escape(tco)}</b>\n" if tco and tco != tname else "")
                + "\n"
                "✍️ بانتظار رسالة التاجر الآن…"
            )

            for aid in ADMIN_IDS:
                try:
                    await context.bot.send_message(
                        chat_id=aid,
                        text=admin_ping,
                        parse_mode="HTML",
                        disable_web_page_preview=True,
                    )
                except Exception as e:
                    _swallow(e)

            try:
                await q.message.reply_text(
                    "✅ تم فتح قناة تواصل مباشر مع الإدارة\n"
                    "اكتب رسالتك الآن (استفسار/شكوى/ملاحظة)…",
                    reply_markup=_support_kb(),
                    disable_web_page_preview=True,
                )
            except Exception as e:
                _swallow(e)
            return
    except Exception as e:
        _swallow(e)

    # نفس شرط support_cmd: لا نفتح منصة لو داخل عملية
    try:
        cur_stage = ud.get(STAGE_KEY)
    except Exception:
        cur_stage = None
    if cur_stage and cur_stage != STAGE_NONE:
        try:
            await q.message.reply_text(
                "⚠️ أنت الآن داخل خطوة/عملية. أكملها أو الغِها ثم أعد فتح منصة\n"
                "(حتى لا تختلط رسائل الطلب برسائل الإدارة)",
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)
        return

    ud["support_open"] = True
    _support_touch(ud)

    # ✅ إشعار فوري للإدارة باسم المستخدم/التاجر
    try:
        tp = get_trader_profile(int(user_id or 0)) or {}
    except Exception:
        tp = {}
    tname = (tp.get("display_name") or "").strip() or (tp.get("company_name") or "").strip() or _user_name(q)
    tco = (tp.get("company_name") or "").strip()

    admin_ping = (
        "📩 <b>تواصل جديد مع الإدارة</b>\n"
        f"🆔 المرسل: <code>{user_id}</code>\n"
        f"👤 الاسم: <b>{html.escape(str(tname or '—'))}</b>\n"
        + (f"🏪 المتجر: <b>{html.escape(tco)}</b>\n" if tco and tco != tname else "")
        + "\n"
        "✍️ بانتظار رسالة المستخدم الآن…"
    )

    for aid in ADMIN_IDS:
        try:
            await context.bot.send_message(
                chat_id=aid,
                text=admin_ping,
                parse_mode="HTML",
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)

    try:
        await q.message.reply_text(
            "✅ تم فتح قناة تواصل مباشر مع الإدارة\n"
            "اكتب رسالتك الآن (استفسار/شكوى/ملاحظة)…",
            reply_markup=_support_kb(),
        )
    except Exception as e:
        _swallow(e)

async def support_close_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    user_id = q.from_user.id
    await _support_close(update, context, user_id)

async def support_admin_reply_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """زر عند الإدارة لبدء الرد على مستخدم داخل /منصة."""
    q = update.callback_query
    await q.answer()
    admin_id = q.from_user.id
    if admin_id not in ADMIN_IDS:
        return

    data = (q.data or "").split("|")
    if len(data) < 2:
        return
    try:
        target_uid = int(data[1] or 0)
    except Exception:
        target_uid = 0
    if not target_uid:
        return

    ud = get_ud(context, admin_id)
    ud[STAGE_KEY] = STAGE_SUPPORT_ADMIN_REPLY
    ud["support_reply_to_uid"] = target_uid
    try:
        await q.message.reply_text(
            f"✉️ اكتب رد الإدارة الآن (سيصل للمستخدم {target_uid})",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("✅ إنهاء", callback_data="pp_support_admin_done")]]),
        )
    except Exception as e:
        _swallow(e)

async def pp_support_reply_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    actor_id = q.from_user.id
    if actor_id not in ADMIN_IDS:
        await _alert(q, "غير مصرح")
        return

    data = (q.data or "").strip()
    parts = data.split("|")

    # يدعم الشكلين:
    # القديم: pp_support_reply|user_id
    # الجديد: pp_support_reply|order_id|user_id
    order_id = ""
    uid_int = 0

    try:
        if len(parts) == 2:
            _, uid = parts
            uid_int = int(uid)
        elif len(parts) == 3:
            _, order_id, uid = parts
            order_id = (order_id or "").strip()
            uid_int = int(uid)
        else:
            return
    except Exception:
        return

    ud = get_ud(context, actor_id)
    ud["support_reply_to_uid"] = uid_int
    ud["support_reply_order_id"] = order_id
    ud[STAGE_KEY] = STAGE_SUPPORT_ADMIN_REPLY

    try:
        await q.message.reply_text(
            f"{_user_name(q)}\nاكتب ردك الآن وسيصل للمستخدم باسم {PP_SUPPORT_LABEL}\n🧾 رقم الطلب: {order_id or '—'}",
            reply_markup=admin_reply_done_kb(),
            disable_web_page_preview=True,
        )
    except Exception as e:
        _swallow(e)

async def support_admin_done_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer("تم")
    admin_id = q.from_user.id
    if admin_id not in ADMIN_IDS:
        return
    ud = get_ud(context, admin_id)
    if ud.get(STAGE_KEY) == STAGE_SUPPORT_ADMIN_REPLY:
        ud[STAGE_KEY] = STAGE_NONE
    ud.pop("support_reply_to_uid", None)
    ud.pop("support_reply_order_id", None)
    try:
        await q.message.reply_text("✅ تم إنهاء وضع الرد")
    except Exception as e:
        _swallow(e)

async def begin_flow(update_or_q, context: ContextTypes.DEFAULT_TYPE):
    user = update_or_q.effective_user if hasattr(update_or_q, "effective_user") else update_or_q.from_user
    user_id = user.id

    # 🔧 وضع الصيانة
    if _is_maintenance_mode() and user_id not in ADMIN_IDS:
        try:
            if hasattr(update_or_q, "message") and update_or_q.message:
                await update_or_q.message.reply_text(
                    _maintenance_block_text(),
                    parse_mode="HTML",
                    disable_web_page_preview=True
                )
            else:
                await update_or_q.edit_message_text(
                    _maintenance_block_text(),
                    parse_mode="HTML",
                    disable_web_page_preview=True
                )
        except Exception as e:
            _swallow(e)
        return
    # ✅ إذا كانت قناة /منصة مفتوحة: نغلقها تلقائياً عند بدء أي طلب
    try:
        ud0 = get_ud(context, user_id)
        if _support_is_open(ud0):
            await _support_close(update_or_q, context, user_id, reason="(تم إغلاقها لأنك بدأت طلباً جديداً)")
    except Exception as e:
        _swallow(e)

    # ✅ بدء الطلب فعلياً
    reset_flow(context, user_id)
    ud = get_ud(context, user_id)
    with _EXCEL_WRITE_LOCK:
        ud["order_id"] = generate_order_id("PP")
    ud["user_id"] = user_id
    ud["user_name"] = user.full_name or ""
    ud["items"] = []
    ud["notes"] = ""
    ud["created_at_utc"] = utc_now_iso()

    set_stage(context, user_id, STAGE_ASK_CAR)

    try:
        text = (
            f"{ud['user_name']}\n"
            "اكتب اسم الشركة واسم السيارة بشكل واضح كما يظهر بالاستمارة\n"
            "مثال: شيري اريزو 8 او تويوتا كامري"
        )
        if hasattr(update_or_q, "message") and update_or_q.message:
            await update_or_q.message.reply_text(text, reply_markup=cancel_only_kb())
        else:
            # fallback: لو ما عندنا message (نادر)، نرسل مباشرة للخاص
            try:
                await context.bot.send_message(chat_id=int(user_id), text=text, reply_markup=cancel_only_kb())
            except Exception:
                await context.bot.send_message(chat_id=int(user_id), text=text)
    except Exception as e:
        _swallow(e)

async def back_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    user_id = q.from_user.id
    ud = get_ud(context, user_id)

    data = (q.data or "").strip()
    parts = data.split("|", 1)
    target = parts[1].strip() if len(parts) > 1 else ""

    # رجوع لاختيار طريقة التسليم
    if target == "delivery":
        set_stage(context, user_id, STAGE_AWAIT_DELIVERY)
        try:
            await q.message.reply_text(
                f"{_user_name(q)}\nاختر طريقة التسليم",
                reply_markup=delivery_kb(),
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)
        return

    # رجوع لمدينة الشحن
    if target == "ship_city":
        set_stage(context, user_id, STAGE_ASK_SHIP_CITY)
        try:
            await q.message.reply_text(
                f"{_user_name(q)}\nاكتب اسم المدينة",
                reply_markup=_flow_nav_kb("delivery"),
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)
        return

    # رجوع للعنوان المختصر
    if target == "ship_street":
        set_stage(context, user_id, STAGE_ASK_SHIP_STREET)
        try:
            await q.message.reply_text(
                f"{_user_name(q)}\n🏠 اكتب العنوان الوطني المختصر.",
                reply_markup=_flow_nav_kb("ship_city"),
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)
        return

    # رجوع لهاتف الشحن
    if target == "ship_phone":
        set_stage(context, user_id, STAGE_ASK_SHIP_PHONE)
        try:
            await q.message.reply_text(
                f"{_user_name(q)}\n📱 اكتب رقم الاتصال. مثال: 05xxxxxxxx",
                reply_markup=_flow_nav_kb("ship_street"),
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)
        return

async def cancel_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, f"تم الالغاء يا {_user_name(q)}")
    user_id = q.from_user.id
    reset_flow(context, user_id)
    await q.message.reply_text("تم الغاء العملية\للبداء بطلب قطع غيار  ارسل كلمة pp فقط")
        
async def skip_notes_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, f"تم يا {_user_name(q)}")
    user_id = q.from_user.id
    ud = get_ud(context, user_id)

    if ud.get(STAGE_KEY) != STAGE_ASK_NOTES:
        await q.message.reply_text(f"{_user_name(q)}\nلا يوجد ملاحظات حاليا")
        return

    ud["notes"] = ""
    set_stage(context, user_id, STAGE_ASK_ITEM_NAME)
    await q.message.reply_text(
        f"{_user_name(q)}\n"
        "اكتب اسم القطعة المطلوبة بدقة\n"
        "واذكر رقم القطعة ان توفر لرفع دقة الطلب"
)

async def prepay_notes_skip_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    user_id = q.from_user.id
    ud = get_ud(context, user_id)

    if ud.get(STAGE_KEY) not in (STAGE_PREPAY_NOTES, STAGE_PREPAY_NOTES_TEXT):
        return

    # ✅ تثبيت الملاحظات في الطلب قبل الانتقال (حتى تظهر بالمجموعة دائمًا)
    # لا نحفظ قيمة فاضية حتى لا نمسح ملاحظة موجودة سابقًا بالطلب
    try:
        order_id = (ud.get("order_id") or "").strip()
        notes = (ud.get("notes") or "").strip()
        if order_id and notes:
            update_order_fields(order_id, {"notes": notes})
    except Exception as e:
        _swallow(e)

    # بعد الملاحظات -> ننتقل للتسليم (العنوان) ثم بعدها الدفع
    set_stage(context, user_id, STAGE_AWAIT_DELIVERY)
    await q.message.reply_text(
        f"{_user_name(q)}\nاختر طريقة التسليم",
        reply_markup=delivery_kb(),
        disable_web_page_preview=True,
    )

async def more_yes_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    user_id = q.from_user.id
    ud = get_ud(context, user_id)

    items = ud.get("items", []) or []
    if len(items) >= MAX_ITEMS:
        try:
            await context.bot.send_message(
                chat_id=q.message.chat_id,
                text=f"{_user_name(q)}\nتم الوصول للحد الاقصى {MAX_ITEMS} قطعة"
            )
        except Exception as e:
            _swallow(e)
        return

    set_stage(context, user_id, STAGE_ASK_ITEM_NAME)
    next_no = len(items) + 1
    try:
        await context.bot.send_message(
            chat_id=q.message.chat_id,
            text=f"{_user_name(q)}\nاكتب اسم القطعة رقم {next_no}"
        )
    except Exception as e:
        _swallow(e)

async def more_no_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    user_id = q.from_user.id
    ud = get_ud(context, user_id)

    items = ud.get("items", []) or []
    if not items:
        await q.message.reply_text(f"{_user_name(q)}\nلا يوجد قطع مضافة اكتب اسم القطعة اولا")
        set_stage(context, user_id, STAGE_ASK_ITEM_NAME)
        return

    # رسوم المنصة: حسب عدد القطع غير الاستهلاكية (الاستهلاكي مجاني بالكامل)
    fee, non_cnt, cons_cnt = _platform_fee_for_items(items)
    ud["price_sar"] = fee
    ud["non_consumable_count"] = non_cnt
    ud["consumable_count"] = cons_cnt

    # حفظ الطلب (مرة واحدة) قبل الانتقال للخطوات التالية
    try:
        _save_order_once(ud)
    except Exception as e:
        _swallow(e)

    order_id = (ud.get("order_id") or "").strip()
    if order_id:
        try:
            update_order_fields(order_id, {
                "price_sar": fee,
                "non_consumable_count": non_cnt,
                "consumable_count": cons_cnt,
            })
        except Exception as e:
            _swallow(e)

    # معاينة أولية (اختياري)
    try:
        await q.message.reply_text(build_order_preview(ud), parse_mode="HTML", disable_web_page_preview=True)
    except Exception as e:
        _swallow(e)

    # ✅ رجّع مرحلة الملاحظات (بدل ما تختفي)
    set_stage(context, user_id, STAGE_PREPAY_NOTES)
    await q.message.reply_text(
        f"{_user_name(q)}\nاذا لديك ملاحظة ارسلها الان او اختر تخطي للانتقال لاختيار طريقة التسليم",
        reply_markup=prepay_notes_kb(),
        disable_web_page_preview=True,
    )
    return

async def partno_skip_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    user_id = q.from_user.id
    ud = get_ud(context, user_id)

    if ud.get(STAGE_KEY) != STAGE_ASK_ITEM_PARTNO:
        return

    pending_name = _norm(ud.get("pending_item_name", ""))
    if not pending_name:
        set_stage(context, user_id, STAGE_ASK_ITEM_NAME)
        await q.message.reply_text(f"{_user_name(q)}\nاكتب اسم القطعة اولا")
        return

    ud.setdefault("items", []).append({
        "name": pending_name,
        "part_no": "",
        "photo_file_id": "",
        "created_at_utc": utc_now_iso(),
    })
    ud.pop("pending_item_name", None)
    ud["pending_item_idx"] = len(ud["items"]) - 1

    set_stage(context, user_id, STAGE_ASK_ITEM_PHOTO)
    item_no = len(ud["items"])
    await q.message.reply_text(
        f"{_user_name(q)}\nتمت اضافة القطعة رقم {item_no}\nارسل صورة الان (اختياري) او اكتب اسم القطعة التالية مباشرة",
        reply_markup=photo_prompt_kb(),
    )

async def skip_photo_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    user_id = q.from_user.id
    ud = get_ud(context, user_id)

    if ud.get(STAGE_KEY) != STAGE_ASK_ITEM_PHOTO:
        await q.message.reply_text(f"{_user_name(q)}\nلا يوجد طلب صورة حاليا")
        return

    items = ud.get("items", []) or []
    idx = ud.get("pending_item_idx")
    try:
        item_no = int(idx) + 1 if isinstance(idx, int) else len(items)
    except Exception:
        item_no = len(items)

    ud.pop("pending_item_idx", None)
    ud.pop("pending_item_name", None)

    set_stage(context, user_id, STAGE_CONFIRM_MORE)
    await q.message.reply_text(
        f"{_user_name(q)}\n"
        f"تم تخطي صورة القطعة رقم {item_no}\n"
        f"عدد القطع الحالي: {len(items)}\n\n"
        "يمكنك الان كتابة اسم قطعة جديدة مباشرة\n"
        "او اختيار انهاء وارسال الطلب",
        reply_markup=more_kb(),
    )
    
async def copy_iban_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, f"تم تجهيز الايبان يا {_user_name(q)}")
    await q.message.reply_text(f"IBAN:\n`{PP_IBAN}`", parse_mode="Markdown")
    
async def copy_beneficiary_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, f"تم تجهيز اسم المستفيد يا {_user_name(q)}")
    await q.message.reply_text(f"اسم المستفيد:\n`{PP_BENEFICIARY}`", parse_mode="Markdown")

async def copy_stc_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, f"تم تجهيز رقم STC Pay يا {_user_name(q)}")
    await q.message.reply_text(f"رقم STC Pay:\n`{PP_STC_PAY}`", parse_mode="Markdown")

async def pay_bank_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, f"تمام يا {_user_name(q)}")
    user_id = q.from_user.id
    ud = get_ud(context, user_id)

    stage = ud.get(STAGE_KEY)
    if stage == STAGE_TRADER_SUB_AWAIT_PAY_METHOD:
        month = str(ud.get("sub_month") or month_key_utc()).strip()
        amount = int(float(ud.get("sub_amount_sar") or 99))
        ud["sub_payment_method"] = "bank_transfer"
        set_stage(context, user_id, STAGE_TRADER_SUB_AWAIT_RECEIPT)
        try:
            upsert_trader_subscription(user_id, month, {
                "amount_sar": amount,
                "payment_method": "bank_transfer",
                "payment_status": "pending",
            })
        except Exception as e:
            _swallow(e)

        await q.message.reply_text(
            f"🤍 اهلا {_user_name(q)}\n\n"
            f"💳 <b>طريقة الدفع: تحويل بنكي</b>\n\n"
            f"المبلغ المطلوب <b>{amount} ريال</b> مقابل <b>رسوم اشتراك المنصة</b> لشهر {month}\n\n"
            f"🏦 <b>المستفيد</b>:\n<i>{PP_BENEFICIARY}</i>\n\n"\
            f"IBAN:\n<code>{PP_IBAN}</code>\n\n"\

            f"🧾 <b>رقم المرجع</b>:\n<code>SUB-{user_id}-{month}</code>\n\n"
            "📸 بعد الدفع أرسل <b>صورة/ملف الإيصال</b> هنا مباشرة (الايصال الزامي)\n",
            parse_mode="HTML",
            disable_web_page_preview=True,
        )
        return

    ud["payment_method"] = "bank_transfer"
    set_stage(context, user_id, STAGE_AWAIT_RECEIPT)

    _save_order_once(ud)
    update_order_fields(ud["order_id"], {"payment_method": "bank_transfer", "payment_status": "pending"})

    await q.message.reply_text(
        f"🤍 اهلا { _user_name(q) }\n\n"
        "💳 <b>طريقة الدفع: تحويل بنكي</b>\n\n"
        f"المبلغ المطلوب <b>{ud.get('price_sar', 0)} ريال</b> هو مقابل خدمة احترافية تشمل\n"
        "البحث الدقيق عن القطع المطلوبة حسب بيانات سيارتك\n"
        "والتحقق من التوافق والتوفر وإصدار تسعيرة واضحة قبل تنفيذ الطلب\n\n"
        "هدفنا ان تصل الى القطعة والتسعيرة من خلال شركات السيارات او وكلاء محليين / عالميين باسرع وقت وتلقي عروض مختلفة\n"
        f"🏦 <b>المستفيد</b>:\n<i>{PP_BENEFICIARY}</i>\n\n"
        f"IBAN:\n<code>{PP_IBAN}</code>\n\n"
        f"🧾 <b>رقم المرجع</b>:\n{ud.get('order_id','')}\n\n"
        "📸 بعد التحويل يرجى ارسال <b>صورة ايصال الدفع</b> هنا مباشرة\n"
        "لاستكمال الطلب (الايصال الزامي)\n\n"
        "✨ سعداء بخدمتك وملتزمون بتقديم تجربة موثوقة وواضحة",
        parse_mode="HTML",
        reply_markup=bank_info_kb()
    )

async def pay_stc_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, f"تمام يا {_user_name(q)}")
    user_id = q.from_user.id
    ud = get_ud(context, user_id)

    stage = ud.get(STAGE_KEY)
    if stage == STAGE_TRADER_SUB_AWAIT_PAY_METHOD:
        month = str(ud.get("sub_month") or month_key_utc()).strip()
        amount = int(float(ud.get("sub_amount_sar") or 99))
        ud["sub_payment_method"] = "stc_pay"
        set_stage(context, user_id, STAGE_TRADER_SUB_AWAIT_RECEIPT)
        try:
            upsert_trader_subscription(user_id, month, {
                "amount_sar": amount,
                "payment_method": "stc_pay",
                "payment_status": "pending",
            })
        except Exception as e:
            _swallow(e)

        await q.message.reply_text(
            f"🤍 اهلا {_user_name(q)}\n\n"
            f"💳 <b>طريقة الدفع: STC Pay</b>\n\n"
            f"المبلغ المطلوب <b>{amount} ريال</b> مقابل <b>رسوم اشتراك المنصة</b> لشهر {month}\n\n"
            f"📱 <b>رقم STC Pay</b>:\n<code>{PP_STC_PAY}</code>\n\n"\

            f"🧾 <b>رقم المرجع</b>:\n<code>SUB-{user_id}-{month}</code>\n\n"
            "📸 بعد الدفع أرسل <b>صورة/ملف الإيصال</b> هنا مباشرة (الايصال الزامي)\n",
            parse_mode="HTML",
            disable_web_page_preview=True,
        )
        return

    ud["payment_method"] = "stc_pay"
    set_stage(context, user_id, STAGE_AWAIT_RECEIPT)

    _save_order_once(ud)
    update_order_fields(ud["order_id"], {"payment_method": "stc_pay", "payment_status": "pending"})

    await q.message.reply_text(
        f"🤍 اهلا { _user_name(q) }\n\n"
        "💳 <b>طريقة الدفع: STC Pay</b>\n\n"
        f"المبلغ المطلوب <b>{ud.get('price_sar', 0)} ريال</b> هو مقابل خدمة احترافية تشمل\n"
        "البحث الدقيق عن القطع المطلوبة حسب بيانات سيارتك\n"
        "والتحقق من التوافق والتوفر وإصدار تسعيرة واضحة قبل تنفيذ الطلب\n\n"
        "نحرص ان تكمل العملية وانت مطمئن تماما 🤝\n\n"
        f"📱 <b>رقم STC Pay</b>:\n<code>{PP_STC_PAY}</code>\n\n"
        f"🧾 <b>رقم المرجع</b>:\n{ud.get('order_id','')}\n\n"
        "📸 بعد التحويل يرجى ارسال <b>صورة ايصال الدفع</b> هنا مباشرة\n"
        "لاستكمال الطلب (الايصال الزامي)\n\n"
        "✨ سعداء بخدمتك وملتزمون بتقديم تجربة موثوقة وواضحة",
        parse_mode="HTML",
        reply_markup=stc_info_kb()
)

async def pay_link_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    user_id = q.from_user.id
    ud = get_ud(context, user_id)

    stage = ud.get(STAGE_KEY)
    if stage == STAGE_TRADER_SUB_AWAIT_PAY_METHOD:
        month = str(ud.get("sub_month") or month_key_utc()).strip()
        amount = int(float(ud.get("sub_amount_sar") or 99))
        ud["sub_payment_method"] = "pay_link"
        set_stage(context, user_id, STAGE_TRADER_SUB_AWAIT_RECEIPT)
        try:
            upsert_trader_subscription(user_id, month, {
                "amount_sar": amount,
                "payment_method": "pay_link",
                "payment_status": "pending",
            })
        except Exception as e:
            _swallow(e)

        if PP_PAY_LINK_URL:
            await q.message.reply_text(
                f"🔗 <b>رابط دفع الاشتراك</b>\n\n{html.escape(PP_PAY_LINK_URL)}\n\n"
                f"المرجع: <code>SUB-{user_id}-{month}</code>\n"
                "بعد الدفع أرسل صورة/ملف الإيصال هنا (الايصال الزامي)",
                parse_mode="HTML",
                disable_web_page_preview=True,
            )
            return

        try:
            cap = (
                "🔗 <b>طلب رابط دفع (اشتراك تاجر)</b>\n"
                f"👤 التاجر: <b>{html.escape(str(q.from_user.full_name or q.from_user.first_name or ''))}</b>\n"
                f"🆔 trader_id: <code>{user_id}</code>\n"
                f"📅 الشهر: <b>{html.escape(month)}</b>\n"
                f"💰 المبلغ: <b>{amount}</b> ريال\n\n"
                "الصق رابط الدفع وارسله للتاجر."
            )
            for aid in ADMIN_IDS:
                try:
                    await context.bot.send_message(chat_id=aid, text=cap, parse_mode="HTML", disable_web_page_preview=True)
                except Exception as e:
                    _swallow(e)
        except Exception as e:
            _swallow(e)

        await q.message.reply_text(
            f"{_user_name(q)}\n"
            "✅ تم تسجيل طلب رابط الدفع للاشتراك\n"
            "سيتم تزويدك بالرابط قريبًا\n"
            "بعد الدفع أرسل الإيصال هنا",
            disable_web_page_preview=True,
        )
        return

    ud["payment_method"] = "pay_link"
    set_stage(context, user_id, STAGE_AWAIT_RECEIPT)

    # حفظ الطلب مرة واحدة
    try:
        _save_order_once(ud)
    except Exception as e:
        _swallow(e)

    order_id = (ud.get("order_id") or "").strip()
    if not order_id:
        await q.message.reply_text(f"{_user_name(q)}\n🟥 تعذر تحديد رقم الطلب")
        return

    try:
        update_order_fields(order_id, {
            "payment_method": "pay_link",
            "payment_status": "pending",
        })
    except Exception as e:
        _swallow(e)

    # ✅ في حال وجود رابط ثابت
    if PP_PAY_LINK_URL:
        await q.message.reply_text(
            "طريقة الدفع: رابط دفع سريع\n\n"
            f"{PP_PAY_LINK_URL}\n\n"
            f"المرجع: {order_id}\n"
            "بعد الدفع ارسل صورة ايصال الدفع هنا (الايصال الزامي)",
            disable_web_page_preview=True,
        )
        return

    # ❗ بدون رابط ثابت → طلب يدوي من الإدارة
    try:
        # جلب نسخة الطلب للمعاينة
        try:
            b = get_order_bundle(order_id)
            order = b.get("order", {}) or {}
        except Exception:
            order = {}

        cap = _build_admin_order_caption(
            order_id,
            ud,
            order,
            "🔗 طلب رابط دفع يدوي (رسوم المنصة)",
            extra_lines=[
                "المطلوب: اضغط الزر ثم الصق رابط الدفع ليتم إرساله للعميل"
            ],
        )
        
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton(
                "📩 إرسال رابط الدفع للعميل",
                callback_data=f"pp_admin_paylink|{order_id}|{user_id}"
            )],
            [InlineKeyboardButton(
                "💬 مراسلة العميل داخل المنصة",
                callback_data=f"pp_admin_reply|{order_id}|{user_id}"
            )],
        ])

        for aid in ADMIN_IDS:
            try:
                await context.bot.send_message(
                    chat_id=aid,
                    text=cap,
                    reply_markup=kb,
                    disable_web_page_preview=True,
                )
            except Exception as e:
                _swallow(e)

    except Exception as e:
        _swallow(e)

    await q.message.reply_text(
        f"{_user_name(q)}\n"
        "✅ تم تسجيل طلب الدفع بالرابط\n"
        "سيتم تزويدك برابط الدفع قريبًا داخل المنصة\n"
        "بعد السداد أرسل صورة الإيصال هنا",
        disable_web_page_preview=True,
    )

async def admin_paylink_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """الإدارة تضغط زر (إرسال رابط الدفع) ثم تلصق الرابط ليتم إرساله للعميل."""
    q = update.callback_query
    await _alert(q, "")

    data = (q.data or "").strip()
    parts = data.split("|")
    if len(parts) < 3:
        return

    order_id = (parts[1] or "").strip()
    try:
        client_id = int(parts[2] or 0)
    except Exception:
        client_id = 0

    actor_id = q.from_user.id
    if actor_id not in ADMIN_IDS:
        await _alert(q, "⛔ غير مصرح")
        return

    if not order_id or not client_id:
        await _alert(q, "تعذر تحديد الطلب/العميل")
        return

    ud = get_ud(context, actor_id)
    ud["paylink_order_id"] = order_id
    ud["paylink_client_id"] = client_id
    set_stage(context, actor_id, STAGE_ADMIN_SEND_PAYLINK)

    await q.message.reply_text(
        f"{_user_name(q)}\n🟦 ارسل الآن رابط الدفع (نص فقط)\n🧾 رقم الطلب: {order_id}",
        disable_web_page_preview=True,
    )

async def quote_ok_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")

    data = (q.data or "").strip()
    parts = data.split("|")
    order_id = parts[1].strip() if len(parts) >= 2 else ""

    cb_tid = 0
    if len(parts) >= 3:
        try:
            cb_tid = int(parts[2] or 0)
        except Exception:
            cb_tid = 0

    if not order_id:
        return

    b = get_order_bundle(order_id)
    order = b.get("order", {}) or {}

    gps = str(order.get("goods_payment_status") or "").strip().lower()
    ost = str(order.get("order_status") or "").strip().lower()
    locked_now = str(order.get("quote_locked") or "").strip().lower() == "yes"

    # ✅ بعد الدفع/القفل: ممنوع قبول جديد
    if gps in ("awaiting_confirm", "confirmed") or locked_now or ost in ("closed", "delivered"):
        await q.message.reply_text(f"{_user_name(q)}\n🔒 الطلب مغلق/مدفوع ولا يمكن قبول عروض جديدة")
        return

    if not cb_tid:
        await q.message.reply_text(
            f"{_user_name(q)}\n"
            "⚠️ هذا زر قديم ولا يحتوي هوية التاجر.\n"
            "افتح آخر رسالة عرض سعر ثم اضغط زر القبول منها."
        )
        return

    tid = int(cb_tid or 0)
    if tid <= 0:
        await q.message.reply_text(f"{_user_name(q)}\nلا يوجد تاجر مرسل عرض سعر لهذا الطلب")
        return

    prev_tid = 0
    try:
        prev_tid = int(order.get("accepted_trader_id") or 0)
    except Exception:
        prev_tid = 0

    # ✅ تثبيت اسم التاجر صراحةً من ملف التاجر (مع حفظ الشركة والليبل أيضًا)
    try:
        tprof = get_trader_profile(tid) or {}
    except Exception:
        tprof = {}

    tname = (tprof.get("display_name") or "").strip()
    tcompany = (tprof.get("company_name") or "").strip()

    if not tname:
        tname = (tcompany or "").strip()
    if not tname:
        tname = (order.get("quoted_trader_name") or "").strip()
    if not tname:
        tname = "التاجر"

    # ✅ بدون أقواس
    tlabel = tname + (f" - {tcompany}" if tcompany else "")

    prev_label = ""
    if prev_tid:
        try:
            pp = get_trader_profile(int(prev_tid)) or {}
            pn = (pp.get("display_name") or "").strip() or (order.get("accepted_trader_name") or "").strip() or "التاجر"
            pc = (pp.get("company_name") or "").strip()
            prev_label = pn + (f" - {pc}" if pc else "")
        except Exception:
            prev_label = (order.get("accepted_trader_name") or "").strip() or "التاجر"

    switched = bool(prev_tid and prev_tid != tid)

    ship_city = (order.get("ship_city") or "").strip()
    city_line = f"\n🏙️ مدينة التسليم: {ship_city}" if ship_city else ""

    # =========================
    # ✅ حساب (قيمة القطع + الشحن) وإظهارها
    # =========================
    def _s(x: object) -> str:
        return ("" if x is None else str(x)).strip()

    def _to_float(x: object) -> float:
        s0 = _s(x)
        if not s0:
            return 0.0
        try:
            return float(str(s0).replace(",", "").strip())
        except Exception:
            return 0.0

    def _money(x: object) -> str:
        s0 = _s(x)
        if not s0:
            return "—"
        try:
            f = float(str(s0).replace(",", "").strip())
            if abs(f - int(f)) < 1e-9:
                return f"{int(f)} ﷼"
            return f"{f:.2f}".rstrip("0").rstrip(".") + " ﷼"
        except Exception:
            return f"{s0} ﷼"

    goods_amount = _s(order.get("goods_amount_sar") or "")
    ship_included_norm = _s(order.get("ship_included")).lower()

    raw_shipping_fee = (
        order.get("shipping_fee_sar")
        or order.get("quote_shipping_fee")
        or order.get("shipping_fee")
        or order.get("ship_fee")
        or ""
    )

    missing_ship = (raw_shipping_fee is None or _s(raw_shipping_fee) == "")

    inc_yes = ship_included_norm in ("yes", "true", "1", "included", "مشمولة", "مشمول")

    goods_txt = _money(goods_amount)

    if inc_yes:
        ship_txt = "مشمول (ضمن الإجمالي)"
        total_val = _to_float(goods_amount)
        total_txt = _money(total_val)
    else:
        if missing_ship:
            ship_txt = "غير محدد"
            total_txt = "غير محدد"
        else:
            ship_txt = _money(raw_shipping_fee)
            total_val = _to_float(goods_amount) + _to_float(raw_shipping_fee)
            total_txt = _money(total_val)

    amount_block_trader = (
        "\n\n📌 تفاصيل المبالغ:\n"
        f"🧩 قيمة القطع: {goods_txt}\n"
        f"🚚 الشحن: {ship_txt}\n"
        f"✅ الإجمالي شامل الشحن: {total_txt}\n"
    )

    amount_block_client = amount_block_trader

    # ✅ إذا تم تبديل التاجر بعد قبول سابق: نظّف آثار التاجر السابق (فاتورة/حالات متقدمة)
    reset_fields = {}
    if prev_tid and prev_tid != tid:
        try:
            # إزالة أي فاتورة/مرفقات من تاجر سابق حتى لا تُقفل خطوات التاجر الجديد
            reset_fields = {
                "seller_invoice_file_id": "",
                "seller_invoice_mime": "",
                "seller_invoice_at": "",
                "shop_invoice_file_id": "",
                "shop_invoice_mime": "",
                "shop_invoice_at": "",
                # إعادة تعيين رقم التتبع/الشحن لو كان موجوداً
                "tracking_number": "",
                "tracking_no": "",
                "tracking": "",
                # إعادة تعيين أختام/تواريخ الشحن/التسليم
                "shipped_at_utc": "",
                "delivered_at_utc": "",
                "closed_at_utc": "",
                # إعادة تعيين مؤقت المحادثة (لأنها مرتبطة بالتسلسل بعد الدفع)
                "chat_expires_at_utc": "",
                # إعادة تعيين الإجمالي المخزن إن كان مبني على فاتورة قديمة
                "total_amount_sar": "",
            }

            # إذا كان هناك حالة دفع بضاعة من تاجر سابق ولم تُؤكد فعلاً، نعيدها فارغة
            gps_now = str(order.get("goods_payment_status") or "").strip().lower()
            if gps_now not in ("confirmed", "paid", "success", "successful", "done", "ok"):
                reset_fields["goods_payment_status"] = ""
                reset_fields["goods_payment_confirmed_at_utc"] = ""
        except Exception:
            reset_fields = {}

    fields = {
        "quote_status": "accepted",
        "accepted_trader_id": tid,
        "accepted_trader_name": tname,
        "accepted_trader_company": tcompany,
        "accepted_trader_label": tlabel,
        "order_status": "accepted",
        "accepted_at_utc": utc_now_iso(),
    }
    if reset_fields:
        fields.update(reset_fields)

    update_order_fields(order_id, fields)

    # إشعار التاجر المقبول
    try:
        await context.bot.send_message(
            chat_id=tid,
            text=(
                "✅ تم قبول عرض السعر من العميل\n"
                f"🧾 رقم الطلب: {order_id}"
                f"{city_line}"
                f"{amount_block_trader}\n"
                "🧰 يمكنك الآن البدء بتجهيز الطلب\n"
                "🟦 حدّث الحالة من الزر بالاسفل\n"
                "🧾 سيتم إشعارك عند إرسال إثبات الدفع\n\n"
                "🔒 ملاحظة: لا يتم عرض العنوان الكامل أو رقم العميل قبل الدفع"
            ),
            reply_markup=trader_status_kb(order_id),
            disable_web_page_preview=True,
        )
    except Exception as e:
        _swallow(e)

    # ✅ إشعار التاجر السابق إذا تم التحويل لعرض آخر (قبل الدفع)
    if switched and prev_tid and prev_tid != tid:
        try:
            await context.bot.send_message(
                chat_id=int(prev_tid),
                text=(
                    "ℹ️ تحديث على عرضك\n"
                    f"🧾 رقم الطلب: {order_id}\n"
                    "تم إلغاء موافقة العميل على عرضك بعد اختيار عرض آخر.\n"
                    "يمكنك إنشاء عرض جديد إذا رغبت."
                ),
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)

    # إشعار العميل
    try:
        msg = (
            f"{_user_name(q)}\n"
            "✅ تم قبول عرض السعر\n"
            f"🧾 رقم الطلب: {order_id}\n"
            f"👤 التاجر: {tlabel}"
            f"{amount_block_client}\n"
            "📌 ملاحظة مهمة:\n"
            "• يمكنك العدول واختيار عرض آخر طالما لم يتم الدفع\n"
            "• عند الدفع سيتم قفل الطلب ومنع العروض الجديدة\n\n"
            "🟦 الخطوة التالية: سيقوم التاجر بتحديث الحالة وارسال فاتورة الدفع داخل الطلب"
        )
        if switched and prev_label:
            msg += f"\n\nℹ️ تم إلغاء الموافقة السابقة تلقائيًا عن: {prev_label}"
        await q.message.reply_text(msg, disable_web_page_preview=True)
    except Exception as e:
        _swallow(e)

    return

async def quote_no_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")

    data = (q.data or "").strip()
    parts = data.split("|")

    order_id = (parts[1] or "").strip() if len(parts) >= 2 else ""
    btn_tid = 0
    if len(parts) >= 3:
        try:
            btn_tid = int(parts[2] or 0)
        except Exception:
            btn_tid = 0

    if not order_id:
        return

    try:
        b = get_order_bundle(order_id)
        order = b.get("order", {}) or {}
        _ = b.get("items", []) or []
    except Exception:
        order = {}

    if not order:
        try:
            await q.message.reply_text(f"{_user_name(q)}\nتعذر قراءة بيانات الطلب")
        except Exception as e:
            _swallow(e)
        return

    gps = str(order.get("goods_payment_status") or "").strip().lower()
    ost = str(order.get("order_status") or "").strip().lower()
    locked_now = str(order.get("quote_locked") or "").strip().lower() == "yes"
    if gps in ("awaiting_confirm", "confirmed") or ost in ("closed", "delivered") or locked_now:
        try:
            await q.message.reply_text(f"{_user_name(q)}\n🔒 لا يمكن رفض العرض بعد الدفع/قفل الطلب")
        except Exception as e:
            _swallow(e)
        return

    # ✅ حماية: لو الزر قديم ولا يحمل هوية التاجر لا نرسل إشعار لتاجر خاطئ
    tid = int(btn_tid or 0)
    if not tid:
        try:
            await q.message.reply_text(
                f"{_user_name(q)}\n"
                "⚠️ هذا زر قديم ولا يحتوي هوية التاجر.\n"
                "افتح آخر رسالة عرض سعر ثم اضغط زر (غير موافق) منها."
            )
        except Exception as e:
            _swallow(e)
        return

    # ✅ مدينة التسليم فقط للتاجر (بدون رقم/تفاصيل)
    ship_city = (order.get("ship_city") or "").strip()
    city_line = f"\n🏙️ مدينة التسليم: {ship_city}" if ship_city else ""

    # ✅ تسجيل الرفض وفتح الباب لعروض أخرى
    try:
        update_order_fields(order_id, {
            "quote_status": "rejected",
            "accepted_trader_id": "",
            "accepted_trader_name": "",
            "quoted_trader_id": "",
            "quoted_trader_name": "",
            "quote_locked": "no",
            "last_group_broadcast_at_utc": utc_now_iso(),
        })
    except Exception as e:
        _swallow(e)

    # ✅ إشعار التاجر صاحب العرض فقط (بدون ذكر أي تاجر آخر، وبدون رقم العميل) + زر عرض جديد
    try:
        await context.bot.send_message(
            chat_id=tid,
            text=(
                "❌ لم يوافق العميل على عرض السعر\n"
                f"🧾 رقم الطلب: {order_id}"
                f"{city_line}\n\n"
                "يمكنك تقديم عرض جديد إذا رغبت (طالما لم يتم الدفع).\n"
                "⚠️ تنبيه: معلومات التواصل/العنوان الكامل لا تُعرض قبل الدفع."
            ),
            reply_markup=trader_quote_start_kb(order_id),
            disable_web_page_preview=True,
        )
    except Exception as e:
        _swallow(e)

    # ✅ إشعار العميل
    try:
        await q.message.reply_text(
            f"{_user_name(q)}\n"
            "تم تسجيل عدم الموافقة.\n"
            "يمكنك اختيار عرض آخر من العروض المتاحة.",
            disable_web_page_preview=True,
        )
    except Exception as e:
        _swallow(e)

    # ✅ لا يوجد أي إرسال للمجموعة نهائيًا
    await _alert(q, "تم")

def trader_quote_items_kb(order_id: str, items: list[dict], selected: set[int]) -> InlineKeyboardMarkup:
    rows = []

    for i, it in enumerate(items or [], start=1):
        nm = (it.get("name") or it.get("item_name") or "").strip() or "قطعة"
        mark = "✅" if i in selected else "⬜"
        label = _wide_btn_label(f"{mark} {i}- {nm[:28]}")
        rows.append([InlineKeyboardButton(label, callback_data=f"ppq_it|{order_id}|{i}")])

    rows.append([InlineKeyboardButton(_wide_btn_label("☑️ تحديد الكل"), callback_data=f"ppq_it_all|{order_id}")])
    rows.append([InlineKeyboardButton(_wide_btn_label("🧹 مسح الكل"), callback_data=f"ppq_it_none|{order_id}")])
    rows.append([InlineKeyboardButton(_wide_btn_label("✅ التالي"), callback_data=f"ppq_it_done|{order_id}")])

    return InlineKeyboardMarkup(rows)

async def ppq_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()

    data = q.data or ""
    parts = data.split("|")
    action = parts[0]

    user_id = q.from_user.id
    name = _user_name(q)
    td = context.user_data.setdefault(user_id, {})

    # ✅ ضمان تحميل الاكسل قبل أي قراءة/تسعير
    try:
        ensure_workbook()
    except Exception as e:
        _swallow(e)

    # ✅ كل خطوات عرض السعر تكون بالخاص فقط
    if q.message and q.message.chat and q.message.chat.type != "private":
        await _alert(q, "افتح الخاص لإكمال عرض السعر")
        return

    # ✅ منع التاجر الموقوف
    if _trader_is_disabled(user_id):
        await _deny_disabled_trader_q(q, "لا يمكنك تقديم أو تعديل عروض السعر لأن حسابك موقوف")
        return

    # ✅ لا يبدأ/يكمل عرض سعر إلا بعد اكتمال ملف التاجر
    try:
        tp = get_trader_profile(int(user_id or 0)) or {}
    except Exception:
        tp = {}

    pay_mode = (str(tp.get("payment_mode") or "").strip().lower())
    required_fields = ["display_name","company_name","shop_phone","cr_no","vat_no"]
    if pay_mode != "link":
        required_fields += ["bank_name","iban","stc_pay"]
    if not all((tp.get(f) or "").strip() for f in required_fields):
        await _need_complete_trader_profile_notice(
            context,
            user_id,
            _user_name(q),
            td.get("quote_order_id") or "",
        )
        try:
            await q.message.reply_text("أكمل ملف التاجر ثم ارجع وابدأ عرض السعر.")
        except Exception as e:
            _swallow(e)
        return

    # =========================
    # Helpers: Snapshot + Header
    # =========================
    def _safe_int(x) -> int:
        try:
            return int(x or 0)
        except Exception:
            return 0

    def _to_float(x) -> float:
        try:
            s = str(x or "").strip()
            s = s.replace(",", "")
            s = re.sub(r"[^0-9.]+", "", s)
            return float(s or 0)
        except Exception:
            return 0.0

    def _norm_money_str(x) -> str:
        """
        يرجع رقم كنص بدون عملة (مثال: '120' أو '120.5') أو '' إذا فاضي.
        """
        s = str(x or "").strip()
        if not s:
            return ""
        s = s.replace(",", "")
        s = re.sub(r"[^0-9.]+", "", s)
        s = s.strip()
        return s

    def _fmt_money_num(x) -> str:
        """
        يرجع رقم منسق كنص (بدون ر.س) مثال: '120' أو '120.50'
        """
        v = _to_float(x)
        if abs(v - int(v)) < 1e-9:
            return str(int(v))
        return f"{v:.2f}".rstrip("0").rstrip(".")

    def _money_txt(v) -> str:
        """
        يطبع بصيغة _money إن أمكن وإلا يطبع "X ر.س"
        """
        s = str(v or "").strip()
        if not s or s in ("0", "0.0", "0.00"):
            return ""
        try:
            return _money(s)
        except Exception:
            return f"{s} ر.س"

    def _order_snapshot(order_id: str) -> dict:
        out = {
            "client_name": "—",
            "goods_amount": "",
            "shipping_fee": "",
            "total_amount": "",
            "ship_included": "",
        }
        oid = (str(order_id or "")).strip()
        if not oid:
            return out

        try:
            ob = get_order_bundle(oid) or {}
            oo = (ob.get("order", {}) or {}) if isinstance(ob, dict) else {}
        except Exception:
            oo = {}

        try:
            out["client_name"] = (str(oo.get("user_name") or "").strip() or "—")
        except Exception as e:
            _swallow(e)

        # ✅ نقرأ القيم الصحيحة التي سنحافظ عليها
        try:
            out["goods_amount"] = str(oo.get("goods_amount_sar") or oo.get("quote_goods_amount") or "").strip()
        except Exception as e:
            _swallow(e)

        try:
            out["shipping_fee"] = str(oo.get("shipping_fee_sar") or oo.get("shipping_fee") or "").strip()
        except Exception as e:
            _swallow(e)

        # ✅ الإجمالي النهائي الحقيقي (بدل price_sar اللي هو رسوم منصة)
        try:
            out["total_amount"] = str(oo.get("total_amount_sar") or "").strip()
        except Exception as e:
            _swallow(e)

        try:
            out["ship_included"] = str(oo.get("ship_included") or oo.get("shipping_included") or "").strip()
        except Exception as e:
            _swallow(e)

        return out

    def _calc_total(items: list, price_map: dict):
        total = 0.0
        for i in range(1, len(items) + 1):
            v = str(price_map.get(str(i), "")).strip()
            if not v:
                continue
            total += _to_float(v)
        return _fmt_money_num(total)

    def _calc_totals(goods_total_str: str, shipping_fee_str: str):
        g = _to_float(goods_total_str)
        s = _to_float(shipping_fee_str)
        t = g + s
        return _fmt_money_num(g), _fmt_money_num(s), _fmt_money_num(t)

    def _save_amounts_to_order(
        order_id: str,
        items: list,
        price_map: dict,
        shipping_fee: str = "",
        ship_included: str = "",
    ):
        """
        ✅ أهم إصلاح: نحفظ في الإكسل أرقام صحيحة ليستفيد منها (العميل/الفاتورة/الإشعارات)
        - quote_item_prices: خريطة أسعار القطع
        - goods_amount_sar: إجمالي القطع
        - shipping_fee_sar: رسوم الشحن
        - total_amount_sar: إجمالي (قطع + شحن)
        - ship_included: yes/no
        """
        try:
            goods_total = _calc_total(items, price_map)
        except Exception:
            goods_total = ""

        ship_fee_norm = _norm_money_str(shipping_fee)
        if ship_fee_norm == "":
            # إذا مشمول نعتبره 0، غير ذلك نخليه فاضي
            inc = str(ship_included or "").strip().lower()
            if inc in ("yes", "true", "1", "included", "مشمولة", "مشمول"):
                ship_fee_norm = "0"

        g_str, s_str, t_str = _calc_totals(goods_total, ship_fee_norm or "0")

        payload = {
            "quote_item_prices": dict(price_map or {}),
            "goods_amount_sar": g_str,
            "shipping_fee_sar": (s_str if ship_fee_norm != "" else ""),  # نخليها فاضية إذا غير محددة
            "total_amount_sar": (t_str if ship_fee_norm != "" else g_str),  # إذا ما تحدد شحن -> الإجمالي = القطع
        }

        if ship_included:
            payload["ship_included"] = str(ship_included).strip().lower()

        try:
            update_order_fields(order_id, payload)
        except Exception as e:
            _swallow(e)

        return g_str, (s_str if ship_fee_norm != "" else ""), (t_str if ship_fee_norm != "" else g_str)

    def _prices_lines(items: list, price_map: dict, limit: int = 12) -> str:
        """
        ✅ سطور تعرض القطع المسعرة حتى الآن + سعر كل قطعة (للرسائل)
        """
        lines = []
        for i, it in enumerate(items, start=1):
            nm = (it.get("name") or "").strip() or f"قطعة {i}"
            pn = (it.get("part_no") or "").strip()
            pv = str(price_map.get(str(i), "")).strip()
            if not pv:
                continue
            tail = f"{pv} ر.س"
            label = f"{i}) {nm}"
            if pn:
                label += f" ({pn})"
            label += f" — {tail}"
            lines.append(label)
            if len(lines) >= limit:
                break
        return "\n".join(lines) if lines else "—"

    def _hdr(
        order_id: str,
        snap: dict = None,
        goods_total: str = "",
        ship_fee: str = "",
        total_amt: str = "",
    ) -> str:
        snap = snap or _order_snapshot(order_id)
        client_name = (snap.get("client_name") or "—").strip()

        # ✅ نأخذ الأحدث (من الرسالة الحالية)، وإلا من الإكسل
        goods_now = (str(goods_total or "").strip() or str(snap.get("goods_amount") or "").strip())
        ship_now_raw = (str(ship_fee or "").strip())
        total_now_raw = (str(total_amt or "").strip())

        lines = []
        lines.append(f"{name}")
        lines.append(f"🧾 رقم الطلب: {order_id}")
        lines.append(f"👤 العميل: {client_name}")

        if goods_now and _to_float(goods_now) > 0:
            lines.append(f"💰 إجمالي القطع: {_fmt_money_num(goods_now)} ر.س")

        # الشحن
        if ship_now_raw != "":
            if _to_float(ship_now_raw) > 0:
                lines.append(f"🚚 الشحن: {_fmt_money_num(ship_now_raw)} ر.س")
            else:
                # لو صفر نعرضه كصفر (أفضل من الاختفاء حتى لا يلتبس)
                lines.append("🚚 الشحن: 0 ر.س")

        # الإجمالي النهائي
        if total_now_raw != "":
            if _to_float(total_now_raw) > 0:
                lines.append(f"🧾 الإجمالي: {_fmt_money_num(total_now_raw)} ر.س")

        return "\n".join(lines)

    # ===== helpers (داخلية فقط) =====
    def _get_items(order_id: str):
        oid = (str(order_id or "")).strip()
        if not oid:
            return []

        try:
            ensure_workbook()
        except Exception as e:
            _swallow(e)

        try:
            obx = get_order_bundle(oid) or {}
            its = (obx.get("items") or []) if isinstance(obx, dict) else []
        except Exception:
            its = []

        if not isinstance(its, list):
            its = []

        out = []
        for it in its:
            if isinstance(it, dict):
                nm = (it.get("name") or it.get("item_name") or "").strip()
                pn = (it.get("part_no") or it.get("partno") or it.get("part_number") or "").strip()
                out.append({"name": nm, "part_no": pn})
            else:
                s0 = str(it).strip()
                out.append({"name": s0, "part_no": ""})
        return out

    def _get_price_map_for_order(order_id: str):
        qs = _qget(td, order_id, create=True)
        pm = qs.get("item_prices")
        if not isinstance(pm, dict):
            pm = {}
            qs["item_prices"] = pm

        out = {}
        for k, v in pm.items():
            ks = str(k).strip()
            vs = str(v).strip()
            if ks.isdigit() and vs:
                out[ks] = vs

        qs["item_prices"] = out
        qs["last_touch"] = int(_dt_utc_now().timestamp())
        return out

    def _set_price_map_for_order(order_id: str, new_pm: dict):
        qs = _qget(td, order_id, create=True)
        qs["item_prices"] = new_pm if isinstance(new_pm, dict) else {}
        qs["last_touch"] = int(_dt_utc_now().timestamp())
        td["quote_item_prices"] = qs["item_prices"]

    def _items_kb(order_id: str, items: list, price_map: dict):
        oid = (str(order_id or "")).strip()
        rows = []
        for i, it in enumerate(items, start=1):
            nm = (it.get("name") or "").strip() or f"قطعة {i}"
            pn = (it.get("part_no") or "").strip()
            price = str(price_map.get(str(i), "")).strip()
            tail = f" — {price} ر.س" if price else " — اضف سعر"
            label = f"🧩 {i}) {nm}"
            if pn:
                label += f" ({pn})"
            label += tail
            rows.append([InlineKeyboardButton(label, callback_data=f"ppq_it|{oid}|{i}")])

        rows.append([InlineKeyboardButton("✅ اكمال خطوات العرض", callback_data=f"ppq_it_done|{oid}")])
        return InlineKeyboardMarkup(rows)

    # ===========================
    # بدء بناء عرض السعر
    # ===========================
    if action == "ppq_begin":
        if len(parts) < 2:
            return
        order_id = (parts[1] or "").strip()
        if not order_id:
            return

        # منع بناء عرض سعر اذا الطلب مقفول / ملغي / بعد سداد قيمة القطع
        try:
            ob = get_order_bundle(order_id) or {}
            oo = (ob.get("order", {}) or {}) if isinstance(ob, dict) else {}
        except Exception:
            oo = {}

        order_status = str(oo.get("order_status") or "").strip().lower()
        quote_locked = str(oo.get("quote_locked") or "").strip().lower()
        goods_pay_status = str(oo.get("goods_payment_status") or "").strip().lower()
        cancelled_by_client = str(oo.get("cancelled_by_client_id") or "").strip()

        locked_due_paid = goods_pay_status in ("awaiting_confirm", "confirmed")
        locked_due_admin = order_status in ("canceled", "cancelled", "ملغي") and (not cancelled_by_client)
        locked_due_client = bool(cancelled_by_client) and (order_status in ("canceled", "cancelled", "ملغي") or quote_locked in ("1", "true", "yes", "on"))
        locked_due_flag = quote_locked in ("1", "true", "yes", "on")

        if locked_due_client or locked_due_admin or locked_due_paid or locked_due_flag or order_status in ("closed", "delivered"):
            if locked_due_client:
                msg_lock = "⛔ الطلب ملغي من قبل العميل ولا يقبل عروض جديدة"
            elif locked_due_admin:
                msg_lock = "⛔ الطلب معلق من الإدارة ولا يقبل عروض جديدة"
            elif locked_due_paid:
                msg_lock = "🔒 تم دفع قيمة القطع والطلب قيد التحديث ولا يقبل عروض جديدة"
            else:
                msg_lock = "🔒 الطلب مقفول حالياً ولا يقبل عروض جديدة"

            await _alert(q, msg_lock, force=True)
            return

        # ✅ جلسة مستقلة لهذا الطلب
        _qreset(td, order_id)
        qs = _qget(td, order_id, create=True)
        qs["item_prices"] = {}
        qs.pop("pending_item_idx", None)
        qs.pop("pending_item_name", None)
        qs["last_touch"] = int(_dt_utc_now().timestamp())
        _qgc(td, keep_last=50)

        td["quote_order_id"] = order_id
        td["quote_step"] = "it_pick"
        td["quote_started_at_utc"] = utc_now_iso()
        set_stage(context, user_id, STAGE_TRADER_SET_QUOTE)

        td["quote_item_prices"] = {}
        td.pop("quote_pending_item_idx", None)
        td.pop("quote_pending_item_name", None)

        # ✅ تنظيف بقايا عرض سابق حتى لا تظهر قيم شحن/إجمالي لتاجر آخر
        td.pop("quote_goods_amount", None)
        td.pop("quote_shipping_fee", None)
        td.pop("quote_ship_included", None)
        td.pop("quote_ship_eta", None)
        td.pop("quote_availability", None)

        its = _get_items(order_id)
        if not its:
            await q.message.reply_text(
                _hdr(order_id) + "\n\n⚠️ لا توجد بنود داخل هذا الطلب لتسعيرها.",
                disable_web_page_preview=True,
            )
            return

        pm = _get_price_map_for_order(order_id)

        # ✅ حفظ/تصفير القيم الصحيحة في الإكسل من البداية
        g, s_fee, t = _save_amounts_to_order(order_id, its, pm)

        snap0 = _order_snapshot(order_id)
        await q.message.reply_text(
            _hdr(order_id, snap=snap0, goods_total=g, ship_fee=s_fee, total_amt=t)
            + "\n\n🧩 اختر القطعة المراد تسعيرها، ثم اضغط زر «إكمال خطوات العرض» أدناه:",
            reply_markup=_items_kb(order_id, its, pm),
            disable_web_page_preview=True,
        )
        return
    # ===========================
    # كل الاكشنات التالية تتطلب order_id
    # ===========================
    if len(parts) < 2:
        return
    order_id = (parts[1] or "").strip()
    if not order_id:
        return

    td["quote_order_id"] = order_id
    snap = _order_snapshot(order_id)

    # ===========================
    # ✅ أكشنات المعاينة قبل/بعد الإرسال (A+B + Versioning)
    # ===========================

    # ✅ جلسة العرض لهذا الطلب
    s = _qget(td, order_id, create=True)
    if not isinstance(s, dict):
        s = {}
        try:
            td.setdefault("quote_sessions", {})[order_id] = s
        except Exception as e:
            _swallow(e)

    def _q_is_sent_local(ss: dict) -> bool:
        try:
            return str((ss or {}).get("sent") or "").strip().lower() in ("1", "yes", "true", "sent", "done")
        except Exception:
            return False

    def _q_mark_sent_local(ss: dict) -> None:
        try:
            ss["sent"] = "1"
            ss["sent_at_utc"] = utc_now_iso()
        except Exception as e:
            _swallow(e)

    def _q_get_version_local(ss: dict) -> int:
        try:
            v = int((ss or {}).get("version") or 1)
            return v if v >= 1 else 1
        except Exception:
            return 1

    def _q_bump_version_local(ss: dict) -> int:
        try:
            v = _q_get_version_local(ss) + 1
        except Exception:
            v = 2
        try:
            ss["version"] = v
            ss.pop("sent", None)
            ss.pop("sent_at_utc", None)
        except Exception as e:
            _swallow(e)
        return v

    def _locked_kb(oid: str) -> InlineKeyboardMarkup:
        return InlineKeyboardMarkup([
            [InlineKeyboardButton(_wide_btn_label("✅ تم إرسال العرض للعميل"), callback_data="pp_ui_locked")],
            [InlineKeyboardButton(_wide_btn_label("🆕 إنشاء عرض جديد (إصدار جديد)"), callback_data=f"ppq_new_version|{oid}")],
            [InlineKeyboardButton(_wide_btn_label("✖️ إغلاق"), callback_data="pp_ui_close")],
        ])

    async def _start_items_pricing_from_scratch(oid: str):
        # 🔁 الرجوع لأول مرحلة (تسعير القطع)
        _qreset(td, oid)
        qs0 = _qget(td, oid, create=True)
        try:
            qs0["item_prices"] = {}
            qs0.pop("pending_item_idx", None)
            qs0.pop("pending_item_name", None)
            qs0["last_touch"] = int(_dt_utc_now().timestamp())
        except Exception as e:
            _swallow(e)
        _qgc(td, keep_last=50)

        td["quote_order_id"] = oid
        td["quote_step"] = "it_pick"
        td["quote_started_at_utc"] = utc_now_iso()
        set_stage(context, user_id, STAGE_TRADER_SET_QUOTE)

        td["quote_item_prices"] = {}
        td.pop("quote_pending_item_idx", None)
        td.pop("quote_pending_item_name", None)

        its0 = _get_items(oid)
        if not its0:
            try:
                await q.message.reply_text(
                    _hdr(oid) + "\n\n⚠️ لا توجد بنود داخل هذا الطلب لتسعيرها.",
                    disable_web_page_preview=True,
                )
            except Exception as e:
                _swallow(e)
            return

        pm0 = _get_price_map_for_order(oid)

        # ✅ حفظ/تصفير القيم الصحيحة في الإكسل من البداية
        g0, s0, t0 = _save_amounts_to_order(oid, its0, pm0)

        snap0 = _order_snapshot(oid)
        try:
            await q.message.reply_text(
                _hdr(oid, snap=snap0, goods_total=g0, ship_fee=s0, total_amt=t0)
                + "\n\n🧩 اختر القطعة المراد تسعيرها، ثم اضغط زر «إكمال خطوات العرض» أدناه:",
                reply_markup=_items_kb(oid, its0, pm0),
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)

    # ✅ (B+) قفل شامل: لا يسمح بأي تعديل بعد الإرسال حتى لو ضغط أزرار قديمة
    if _q_is_sent_local(s) and action not in ("ppq_new_version", "pp_ui_close"):
        await _alert(q, "⚠️ تم إرسال العرض للعميل ولا يمكن تعديله. أنشئ عرضًا جديدًا (إصدار جديد).")
        try:
            await q.message.edit_reply_markup(reply_markup=_locked_kb(order_id))
        except Exception as e:
            _swallow(e)
        return
    # --------- (Versioning) إنشاء إصدار جديد ----------
    if action == "ppq_new_version":
        if not _q_is_sent_local(s):
            await _alert(q, "ℹ️ العرض الحالي لم يُرسل بعد. يمكنك تعديله من الأزرار الحالية.")
            return

        new_v = _q_bump_version_local(s)

        await _alert(q, f"✅ تم إنشاء عرض جديد (إصدار V{new_v}). ابدأ تسعير القطع من البداية.")

        try:
            await q.message.edit_reply_markup(reply_markup=None)
        except Exception as e:
            _swallow(e)

        await _start_items_pricing_from_scratch(order_id)
        return
    # --------- (A) إرسال/إعادة بناء/إلغاء قبل الإرسال ----------
    if action == "ppq_preview_send":
        await finalize_quote_send(context, user_id, q.message, order_id)

        # ✅ قفل العرض بعد الإرسال
        _q_mark_sent_local(s)
        try:
            await q.message.edit_reply_markup(reply_markup=_locked_kb(order_id))
        except Exception:
            try:
                await q.message.edit_reply_markup(reply_markup=None)
            except Exception as e:
                _swallow(e)
        return

    if action == "ppq_preview_restart":
        await _alert(q, "🔁 تم إعادة بناء العرض من البداية.")
        try:
            await q.message.edit_reply_markup(reply_markup=None)
        except Exception as e:
            _swallow(e)
        await _start_items_pricing_from_scratch(order_id)
        return

    if action == "ppq_preview_cancel":
        _qreset(td, order_id)
        td["quote_step"] = "done"
        set_stage(context, int(user_id or 0), STAGE_NONE)
        await _alert(q, "✖️ تم إلغاء إعداد العرض (لم يتم إرسال شيء للعميل).")
        try:
            await q.message.edit_reply_markup(reply_markup=None)
        except Exception as e:
            _swallow(e)
        return

    # ✅ أكشنات تسعير القطع
    if action in ("ppq_it", "ppq_it_all", "ppq_it_none", "ppq_it_done"):
        its = _get_items(order_id)
        pm = _get_price_map_for_order(order_id)
        n = len(its)

        if not its:
            await _alert(q, "لا توجد بنود")
            return

        if action == "ppq_it":
            if len(parts) < 3:
                return
            try:
                idx = int(parts[2])
            except Exception:
                idx = 0
            if not (1 <= idx <= n):
                await _alert(q, "تعذر تحديد القطعة")
                return

            it = its[idx - 1]
            nm = (it.get("name") or "").strip() or f"قطعة {idx}"
            pn = (it.get("part_no") or "").strip()

            qs = _qget(td, order_id, create=True)
            qs["pending_item_idx"] = idx
            qs["pending_item_name"] = nm
            qs["last_touch"] = int(_dt_utc_now().timestamp())

            td["quote_pending_item_idx"] = idx
            td["quote_pending_item_name"] = nm

            td["quote_step"] = "it_price"
            try:
                await q.message.edit_reply_markup(reply_markup=None)
            except Exception as e:
                _swallow(e)

            # ✅ عرض ملخص الأسعار حتى الآن + الإجمالي الصحيح
            g, s_fee, t = _save_amounts_to_order(
                order_id,
                its,
                pm,
                shipping_fee=snap.get("shipping_fee") or "",
                ship_included=snap.get("ship_included") or "",
            )

            extra = f"\n🏷️ رقم القطعة: {pn}" if pn else ""
            await q.message.reply_text(
                _hdr(order_id, snap=snap, goods_total=g, ship_fee=s_fee, total_amt=t) +
                "\n\n🧾 الأسعار المدخلة حتى الآن:\n" +
                _prices_lines(its, pm) +
                "\n\n💬 اكتب سعر هذه القطعة (أرقام فقط)\n"
                f"🧩 القطعة: {nm}{extra}\n\n"
                "مثال: 120 أو 120.50",
                disable_web_page_preview=True,
            )
            return

        elif action == "ppq_it_all":
            td["quote_step"] = "it_all_price"
            try:
                await q.message.edit_reply_markup(reply_markup=None)
            except Exception as e:
                _swallow(e)

            g, s_fee, t = _save_amounts_to_order(
                order_id,
                its,
                pm,
                shipping_fee=snap.get("shipping_fee") or "",
                ship_included=snap.get("ship_included") or "",
            )

            await q.message.reply_text(
                _hdr(order_id, snap=snap, goods_total=g, ship_fee=s_fee, total_amt=t) +
                "\n\n💬 اكتب سعر واحد لتطبيقه على جميع القطع (أرقام فقط)\n"
                "مثال: 50 أو 75.5",
                disable_web_page_preview=True,
            )
            return

        elif action == "ppq_it_none":
            pm = {}
            _set_price_map_for_order(order_id, pm)

            qs = _qget(td, order_id, create=True)
            qs.pop("pending_item_idx", None)
            qs.pop("pending_item_name", None)
            qs["last_touch"] = int(_dt_utc_now().timestamp())

            td.pop("quote_pending_item_idx", None)
            td.pop("quote_pending_item_name", None)
            td["quote_step"] = "it_pick"

            # ✅ تصفير الحقول الصحيحة في الإكسل
            try:
                update_order_fields(order_id, {
                    "quote_item_prices": {},
                    "goods_amount_sar": "0",
                    "total_amount_sar": "0",
                })
            except Exception as e:
                _swallow(e)

        elif action == "ppq_it_done":
            has_any = any(str(pm.get(str(i), "")).strip() for i in range(1, n + 1))
            if not has_any:
                await _alert(q, "سعّر قطعة واحدة على الأقل")
                return

            # ✅ إجمالي القطع الصحيح + حفظه في الإكسل
            ship_fee_existing = snap.get("shipping_fee") or ""
            ship_inc = snap.get("ship_included") or ""
            g, s_fee, t = _save_amounts_to_order(order_id, its, pm, shipping_fee=ship_fee_existing, ship_included=ship_inc)

            _set_price_map_for_order(order_id, pm)
            td["quote_goods_amount"] = g
            td["quote_step"] = "type"

            try:
                await q.message.edit_reply_markup(reply_markup=None)
            except Exception as e:
                _swallow(e)

            await q.message.reply_text(
                _hdr(order_id, snap=snap, goods_total=g, ship_fee=s_fee, total_amt=t) +
                "\n\n🧾 الأسعار النهائية للقطع:\n" +
                _prices_lines(its, pm) +
                "\n\n🧩 اختر نوع القطع من الأزرار:",
                reply_markup=trader_quote_type_kb(order_id),
                disable_web_page_preview=True,
            )
            return

        # إعادة عرض كيبورد القطع بعد أي تحديث
        try:
            pm2 = _get_price_map_for_order(order_id)
            await q.message.edit_reply_markup(reply_markup=_items_kb(order_id, its, pm2))
        except Exception as e:
            _swallow(e)

        await _alert(q, "تم التحديث")
        return

    # ========= باقي التدفق: نوع القطع / الشحن / التجهيز / مدة الشحن =========
    if action == "ppq_type":
        if len(parts) < 3:
            return
        td["quote_parts_type"] = parts[2]
        td["quote_step"] = "shipping_method"
        try:
            await q.message.edit_reply_markup(reply_markup=None)
        except Exception as e:
            _swallow(e)

        # ✅ تحديث سريع للهيدر من الأرقام المخزنة
        snap = _order_snapshot(order_id)
        await q.message.reply_text(
            _hdr(order_id, snap=snap) + "\n\n🚚 اختر طريقة التسليم:",
            reply_markup=trader_quote_shipping_method_kb(order_id),
            disable_web_page_preview=True,
        )
        return

    if action == "ppq_ship":
        if len(parts) < 3:
            return
        td["quote_ship_method"] = parts[2]
        td.pop("quote_ship_carrier", None)
        td.pop("quote_shipping_fee", None)
        td["quote_step"] = "shipinc"
        try:
            await q.message.edit_reply_markup(reply_markup=None)
        except Exception as e:
            _swallow(e)

        snap = _order_snapshot(order_id)
        await q.message.reply_text(
            _hdr(order_id, snap=snap) + "\n\n🚚 هل السعر يشمل الشحن؟",
            reply_markup=trader_quote_shipping_included_kb(order_id),
            disable_web_page_preview=True,
        )
        return

    if action == "ppq_shipinc":
        if len(parts) < 3:
            return
        v_inc = parts[2]
        td["quote_ship_included"] = v_inc

        # ✅ نضمن حفظ ship_included + رسوم الشحن + الإجمالي النهائي الصحيح
        its = _get_items(order_id)
        pm = _get_price_map_for_order(order_id)

        if v_inc == "yes":
            td["quote_shipping_fee"] = "0"
            td["quote_step"] = "availability"

            # ✅ (تعديل سطرين) حفظ مباشر في الإكسل لضمان عدم ظهور الشحن فارغ لاحقًا
            try:
                update_order_fields(order_id, {"ship_included": "yes", "shipping_fee_sar": "0"})
            except Exception as e:
                _swallow(e)

            g, s_fee, t = _save_amounts_to_order(
                order_id,
                its,
                pm,
                shipping_fee="0",
                ship_included="yes",
            )

            try:
                await q.message.edit_reply_markup(reply_markup=None)
            except Exception as e:
                _swallow(e)

            await q.message.reply_text(
                _hdr(order_id, snap=_order_snapshot(order_id), goods_total=g, ship_fee=s_fee, total_amt=t) +
                "\n\n⏳ اختر مدة التجهيز:",
                reply_markup=trader_quote_availability_kb(order_id),
                disable_web_page_preview=True,
            )
            return

        # غير مشمول: نطلب قيمة الشحن
        td["quote_step"] = "shipping_fee"
        # نخلي shipping_fee تُدخل بالنص داخل text_handler

        try:
            update_order_fields(order_id, {"ship_included": "no", "shipping_fee_sar": ""})
        except Exception as e:
            _swallow(e)

        g, s_fee, t = _save_amounts_to_order(
            order_id,
            its,
            pm,
            shipping_fee="",
            ship_included="no",
        )

        try:
            await q.message.edit_reply_markup(reply_markup=None)
        except Exception as e:
            _swallow(e)

        await q.message.reply_text(
            _hdr(order_id, snap=_order_snapshot(order_id), goods_total=g, ship_fee=s_fee, total_amt=t) +
            "\n\n💬 اكتب قيمة الشحن (أرقام فقط)\n"
            "مثال: 25 أو 40.5",
            disable_web_page_preview=True,
        )
        return

    if action == "ppq_avail":
        if len(parts) < 3:
            return
        v = parts[2]
        if v == "custom":
            td["quote_step"] = "avail_custom"
            try:
                await q.message.edit_reply_markup(reply_markup=None)
            except Exception as e:
                _swallow(e)
            await q.message.reply_text(
                _hdr(order_id, snap=_order_snapshot(order_id)) + "\n\n💬 اكتب مدة التجهيز (مثال: 5 أيام)",
                disable_web_page_preview=True,
            )
            return

        td["quote_availability"] = v
        td["quote_step"] = "eta"
        try:
            await q.message.edit_reply_markup(reply_markup=None)
        except Exception as e:
            _swallow(e)
        await q.message.reply_text(
            _hdr(order_id, snap=_order_snapshot(order_id)) + "\n\n🚚 اختر مدة الشحن:",
            reply_markup=trader_quote_eta_kb(order_id),
            disable_web_page_preview=True,
        )
        return

    if action == "ppq_eta":
        if len(parts) < 3:
            return
        v = parts[2]
        if v == "custom":
            td["quote_step"] = "eta_custom"
            try:
                await q.message.edit_reply_markup(reply_markup=None)
            except Exception as e:
                _swallow(e)
            await q.message.reply_text(
                _hdr(order_id, snap=_order_snapshot(order_id)) + "\n\n💬 اكتب مدة الشحن (مثال: 2-3 أيام)",
                disable_web_page_preview=True,
            )
            return

        td["quote_ship_eta"] = v
        td["quote_step"] = "done"
        try:
            await q.message.edit_reply_markup(reply_markup=None)
        except Exception as e:
            _swallow(e)

        # ✅ قبل المعاينة: نثبت آخر أرقام صحيحة (قطع + شحن + إجمالي)
        try:
            its = _get_items(order_id)
            pm = _get_price_map_for_order(order_id)
            ship_fee = str(td.get("quote_shipping_fee") or _order_snapshot(order_id).get("shipping_fee") or "").strip()
            ship_inc = str(td.get("quote_ship_included") or _order_snapshot(order_id).get("ship_included") or "").strip()
            _save_amounts_to_order(order_id, its, pm, shipping_fee=ship_fee, ship_included=ship_inc)
        except Exception as e:
            _swallow(e)

        # ✅ بدل الإرسال النهائي مباشرة: نفتح معاينة مع أزرار (إرسال/إعادة/إلغاء)
        td["quote_step"] = "preview"
        await show_quote_preview(context, user_id, q.message, order_id)
        return

async def show_quote_preview(context: ContextTypes.DEFAULT_TYPE, trader_id: int, message, order_id: str):
    # نبني نفس نص العرض الرسمي لكن نعرضه للتاجر كمعاينة
    try:
        b0 = get_order_bundle(order_id)
        o0 = b0.get("order", {}) or {}
    except Exception:
        o0 = {}

    # ✅ اسم التاجر (بدون مخاطرة على _user_name(message))
    trader_name = "التاجر"
    try:
        trader_name = (getattr(getattr(message, "from_user", None), "full_name", "") or "").strip() or "التاجر"
    except Exception:
        trader_name = "التاجر"

    client_name = (str(o0.get("user_name") or "").strip() or "—")

    # ✅ خذ قيم الجلسة إن كانت موجودة (أحدث من الإكسل غالباً وقت المعاينة)
    td = {}
    try:
        td = (context.user_data or {}).setdefault(int(trader_id or 0), {}) or {}
    except Exception:
        td = {}

    def _pick(*vals, default=""):
        for v in vals:
            s = ("" if v is None else str(v)).strip()
            if s != "":
                return s
        return default

    # -----------------------------
    # ✅ مصادر المعاينة (الأولوية للجلسة td لأنها الأحدث)
    # -----------------------------

    goods_amount_sar = _pick(
        td.get("quote_goods_amount"),
        o0.get("goods_amount_sar"),
        default="",
    )

    shipping_fee_sar = _pick(
        td.get("quote_shipping_fee"),
        o0.get("shipping_fee_sar"),
        o0.get("quote_shipping_fee"),
        o0.get("shipping_fee"),
        o0.get("ship_fee"),
        default="",
    )

    # ✅ نوع القطع: خذ اختيار التاجر أولاً حتى لا يتحول لـ "غير محدد" في (تجاري/مختلط/وكلاء)
    parts_type = _pick(
        td.get("quote_parts_type"),
        o0.get("quote_parts_type"),
        o0.get("parts_type"),
        default="mixed",
    )

    # ✅ طريقة الشحن: لا تأخذ o0["ship_method"] نهائياً لأنها قد تكون (شحن/استلام) وتخرب label
    ship_method = _pick(
        td.get("quote_ship_method"),
        o0.get("quote_ship_method"),
        default="local",
    )

    ship_eta = _pick(
        td.get("quote_ship_eta"),
        o0.get("quote_ship_eta"),
        o0.get("ship_eta"),
        default="—",
    )

    ship_included = _pick(
        td.get("quote_ship_included"),
        o0.get("quote_ship_included"),
        o0.get("ship_included"),
        default="no",
    )

    availability = _pick(
        td.get("quote_availability"),
        o0.get("quote_availability"),
        o0.get("availability"),
        default="—",
    )

    ship_block = build_legal_shipping_block(
        method=ship_method,
        fee_sar=shipping_fee_sar,
        eta=ship_eta,
        included=ship_included,
    )

    official = build_official_quote_text(
        order_id=order_id,
        client_name=client_name,
        goods_amount_sar=goods_amount_sar,
        parts_type=parts_type,
        ship_block=ship_block,
        availability=availability,
        shipping_fee_sar=shipping_fee_sar,
        ship_included=(str(ship_included).lower() in ("yes", "true", "1", "included", "مشمولة", "مشمول")),
    )

    try:
        await message.reply_text(
            f"{trader_name}\n🧾 معاينة عرض السعر قبل الإرسال:\n\n{official}",
            reply_markup=trader_quote_preview_kb(order_id),
            disable_web_page_preview=True,
        )
    except Exception as e:
        _swallow(e)

async def finalize_quote_send(context: ContextTypes.DEFAULT_TYPE, trader_id: int, message, order_id: str):
    td = context.user_data.setdefault(int(trader_id or 0), {})

    # 🔧 وضع الصيانة: منع ارسال عروض جديدة (لغير الادمن)
    if _is_maintenance_mode() and int(trader_id or 0) not in ADMIN_IDS:
        try:
            await message.reply_text(
                f"{_user_name(message)}\n🟧 المنصة في وضع الصيانة حاليا\nتم ايقاف ارسال عروض السعر مؤقتا"
            )
        except Exception as e:
            _swallow(e)
        return

    # ✅ حماية: لا تسمح بإرسال عرض إذا الطلب مقفول/مقبول/مدفوع
    try:
        b0 = get_order_bundle(order_id)
        o0 = b0.get("order", {}) or {}
    except Exception:
        o0 = {}

    try:
        accepted_tid0 = int(o0.get("accepted_trader_id") or 0)
    except Exception:
        accepted_tid0 = 0

    locked0 = str(o0.get("quote_locked") or "").strip().lower() == "yes"
    qst0 = str(o0.get("quote_status") or "").strip().lower()
    gps0 = str(o0.get("goods_payment_status") or "").strip().lower()
    ost0 = str(o0.get("order_status") or "").strip().lower()

    if locked0 or gps0 in ("awaiting_confirm", "confirmed") or ost0 in ("closed", "delivered"):
        try:
            tname = (message.from_user.full_name or "").strip() if message and message.from_user else ""
            await message.reply_text(
                f"{_user_name(message)}\n"
                "⛔ هذا الطلب مقفول ولا يقبل عروض جديدة.\n"
                f"رقم الطلب: {order_id}"
            )
        except Exception as e:
            _swallow(e)
        return

    goods_amount = str(td.get("quote_goods_amount") or "").strip()
    parts_type = str(td.get("quote_parts_type") or "").strip()
    ship_method = str(td.get("quote_ship_method") or "").strip()
    ship_inc = str(td.get("quote_ship_included") or "").strip().lower() or "no"
    fee_sar = str(td.get("quote_shipping_fee") or "").strip()

    # ✅ لا يوجد رقم افتراضي للشحن — يحدده التاجر
    ship_inc_yes = ship_inc in ("yes", "true", "1", "included", "مشمولة", "مشمول")
    if not fee_sar:
        fee_sar = "0" if ship_inc_yes else ""

    ship_eta = str(td.get("quote_ship_eta") or "").strip() or "غير محدد"
    availability = str(td.get("quote_availability") or "").strip() or ship_eta

    if not goods_amount or not parts_type or not ship_method:
        try:
            await message.reply_text(f"{_user_name(message)}\nنقص في بيانات العرض اعد المحاولة من زر البدء")
        except Exception as e:
            _swallow(e)
        return

    ship_block = build_legal_shipping_block(ship_method, fee_sar, ship_eta, ship_inc)

    # ✅ اسم العميل من الطلب (وليس من رسالة التاجر)
    client_name = "—"
    try:
        ob0 = get_order_bundle(order_id) or {}
        o0 = (ob0.get("order") or {}) if isinstance(ob0, dict) else {}
        client_name = (o0.get("user_name") or o0.get("client_name") or o0.get("name") or "").strip() or "—"
    except Exception:
        client_name = "—"

    # ✅ مهم: تمرير الشحن + هل هو مشمول للدالة (عشان الإجمالي يكون صحيح وواضح)
    official = build_official_quote_text(
        order_id,
        client_name,
        goods_amount,
        parts_type,
        ship_block,
        availability,
        shipping_fee_sar=fee_sar,
        ship_included=bool(ship_inc_yes),
    )

    # ✅ تفصيل تسعير القطع داخل العرض: تفاصيل القطع فقط (بدون تكرار الشحن/الإجمالي)
    try:
        obx = get_order_bundle(order_id) or {}
        itemsx = (obx.get("items") or []) if isinstance(obx, dict) else []
    except Exception:
        itemsx = []

    pm0 = td.get("quote_item_prices")
    if not isinstance(pm0, dict):
        pm0 = {}

    pm = {}
    for k, v in pm0.items():
        ks = str(k).strip()
        vs = str(v).strip()
        if ks.isdigit() and vs:
            pm[ks] = vs

    items_block = ""
    try:
        if isinstance(itemsx, list) and itemsx:
            lines = [
                "━━━━━━━━━━━━",
                "🧩 تفاصيل تسعير القطع",
            ]

            for i in range(1, len(itemsx) + 1):
                it = itemsx[i - 1] if isinstance(itemsx[i - 1], dict) else {}
                nm = (it.get("name") or it.get("item_name") or "").strip() or f"قطعة {i}"
                pn = (it.get("part_no") or it.get("partno") or it.get("item_part_no") or "").strip()

                label = f"{i}) {nm}"
                if pn:
                    label += f" ({pn})"

                vv = str(pm.get(str(i), "")).strip()
                if vv:
                    lines.append(f"✅ {label}: {vv} ر.س")
                else:
                    lines.append(f"🟥 {label}: غير متوفرة")

                # حد بصري حتى ما تتكدس الرسالة
                if len(lines) >= 18:
                    remain = (len(itemsx) - i)
                    if remain > 0:
                        lines.append(f"… (+{remain} قطع أخرى)")
                    break

            lines.append("━━━━━━━━━━━━")
            items_block = "\n".join(lines) + "\n"
    except Exception:
        items_block = ""

    official2 = official
    try:
        key = f"🧾 رقم الطلب: {order_id}"
        if items_block:
            if key in official2:
                official2 = official2.replace(key, key + "\n" + items_block, 1)
            else:
                official2 = items_block + "\n" + official2
    except Exception:
        official2 = official

    # ✅ بيانات التاجر من لوحة التاجر (الاسم + الشركة)
    trader_profile = {}
    try:
        trader_profile = get_trader_profile(int(trader_id or 0)) or {}
    except Exception:
        trader_profile = {}

    trader_display = (trader_profile.get("display_name") or "").strip()
    if not trader_display:
        trader_display = (message.from_user.full_name or "").strip() if message and message.from_user else "تاجر"

    trader_company = (trader_profile.get("company_name") or "").strip()

    trader_header = f"👤 التاجر: {trader_display}"
    if trader_company:
        trader_header += f"\n🏢 المتجر: {trader_company}"

    # ✅ نقل اسم التاجر: قبل سطر "يرجى مراجعة العرض"
    official_with_trader = official2
    try:
        anchor = "يرجى مراجعة العرض"
        if anchor in official_with_trader:
            official_with_trader = official_with_trader.replace(
                anchor,
                f"{trader_header}\n\n{anchor}",
                1
            )
        else:
            anchor2 = "يرجى مراجعة"
            if anchor2 in official_with_trader:
                official_with_trader = official_with_trader.replace(
                    anchor2,
                    f"{trader_header}\n\n{anchor2}",
                    1
                )
            else:
                official_with_trader = official_with_trader.rstrip() + "\n\n" + trader_header
    except Exception:
        official_with_trader = official2

    # ✅ حفظ: ثبّت quoted_trader_id = trader_id (مو message.from_user)
    fields_to_update = {
        "goods_amount_sar": goods_amount,
        "parts_type": _ppq_type_label(parts_type),
        "ship_method": _ppq_ship_label(ship_method),
        "shipping_fee_sar": fee_sar,
        "ship_eta": ship_eta,
        "ship_included": "مشمولة" if ship_inc_yes else "غير مشمولة",
        "availability_days": availability,
        "quoted_trader_id": int(trader_id or 0),
        "quoted_trader_name": trader_display,
        "quote_item_prices": json.dumps(pm, ensure_ascii=False),
    }

    if str(o0.get("quote_status") or "").strip().lower() != "accepted":
        fields_to_update["quote_status"] = "sent"
        fields_to_update["order_status"] = "quoted"

    update_order_fields(order_id, fields_to_update)

    # ✅ نسخة احتياطية ذكية بعد حفظ العرض في الاكسل (بدون بطء + بدون تكرار سريع)
    try:
        app = getattr(context, "application", None)
        if app:
            if not app.bot_data.get("_backup_touch_finalize_quote"):
                app.bot_data["_backup_touch_finalize_quote"] = True

                async def _bk_job():
                    try:
                        await asyncio.sleep(5)
                        await _send_backup_excel(app, reason="quote_sent")
                    finally:
                        try:
                            app.bot_data["_backup_touch_finalize_quote"] = False
                        except Exception as e:
                            _swallow(e)

                asyncio.create_task(_bk_job())
    except Exception as e:
        _swallow(e)

    # ✅ ارسال للعميل + كيبورد يحمل trader_id
    client_id = 0
    try:
        b = get_order_bundle(order_id)
        order = b.get("order", {}) or {}
        client_id = int(order.get("user_id") or 0)
    except Exception:
        client_id = 0

    client_name = (order.get("user_name") or "").strip()
    if not client_name:
        client_name = "عزيزي العميل"

    if client_id:
        try:
            await context.bot.send_message(
                chat_id=client_id,
                text=(
                    f"👋 {client_name}\n"
                    "✅ وصلك عرض السعر الرسمي للطلب أدناه:\n\n"
                    f"{official_with_trader}\n\n"
                ),
                reply_markup=quote_client_kb(order_id, int(trader_id or 0)),
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)

    # ✅ نسخة للتاجر (مخصصة): اسم العميل + استبدال نص العميل بالكامل بنص التاجر
    trader_copy = official_with_trader
    try:
        client_real_name = (o0.get("user_name") or "").strip() or "غير محدد"
        trader_copy = f"👤 اسم العميل: {client_real_name}\n\n" + trader_copy

        old_block = (
            "يرجى مراجعة العرض ثم اختيار القرار من الأزرار بالأسفل.\n"
            "في حال قبول العرض سيتم فتح قناة تواصل مباشرة بين التاجر والعميل."
        )
        if old_block in trader_copy:
            trader_copy = trader_copy.replace(
                old_block,
                "📌 عند قبول العرض استخدم الأزرار المرفقة مع إشعارات الطلب لتحديث الحالة وإكمال البيع.",
                1
            )
    except Exception as e:
        _swallow(e)

    try:
        await message.reply_text(
            f"{_user_name(message)}\n"
            "✅ تم إرسال عرض السعر للعميل بنجاح.\n"
            f"{trader_copy}",
            reply_markup=trader_quote_preview_kb_locked(order_id),
            disable_web_page_preview=True,
        )
    except Exception as e:
        _swallow(e)

    td["quote_step"] = "done"
    set_stage(context, int(trader_id or 0), STAGE_NONE)
    return

async def trader_status_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    try:
        await q.answer()
    except Exception as e:
        _swallow(e)

    data = q.data or ""
    try:
        _, st, order_id = data.split("|", 2)
    except Exception:
        return

    order_id = (order_id or "").strip()
    if not order_id:
        return

    # ✅ تحويل زر إلى حالة داخلية
    st_norm = (st or "").strip().lower()

    # ✅ زر تخطي رقم التتبع (من نفس الهاندلر)
    skip_tracking_btn = (st_norm == "trk_skip")

    _st_map = {
        "prep": "preparing",
        "ready": "ready_to_ship",
        "shipped": "shipped",
        "delivered": "delivered",
        "closed": "closed",
        # زر خاص للتخطي
        "trk_skip": "shipped",
    }
    new_status = _st_map.get(st_norm, st_norm)
    if not new_status:
        return

    b = get_order_bundle(order_id)
    order = b.get("order", {}) or {}
    items = b.get("items", []) or []

    # IDs
    try:
        accepted_tid = int(order.get("accepted_trader_id") or 0)
    except Exception:
        accepted_tid = 0

    actor_id = q.from_user.id

    # ✅ منع التاجر المعطّل
    if actor_id not in ADMIN_IDS and _trader_is_disabled(actor_id):
        await _deny_disabled_trader_q(q, "لا يمكنك تحديث حالة الطلب لأن حسابك موقوف")
        return

    # ✅ سماح للتاجر المقبول فقط أو الأدمن
    if actor_id not in ADMIN_IDS and actor_id != accepted_tid:
        return

    # ===== أسماء العميل/التاجر/المتجر =====
    client_name = (order.get("user_name") or order.get("client_name") or order.get("customer_name") or "").strip() or "العميل"

    accepted_name = (order.get("accepted_trader_name") or order.get("quoted_trader_name") or "").strip()
    trader_store = (order.get("accepted_store_name") or order.get("shop_name") or order.get("store_name") or "").strip()

    tp = {}
    if (not accepted_name or not trader_store) and accepted_tid:
        try:
            tp = get_trader_profile(int(accepted_tid)) or {}
        except Exception:
            tp = {}
        if not accepted_name:
            accepted_name = (tp.get("display_name") or "").strip() or (tp.get("company_name") or "").strip()
        if not trader_store:
            trader_store = (tp.get("company_name") or "").strip() or (tp.get("store_name") or "").strip() or (tp.get("display_name") or "").strip()

    accepted_name = accepted_name or "التاجر"
    trader_store = trader_store or "المتجر"

    # ===== بيانات السيارة =====
    car_name = (
        (order.get("car_name") or order.get("vehicle_name") or order.get("car") or order.get("car_model") or "")
    )
    car_name = (str(car_name).strip() or "—")

    # ===== مدد (التجهيز / الشحن) =====
    availability_txt = (str(order.get("availability_days") or order.get("quote_availability") or order.get("availability") or "")).strip()
    ship_eta_txt = (str(order.get("ship_eta") or order.get("shipping_eta") or order.get("ship_days") or "")).strip()

    # ===== الإجماليات (القطع + الشحن) =====
    goods_total = 0
    shipping_fee_raw = 0

    try:
        goods_total = int(float(order.get("goods_total_sar") or order.get("goods_amount_sar") or 0))
    except Exception:
        goods_total = 0

    try:
        shipping_fee_raw = int(float(order.get("shipping_fee_sar") or 0))
    except Exception:
        shipping_fee_raw = 0

    try:
        ship_inc_raw = str(
            order.get("shipping_included")
            or order.get("ship_included")
            or order.get("shipping_fee_included")
            or ""
        ).strip().lower()
    except Exception:
        ship_inc_raw = ""

    shipping_included = ship_inc_raw in ("yes", "1", "true", "on", "included", "مشمولة", "مشمول")
    shipping_fee_effective = 0 if shipping_included else int(shipping_fee_raw or 0)

    grand_total = int((goods_total or 0) + (shipping_fee_effective or 0))

    # نصوص مالية موحدة
    goods_line = f"💰 مبلغ القطع: {_money(goods_total) if goods_total else '—'}"
    if shipping_included:
        ship_line = "🚚 الشحن: مشمولة ✅"
    else:
        ship_line = f"🚚 الشحن: {_money(shipping_fee_effective) if shipping_fee_effective else '—'}"
    total_line = f"🧾 الإجمالي: {_money(grand_total) if grand_total else '—'}"

    # ===== حالة دفع البضاعة (منع الشحن 100%) =====
    gps = str(order.get("goods_payment_status") or "").strip().lower()
    goods_paid = gps in ("confirmed", "paid", "success", "successful", "done", "ok")

    # ===== الحالة الحالية =====
    ost_now = str(order.get("order_status") or "").strip().lower()

    # ===== سيناريو التسلسل =====

    # 1) جاري تجهيز: مسموح فقط من البداية
    if new_status == "preparing":
        if ost_now not in ("", "new", "accepted", "quoted"):
            return

    # 2) جاهز للشحن: يطلب رفع فاتورة المتجر (إجباري) في الخاص
    if new_status == "ready_to_ship":
        if ost_now not in ("preparing", "prep"):
            return

        # ✅ تحديد وضع الدفع للتاجر (تحويل / رابط)
        try:
            tp_mode = get_trader_profile(int(actor_id or 0)) or {}
        except Exception:
            tp_mode = {}
        pay_mode = (str(tp_mode.get("payment_mode") or "").strip().lower())
        if pay_mode not in ("link", "bank"):
            pay_mode = "bank"

        # ✅ في وضع الرابط: نطلب رابط الدفع لهذا الطلب (مرة واحدة) بدل فاتورة المتجر
        pay_link_existing = (str(order.get("goods_payment_link") or "")).strip()

        inv_file = (str(order.get("seller_invoice_file_id") or order.get("shop_invoice_file_id") or "")).strip()
        pay_method = str(order.get("goods_payment_method") or "").strip().lower()
        if (not inv_file) and pay_link_existing and pay_method in ("pay_link", "link", "payment_link"):
            inv_file = "__PAYLINK__"

        if pay_mode == "link":
            if not pay_link_existing:
                # افتح مرحلة إدخال رابط الدفع في الخاص
                try:
                    ud2 = get_ud(context, actor_id)
                    ud2["tsu_kind"] = "goods_paylink"
                    ud2["tsu_order_id"] = order_id
                    set_stage(context, actor_id, STAGE_TRADER_STATUS_UPDATE)
                except Exception as e:
                    _swallow(e)

                sent_private = False
                try:
                    msg_private = (
                        "🔗 <b>إرسال رابط الدفع مطلوب</b>\n\n"
                        f"🧾 رقم الطلب: <b>{_order_id_link_html(order_id)}</b>\n"
                        f"👤 العميل: <b>{client_name}</b>\n"
                        f"🧑‍💼 التاجر: <b>{accepted_name}</b>\n"
                        f"🏪 المتجر: <b>{trader_store}</b>\n"
                        f"🚗 السيارة: <b>{car_name}</b>\n"
                        f"📌 الحالة: <b>{_pay_status_ar('ready_to_ship')}</b>\n"
                        + (f"🛠 مدة التجهيز: <b>{availability_txt}</b>\n" if availability_txt else "")
                        + (f"⏱ مدة الشحن: <b>{ship_eta_txt}</b>\n" if ship_eta_txt else "")
                        + "\n"
                        + f"{goods_line}\n{ship_line}\n{total_line}\n\n"
                        + "➡️ <b>الخطوة القادمة</b>:\n"
                        + "ارسل رابط الدفع الإلكتروني لهذا الطلب هنا في الخاص (نص يبدأ بـ https://).\n"
                        + "بعد إرسال الرابط ننتظر إيصال السداد من العميل."
                    )
                    await context.bot.send_message(
                        chat_id=actor_id,
                        text=msg_private,
                        parse_mode="HTML",
                        disable_web_page_preview=True,
                    )
                    sent_private = True
                except Exception as e:
                    _swallow(e)

                if not sent_private:
                    try:
                        await q.message.reply_text(
                            "تعذر ارسال رسالة لك في الخاص\n"
                            "افتح محادثة البوت ثم اعد ضغط زر الطلب جاهز للشحن",
                            disable_web_page_preview=True,
                        )
                    except Exception as e:
                        _swallow(e)
                    return

                # حدّث كيبورد رسالة المجموعة/اللوحة (يبقى زر جاهز للشحن فقط)
                try:
                    await q.message.edit_reply_markup(reply_markup=trader_status_kb(order_id))
                except Exception as e:
                    _swallow(e)
                return

            # إذا الرابط موجود أصلاً: التحويل إلى ready_to_ship يتم من هاندر إدخال الرابط
            return

        if not inv_file:
            # افتح مرحلة رفع الفاتورة في الخاص
            try:
                ud2 = get_ud(context, actor_id)
                ud2["tsu_kind"] = "seller_invoice"
                ud2["tsu_order_id"] = order_id
                set_stage(context, actor_id, STAGE_TRADER_STATUS_UPDATE)
            except Exception as e:
                _swallow(e)

            sent_private = False
            try:
                msg_private = (
                    "🧾 <b>رفع فاتورة المتجر مطلوب</b>\n\n"
                    f"🧾 رقم الطلب: <b>{_order_id_link_html(order_id)}</b>\n"
                    f"👤 العميل: <b>{client_name}</b>\n"
                    f"🧑‍💼 التاجر: <b>{accepted_name}</b>\n"
                    f"🏪 المتجر: <b>{trader_store}</b>\n"
                    f"🚗 السيارة: <b>{car_name}</b>\n"
                    f"📌 الحالة: <b>{_pay_status_ar('ready_to_ship')}</b>\n"
                    + (f"🛠 مدة التجهيز: <b>{availability_txt}</b>\n" if availability_txt else "")
                    + (f"⏱ مدة الشحن: <b>{ship_eta_txt}</b>\n" if ship_eta_txt else "")
                    + "\n"
                    + f"{goods_line}\n{ship_line}\n{total_line}\n\n"
                    + "➡️ <b>الخطوة القادمة</b>:\n"
                    + "ارسل فاتورة المتجر الرسمية (PDF أو صورة) هنا في الخاص.\n"
                    + "بعد رفع الفاتورة ننتظر إيصال السداد من العميل."
                )
                await context.bot.send_message(
                    chat_id=actor_id,
                    text=msg_private,
                    parse_mode="HTML",
                    disable_web_page_preview=True,
                )
                sent_private = True
            except Exception as e:
                _swallow(e)

            if not sent_private:
                try:
                    await q.message.reply_text(
                        "تعذر ارسال رسالة لك في الخاص\n"
                        "افتح محادثة البوت ثم اعد ضغط زر الطلب جاهز للشحن",
                        disable_web_page_preview=True,
                    )
                except Exception as e:
                    _swallow(e)
                return

            # حدّث كيبورد رسالة المجموعة/اللوحة (يبقى زر جاهز للشحن فقط)
            try:
                await q.message.edit_reply_markup(reply_markup=trader_status_kb(order_id))
            except Exception as e:
                _swallow(e)
            return

        # إذا الفاتورة موجودة أصلاً: التحويل إلى ready_to_ship يتم من هاندر رفع الفاتورة
        return

    # 3) تم الشحن: ممنوع 100% قبل السداد + لازم تكون الحالة ready_to_ship + لازم تكون الفاتورة موجودة
    if new_status == "shipped":
        if ost_now not in ("ready_to_ship", "ready"):
            return

        inv_file = (str(order.get("seller_invoice_file_id") or order.get("shop_invoice_file_id") or "")).strip()
        pay_method = str(order.get("goods_payment_method") or "").strip().lower()
        pay_link_existing = (str(order.get("goods_payment_link") or "")).strip()
        if (not inv_file) and pay_link_existing and pay_method in ("pay_link", "link", "payment_link"):
            inv_file = "__PAYLINK__"
        if not inv_file:
            return

        # ✅ ممنوع حتى لو الأدمن ضغط
        if not goods_paid:
            return

        # ✅ زر التخطي: ينفّذ الشحن فورًا بدون رقم تتبع (بدون كتابة)
        if skip_tracking_btn:
            # نكمّل تحت في "تنفيذ التحديث" مع حقول إضافية للتتبع الفارغ
            pass
        else:
            # ✅ فتح مرحلة إدخال رقم التتبع (اختياري) في الخاص + زر تخطي
            try:
                udt = get_ud(context, actor_id)
                udt["tsu_kind"] = "tracking"      # ✅ مطابق للـ text_handler (kind == "tracking")
                udt["tsu_order_id"] = order_id
                set_stage(context, actor_id, STAGE_TRADER_STATUS_UPDATE)
            except Exception as e:
                _swallow(e)

            try:
                kb = InlineKeyboardMarkup([
                    [InlineKeyboardButton("⏭️ تخطي رقم التتبع", callback_data=f"pp_trader_status|trk_skip|{order_id}")],
                ])
                await context.bot.send_message(
                    chat_id=actor_id,
                    text=(
                        "🚚 <b>تم اختيار تحديث الحالة إلى: تم الشحن</b>\n\n"
                        f"🧾 رقم الطلب: <b>{_order_id_link_html(order_id)}</b>\n"
                        f"👤 العميل: <b>{client_name}</b>\n"
                        f"🧑‍💼 التاجر: <b>{accepted_name}</b>\n"
                        f"🏪 المتجر: <b>{trader_store}</b>\n"
                        f"🚗 السيارة: <b>{car_name}</b>\n"
                        f"📌 الحالة: <b>{_pay_status_ar('shipped')}</b>\n"
                        + (f"🛠 مدة التجهيز: <b>{availability_txt}</b>\n" if availability_txt else "")
                        + (f"⏱ مدة الشحن: <b>{ship_eta_txt}</b>\n" if ship_eta_txt else "")
                        + "\n"
                        + f"{goods_line}\n{ship_line}\n{total_line}\n\n"
                        + "➡️ <b>أرسل رقم التتبع</b> الآن (اختياري)\n"
                        + "أو اضغط زر <b>(تخطي رقم التتبع)</b>."
                    ),
                    parse_mode="HTML",
                    disable_web_page_preview=True,
                    reply_markup=kb,
                )
            except Exception as e:
                _swallow(e)

            # ✅ مهم جدًا: لا ننفّذ shipped هنا — ننتظر رقم التتبع أو زر التخطي
            return

    # 4) تأكيد التسليم: فقط بعد shipped
    if new_status == "delivered":
        if ost_now not in ("shipped",):
            return

    # ===== تنفيذ التحديث =====
    fields: dict = {"order_status": new_status}

    if new_status == "shipped":
        fields["shipped_at_utc"] = utc_now_iso()

        # ✅ عند زر التخطي: سجل التتبع فارغ + وقت الشحن (متوافق مع text_handler)
        if skip_tracking_btn:
            fields["shipping_tracking"] = ""
            fields["shipping_at"] = utc_now_iso()
            try:
                udx = get_ud(context, actor_id)
                udx.pop("tsu_kind", None)
                udx.pop("tsu_order_id", None)
                set_stage(context, actor_id, STAGE_NONE)
            except Exception as e:
                _swallow(e)

            # ✅ إرسال فاتورة الشحن للتاجر (بدون رقم تتبع)
            try:
                await send_invoice_pdf(
                    context,
                    order_id,
                    kind="shipping",
                    invoice_for="trader",
                    include_admins=False,
                    tracking_number="",
                )
            except Exception as e:
                _swallow(e)

        # ⏳ مؤقت المراسلة 7 أيام (أول مرة)
        try:
            if not (order.get("chat_expires_at_utc") or "").strip():
                expires = datetime.now(timezone.utc) + timedelta(days=7)
                fields["chat_expires_at_utc"] = expires.isoformat()
        except Exception as e:
            _swallow(e)

    if new_status in ("delivered", "closed"):
        fields["closed_at_utc"] = utc_now_iso()

    update_order_fields(order_id, fields)

    try:
        log_event(
            order_id,
            "status_updated",
            actor_role="trader" if actor_id == accepted_tid else "admin",
            actor_id=actor_id,
            actor_name=_user_name(q),
            payload={"order_status": new_status},
        )
    except Exception as e:
        _swallow(e)

    # ✅ تحديث كيبورد رسالة التاجر الأصلية في المجموعة حسب الحالة الجديدة
    try:
        await q.message.edit_reply_markup(reply_markup=trader_status_kb(order_id))
    except Exception as e:
        _swallow(e)

    # ---------- تجهيز نص القطع ----------
    def _s(x: object) -> str:
        return ("" if x is None else str(x)).strip()

    def _to_float(x: object) -> float:
        try:
            return float(str(x or 0).replace(",", "").strip() or 0)
        except Exception:
            return 0.0

    def _money_tail(x: object, fb: str = "0") -> str:
        s = _s(x)
        if not s:
            s = fb
        s = re.sub(r'^\s*(ر\.?\s*س|ر\.س|SAR|SR|s\.r|s\.r\.?)\s*', '', s, flags=re.I)
        s = re.sub(r'\s*(ر\.?\s*س|ر\.س|SAR|SR|s\.r|s\.r\.?)\s*$', '', s, flags=re.I)
        s = (s.strip() or fb)
        return f"{s} ﷼"

    def _load_item_prices_map_from_order(o: dict) -> dict:
        pm = {}
        try:
            import ast
        except Exception:
            ast = None

        for k in ("quote_item_prices", "item_prices", "goods_item_prices", "quote_items_prices"):
            raw = o.get(k)
            if raw is None or raw == "":
                continue

            src = None
            try:
                if isinstance(raw, dict):
                    src = raw
                else:
                    sraw = str(raw).strip()
                    if sraw.startswith("{") and sraw.endswith("}"):
                        try:
                            src = json.loads(sraw)
                        except Exception:
                            src = None
                        if src is None and ast:
                            try:
                                src = ast.literal_eval(sraw)
                            except Exception:
                                src = None
            except Exception:
                src = None

            if isinstance(src, dict):
                for kk, vv in src.items():
                    ks = str(kk).strip()
                    vs = str(vv).strip()
                    if ks.isdigit() and vs and vs not in ("0", "0.0", "0.00"):
                        pm[ks] = vs

        return pm

    def _pick_item_price_local(i: int, it: dict, pm: dict) -> str:
        cand_keys = ("price_sar", "item_price", "price", "unit_price", "amount_sar", "cost_sar", "cost", "sar")
        for ck in cand_keys:
            v = it.get(ck) if isinstance(it, dict) else None
            vs = _s(v)
            if vs and vs not in ("0", "0.0", "0.00"):
                return vs
        vs = _s(pm.get(str(i)))
        if vs and vs not in ("0", "0.0", "0.00"):
            return vs
        return ""

    pm = _load_item_prices_map_from_order(order)

    priced_lines = []
    unavail_lines = []
    priced_count = 0
    unavail_count = 0
    parts_total = 0.0

    try:
        if isinstance(items, list) and items:
            for i, it in enumerate(items, start=1):
                nm = _s((it or {}).get("name") or (it or {}).get("item_name") or "—")
                pn = _s((it or {}).get("part_no") or (it or {}).get("item_part_no") or (it or {}).get("number") or "")
                pr = _pick_item_price_local(i, it if isinstance(it, dict) else {}, pm)

                title = f"{i}- {nm}"
                if pn:
                    title += f" ({pn})"

                if pr:
                    priced_count += 1
                    parts_total += _to_float(pr)
                    priced_lines.append(f"✅ {title}\n   💰 {_money_tail(pr)}")
                else:
                    unavail_count += 1
                    unavail_lines.append(f"⚠️ {title}\n   🚫 غير متوفرة لدى التاجر")
    except Exception:
        priced_lines, unavail_lines = [], []
        priced_count = unavail_count = 0
        parts_total = 0.0

    total_items = len(items) if isinstance(items, list) else 0
    parts_block = ""
    try:
        if total_items > 0:
            parts_block_lines = []
            parts_block_lines.append("🧩 <b>تفاصيل القطع</b>")
            parts_block_lines.append(f"✅ المتوفر: <b>{priced_count}</b> / {total_items}")
            if unavail_count:
                parts_block_lines.append(f"⚠️ غير متوفر: <b>{unavail_count}</b> / {total_items}")
            if priced_count:
                parts_block_lines.append(f"💰 إجمالي القطع المتوفرة: <b>{_money_tail(parts_total, fb='0')}</b>")
            parts_block_lines.append("")
            if priced_lines:
                parts_block_lines.append("✅ <b>القطع المتوفرة:</b>")
                parts_block_lines.extend(priced_lines)
            if unavail_lines:
                parts_block_lines.append("")
                parts_block_lines.append("⚠️ <b>القطع غير المتوفرة:</b>")
                parts_block_lines.extend(unavail_lines)
            parts_block = "\n".join(parts_block_lines)
    except Exception:
        parts_block = ""

    # ===== عرض الحالة =====
    display_status = _pay_status_ar(new_status)

    # ===== خطوة تالية (آخر الرسالة لكل جهة) =====
    next_step_client = ""
    next_step_trader = ""
    next_step_admin = ""

    if new_status in ("preparing", "prep"):
        next_step_client = "➡️ <b>الخطوة القادمة</b>: انتظر تجهيز الطلب من التاجر."
        next_step_trader = "➡️ <b>الخطوة القادمة</b>: بعد التجهيز اضغط (🟢 الطلب جاهز للشحن)."
        next_step_admin = "➡️ متابعة: التاجر بدأ التجهيز."
    elif new_status in ("ready_to_ship", "ready"):
        next_step_client = "➡️ <b>الخطوة القادمة</b>: سيتم تزويدك بفاتورة المتجر، ثم أرسل إيصال السداد لإكمال الشحن."
        next_step_trader = "➡️ <b>الخطوة القادمة</b>: ارفع فاتورة المتجر الرسمية ثم انتظر إيصال السداد من العميل."
        next_step_admin = "➡️ متابعة: بانتظار فاتورة المتجر/إيصال السداد."
    elif new_status == "shipped":
        next_step_client = (
            "➡️ <b>الخطوة القادمة</b>: عند استلام الطلب، يُتاح لك التواصل مع التاجر في حال وجود أي ملاحظات. "
            "سيظل زر التواصل نشطًا لمدة <b>7 أيام</b>، وبعدها يُغلق الطلب تلقائيًا وفق سياسة المنصة."
        )
        next_step_trader = "➡️ <b>الخطوة القادمة</b>: بعد التسليم اضغط (✅ تأكيد التسليم بنجاح)."
        next_step_admin = "➡️ متابعة: تم الشحن."
    elif new_status in ("delivered", "closed"):
        next_step_client = "✅ <b>تم إغلاق الطلب بنجاح</b>."
        next_step_trader = "✅ <b>تم إغلاق الطلب</b>."
        next_step_admin = "✅ تم إغلاق الطلب."

    # ===== كيبورد التاجر: إزالة أي زر مراسلة للعميل من داخل إشعارات الحالة =====
    def _strip_chat_buttons(kb: InlineKeyboardMarkup) -> InlineKeyboardMarkup:
        try:
            if not kb or not getattr(kb, "inline_keyboard", None):
                return kb
            out = []
            for row in kb.inline_keyboard:
                new_row = []
                for btn in row:
                    cd = (getattr(btn, "callback_data", None) or "")
                    if cd.startswith("pp_chat_open") or cd.startswith("pp_chat_"):
                        continue
                    new_row.append(btn)
                if new_row:
                    out.append(new_row)
            return InlineKeyboardMarkup(out) if out else None
        except Exception:
            return kb

    # ===== إشعار العميل =====
    client_id = 0
    try:
        client_id = int(order.get("user_id") or 0)
    except Exception:
        client_id = 0

    if client_id:
        try:
            client_msg_lines = []
            client_msg_lines.append("📦 <b>تحديث حالة الطلب</b>")
            client_msg_lines.append("")
            client_msg_lines.append(f"🧾 رقم الطلب: <b>{order_id}</b>")
            client_msg_lines.append(f"👤 العميل: <b>{client_name}</b>")
            client_msg_lines.append(f"🧑‍💼 التاجر: <b>{accepted_name}</b>")
            client_msg_lines.append(f"🏪 المتجر: <b>{trader_store}</b>")
            client_msg_lines.append(f"🚗 السيارة: <b>{car_name}</b>")
            client_msg_lines.append(f"📌 الحالة: <b>{display_status}</b>")
            if availability_txt:
                client_msg_lines.append(f"🛠 مدة التجهيز: <b>{availability_txt}</b>")
            if ship_eta_txt:
                client_msg_lines.append(f"⏱ مدة الشحن: <b>{ship_eta_txt}</b>")
            client_msg_lines.append("")
            client_msg_lines.append(f"{goods_line}")
            client_msg_lines.append(f"{ship_line}")
            client_msg_lines.append(f"{total_line}")

            if parts_block:
                client_msg_lines.append("")
                client_msg_lines.append(parts_block)

            if next_step_client:
                client_msg_lines.append("")
                client_msg_lines.append(next_step_client)

            msg = "\n".join(client_msg_lines).strip()

            show_client_chat = (new_status in ("preparing", "prep", "ready_to_ship", "ready", "shipped", "delivered", "closed")) and bool(_assigned_trader_id(order_id))
            await context.bot.send_message(
                chat_id=client_id,
                text=msg,
                parse_mode="HTML",
                reply_markup=client_trader_chat_kb(order_id) if show_client_chat else None,
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)

    # ===== إشعار الإدارة =====
    try:
        admin_kb = None
        try:
            admin_kb = admin_contact_kb(order_id)
        except Exception:
            admin_kb = None

        admin_msg_lines = []
        admin_msg_lines.append("📌 <b>تحديث حالة من التاجر</b>")
        admin_msg_lines.append("")
        admin_msg_lines.append(f"🧾 رقم الطلب: <b>{order_id}</b>")
        admin_msg_lines.append(f"👤 العميل: <b>{client_name}</b>")
        admin_msg_lines.append(f"🚗 السيارة: <b>{car_name}</b>")
        admin_msg_lines.append(f"🧑‍💼 التاجر: <b>{accepted_name}</b>")
        admin_msg_lines.append(f"🏪 المتجر: <b>{trader_store}</b>")
        admin_msg_lines.append(f"📌 الحالة: <b>{display_status}</b>")
        if availability_txt:
            admin_msg_lines.append(f"🛠 مدة التجهيز: <b>{availability_txt}</b>")
        if ship_eta_txt:
            admin_msg_lines.append(f"⏱ مدة الشحن: <b>{ship_eta_txt}</b>")
        admin_msg_lines.append("")
        admin_msg_lines.append(f"{goods_line}")
        admin_msg_lines.append(f"{ship_line}")
        admin_msg_lines.append(f"{total_line}")
        admin_msg_lines.append("")
        admin_msg_lines.append(f"🧑‍💼 تم التحديث بواسطة: <b>{_user_name(q)}</b>")

        if parts_block:
            admin_msg_lines.append("")
            admin_msg_lines.append(parts_block)

        if next_step_admin:
            admin_msg_lines.append("")
            admin_msg_lines.append(f"🧭 <b>متابعة</b>: {next_step_admin}")

        admin_msg = "\n".join(admin_msg_lines).strip()

        for aid in ADMIN_IDS:
            try:
                await context.bot.send_message(
                    chat_id=aid,
                    text=admin_msg,
                    parse_mode="HTML",
                    reply_markup=admin_kb,
                    disable_web_page_preview=True,
                )
            except Exception as e:
                _swallow(e)
    except Exception as e:
        _swallow(e)

    # ===== إشعار للتاجر (بدون زر مراسلة العميل نهائياً هنا) =====
    try:
        trader_msg_lines = []
        trader_msg_lines.append(f"{_user_name(q)}")
        trader_msg_lines.append("✅ <b>تم تحديث حالة الطلب</b>")
        trader_msg_lines.append("")
        trader_msg_lines.append(f"🧾 رقم الطلب: <b>{order_id}</b>")
        trader_msg_lines.append(f"👤 العميل: <b>{client_name}</b>")
        trader_msg_lines.append(f"🧑‍💼 التاجر: <b>{accepted_name}</b>")
        trader_msg_lines.append(f"🏪 المتجر: <b>{trader_store}</b>")
        trader_msg_lines.append(f"🚗 السيارة: <b>{car_name}</b>")
        trader_msg_lines.append(f"📌 الحالة: <b>{display_status}</b>")
        if availability_txt:
            trader_msg_lines.append(f"🛠 مدة التجهيز: <b>{availability_txt}</b>")
        if ship_eta_txt:
            trader_msg_lines.append(f"⏱ مدة الشحن: <b>{ship_eta_txt}</b>")
        trader_msg_lines.append("")
        trader_msg_lines.append(f"{goods_line}")
        trader_msg_lines.append(f"{ship_line}")
        trader_msg_lines.append(f"{total_line}")

        if parts_block:
            trader_msg_lines.append("")
            trader_msg_lines.append(parts_block)

        if next_step_trader:
            trader_msg_lines.append("")
            trader_msg_lines.append(next_step_trader)

        trader_msg = "\n".join(trader_msg_lines).strip()

        kb_trader = None
        try:
            kb_trader = _strip_chat_buttons(trader_status_kb(order_id))
        except Exception:
            kb_trader = None

        await context.bot.send_message(
            chat_id=actor_id,
            text=trader_msg,
            parse_mode="HTML",
            reply_markup=kb_trader,
            disable_web_page_preview=True,
        )
    except Exception as e:
        _swallow(e)
        

async def tsu_skip_tracking_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")

    user_id = q.from_user.id
    data = (q.data or "").strip()

    try:
        _, order_id = data.split("|", 1)
    except Exception:
        return

    order_id = (order_id or "").strip()
    if not order_id:
        return

    # ✅ اقرأ الطلب
    try:
        b = get_order_bundle(order_id) or {}
        order = b.get("order", {}) or {}
    except Exception:
        order = {}

    # ✅ سماح للتاجر المعين أو الأدمن فقط
    try:
        accepted_tid = int(order.get("accepted_trader_id") or 0)
    except Exception:
        accepted_tid = 0

    if user_id not in ADMIN_IDS and accepted_tid and user_id != accepted_tid:
        await _alert(q, "غير مصرح")
        return

    # ✅ منع التنفيذ المتكرر: إذا كان الطلب مشحون مسبقاً، لا نعيد الإرسال
    ost = str(order.get("order_status") or "").strip().lower()
    already_shipped = ost in ("shipped", "delivered", "closed")

    # ===== بيانات مختصرة للرسائل =====
    client_id = 0
    try:
        client_id = int(order.get("user_id") or 0)
    except Exception:
        client_id = 0

    client_name = (order.get("user_name") or order.get("client_name") or "").strip() or "العميل"

    tname = (order.get("accepted_trader_name") or order.get("quoted_trader_name") or "").strip() or "التاجر"
    trader_store = (order.get("accepted_trader_store") or order.get("trader_store") or "").strip()

    try:
        if accepted_tid:
            tp = get_trader_profile(int(accepted_tid)) or {}
            if not (order.get("accepted_trader_name") or "").strip():
                tname = (tp.get("display_name") or "").strip() or (tp.get("company_name") or "").strip() or tname
            if not trader_store:
                trader_store = (tp.get("company_name") or "").strip()
    except Exception as e:
        _swallow(e)

    # ✅ حساب المبلغ الإجمالي (القطع + الشحن) للعرض (إن وجد)
    goods_total = 0
    shipping_fee = 0
    try:
        goods_total = int(float(order.get("goods_total_sar") or order.get("goods_amount_sar") or 0))
    except Exception:
        goods_total = 0
    try:
        shipping_fee = int(float(order.get("shipping_fee_sar") or order.get("shipping_fee") or 0))
    except Exception:
        shipping_fee = 0

    ship_included = False
    try:
        ship_included = str(order.get("ship_included") or order.get("shipping_included") or "").strip().lower() in (
            "yes", "y", "true", "1", "مشمول", "included"
        )
    except Exception:
        ship_included = False

    grand_total = goods_total + (0 if ship_included else shipping_fee)
    ship_txt = "مشمول" if ship_included else (f"{shipping_fee} ر.س" if shipping_fee > 0 else "—")

    # ✅ نفّذ “تم الشحن بدون تتبع” مرة واحدة فقط
    if not already_shipped:
        fields = {
            "order_status": "shipped",
            "shipping_tracking": "",
            "shipping_at": utc_now_iso(),
            "shipped_at_utc": utc_now_iso(),
        }

        # ✅ حافظ على نظام المراسلة 7 أيام (إن لم يكن مثبت)
        try:
            if not (order.get("chat_expires_at_utc") or "").strip():
                expires = datetime.now(timezone.utc) + timedelta(days=7)
                fields["chat_expires_at_utc"] = expires.isoformat()
        except Exception as e:
            _swallow(e)

        try:
            update_order_fields(order_id, fields)
        except Exception as e:
            _swallow(e)

        # ✅ فاتورة الشحن للتاجر فقط (بدون الإدارة) — بدون رقم تتبع
        try:
            await send_invoice_pdf(
                context,
                order_id,
                kind="shipping",
                invoice_for="trader",
                include_admins=False,
                tracking_number="",
            )
        except Exception as e:
            try:
                await _notify_invoice_error(context, order_id, "فاتورة الشحن", e)
            except Exception as e:
                _swallow(e)

        # ✅ إشعار العميل: تم الشحن بدون تتبع (مع زر مراسلة التاجر)
        if client_id:
            try:
                lines = [
                    "✅ <b>تم شحن طلبك بنجاح</b>",
                    f"🧾 رقم الطلب: {html.escape(order_id)}",
                    "📦 رقم التتبع: <b>غير متوفر</b>",
                    f"🧑‍💼 التاجر: <b>{html.escape(tname)}</b>",
                ]
                if trader_store:
                    lines.append(f"🏪 المتجر: <b>{html.escape(trader_store)}</b>")
                lines.extend([
                    "",
                    f"🧩 قيمة القطع: <b>{goods_total} ر.س</b>" if goods_total > 0 else "🧩 قيمة القطع: —",
                    f"🚚 الشحن: <b>{html.escape(ship_txt)}</b>",
                    f"💰 الإجمالي (قطع + شحن): <b>{grand_total} ر.س</b>" if (goods_total > 0 or shipping_fee > 0 or ship_included) else "💰 الإجمالي (قطع + شحن): —",
                    "",
                    "يمكنك مراسلة التاجر أو المتابعة من الزر بالأسفل.",
                ])

                await context.bot.send_message(
                    chat_id=client_id,
                    text="\n".join(lines),
                    parse_mode="HTML",
                    reply_markup=chat_nav_kb_for(context, to_uid, order_id, "pp_chat_trader_done"),
                    disable_web_page_preview=True,
                )
            except Exception as e:
                _swallow(e)

        # ✅ إشعار الإدارة: تم الشحن بدون تتبع (بدون PDF)
        for aid in ADMIN_IDS:
            try:
                alines = [
                    "📦 <b>تحديث حالة: تم الشحن</b>",
                    f"🧾 الطلب: {html.escape(order_id)}",
                    f"👤 العميل: <b>{html.escape(client_name)}</b> (<code>{client_id}</code>)" if client_id else f"👤 العميل: <b>{html.escape(client_name)}</b>",
                    f"🧑‍💼 التاجر: <b>{html.escape(tname)}</b> (<code>{accepted_tid or user_id}</code>)",
                ]
                if trader_store:
                    alines.append(f"🏪 المتجر: <b>{html.escape(trader_store)}</b>")
                alines.extend([
                    "📦 التتبع: <b>غير متوفر</b>",
                    f"🧩 قيمة القطع: <b>{goods_total} ر.س</b>" if goods_total > 0 else "🧩 قيمة القطع: —",
                    f"🚚 الشحن: <b>{html.escape(ship_txt)}</b>",
                    f"💰 الإجمالي (قطع + شحن): <b>{grand_total} ر.س</b>" if (goods_total > 0 or shipping_fee > 0 or ship_included) else "💰 الإجمالي (قطع + شحن): —",
                ])

                await context.bot.send_message(
                    chat_id=int(aid),
                    text="\n".join(alines),
                    parse_mode="HTML",
                    disable_web_page_preview=True,
                )
            except Exception as e:
                _swallow(e)

        # ✅ إشعار التاجر (هذه هي الرسالة اللي كانت ناقصة) + 3 أزرار
        try:
            to_trader = int(accepted_tid or user_id or 0)
        except Exception:
            to_trader = 0

        if to_trader:
            try:
                tlines = [
                    f"{_user_name(q)}",
                    "✅ <b>تم تحديث الحالة إلى: تم الشحن</b>",
                    "",
                    f"🧾 رقم الطلب: <b>{html.escape(order_id)}</b>",
                    f"👤 العميل: <b>{html.escape(client_name)}</b>",
                    f"🧑‍💼 التاجر: <b>{html.escape(tname)}</b>",
                ]
                if trader_store:
                    tlines.append(f"🏪 المتجر: <b>{html.escape(trader_store)}</b>")
                tlines.extend([
                    "📦 التتبع: <b>غير متوفر</b>",
                    "",
                    f"🧩 قيمة القطع: <b>{goods_total} ر.س</b>" if goods_total > 0 else "🧩 قيمة القطع: —",
                    f"🚚 الشحن: <b>{html.escape(ship_txt)}</b>",
                    f"💰 الإجمالي (قطع + شحن): <b>{grand_total} ر.س</b>" if (goods_total > 0 or shipping_fee > 0 or ship_included) else "💰 الإجمالي (قطع + شحن): —",
                ])

                await context.bot.send_message(
                    chat_id=to_trader,
                    text="\n".join(tlines),
                    parse_mode="HTML",
                    reply_markup=trader_received_notice_kb(order_id),
                    disable_web_page_preview=True,
                )
            except Exception as e:
                _swallow(e)

    # ✅ تنظيف مرحلة انتظار إدخال التتبع (لو كانت مستخدمة)
    try:
        ud = get_ud(context, user_id)
        ud.pop("tsu_kind", None)
        ud.pop("tsu_order_id", None)
        ud.pop("track_order_id", None)
        set_stage(context, user_id, STAGE_NONE)
    except Exception as e:
        _swallow(e)

    # ✅ الأكثر نظافة: عدّل نفس رسالة التاجر + اقفل الأزرار
    done_text = (
        "✅ تم إرسال تحديث (تم الشحن) بدون رقم تتبع\n"
        f"🧾 رقم الطلب: {order_id}\n"
        "🔎 التتبع: غير متوفر"
    )
    if goods_total > 0 or shipping_fee > 0 or ship_included:
        done_text += f"\n🧩 القطع: {goods_total} ر.س"
        done_text += f"\n🚚 الشحن: {ship_txt}"
        done_text += f"\n💰 الإجمالي (قطع + شحن): {grand_total} ر.س"

    try:
        await q.message.edit_text(
            done_text,
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("✖️ إغلاق", callback_data="pp_ui_close")],
            ]),
            disable_web_page_preview=True,
        )
    except Exception:
        # fallback لو ما قدر يعدّل (رسالة قديمة/محذوفة..)
        try:
            await q.message.reply_text(
                done_text,
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("✖️ إغلاق", callback_data="pp_ui_close")],
                ]),
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)

    await _alert(q, "تم ✅")


async def order_finish_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(
        q,
        "🔒 إقفال الطلب غير متاح حالياً.\n\n"
        "⏳ يمكن إقفال الطلب بعد مرور 7 أيام من تاريخ الشحن أو التسليم.",
        force=True
    )

    data = (q.data or "").strip()
    parts = data.split("|", 1)
    order_id = parts[1].strip() if len(parts) > 1 else ""
    if not order_id:
        return

    actor_id = q.from_user.id
    actor_name = _user_name(q)

    # جلب الطلب
    try:
        b = get_order_bundle(order_id) or {}
        order = b.get("order", {}) or {}
    except Exception:
        order = {}

    # صلاحيات: التاجر المقبول أو الأدمن فقط
    accepted_tid = 0
    try:
        accepted_tid = int(order.get("accepted_trader_id") or 0)
    except Exception:
        accepted_tid = 0

    if actor_id not in ADMIN_IDS and actor_id != accepted_tid:
        try:
            await _alert(
                q,
                f"🔒 هذا الزر مخصص للتاجر المستلم أو الإدارة فقط\n🧾 رقم الطلب: {order_id}",
                force=True
            )
        except Exception as e:
            _swallow(e)
        return
    # ---------- helpers ----------
    def _to_float(v):
        try:
            return float(str(v or 0).replace(",", "").strip() or 0)
        except Exception:
            return 0.0

    def _parse_iso(s: object):
        try:
            import datetime as _dt
            ss = ("" if s is None else str(s)).strip()
            if not ss:
                return None
            # 2026-02-10T10:19:41.070+00:00 / 2026-02-10T10:19:41Z
            ss = ss.replace("Z", "+00:00")
            dt = _dt.datetime.fromisoformat(ss)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=_dt.timezone.utc)
            return dt.astimezone(_dt.timezone.utc)
        except Exception:
            return None

    def _now_utc():
        import datetime as _dt
        return _dt.datetime.now(_dt.timezone.utc)

    # حساب سطر المبلغ (سطر واحد) — إجمالي شامل الشحن إذا يوجد/مشمول، وإلا مبلغ القطع فقط
    goods_amt = _to_float(order.get("goods_amount_sar") or order.get("goods_total_sar") or 0)
    ship_fee = _to_float(order.get("shipping_fee_sar") or 0)

    try:
        ship_inc_raw = str(
            order.get("shipping_included")
            or order.get("ship_included")
            or order.get("shipping_fee_included")
            or ""
        ).strip().lower()
    except Exception:
        ship_inc_raw = ""

    shipping_included = ship_inc_raw in ("yes", "1", "true", "on", "included", "مشمولة", "مشمول")
    has_shipping = (ship_fee > 0.0) or shipping_included

    if has_shipping:
        total_num = goods_amt + (ship_fee if ship_fee > 0.0 else 0.0)
        amount_line = f"💰 إجمالي الطلب (شامل الشحن): {_money(total_num)}"
    else:
        amount_line = f"💰 مبلغ القطع: {_money(goods_amt)}"

    # الحالة الحالية
    st_now = str(order.get("order_status") or "").strip().lower()

    # لو مقفول سابقاً
    if st_now == "closed":
        try:
            await _alert(
                q,
                "🔒 الطلب مقفول بالفعل\n"
                f"🧾 رقم الطلب: {order_id}\n"
                f"{amount_line}",
                force=True
            )
        except Exception as e:
            _swallow(e)
        try:
            if q.message:
                await q.message.edit_reply_markup(reply_markup=trader_status_kb(order_id))
        except Exception as e:
            _swallow(e)
        return

    # ✅ شرط 7 أيام (نفس سياسة الإغلاق) — نعتمد على تاريخ الشحن (أولوية) أو التسليم
    shipped_dt = _parse_iso(order.get("shipped_at_utc") or order.get("shipped_at") or "")
    delivered_dt = _parse_iso(order.get("delivered_at_utc") or order.get("delivered_at") or "")
    ref_dt = shipped_dt or delivered_dt

    if not ref_dt:
        try:
            await _alert(
                q,
                "⚠️ لا يمكن إنهاء الطلب قبل وجود (تاريخ شحن/تسليم)\n"
                f"🧾 رقم الطلب: {order_id}\n"
                f"{amount_line}",
                force=True
            )
        except Exception as e:
            _swallow(e)
        return

    try:
        delta_sec = (_now_utc() - ref_dt).total_seconds()
    except Exception:
        delta_sec = 0

    if delta_sec < 7 * 24 * 3600:
        try:
            await _alert(
                q,
                "⏳ لا يمكن إقفال الطلب الآن\n"
                "🔒 الإقفال متاح بعد مرور 7 أيام من (الشحن/التسليم)\n"
                f"🧾 رقم الطلب: {order_id}\n"
                f"{amount_line}",
                force=True
            )
        except Exception as e:
            _swallow(e)
        return

    # (اختياري) منع الإقفال قبل الشحن/التسليم حسب سياستك العامة
    # نسمح هنا طالما تحقق شرط 7 أيام من الشحن/التسليم.
    # إذا تبغى تقييدها أكثر، فعّل هذا الشرط:
    # if st_now not in ("shipped", "delivered"):
    #     ...

    # حدّث الحقول (إقفال الطلب)
    try:
        update_order_fields(order_id, {
            "order_status": "closed",
            "closed_at_utc": utc_now_iso(),
        })
    except Exception as e:
        _swallow(e)

    # ✅ لا نحذف الكيبورد — نحدثه فقط
    try:
        if q.message:
            await q.message.edit_reply_markup(reply_markup=trader_status_kb(order_id))
    except Exception as e:
        _swallow(e)

    # ✅ إشعار العميل بإغلاق الطلب نهائياً (لاكتمال مدة الطلب)
    try:
        client_chat_id = int(order.get("user_id") or order.get("client_id") or 0)
    except Exception:
        client_chat_id = 0

    if client_chat_id:
        try:
            await context.bot.send_message(
                chat_id=client_chat_id,
                text=(
                    f"{_user_name(update)}\n"
                    "🔒 تم إقفال الطلب نهائيًا لاكتمال مدة الطلب (7 أيام).\n"
                    f"🧾 رقم الطلب: {html.escape(order_id)}\n"
                    f"{amount_line}\n\n"
                    "✅ في حال وجود ملاحظة إضافية تواصل مع الإدارة."
                ),
                parse_mode="HTML",
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)

    # سجل حدث (إن أردت)
    try:
        log_event(
            order_id,
            "order_finished",
            actor_role="admin" if actor_id in ADMIN_IDS else "trader",
            actor_id=actor_id,
            actor_name=actor_name,
            payload={"order_status": "closed", "policy": "7_days"},
        )
    except Exception as e:
        _swallow(e)

    # Popup واضح (حسناً)
    try:
        await _alert(
            q,
            "✅ تم إقفال الطلب بنجاح (بعد اكتمال 7 أيام)\n"
            f"🧾 رقم الطلب: {order_id}\n"
            f"{amount_line}",
            force=True
        )
    except Exception as e:
        _swallow(e)

async def _open_chat_session(context: ContextTypes.DEFAULT_TYPE, order_id: str, client_id: int, trader_id: int):
    if not (client_id and trader_id):
        return

    # استخراج مبلغ القطع (بدون تغيير أي منطق للتدفق)
    amt_txt = ""
    try:
        b = get_order_bundle(order_id) or {}
        o = b.get("order", {}) or {}
        raw_amt = o.get("goods_amount_sar") or o.get("quote_goods_amount") or ""
        amt_txt = _money(raw_amt)
    except Exception:
        amt_txt = ""

    # أسماء الأطراف
    try:
        cn, tn = _order_parties(order_id)
    except Exception:
        cn, tn = ("—", "—")

    # ⏱️ Timeout (افتراضي 30 دقيقة خمول / 6 ساعات كحد أقصى)
    try:
        idle_secs = int(os.getenv("PP_CHAT_IDLE_SECS", "1800") or 1800)
    except Exception:
        idle_secs = 1800
    try:
        max_secs = int(os.getenv("PP_CHAT_MAX_SECS", "21600") or 21600)
    except Exception:
        max_secs = 21600

    try:
        now_ts = int(time.time())
    except Exception:
        now_ts = 0

    kb_end = InlineKeyboardMarkup([[InlineKeyboardButton("✖️ إنهاء المراسلة", callback_data=f"pp_chat_end|{order_id}")]])

    client_open_txt = (
        "💬 تم فتح المراسلة الداخلية\n"
        f"⬅️ إلى: {tn}\n"
        f"{_order_tag_plain(order_id)}\n"
        f"💰 مبلغ القطع: {amt_txt if amt_txt else '—'}\n"
        f"⏱️ تنتهي تلقائيا بعد {int(idle_secs/60)} دقيقة خمول\n"
        "اكتب رسالتك (نص/وسائط) وسيتم تمريرها للطرف الاخر."
    )

    trader_open_txt = (
        "💬 تم فتح المراسلة الداخلية\n"
        f"⬅️ إلى: {cn}\n"
        f"{_order_tag_plain(order_id)}\n"
        f"💰 مبلغ القطع: {amt_txt if amt_txt else '—'}\n"
        f"⏱️ تنتهي تلقائيا بعد {int(idle_secs/60)} دقيقة خمول\n"
        "اكتب رسالتك (نص/وسائط) وسيتم تمريرها للطرف الاخر."
    )

    try:
        await context.bot.send_message(chat_id=client_id, text=client_open_txt, reply_markup=kb_end, disable_web_page_preview=True)
    except Exception as e:
        _swallow(e)
    try:
        await context.bot.send_message(chat_id=trader_id, text=trader_open_txt, reply_markup=kb_end, disable_web_page_preview=True)
    except Exception as e:
        _swallow(e)

    try:
        sessions = context.bot_data.setdefault("pp_chat_sessions", {})
        sessions[str(client_id)] = {"order_id": order_id, "peer_id": trader_id, "role": "client", "started_at": now_ts, "last_touch": now_ts}
        sessions[str(trader_id)] = {"order_id": order_id, "peer_id": client_id, "role": "trader", "started_at": now_ts, "last_touch": now_ts}
        context.bot_data["pp_chat_sessions"] = sessions
    except Exception:
        pass

# ==============================
# ✅ نظام مراسلة محكم (مختصر)
# Admin ↔ Client  |  Admin ↔ Trader
# ==============================

STAGE_ADMIN_CHAT = "pp_admin_chat"
STAGE_TRADER_CHAT_ADMIN = "pp_trader_chat_admin"

def admin_contact_kb(order_id: str) -> InlineKeyboardMarkup:
    # زر مراسلة العميل + التاجر (من رسالة الإدارة)
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("💬 مراسلة العميل", callback_data=f"pp_admin_chat_client|{order_id}"),
            InlineKeyboardButton("🧑‍🔧 مراسلة التاجر", callback_data=f"pp_admin_chat_trader|{order_id}"),
        ],
        [InlineKeyboardButton("✖️ إنهاء", callback_data="pp_admin_chat_done")],
    ])

def trader_chat_admin_kb(order_id: str, admin_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✍️ رد للإدارة", callback_data=f"pp_trader_chat_admin|{order_id}|{admin_id}")],
        [InlineKeyboardButton("✖️ إنهاء", callback_data="pp_trader_chat_admin_done")],
    ])

async def admin_chat_client_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    actor_id = q.from_user.id

    if actor_id not in ADMIN_IDS:
        await _alert(q, "غير مصرح")
        return

    data = (q.data or "").strip()
    parts = data.split("|", 1)
    if len(parts) != 2:
        return

    order_id = (parts[1] or "").strip()
    if not order_id:
        return

    # جلب معرف العميل المرتبط بالطلب
    uid = get_order_user_id(order_id)
    if not uid:
        await _alert(q, "لا يوجد عميل مرتبط بالطلب")
        return

    # تجهيز جلسة المراسلة
    ud = get_ud(context, actor_id)
    ud["admin_chat_order_id"] = order_id
    ud["admin_chat_peer_id"] = int(uid)
    ud["admin_chat_role"] = "client"

    set_stage(context, actor_id, STAGE_ADMIN_CHAT)

    # رسالة بدء المراسلة
    await q.message.reply_text(
        f"👤 {_admin_public_name()}\n"
        f"💬 مراسلة العميل\n"
        f"{_order_tag_plain(order_id)}\n"
        f"اكتب رسالتك الآن وسيتم إرسالها للعميل مباشرة.",
        reply_markup=chat_nav_kb(order_id, "pp_admin_chat_done"),
        disable_web_page_preview=True,
    )

async def admin_chat_trader_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    actor_id = q.from_user.id
    if actor_id not in ADMIN_IDS:
        await _alert(q, "غير مصرح")
        return

    data = (q.data or "").strip()
    parts = data.split("|", 1)
    if len(parts) != 2:
        return
    order_id = (parts[1] or "").strip()
    if not order_id:
        return

    tid = _assigned_trader_id(order_id)
    if not tid:
        await _alert(q, "لا يوجد تاجر مرتبط بالطلب")
        return

    ud = get_ud(context, actor_id)
    ud["admin_chat_order_id"] = order_id
    ud["admin_chat_peer_id"] = int(tid)
    ud["admin_chat_role"] = "trader"
    set_stage(context, actor_id, STAGE_ADMIN_CHAT)

    await q.message.reply_text(
        f"👤 {_admin_public_name()}\n🧑‍🔧 مراسلة التاجر\n{_order_tag_plain(order_id)}\nاكتب رسالتك الآن وسيتم إرسالها للتاجر.",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("✖️ إنهاء", callback_data="pp_admin_chat_done")]]),
        disable_web_page_preview=True,
    )

async def admin_chat_done_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    actor_id = q.from_user.id
    if actor_id not in ADMIN_IDS:
        return
    ud = get_ud(context, actor_id)
    ud.pop("admin_chat_order_id", None)
    ud.pop("admin_chat_peer_id", None)
    ud.pop("admin_chat_role", None)
    set_stage(context, actor_id, STAGE_NONE)
    try:
        await q.message.reply_text("تم إنهاء وضع المراسلة.")
    except Exception as e:
        _swallow(e)

async def trader_chat_admin_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    actor_id = q.from_user.id

    data = (q.data or "").strip()
    parts = data.split("|")
    if len(parts) != 3:
        return
    _, order_id, admin_id = parts
    order_id = (order_id or "").strip()
    try:
        admin_id = int(admin_id)
    except Exception:
        admin_id = 0

    if not order_id or not admin_id:
        await _alert(q, "بيانات غير صحيحة")
        return

    ud = get_ud(context, actor_id)
    ud["trader_chat_order_id"] = order_id
    ud["trader_chat_admin_id"] = admin_id
    set_stage(context, actor_id, STAGE_TRADER_CHAT_ADMIN)

    await q.message.reply_text(
        f"{_user_name(q)}\n🟨 رد للإدارة\n🧾 رقم الطلب: {order_id}\nاكتب ردك الآن وسيصل للإدارة.",
        reply_markup=chat_nav_kb(order_id, "pp_trader_chat_admin_done"),
        disable_web_page_preview=True,
    )

async def trader_chat_admin_done_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    actor_id = q.from_user.id
    ud = get_ud(context, actor_id)
    ud.pop("trader_chat_order_id", None)
    ud.pop("trader_chat_admin_id", None)
    set_stage(context, actor_id, STAGE_NONE)
    try:
        await q.message.reply_text("تم إنهاء وضع الرد للإدارة.")
    except Exception as e:
        _swallow(e)

async def chat_open_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    data = (q.data or "").strip()

    order_id = data.split("|", 1)[1] if "|" in data else ""
    order_id = (order_id or "").strip()
    if not order_id:
        return

    actor_id = q.from_user.id
    actor_name = (q.from_user.full_name or "").strip()
    actor_first = (q.from_user.first_name or actor_name or "").strip()

    def _to_int(v):
        try:
            return int(v)
        except Exception:
            try:
                return int(float(str(v).strip()))
            except Exception:
                return 0

    def _to_float(v):
        try:
            return float(str(v or 0).replace(",", "").strip() or 0)
        except Exception:
            return 0.0

    # جلب الطلب
    try:
        b = get_order_bundle(order_id)
        order = b.get("order", {}) or {}
        items = b.get("items", []) or []
    except Exception:
        order = {}
        items = []

    client_id = _to_int(order.get("user_id"))
    trader_id = _to_int(order.get("accepted_trader_id"))

    if not client_id:
        await _alert(q, "لا يوجد عميل مرتبط بالطلب")
        return

    # لا يوجد تاجر مقبول بعد
    if not trader_id and actor_id not in ADMIN_IDS:
        if actor_id == client_id:
            await _alert(q, "لم يتم تحديد تاجر لهذا الطلب بعد")
        else:
            await _alert(q, "لا يوجد تاجر مقبول مرتبط بالطلب")
        return

    # السماح فقط: العميل صاحب الطلب / التاجر المقبول / الادمن
    if actor_id not in ADMIN_IDS and actor_id not in (client_id, trader_id):
        intruder = actor_first or actor_name or "المستخدم"
        await _alert(
            q,
            f"🔒 غير مصرح\n\n"
            f"👤 {intruder}\n"
            "هذا الزر مخصص لصاحب الطلب أو للتاجر المستلم فقط.",
            force=True
        )
        return
    # ============================
    # ✅ قفل الشات قبل الوقت الصحيح
    # ============================
    st = str(order.get("order_status") or "").strip().lower()
    gps = str(order.get("goods_payment_status") or "").strip().lower()
    goods_paid = gps in ("confirmed", "paid", "success", "successful", "done", "ok")

    # العميل: لا يفتح الشات قبل أول تحديث (جاري تجهيز)
    # (حتى لو ضغط زر قديم)
    if actor_id == client_id and actor_id not in ADMIN_IDS:
        if st not in ("preparing", "prep", "ready_to_ship", "ready", "shipped", "delivered", "closed"):
            await _alert(q, "⏳ لا يمكن مراسلة التاجر قبل بدء تجهيز الطلب")
            return

    # التاجر: ممنوع قبل الدفع (شرطك الأساسي)
    if actor_id == trader_id and actor_id not in ADMIN_IDS:
        if not goods_paid:
            await _alert(q, "🔒 لا يمكن مراسلة العميل أو عرض بياناته قبل تأكيد سداد قيمة القطع")
            return

    # ============================
    # ✅ سطر/كتلة مالية موحدة (قطع/شحن/إجمالي)
    # ============================
    try:
        goods_amt_num = _to_float(order.get("goods_total_sar") or order.get("goods_amount_sar") or 0)
    except Exception:
        goods_amt_num = 0.0

    try:
        ship_fee_raw = _to_float(order.get("shipping_fee_sar") or 0)
    except Exception:
        ship_fee_raw = 0.0

    try:
        ship_inc_raw = str(
            order.get("shipping_included")
            or order.get("ship_included")
            or order.get("shipping_fee_included")
            or ""
        ).strip().lower()
    except Exception:
        ship_inc_raw = ""

    shipping_included = ship_inc_raw in ("yes", "1", "true", "on", "included", "مشمولة", "مشمول")

    ship_fee_effective = 0.0 if shipping_included else max(0.0, ship_fee_raw)
    total_num = float(goods_amt_num or 0.0) + float(ship_fee_effective or 0.0)

    goods_line = f"💰 مبلغ القطع: {_money(goods_amt_num) if goods_amt_num else '—'}"
    ship_line = "🚚 الشحن: مشمولة ✅" if shipping_included else f"🚚 الشحن: {_money(ship_fee_effective) if ship_fee_effective else '—'}"
    total_line = f"🧾 الإجمالي: {_money(total_num) if total_num else '—'}"

    # ============================
    # ⏳ مؤقت المراسلة 7 أيام (للعميل/التاجر) — الأدمن مستثنى
    # ============================
    if actor_id not in ADMIN_IDS:
        now_utc = datetime.now(timezone.utc)

        expires_raw = (order.get("chat_expires_at_utc") or "").strip()
        expires_dt = None
        if expires_raw:
            try:
                expires_dt = datetime.fromisoformat(expires_raw.replace("Z", "+00:00")).astimezone(timezone.utc)
            except Exception:
                expires_dt = None

        if not expires_dt:
            base_raw = (
                (order.get("shipped_at_utc") or "").strip()
                or (order.get("delivered_at_utc") or "").strip()
                or (order.get("closed_at_utc") or "").strip()
                or (order.get("goods_payment_confirmed_at_utc") or "").strip()
                or (order.get("delivered_at") or "").strip()
                or (order.get("closed_at") or "").strip()
            )
            base_dt = None
            if base_raw:
                try:
                    base_dt = datetime.fromisoformat(base_raw.replace("Z", "+00:00")).astimezone(timezone.utc)
                except Exception:
                    base_dt = None

            # ✅ بداية نافذة 7 أيام: من لحظة تأكيد سداد القطع أو من (shipped/closed...) إن وجدت
            if not base_dt and (goods_paid or st in ("shipped", "delivered", "closed")):
                base_dt = now_utc

            if base_dt:
                expires_dt = base_dt + timedelta(days=7)
                try:
                    update_order_fields(order_id, {"chat_expires_at_utc": expires_dt.isoformat()})
                except Exception as e:
                    _swallow(e)

        if expires_dt and now_utc > expires_dt:
            await _alert(q, "🔒 انتهت مدة المتابعة/الاسترجاع (7 أيام) لهذا الطلب")
            return

    # ============================
    # ✅ مسار العميل: افتح مراسلة للتاجر
    # ============================
    if actor_id == client_id and actor_id not in ADMIN_IDS:
        ud = get_ud(context, actor_id)
        ud["chat_trader_order_id"] = order_id
        set_stage(context, actor_id, STAGE_CHAT_TRADER)

        msg_client = (
            f"{_user_name(q)}\n"
            f"🧾 رقم الطلب: {order_id}\n"
            f"{goods_line}\n"
            f"{ship_line}\n"
            f"{total_line}\n\n"
            "✍️ اكتب رسالتك للتاجر الآن وسيتم إرسالها عبر المنصة.\n"
            "🔔 تذكير: تجنب إرسال بيانات حساسة خارج نطاق الطلب."
        )

        try:
            await q.message.reply_text(
                msg_client,
                reply_markup=chat_nav_kb_for(context, user_id, order_id, "pp_chat_trader_done"),
                disable_web_page_preview=True,
            )
        except Exception:
            try:
                await context.bot.send_message(
                    chat_id=actor_id,
                    text=msg_client,
                    reply_markup=chat_nav_kb_for(context, user_id, order_id, "pp_chat_trader_done"),
                    disable_web_page_preview=True,
                )
            except Exception as e:
                _swallow(e)
        return

    # ============================
    # ✅ مسار التاجر/الأدمن: تفعيل وضع الرد (Relay)
    # ============================
    td = context.user_data.setdefault(actor_id, {})
    td["trader_reply_user_id"] = client_id
    td["trader_reply_order_id"] = order_id
    set_stage(context, actor_id, STAGE_TRADER_REPLY)

    # اسم التاجر
    try:
        tp = get_trader_profile(actor_id) or {}
    except Exception:
        tp = {}
    tname = (tp.get("display_name") or "").strip() or actor_first or actor_name or "التاجر"
    tco = (tp.get("company_name") or "").strip()
    tline = f"👤 <b>{html.escape(tname)}</b>" + (f" • 🏢 <b>{html.escape(tco)}</b>" if tco else "")

    # ملخص الطلب
    car = (order.get("car_name") or "").strip()
    model = (order.get("car_model") or "").strip()

    parts_lines = []
    try:
        for i, it in enumerate(items, start=1):
            if not isinstance(it, dict):
                continue
            nm = (it.get("name") or it.get("item_name") or "").strip()
            pn = (it.get("part_no") or it.get("item_part_no") or it.get("number") or "").strip()
            if nm and pn:
                parts_lines.append(f"{i}- {nm} (رقم: {pn})")
            elif nm:
                parts_lines.append(f"{i}- {nm}")
    except Exception:
        parts_lines = []

    parts_txt = "\n".join(parts_lines) if parts_lines else "—"

    # ✅ رسالة منظمة + الخطوة القادمة آخرها
    msg = (
        "💬 <b>تم فتح قناة المراسلة</b>\n"
        f"{tline}\n\n"
        f"🧾 رقم الطلب: <b>{html.escape(order_id)}</b>\n"
        f"📌 الحالة: <b>{html.escape(_pay_status_ar(st))}</b>\n\n"
        f"{html.escape(goods_line)}\n"
        f"{html.escape(ship_line)}\n"
        f"{html.escape(total_line)}\n"
        + (f"\n🚗 السيارة: <b>{html.escape((car + ' ' + model).strip())}</b>\n" if (car or model) else "\n")
        + "\n🧩 <b>ملخص القطع</b>\n"
        + f"<pre>{html.escape(parts_txt)}</pre>\n"
        + "➡️ <b>الخطوة القادمة</b>: اكتب رسالتك الآن وسيتم إرسالها للعميل عبر المنصة."
    )

    await context.bot.send_message(
        chat_id=actor_id,
        text=msg,
        parse_mode="HTML",
        disable_web_page_preview=True,
    )
async def pp_chat_end_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    try:
        await q.answer()
    except Exception:
        pass

    actor_id = int(getattr(q.from_user, "id", 0) or 0)
    if not actor_id:
        return

    data = (q.data or "").strip()
    order_id = ""
    try:
        parts = data.split("|", 1)
        if len(parts) == 2:
            order_id = (parts[1] or "").strip()
    except Exception:
        order_id = ""

    try:
        sessions = context.bot_data.get("pp_chat_sessions") or {}
    except Exception:
        sessions = {}

    sess = None
    try:
        sess = sessions.get(str(actor_id))
    except Exception:
        sess = None

    if not isinstance(sess, dict):
        try:
            await q.message.reply_text("لا توجد مراسلة فعالة حاليا.")
        except Exception:
            pass
        return

    peer_id = _safe_int(sess.get("peer_id"))
    sess_order = (sess.get("order_id") or "").strip()

    if order_id and sess_order and order_id != sess_order:
        # لا ننهي جلسة خاطئة
        try:
            await q.message.reply_text("هذه المراسلة لم تعد فعالة.")
        except Exception:
            pass
        return

    # اغلاق للطرفين
    try:
        sessions.pop(str(actor_id), None)
        if peer_id:
            sessions.pop(str(peer_id), None)
        context.bot_data["pp_chat_sessions"] = sessions
    except Exception:
        pass

    txt = f"✅ تم إنهاء المراسلة\n{_order_tag_plain(sess_order or order_id or '—')}"
    try:
        await q.message.reply_text(txt, disable_web_page_preview=True)
    except Exception:
        pass
    if peer_id:
        try:
            await context.bot.send_message(chat_id=peer_id, text=txt, disable_web_page_preview=True)
        except Exception:
            pass


async def goods_pay_bank_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, f"تمام يا {_user_name(q)}")
    user_id = q.from_user.id
    order_id = (q.data or "").split("|", 1)[1] if "|" in (q.data or "") else ""
    if not order_id:
        return

    # اجلب الطلب
    try:
        b = get_order_bundle(order_id)
        order = b.get("order", {}) or {}
    except Exception:
        order = {}

    # =========================
    # Helpers (مبالغ واضحة)
    # =========================
    def _s(x: object) -> str:
        return ("" if x is None else str(x)).strip()

    def _is_yes(x: object) -> bool:
        v = _s(x).lower()
        return v in ("yes", "y", "true", "1", "مشمول", "included")

    def _money_txt(x: object) -> str:
        try:
            return (_fmt_money(_s(x)) or "").strip()
        except Exception:
            v = _s(x)
            return (f"{v} ريال" if v else "")

    # قيم: قطع + شحن + إجمالي / ورسوم منصة منفصلة
    goods_str = _s(order.get("goods_amount_sar") or "")
    ship_str = _s(order.get("shipping_fee_sar") or order.get("shipping_fee") or "")
    ship_included = _is_yes(order.get("ship_included") or order.get("shipping_included") or "")

    platform_fee_str = _s(order.get("price_sar") or "")
    platform_fee_txt = _money_txt(platform_fee_str) or "0 ريال"

    goods_txt = _money_txt(goods_str) or "—"
    ship_txt = "مشمول" if ship_included else (_money_txt(ship_str) or "0 ريال")

    total_txt = ""
    try:
        ship_for_total = "0" if ship_included else (ship_str or "0")
        g_num, s_num, t_num = _calc_totals(goods_str or "0", ship_for_total or "0")
        total_txt = _money_txt(t_num) or ""
    except Exception:
        total_txt = ""

    if not total_txt:
        # fallback بسيط
        try:
            g = float(goods_str or 0)
        except Exception:
            g = 0.0
        try:
            s = 0.0 if ship_included else float(ship_str or 0)
        except Exception:
            s = 0.0
        total_txt = _money_txt(g + s) or "—"

    # ✅ التاجر المقبول
    tid = 0
    try:
        tid = int(order.get("accepted_trader_id") or 0)
    except Exception:
        tid = 0

    tp = {}
    if tid:
        try:
            tp = get_trader_profile(int(tid)) or {}
        except Exception:
            tp = {}

    t_bank = (tp.get("bank_name") or "").strip()
    t_iban = (tp.get("iban") or "").strip()
    # اسم المستفيد: الشركة ثم اسم التاجر ثم الافتراضي
    t_benef = (tp.get("company_name") or "").strip() or (tp.get("display_name") or "").strip() or ""

    # لو بيانات التاجر ناقصة: نرجع للمنصة مع تنبيه واضح
    beneficiary = t_benef if t_benef else (PP_BENEFICIARY or "—")
    iban = t_iban if t_iban else (PP_IBAN or "—")
    bank_line = f"🏦 <b>البنك</b>:\n<i>{html.escape(t_bank)}</i>\n\n" if t_bank else ""

    ud = get_ud(context, user_id)
    ud["goods_order_id"] = order_id

    try:
        update_order_fields(order_id, {
            "goods_payment_method": "bank_transfer",
            "goods_payment_status": "awaiting_receipt",
        })
    except Exception as e:
        _swallow(e)

    set_stage(context, user_id, STAGE_AWAIT_GOODS_RECEIPT)

    warn = ""
    if tid and (not t_iban):
        warn = "\n⚠️ <b>تنبيه</b>: بيانات تحويل التاجر غير مكتملة، تم عرض بيانات المنصة مؤقتًا.\n"

    await q.message.reply_text(
        f"🤍 اهلا { _user_name(q) }\n\n"
        "💳 <b>دفع قيمة البضاعة: تحويل بنكي</b>\n\n"
        f"🧾 <b>رقم الطلب</b>: {html.escape(order_id)}\n\n"
        "📌 <b>تفاصيل المبالغ</b>:\n"
        f"🧩 قيمة القطع: <b>{html.escape(goods_txt)}</b>\n"
        f"🚚 الشحن: <b>{html.escape(ship_txt)}</b>\n"
        f"💰 إجمالي القطع + الشحن: <b>{html.escape(total_txt)}</b>\n"
        f"{bank_line}"
        f"🏦 <b>المستفيد</b>:\n<i>{html.escape(beneficiary)}</i>\n\n"
        f"IBAN:\n<code>{html.escape(iban)}</code>\n\n"
        f"{warn}"
        "📸 بعد التحويل يرجى ارسال <b>صورة ايصال الدفع</b> هنا مباشرة\n"
        "لاستكمال تجهيز الطلب (الايصال الزامي)",
        parse_mode="HTML",
        reply_markup=bank_info_kb(),
    )

async def goods_pay_stc_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, f"تمام يا {_user_name(q)}")
    user_id = q.from_user.id
    order_id = (q.data or "").split("|", 1)[1] if "|" in (q.data or "") else ""
    if not order_id:
        return

    b = get_order_bundle(order_id)
    order = b.get("order", {}) or {}

    # =========================
    # Helpers (مبالغ واضحة)
    # =========================
    def _s(x: object) -> str:
        return ("" if x is None else str(x)).strip()

    def _is_yes(x: object) -> bool:
        v = _s(x).lower()
        return v in ("yes", "y", "true", "1", "مشمول", "included")

    def _money_txt(x: object) -> str:
        try:
            return (_fmt_money(_s(x)) or "").strip()
        except Exception:
            v = _s(x)
            return (f"{v} ريال" if v else "")

    goods_str = _s(order.get("goods_amount_sar") or "")
    ship_str = _s(order.get("shipping_fee_sar") or order.get("shipping_fee") or "")
    ship_included = _is_yes(order.get("ship_included") or order.get("shipping_included") or "")

    platform_fee_str = _s(order.get("price_sar") or "")
    platform_fee_txt = _money_txt(platform_fee_str) or "0 ريال"

    goods_txt = _money_txt(goods_str) or "—"
    ship_txt = "مشمول" if ship_included else (_money_txt(ship_str) or "0 ريال")

    total_txt = ""
    try:
        ship_for_total = "0" if ship_included else (ship_str or "0")
        g_num, s_num, t_num = _calc_totals(goods_str or "0", ship_for_total or "0")
        total_txt = _money_txt(t_num) or ""
    except Exception:
        total_txt = ""

    if not total_txt:
        try:
            g = float(goods_str or 0)
        except Exception:
            g = 0.0
        try:
            s = 0.0 if ship_included else float(ship_str or 0)
        except Exception:
            s = 0.0
        total_txt = _money_txt(g + s) or "—"

    # ✅ نحدد التاجر المرتبط بالطلب (المقبول)
    try:
        tid = int(order.get("accepted_trader_id") or 0)
    except Exception:
        tid = 0

    stc_number = ""
    if tid:
        try:
            tp = get_trader_profile(int(tid)) or {}
            stc_number = (tp.get("stc_pay") or "").strip()
        except Exception:
            stc_number = ""

    # fallback على رقم المنصة إذا التاجر ما حط رقم
    if not stc_number:
        stc_number = (PP_STC_PAY or "").strip()

    ud = get_ud(context, user_id)
    ud["goods_order_id"] = order_id

    update_order_fields(order_id, {"goods_payment_method": "stc_pay", "goods_payment_status": "awaiting_receipt"})
    set_stage(context, user_id, STAGE_AWAIT_GOODS_RECEIPT)

    await q.message.reply_text(
        f"🤍 اهلا { _user_name(q) }\n\n"
        "💳 <b>دفع قيمة البضاعة: STC Pay</b>\n\n"
        f"🧾 <b>رقم الطلب</b>: {html.escape(str(order_id))}\n\n"
        "📌 <b>تفاصيل المبالغ</b>:\n"
        f"🧩 قيمة القطع: <b>{html.escape(goods_txt)}</b>\n"
        f"🚚 الشحن: <b>{html.escape(ship_txt)}</b>\n"
        f"💰 إجمالي القطع + الشحن: <b>{html.escape(total_txt)}</b>\n"
        f"رقم STC Pay:\n<code>{html.escape(str(stc_number))}</code>\n\n"
        "📸 بعد التحويل يرجى ارسال <b>صورة ايصال الدفع</b> هنا مباشرة\n"
        "لاستكمال تجهيز الطلب (الايصال الزامي)",
        parse_mode="HTML",
        reply_markup=stc_info_kb()
    )

async def goods_pay_link_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    actor_id = q.from_user.id

    data = (q.data or "").strip()
    parts = data.split("|", 1)
    order_id = parts[1].strip() if len(parts) >= 2 else ""
    if not order_id:
        return

    # اجلب الطلب
    try:
        b = get_order_bundle(order_id)
        order = b.get("order", {}) or {}
    except Exception:
        order = {}

    # =========================
    # Helpers (مبالغ واضحة)
    # =========================
    def _s(x: object) -> str:
        return ("" if x is None else str(x)).strip()

    def _is_yes(x: object) -> bool:
        v = _s(x).lower()
        return v in ("yes", "y", "true", "1", "مشمول", "included")

    def _money_txt(x: object) -> str:
        try:
            return (_fmt_money(_s(x)) or "").strip()
        except Exception:
            v = _s(x)
            return (f"{v} ريال" if v else "")

    goods_str = _s(order.get("goods_amount_sar") or "")
    if goods_str in ("", "0", "0.0"):
        await _alert(q, "لا يوجد مبلغ قيمة بضاعة مسجل لهذا الطلب")
        return

    ship_str = _s(order.get("shipping_fee_sar") or order.get("shipping_fee") or "")
    ship_included = _is_yes(order.get("ship_included") or order.get("shipping_included") or "")

    platform_fee_str = _s(order.get("price_sar") or "")
    platform_fee_txt = _money_txt(platform_fee_str) or "0 ريال"

    goods_txt = _money_txt(goods_str) or "—"
    ship_txt = "مشمول" if ship_included else (_money_txt(ship_str) or "0 ريال")

    total_txt = ""
    try:
        ship_for_total = "0" if ship_included else (ship_str or "0")
        g_num, s_num, t_num = _calc_totals(goods_str or "0", ship_for_total or "0")
        total_txt = _money_txt(t_num) or ""
    except Exception:
        total_txt = ""

    if not total_txt:
        try:
            g = float(goods_str or 0)
        except Exception:
            g = 0.0
        try:
            s = 0.0 if ship_included else float(ship_str or 0)
        except Exception:
            s = 0.0
        total_txt = _money_txt(g + s) or "—"

    # اربط مرحلة ايصال قيمة القطع عند العميل
    ud = get_ud(context, actor_id)
    ud["goods_order_id"] = order_id
    set_stage(context, actor_id, STAGE_AWAIT_GOODS_RECEIPT)

    # خزّن طريقة الدفع
    try:
        update_order_fields(order_id, {
            "goods_payment_method": "pay_link",
            "goods_payment_status": "awaiting_receipt",
        })
    except Exception as e:
        _swallow(e)

    # رابط الدفع: أولوية لرابط الطلب (من التاجر) ثم رابط ثابت (اختياري)
    link = (str(order.get("goods_payment_link") or "")).strip() or (PP_PAY_LINK_URL or "").strip()

    if link and (link.startswith("http://") or link.startswith("https://")):
        try:
            await q.message.reply_text(
                "💳 <b>دفع قيمة القطع عبر رابط</b>\n\n"
                f"🧾 <b>رقم الطلب</b>: {html.escape(order_id)}\n\n"
                "📌 <b>تفاصيل المبالغ</b>:\n"
                f"🧩 قيمة القطع: <b>{html.escape(goods_txt)}</b>\n"
                f"🚚 الشحن: <b>{html.escape(ship_txt)}</b>\n"
                f"💰 إجمالي القطع + الشحن: <b>{html.escape(total_txt)}</b>\n\n"
                "بعد الدفع ارسل صورة/‏PDF إيصال الدفع هنا (الإيصال إلزامي).",
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("💳 دفع الآن", url=link)]]),
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)
        return

    # إذا لا يوجد رابط فعلي لهذا الطلب: نغلق المسار بدون تحويل المستخدم لمسارات التحويل
    await _alert(q, "🔗 رابط الدفع غير متوفر لهذا الطلب حاليا")

async def goods_receipt_photo_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    ud = get_ud(context, user_id)

    if ud.get(STAGE_KEY) != STAGE_AWAIT_GOODS_RECEIPT:
        return

    order_id = (ud.get("goods_order_id") or "").strip()
    if not order_id:
        await update.message.reply_text(f"{_user_name(update)}\nلا يوجد طلب مرتبط بالايصال حاليا")
        set_stage(context, user_id, STAGE_NONE)
        return

    photos = update.message.photo or []
    if not photos:
        await update.message.reply_text(f"{_user_name(update)}\nالايصال الزامي ارسل صورة او PDF فقط")
        return

    file_id = photos[-1].file_id

    try:
        update_order_fields(order_id, {
            "goods_receipt_file_id": file_id,
            "goods_receipt_mime": "image/jpeg",
            "goods_payment_status": "awaiting_confirm",
        })
    except Exception as e:
        _swallow(e)

    # 🔒 قفل استقبال عروض جديدة فور ارسال الايصال
    try:
        update_order_fields(order_id, {"quote_locked": "yes"})
    except Exception as e:
        _swallow(e)
    try:
        await _lock_team_post_keyboard(context, order_id, reason="🔒 تم إيقاف العروض الطلب منتهي")
    except Exception as e:
        _swallow(e)

    tid = _assigned_trader_id(order_id)

    # ✅ بعد دفع قيمة القطع: نرسل للتاجر العنوان كامل (بدون رقم الهاتف) + زر مراسلة العميل
    try:
        b_addr = get_order_bundle(order_id) or {}
        o_addr = b_addr.get("order", {}) or {}
    except Exception:
        o_addr = {}

    ship_city = (o_addr.get("ship_city") or o_addr.get("pickup_city") or "").strip()
    ship_district = (o_addr.get("ship_district") or "").strip()
    ship_short = (o_addr.get("ship_short_address") or "").strip()
    ship_method = (o_addr.get("delivery_type") or o_addr.get("ship_method") or o_addr.get("delivery_choice") or "").strip()
    delivery_details = (o_addr.get("delivery_details") or "").strip()

    addr_lines = []
    if ship_method:
        addr_lines.append(f"🚚 نوع التسليم: {ship_method}")
    if ship_city:
        addr_lines.append(f"🏙 المدينة: {ship_city}")
    if ship_district:
        addr_lines.append(f"📍 الحي: {ship_district}")
    if ship_short:
        addr_lines.append(f"🧭 العنوان المختصر: {ship_short}")
    if delivery_details:
        # لا نرسل رقم الجوال هنا (يبقى سري) — لكن نرسل بقية تفاصيل العنوان
        safe_details = re.sub(r"(\+?9665\d{8}|9665\d{8}|05\d{8})", "*********", delivery_details)
        addr_lines.append(f"📝 تفاصيل العنوان: {safe_details}")

    addr_block = "\n".join(addr_lines) if addr_lines else "—"

    # ✅ مبالغ: إجمالي (قيمة القطع + الشحن)
    def _to_f(x):
        try:
            return float(str(x or "").replace(",", "").strip() or 0)
        except Exception:
            return 0.0

    def _money_sar(x):
        try:
            v = _to_f(x)
            if abs(v) < 1e-9:
                return "—"
            if abs(v - int(v)) < 1e-9:
                return f"{int(v)} ﷼"
            s = f"{v:.2f}".rstrip("0").rstrip(".")
            return f"{s} ﷼"
        except Exception:
            return "—"

    goods_amt_raw = o_addr.get("goods_amount_sar") or ""
    ship_included_norm = str(o_addr.get("ship_included") or "").strip().lower()
    ship_fee_raw = o_addr.get("shipping_fee_sar")

    # ✅ الشحن يحدده التاجر — لا يوجد رقم افتراضي
    if ship_fee_raw is None or str(ship_fee_raw).strip() == "":
        ship_fee_raw = 0 if ship_included_norm in ("yes", "true", "1", "included", "مشمولة", "مشمول") else ""

    goods_val = _to_f(goods_amt_raw)
    ship_val = _to_f(ship_fee_raw)
    total_val = goods_val + ship_val

    ship_line = ""
    if ship_included_norm in ("yes", "true", "1", "included", "مشمولة"):
        ship_line = "🚚 الشحن: مشمول"
    else:
        ship_txt = _money_sar(ship_val)
        if ship_txt != "—":
            ship_line = f"🚚 الشحن: {ship_txt}"

    total_txt = _money_sar(total_val)

    client_name_only = (ud.get("user_name") or "").strip() or "—"

        # ✅ بيانات إضافية للمعاينة
    car_local = (o_addr.get("car_name") or o_addr.get("vehicle_name") or o_addr.get("car_model") or "").strip() or "—"
    availability_local = (str(o_addr.get("availability_days") or o_addr.get("quote_availability") or o_addr.get("availability") or "")).strip()
    ship_eta_local = (str(o_addr.get("ship_eta") or o_addr.get("shipping_eta") or o_addr.get("ship_days") or "")).strip()

    trader_name_local = (o_addr.get("accepted_trader_name") or o_addr.get("quoted_trader_name") or "").strip()
    trader_store_local = (o_addr.get("accepted_trader_store") or o_addr.get("accepted_store_name") or o_addr.get("trader_store") or o_addr.get("store_name") or "").strip()
    try:
        if tid:
            tp0 = get_trader_profile(int(tid)) or {}
            if not trader_name_local:
                trader_name_local = (str(tp0.get("display_name") or tp0.get("company_name") or "")).strip()
            if not trader_store_local:
                trader_store_local = (str(tp0.get("company_name") or tp0.get("store_name") or "")).strip()
    except Exception:
        pass

    trader_name_local = trader_name_local or "—"
    trader_store_local = trader_store_local or "—"

    caption = (
        f"💳 إيصال سداد قيمة القطع\n"
        f"🧾 رقم الطلب: {order_id}\n"
        f"👤 العميل: {client_name_only}\n"
        f"🧑‍💼 التاجر: {trader_name_local}\n"
        f"🏪 المتجر: {trader_store_local}\n"
        f"🚗 السيارة: {car_local}\n"
        + (f"🛠 مدة التجهيز: {availability_local}\n" if availability_local else "")
        + (f"⏱ مدة الشحن: {ship_eta_local}\n" if ship_eta_local else "")
        + "\n"
        f"{ship_line}\n"
        f"💰 مبلغ القطع: {_money_sar(goods_val)}\n"
        f"🧾 الإجمالي (قطع + شحن): {total_txt}\n\n"
        f"📦 عنوان/تسليم:\n{addr_block}\n\n"
        f"✅ الخطوة التالية: اضغط (تأكيد الاستلام)"
    )


    try:
        await notify_admins_goods_receipt(context, ud, file_id, mime="image/jpeg")
    except Exception as e:
        _swallow(e)

    if tid:
        try:
            await context.bot.send_photo(
                chat_id=tid,
                photo=file_id,
                caption=caption,
                reply_markup=trader_goods_receipt_kb(order_id, user_id),
            )
        except Forbidden:
            for aid in ADMIN_IDS:
                try:
                    await context.bot.send_message(
                        chat_id=aid,
                        text=(
                            "⛔ تعذر ارسال ايصال القطع للتاجر (403 Forbidden)\n"
                            f"رقم الطلب: {order_id}\n"
                            f"التاجر: {tid}"
                        ),
                    )
                except Exception as e:
                    _swallow(e)
        except BadRequest:
            for aid in ADMIN_IDS:
                try:
                    await context.bot.send_message(
                        chat_id=aid,
                        text=(
                            "⛔ تعذر ارسال ايصال القطع (BadRequest)\n"
                            f"رقم الطلب: {order_id}\n"
                            "file_id غير صالح"
                        ),
                    )
                except Exception as e:
                    _swallow(e)
        except Exception as e:
            _swallow(e)

    try:
        await _send_client_payment_preview(context, user_id, order_id, pay_scope="goods")
    except Exception as e:
        _swallow(e)

    set_stage(context, user_id, STAGE_DONE)

async def goods_receipt_document_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    ud = get_ud(context, user_id)

    if ud.get(STAGE_KEY) != STAGE_AWAIT_GOODS_RECEIPT:
        return

    order_id = ud.get("goods_order_id", "")
    if not order_id:
        await update.message.reply_text(f"{_user_name(update)}\nلا يوجد طلب مرتبط بالايصال حاليا")
        set_stage(context, user_id, STAGE_NONE)
        return

    doc = update.message.document
    if not doc:
        await update.message.reply_text(f"{_user_name(update)}\nالايصال الزامي ارسل صورة او PDF فقط")
        return

    mime = (doc.mime_type or "").lower()
    fname = (doc.file_name or "").lower()
    is_ok = mime.startswith("image/") or mime.startswith("application/pdf") or fname.endswith((".jpg", ".jpeg", ".png", ".webp", ".pdf"))

    if not is_ok:
        await update.message.reply_text(f"{_user_name(update)}\nالايصال الزامي ارسل صورة او PDF فقط")
        return

    file_id = doc.file_id

    try:
        update_order_fields(order_id, {
            "goods_receipt_file_id": file_id,
            "goods_receipt_mime": mime,
            "goods_payment_status": "awaiting_confirm",
        })
    except Exception as e:
        _swallow(e)

    # 🔒 قفل استقبال عروض جديدة فور ارسال الايصال
    try:
        update_order_fields(order_id, {"quote_locked": "yes"})
    except Exception as e:
        _swallow(e)
    try:
        await _lock_team_post_keyboard(context, order_id, reason="🔒 تم إيقاف العروض الطلب منتهي")
    except Exception as e:
        _swallow(e)

    tid = _assigned_trader_id(order_id)

    # ✅ مبالغ: إجمالي (قيمة القطع + الشحن)
    try:
        b_amt = get_order_bundle(order_id) or {}
        o_amt = b_amt.get("order", {}) or {}
    except Exception:
        o_amt = {}

    def _to_f(x):
        try:
            return float(str(x or "").replace(",", "").strip() or 0)
        except Exception:
            return 0.0

    def _money_sar(x):
        try:
            v = _to_f(x)
            if abs(v) < 1e-9:
                return "—"
            if abs(v - int(v)) < 1e-9:
                return f"{int(v)} ﷼"
            s = f"{v:.2f}".rstrip("0").rstrip(".")
            return f"{s} ﷼"
        except Exception:
            return "—"

    goods_amt_raw = o_amt.get("goods_amount_sar") or ""
    ship_included_norm = str(o_amt.get("ship_included") or "").strip().lower()
    ship_fee_raw = o_amt.get("shipping_fee_sar")

    # ✅ الشحن يحدده التاجر — لا رقم افتراضي
    if ship_fee_raw is None or str(ship_fee_raw).strip() == "":
        ship_fee_raw = 0 if ship_included_norm in ("yes", "true", "1", "included", "مشمولة", "مشمول") else ""

    goods_val = _to_f(goods_amt_raw)
    ship_val = _to_f(ship_fee_raw)
    total_val = goods_val + ship_val

    ship_line = ""
    if ship_included_norm in ("yes", "true", "1", "included", "مشمولة"):
        ship_line = "🚚 الشحن: مشمول"
    else:
        ship_txt = _money_sar(ship_val)
        if ship_txt != "—":
            ship_line = f"🚚 الشحن: {ship_txt}"

    total_txt = _money_sar(total_val)

    client_name_only = (ud.get("user_name") or "").strip() or "—"

    # ✅ بيانات إضافية للمعاينة
    car_local = (o_amt.get("car_name") or o_amt.get("vehicle_name") or o_amt.get("car_model") or "").strip() or "—"
    availability_local = (str(o_amt.get("availability_days") or o_amt.get("quote_availability") or o_amt.get("availability") or "")).strip()
    ship_eta_local = (str(o_amt.get("ship_eta") or o_amt.get("shipping_eta") or o_amt.get("ship_days") or "")).strip()

    trader_name_local = (o_amt.get("accepted_trader_name") or o_amt.get("quoted_trader_name") or "").strip()
    trader_store_local = (o_amt.get("accepted_trader_store") or o_amt.get("accepted_store_name") or o_amt.get("trader_store") or o_amt.get("store_name") or "").strip()
    try:
        if tid:
            tp0 = get_trader_profile(int(tid)) or {}
            if not trader_name_local:
                trader_name_local = (str(tp0.get("display_name") or tp0.get("company_name") or "")).strip()
            if not trader_store_local:
                trader_store_local = (str(tp0.get("company_name") or tp0.get("store_name") or "")).strip()
    except Exception:
        pass

    trader_name_local = trader_name_local or "—"
    trader_store_local = trader_store_local or "—"

    caption = (
        f"💳 إيصال سداد قيمة القطع\n"
        f"🧾 رقم الطلب: {order_id}\n"
        f"👤 العميل: {client_name_only}\n"
        f"🧑‍💼 التاجر: {trader_name_local}\n"
        f"🏪 المتجر: {trader_store_local}\n"
        f"🚗 السيارة: {car_local}\n"
        + (f"🛠 مدة التجهيز: {availability_local}\n" if availability_local else "")
        + (f"⏱ مدة الشحن: {ship_eta_local}\n" if ship_eta_local else "")
        + "\n"
        + (f"{ship_line}\n" if ship_line else "")
        + f"💰 مبلغ القطع: {_money_sar(goods_val)}\n"
        + f"🧾 الإجمالي (قطع + شحن): {total_txt}\n\n"
        + "✅ الخطوة التالية: اضغط (تأكيد الاستلام)"
    )

    try:
        await notify_admins_goods_receipt(context, ud, file_id, mime=mime)
    except Exception as e:
        _swallow(e)

    if tid:
        try:
            await context.bot.send_document(
                chat_id=tid,
                document=file_id,
                caption=caption,
                reply_markup=team_goods_confirm_kb(order_id),
            )
        except Exception as e:
            _swallow(e)

    try:
        await _send_client_payment_preview(context, user_id, order_id, pay_scope="goods")
    except Exception as e:
        _swallow(e)

    set_stage(context, user_id, STAGE_DONE)

async def notify_team(context: ContextTypes.DEFAULT_TYPE, ud: dict):
    if not TEAM_CHAT_ID:
        return

    order_id = (ud.get("order_id") or "").strip()
    if not order_id:
        return

    try:
        b = get_order_bundle(order_id)
        order = b.get("order", {}) or {}
        bundle_items = b.get("items", []) or []
    except Exception:
        order = {}
        bundle_items = []

    # ✅ ملاحظات العميل (fallback)
    notes = _norm(ud.get("notes") or order.get("notes") or "")
    notes_html = f"<i>{html.escape(notes)}</i>" if notes else "<i>—</i>"

    # القطع: من ud أو من bundle
    items = ud.get("items") or bundle_items or []

    # بيانات أساسية
    user_name = (ud.get("user_name") or order.get("user_name") or "").strip()
    user_id = ud.get("user_id") or order.get("user_id") or ""
    car_name = (ud.get("car_name") or order.get("car_name") or "").strip()
    car_model = (ud.get("car_model") or order.get("car_model") or "").strip()
    vin = (ud.get("vin") or order.get("vin") or "").strip()

    # مبالغ
    fee = ud.get("price_sar") or order.get("price_sar") or ""
    goods_amount = order.get("goods_amount_sar") or ""
    ship_fee = order.get("shipping_fee_sar") or ""

    fee_txt = f"{fee} ريال" if str(fee).strip() not in ("", "0", "0.0") else "—"
    goods_txt = f"{goods_amount} ريال" if str(goods_amount).strip() not in ("", "0", "0.0") else "—"
    ship_txt = f"{ship_fee} ريال" if str(ship_fee).strip() not in ("", "0", "0.0") else "—"

    # ✅ بلوك التسليم (كما هو)
    delivery_block = _delivery_brief(order, ud) or "<i>—</i>"

    # ✅ عرض القطع (فخم + مختصر)
    items_lines = []
    media_count = 0
    shown = 0
    for i, it in enumerate(items, start=1):
        nm = (it.get("name") or "").strip()
        if not nm:
            continue

        pn = (it.get("part_no") or it.get("item_part_no") or "").strip()
        has_media = bool(it.get("photo_file_id") or it.get("file_id"))
        if has_media:
            media_count += 1

        # أيقونة حسب وجود صورة
        badge = "🖼️" if has_media else "📄"
        pn_txt = f" <code>{html.escape(pn)}</code>" if pn else ""

        tail = ""
        items_lines.append(f"{badge} <b>{shown+1}.</b> {html.escape(nm)}{pn_txt}{tail}")
        shown += 1

        if shown >= 10:
            break

    parts_html = "\n".join(items_lines) if items_lines else "<i>—</i>"
    if len(items) > 10:
        parts_html += f"\n<i>✨ قطع إضافية: {len(items) - 10}</i>"

    # شارات سريعة
    car_txt = html.escape((car_name + " " + car_model).strip()) if (car_name or car_model) else "—"
    uname_txt = html.escape(user_name) if user_name else "—"
    uid_txt = html.escape(str(user_id)) if str(user_id).strip() else "—"

    # ✅ رسالة فخمة (بدون خطوط)
    # ✅ وضع إعادة النشر (Reminders)
    is_reminder = bool(ud.get("_reminder"))
    reb_no = 0
    try:
        reb_no = int(ud.get("rebroadcast_no") or 0)
    except Exception:
        reb_no = 0

    hdr = "🚀 <b>طلب قطع غيار جديد</b> ✨"
    if is_reminder:
        hdr = "🔁 <b>إعادة نشر الطلب</b>"
        if reb_no > 0:
            hdr = f"🔁 <b>إعادة نشر الطلب</b> (الإعادة رقم {reb_no})"

    txt = (
        f"{hdr}\n"
        f"🧾 <b>رقم الطلب</b>: {html.escape(order_id)}\n\n"

        f"👤 <b>العميل</b>: <b>{uname_txt}</b> <code>({uid_txt})</code>\n"
        f"🚗 <b>السيارة</b>: <b>{car_txt}</b>\n"
        + (f"🔎 <b>VIN</b>: <code>{html.escape(vin)}</code>\n" if vin else "")
        + "\n"

        "📝 <b>ملاحظات العميل</b> 🧠\n"
        f"{notes_html}\n\n"

        "📦 <b>التسليم</b> 🏷️\n"
        f"{delivery_block}\n\n"

        f"🧩 <b>القطع المطلوبة</b> 🛠️  <b>({len(items)})</b>\n"
        f"📸 <b>عدد الصور</b>: <b>{media_count}</b>\n"
        f"{parts_html}\n\n"
    )

    team_msg_id = None
    try:
        sent = await context.bot.send_message(
            chat_id=TEAM_CHAT_ID,
            text=txt,
            parse_mode="HTML",
            reply_markup=team_group_kb(order_id, context.bot.username),
            disable_web_page_preview=True,
        )
        team_msg_id = getattr(sent, "message_id", None)
        if team_msg_id:
            try:
                # ✅ حفظ message_id + ✅ توثيق أول نشر للمجموعة (forwarded_to_team_at_utc)
                try:
                    update_order_fields(order_id, {"team_message_id": team_msg_id})
                except Exception as e:
                    _swallow(e)

                try:
                    # لا تكتبها إلا إذا كانت فاضية (أول إرسال فعلي للمجموعة)
                    fwd0 = str(order.get("forwarded_to_team_at_utc") or "").strip()
                    if not fwd0:
                        fields = {"forwarded_to_team_at_utc": utc_now_iso()}

                        # اختياري: إذا عندك بيانات الأدمن داخل ud
                        try:
                            aid = int(
                                ud.get("assigned_admin_id")
                                or ud.get("forwarded_by_admin_id")
                                or 0
                            )
                        except Exception:
                            aid = 0
                        aname = (ud.get("assigned_admin_name") or ud.get("forwarded_by_admin_name") or "").strip()

                        if aid:
                            fields["forwarded_by_admin_id"] = aid
                        if aname:
                            fields["forwarded_by_admin_name"] = aname

                        update_order_fields(order_id, fields)
                except Exception as e:
                    _swallow(e)

            except Exception as e:
                _swallow(e)
    except Exception:
        return

    # ✅ إرسال الوسائط كرد (Album)
    media: list = []
    for i, it in enumerate(items, start=1):
        fid = it.get("photo_file_id") or it.get("file_id") or ""
        if not fid:
            continue

        nm = (it.get("name") or "").strip()
        pn = (it.get("part_no") or it.get("item_part_no") or "").strip()
        caption = f"🧩 قطعة {i}: {nm}" if nm else f"🧩 قطعة {i}"
        if pn:
            caption += f" ({pn})"

        mt = (it.get("media_type") or "photo").strip().lower()
        if mt in ("video", "video_note"):
            media.append(InputMediaVideo(media=fid, caption=caption))
        elif mt in ("document", "audio", "voice"):
            media.append(InputMediaDocument(media=fid, caption=caption))
        else:
            media.append(InputMediaPhoto(media=fid, caption=caption))

    if not media:
        return

    for chunk_start in range(0, len(media), 10):
        chunk = media[chunk_start:chunk_start + 10]
        try:
            await context.bot.send_media_group(
                chat_id=TEAM_CHAT_ID,
                media=chunk,
                reply_to_message_id=team_msg_id,
            )
        except Exception as e:
            _swallow(e)
            
def _parse_item_name_partno(raw: str) -> tuple[str, str]:
    """
    Accept formats:
    - "فلتر زيت | 26300-2J000"
    - "فلتر زيت رقم 26300-2J000"
    - "فلتر زيت #26300-2J000"
    Returns (name, part_no).
    """
    s = (raw or "").strip()
    if not s:
        return "", ""
    # normalize separators
    if "|" in s:
        a, b = s.split("|", 1)
        return a.strip(), b.strip()
    m = re.search(r"(.*?)(?:\s*(?:رقم|#)\s*)([A-Za-z0-9\-_/\.]+)\s*$", s)
    if m:
        return (m.group(1) or "").strip(), (m.group(2) or "").strip()
    # try last token as part number if it has digits and letters or dashes and is long enough
    toks = s.split()
    if len(toks) >= 2:
        last = toks[-1].strip()
        if re.search(r"\d", last) and (len(last) >= 5) and re.fullmatch(r"[A-Za-z0-9\-_/\.]+", last):
            name = " ".join(toks[:-1]).strip()
            return name, last
    return s, ""

def _mask_phone_in_delivery(details: str) -> str:
    """Hide phone number line in delivery details."""
    if not details:
        return details or ""
    out_lines = []
    for ln in str(details).splitlines():
        if ln.strip().startswith("رقم الاتصال"):
            out_lines.append("رقم الاتصال: مخفي")
        else:
            out_lines.append(ln)
    return "\n".join(out_lines).strip()

# =========================
# Jobs: إعادة نشر الطلبات بدون عروض + تنبيه 24 ساعة
# =========================
def _parse_utc_iso(s: str):
    s = (s or "").strip()
    if not s:
        return None
    try:
        # Accept: 2026-02-01T00:00:00Z or with offset or without
        if s.endswith("Z"):
            s = s[:-1] + "+00:00"  # خليها offset-aware صريح
        dt = datetime.fromisoformat(s)

        # لو جاك بدون tz => اعتبره UTC
        if getattr(dt, "tzinfo", None) is None:
            dt = dt.replace(tzinfo=timezone.utc)

        # رجّعه دائمًا UTC
        return dt.astimezone(timezone.utc)
    except Exception:
        return None

def _dt_utc_now():
    # دائمًا aware
    return datetime.now(timezone.utc)

async def _rebroadcast_noquote_orders_job(context: ContextTypes.DEFAULT_TYPE):
    try:
        orders = list_orders() or []
    except Exception:
        orders = []

    now = _dt_utc_now()
    one_day = timedelta(hours=24)

    admin_need_list = []

    for o in orders:
        try:
            order_id = str(o.get("order_id") or "").strip()
        except Exception:
            order_id = ""
        if not order_id:
            continue

        # فقط الطلبات التي تم إرسالها لمجموعة التجار
        fwd = str(o.get("forwarded_to_team_at_utc") or "").strip()
        if not fwd:
            continue

        # استثناء الطلبات المقفلة/المكتملة
        ost = str(o.get("order_status") or "").strip().lower()
        if ost in ("closed", "delivered"):
            continue

        # إيقاف إعادة النشر (يدوي من الإدارة)
        rb_off = str(o.get("rebroadcast_disabled") or "").strip().lower()
        if rb_off in ("1", "yes", "true", "on", "stop", "stopped"):
            continue

        # بدون عروض فقط (إذا فيه عرض/قبول نخرج)
        try:
            qtid = int(o.get("quoted_trader_id") or 0)
        except Exception:
            qtid = 0
        qs = str(o.get("quote_status") or "").strip().lower()

        if qtid > 0 or qs in ("sent", "accepted"):
            continue

        base_ts = _parse_utc_iso(fwd) or _parse_utc_iso(str(o.get("created_at_utc") or "")) or None
        base_ts = _as_utc_aware(base_ts)
        if not base_ts:
            continue

        # =========================================================
        # ✅ إعادة النشر بعد 24 ساعة فقط (وتتكرر كل 24 ساعة)
        # =========================================================
        if (now - base_ts) >= one_day:
            last_b = _parse_utc_iso(str(o.get("last_group_broadcast_at_utc") or "")) or None
            last_b = _as_utc_aware(last_b)

            if (not last_b) or ((now - last_b) >= one_day):
                # جلب الطلب + القطع
                try:
                    b = get_order_bundle(order_id)
                    order = b.get("order", {}) or {}
                    items = b.get("items", []) or []
                except Exception:
                    order, items = {}, []

                # ✅ ترقيم الإعادة (يزيد كل مرة)
                rb_no = 0
                try:
                    rb_no = int(o.get("rebroadcast_count") or 0)
                except Exception:
                    rb_no = 0
                rb_no = max(0, rb_no) + 1

                ud_payload = {
                    "order_id": str(order_id),
                    "user_id": int(order.get("user_id") or 0),
                    "user_name": str(order.get("user_name") or ""),
                    "car_name": str(order.get("car_name") or ""),
                    "car_model": str(order.get("car_model") or ""),
                    "vin": str(order.get("vin") or ""),
                    "notes": str(order.get("notes") or ""),
                    "price_sar": float(order.get("price_sar") or 0),
                    "items": items,
                    "_reminder": True,
                    "rebroadcast_no": rb_no,
                }

                try:
                    log_event("إعادة نشر طلب بدون عروض (بعد 24 ساعة)", order_id=order_id)
                except Exception as e:
                    _swallow(e)

                # إعادة نشر للمجموعة
                try:
                    await notify_team(context, ud_payload)
                except Exception as e:
                    try:
                        log_event("فشل إعادة نشر الطلب لمجموعة التجار", order_id=order_id, error=e)
                    except Exception as e:
                        _swallow(e)
                # تثبيت وقت آخر إعادة + عداد الإعادة
                try:
                    update_order_fields(order_id, {
                        "last_group_broadcast_at_utc": utc_now_iso(),
                        "rebroadcast_count": str(rb_no),
                    })
                except Exception as e:
                    _swallow(e)

                # إشعار العميل (مرة كل 24 ساعة فقط)
                client_id = 0
                try:
                    client_id = int(order.get("user_id") or 0)
                except Exception:
                    client_id = 0

                if client_id:
                    last_ping = _parse_utc_iso(str(o.get("last_noquote_user_ping_at_utc") or "")) or None
                    last_ping = _as_utc_aware(last_ping)

                    if (not last_ping) or ((now - last_ping) >= one_day):
                        try:
                            await context.bot.send_message(
                                chat_id=client_id,
                                text=(
                                    "🔁 تم إعادة طرح طلبك للتجار\n"
                                    f"🧾 رقم الطلب: {_order_id_link_html(order_id)}\n\n"
                                    "لم يصلنا عرض سعر خلال 24 ساعة، لذلك تم إعادة نشر الطلب للمجموعة\n"
                                    "ومنح طلبك أولوية في المتابعة.\n\n"
                                    "بمجرد وصول أي عرض سيصلك إشعار فورًا.\n"
                                    "🛟 للتواصل مع الإدارة اكتب: منصة"
                                ),
                                reply_markup=track_kb(order_id),
                                disable_web_page_preview=True,
                            )
                        except Exception as e:
                            _swallow(e)
                        try:
                            update_order_fields(order_id, {"last_noquote_user_ping_at_utc": utc_now_iso()})
                        except Exception as e:
                            _swallow(e)

        # =========================================================
        # ✅ تنبيه الأدمن بعد 24 ساعة (ويتكرر كل 24 ساعة)
        # =========================================================
        if (now - base_ts) >= one_day:
            last_admin = _parse_utc_iso(str(o.get("admin_noquote_24h_sent_at_utc") or "")) or None
            last_admin = _as_utc_aware(last_admin)

            if (not last_admin) or ((now - last_admin) >= one_day):
                admin_need_list.append(order_id)
                try:
                    update_order_fields(order_id, {"admin_noquote_24h_sent_at_utc": utc_now_iso()})
                except Exception as e:
                    _swallow(e)

    if admin_need_list:
        admin_need_list = list(dict.fromkeys(admin_need_list))[:25]
        for oid in admin_need_list:
            msg_txt = (
                "⏰ <b>تنبيه إداري</b>\n"
                f"🧾 الطلب: <b>{html.escape(oid)}</b>\n\n"
                "هذا الطلب مضى عليه 24 ساعة بدون أي عروض، وسيتم إعادة نشره تلقائيًا كل 24 ساعة.\n"
                "إذا كان الطلب مستحيل/غير مناسب اضغط الزر لإيقاف إعادة النشر."
            )
            kb = InlineKeyboardMarkup(
                [[InlineKeyboardButton("⛔ إيقاف إعادة النشر", callback_data=f"pp_rb_stop|{oid}")]]
            )
            for aid in ADMIN_IDS:
                try:
                    await context.bot.send_message(
                        chat_id=int(aid),
                        text=msg_txt,
                        parse_mode="HTML",
                        reply_markup=kb,
                        disable_web_page_preview=True,
                    )
                except Exception as e:
                    _swallow(e)


async def notify_admins_goods_receipt(
    context: ContextTypes.DEFAULT_TYPE,
    ud: dict,
    file_id: str,
    mime: str = "",
):
    """Send goods payment receipt to admins only (single attachment + full caption) without duplicates.
    ✅ يعرض (إجمالي القطع + الشحن) فقط ويخفي رسوم المنصة نهائياً.
    """
    if not ADMIN_IDS or not file_id:
        return

    order_id = (ud.get("goods_order_id") or ud.get("order_id") or "").strip()
    user_name = (ud.get("user_name") or "").strip()

    # ✅ المطلوب: يوزر العميل بدل رقمه
    username = (ud.get("username") or ud.get("user_username") or "").strip()
    if username and not username.startswith("@"):
        username = f"@{username}"

    # بيانات إضافية (قد تكون مخزنة داخل ud)
    phone = (ud.get("phone") or ud.get("mobile") or ud.get("user_phone") or "").strip()
    city = (ud.get("city") or ud.get("user_city") or "").strip()

    # اجلب بيانات الطلب للتفاصيل (مبلغ/تاجر/حالة/إجمالي)
    trader_name = ""
    goods_amt = ""      # قيمة القطع
    ship_fee = ""       # رسوم الشحن
    ship_included = False
    total_amt = ""      # إجمالي القطع + الشحن فقط
    status_txt = ""

    def _s(x: object) -> str:
        return ("" if x is None else str(x)).strip()

    def _is_yes(x: object) -> bool:
        v = _s(x).lower()
        return v in ("yes", "y", "true", "1", "مشمول", "included")

    def _money_txt(x: object) -> str:
        try:
            t = _fmt_money(_s(x))
            return (t or "").strip()
        except Exception:
            v = _s(x)
            return (f"{v} ر.س" if v else "")

    try:
        b = get_order_bundle(order_id)
        o = (b.get("order", {}) or {}) if isinstance(b, dict) else {}

        goods_amt = _s(o.get("goods_amount_sar") or "")
        ship_fee = _s(o.get("shipping_fee_sar") or o.get("shipping_fee") or "")
        ship_included = _is_yes(o.get("ship_included") or o.get("shipping_included") or "")

        # ===== الجوال + المدينة من تفاصيل العنوان (fallback) =====
        if not city:
            city = _s(
                o.get("ship_city")
                or o.get("delivery_city")
                or o.get("pickup_city")
                or o.get("city")
                or ""
            )

        delivery_blob = _s(
            o.get("delivery_details")
            or o.get("address_text")
            or o.get("full_address")
            or o.get("address")
            or ""
        )

        if not phone and delivery_blob:
            try:
                m = re.search(r"(\+?9665\d{8}|9665\d{8}|05\d{8})", delivery_blob)
                if m:
                    phone = (m.group(1) or "").strip()
            except Exception as e:
                _swallow(e)

        if not city and delivery_blob:
            try:
                mm = re.search(r"(?:المدينة|مدينة)\s*[:：\-]\s*([^\n,]+)", delivery_blob)
                if mm:
                    city = _s(mm.group(1))
            except Exception as e:
                _swallow(e)
        # ===== نهاية الإضافة =====

        ship_for_total = "0" if ship_included else (ship_fee or "0")

        try:
            g_num, s_num, t_num = _calc_totals(goods_amt or "0", ship_for_total or "0")
            total_amt = _s(t_num)
        except Exception:
            try:
                g = float(goods_amt or 0)
            except Exception:
                g = 0.0
            try:
                s = 0.0 if ship_included else float(ship_fee or 0)
            except Exception:
                s = 0.0
            total_amt = _s(g + s)

        status_txt = _s(o.get("goods_payment_status") or "")
        if not status_txt:
            status_txt = _s(o.get("order_status") or "")

        trader_name = _s(o.get("accepted_trader_name") or o.get("quoted_trader_name") or "")
        if not trader_name:
            tid = 0
            try:
                tid = int(o.get("accepted_trader_id") or 0)
            except Exception:
                tid = 0
            if tid:
                tp = get_trader_profile(int(tid)) or {}
                trader_name = (_s(tp.get("display_name")) or _s(tp.get("company_name")))

    except Exception as e:
        _swallow(e)

    trader_name = trader_name or "—"
    goods_txt = _money_txt(goods_amt) or "—"

    ship_txt = "مشمول" if ship_included else (_money_txt(ship_fee) or "—")
    total_txt = _money_txt(total_amt) or "—"

    try:
        status_ar = _pay_status_ar(status_txt)
    except Exception:
        status_ar = status_txt or "بانتظار تأكيد الاستلام"

    # ✅ بيانات إضافية
    car_local = "—"
    availability_local = ""
    ship_eta_local = ""
    trader_store_local = ""
    try:
        b = get_order_bundle(order_id)
        o = (b.get("order", {}) or {}) if isinstance(b, dict) else {}
        car_local = (str(o.get("car_name") or o.get("vehicle_name") or o.get("car_model") or o.get("car") or "")).strip() or "—"
        availability_local = (str(o.get("availability_days") or o.get("quote_availability") or o.get("availability") or "")).strip()
        ship_eta_local = (str(o.get("ship_eta") or o.get("shipping_eta") or o.get("ship_days") or "")).strip()
        trader_store_local = (str(o.get("accepted_trader_store") or o.get("accepted_store_name") or o.get("trader_store") or o.get("store_name") or "")).strip()
    except Exception:
        pass

    caption = (
        "📌 إشعار إداري\n"
        "🧾 تم استلام إيصال قيمة القطع\n"
        f"🧾 رقم الطلب: {order_id or '—'}\n"
        f"👤 العميل: {user_name or '—'} {f'({username})' if username else ''}\n"
        f"📞 الجوال: {phone or '—'}\n"
        f"🏙 المدينة: {city or '—'}\n"
        f"🚗 السيارة: {car_local or '—'}\n"
        f"🧑‍💼 التاجر: {trader_name}\n"
        + (f"🏪 المتجر: {trader_store_local}\n" if trader_store_local else "")
        + (f"🛠 مدة التجهيز: {availability_local}\n" if availability_local else "")
        + (f"⏱ مدة الشحن: {ship_eta_local}\n" if ship_eta_local else "")
        + f"🧩 قيمة القطع: {goods_txt}\n"
        + f"🚚 الشحن: {ship_txt}\n"
        + f"💰 إجمالي القطع + الشحن: {total_txt}\n"
        + f"📦 الحالة: {status_ar}"
    )

    for aid in ADMIN_IDS:
        try:
            m = (mime or "").lower()
            is_img = m.startswith("image/") or m.endswith(("jpg", "jpeg", "png", "webp"))
            if is_img:
                await context.bot.send_photo(
                    chat_id=aid,
                    photo=file_id,
                    caption=caption,
                )
            else:
                await context.bot.send_document(
                    chat_id=aid,
                    document=file_id,
                    caption=caption,
                )
        except Exception as e:
            _swallow(e)

def admin_forward_kb(order_id: str, client_id: int = 0) -> InlineKeyboardMarkup:
    rows = [
        [InlineKeyboardButton("📤 ارسال الطلب للتاجر", callback_data=f"pp_admin_forward|{order_id}")],
    ]

    if client_id:
        rows.append(
            [InlineKeyboardButton("💬 مراسلة العميل", callback_data=f"pp_admin_reply|{order_id}|{client_id}")]
        )

    rows.append(
        [InlineKeyboardButton("⛔ الغاء الطلب", callback_data=f"pp_admin_cancel|{order_id}")]
    )

    return InlineKeyboardMarkup(rows)


# ✅ كيبورد مقفول بصريًا بعد الإرسال (يوضح للإدمن أن الطلب تم التعامل معه)
def admin_forward_kb_locked(order_id: str, client_id: int = 0) -> InlineKeyboardMarkup:
    rows = [
        [InlineKeyboardButton("✅ تم إرسال الطلب للتاجر مسبقًا", callback_data="pp_ui_close")],
    ]

    if client_id:
        rows.append(
            [InlineKeyboardButton("💬 مراسلة العميل", callback_data=f"pp_admin_reply|{order_id}|{client_id}")]
        )

    return InlineKeyboardMarkup(rows)

async def notify_admins_receipt(
    context: ContextTypes.DEFAULT_TYPE,
    ud: dict,
    receipt_file_id: str,
    receipt_is_photo: bool = True,
    client_id: int = 0,
) -> None:
    """
    اشعار الإيصال للإدارة برسالة واحدة فقط (بدون تشوه بصري):
    - نفس المعاينة (build_order_preview)
    - تفاصيل التسليم داخل صندوق <pre>
    - الإيصال مدمج مع الرسالة (كـ Photo أو Document)
    - أزرار: ارسال للتاجر + مراسلة العميل + الغاء
    """
    if not ADMIN_IDS:
        return

    order_id = (ud.get("order_id") or "").strip()
    if not order_id:
        return

    # جلب نسخة الطلب من الاكسل (للتأكد من البيانات)
    try:
        b = get_order_bundle(order_id)
        order = b.get("order", {}) or {}
    except Exception:
        order = {}

    preview_html = build_order_preview(ud)

    ship_method = (ud.get("delivery_choice") or order.get("delivery_choice") or ud.get("ship_method") or order.get("ship_method") or "").strip()
    delivery_details = (ud.get("delivery_details") or order.get("delivery_details") or "").strip()
    fee = ud.get("price_sar", order.get("price_sar", ""))

    fee_txt = ""
    try:
        if str(fee).strip() not in ("", "0", "0.0"):
            fee_txt = f"\n💰 <b>رسوم المنصة</b>: <b>{html.escape(str(fee), quote=False)}</b> ريال"
    except Exception:
        fee_txt = ""

    details_block = ""
    if ship_method or delivery_details:
        safe_method = html.escape(ship_method, quote=False) if ship_method else ""
        safe_details = html.escape(delivery_details or "", quote=False)
        details_block = (
            "\n\n📦 <b>طريقة التسليم</b>: "
            + (f"<b>{safe_method}</b>" if safe_method else "—")
            + "\n<b>تفاصيل التسليم</b>:\n"
            + f"<pre>{safe_details or '—'}</pre>"
        )

    msg_html = (
        "💳 <b>إيصال دفع جديد</b>\n"
        f"🧾 <b>رقم الطلب</b>: {html.escape(order_id, quote=False)}"
        f"{fee_txt}\n\n"
        f"{preview_html}"
        f"{details_block}\n\n"
        "⬇️ <b>اعتماد الطلب:</b> اضغط (ارسال الطلب للتاجر) أو (الغاء الطلب)"
    )

    # قصّ بسيط عشان لا تتجاوز حدود caption
    def _trim(s: str, limit: int = 950) -> str:
        s = (s or "").strip()
        return s if len(s) <= limit else (s[: max(0, limit - 1)].rstrip() + "…")

    msg_html = _trim(msg_html, 950)

    # ✅ هنا المهم: تمرير client_id لكيبورد الادمن لإظهار زر مراسلة العميل
    kb = admin_forward_kb(order_id, int(client_id or 0))

    # fallback: نص عادي بدون HTML إذا فشل parse
    def _plain_fallback(html_text: str) -> str:
        # نحولها لنص بسيط (بدون ما نحتاج imports إضافية)
        t = html_text or ""
        for tag in ("<b>", "</b>", "<i>", "</i>", "<code>", "</code>", "<pre>", "</pre>"):
            t = t.replace(tag, "")
        return t

    for aid in ADMIN_IDS:
        try:
            if receipt_file_id:
                if receipt_is_photo:
                    await context.bot.send_photo(
                        chat_id=aid,
                        photo=receipt_file_id,
                        caption=msg_html,
                        parse_mode="HTML",
                        reply_markup=kb,
                    )
                else:
                    await context.bot.send_document(
                        chat_id=aid,
                        document=receipt_file_id,
                        caption=msg_html,
                        parse_mode="HTML",
                        reply_markup=kb,
                    )
            else:
                await context.bot.send_message(
                    chat_id=aid,
                    text=msg_html,
                    parse_mode="HTML",
                    reply_markup=kb,
                    disable_web_page_preview=True,
                )

        except Exception:
            # ✅ لا نسكت: نرسل fallback نصي بدون parse_mode (عشان ما يضيع الإيصال)
            try:
                plain = _trim(_plain_fallback(msg_html), 3500)
                if receipt_file_id:
                    # لو الإيصال موجود، نعيده بدون parse_mode وبدون HTML caption
                    if receipt_is_photo:
                        await context.bot.send_photo(
                            chat_id=aid,
                            photo=receipt_file_id,
                            caption=plain,
                            reply_markup=kb,
                        )
                    else:
                        await context.bot.send_document(
                            chat_id=aid,
                            document=receipt_file_id,
                            caption=plain,
                            reply_markup=kb,
                        )
                else:
                    await context.bot.send_message(
                        chat_id=aid,
                        text=plain,
                        reply_markup=kb,
                        disable_web_page_preview=True,
                    )
            except Exception as e:
                _swallow(e)

async def notify_admins_free_order(
    context: ContextTypes.DEFAULT_TYPE,
    ud: dict,
    client_id: int = 0,
) -> None:
    """اشعار الإدارة بطلب مجاني (رسوم المنصة=0) برسالة واحدة: معاينة + مراسلة العميل + الغاء الطلب."""
    if not ADMIN_IDS:
        return

    order_id = (ud.get("order_id") or "").strip()
    if not order_id:
        return

    preview_html = build_order_preview(ud)
    cname = html.escape((ud.get("user_name") or "").strip())

    # ✅ بلوك العنوان الكامل للإدارة فقط
    try:
        b = get_order_bundle(order_id)
        order = b.get("order", {}) or {}
    except Exception:
        order = {}

    delivery_choice = (ud.get("delivery_choice") or order.get("delivery_choice") or "").strip()
    delivery_details = (ud.get("delivery_details") or order.get("delivery_details") or "").strip()

    uid_raw = client_id or ud.get("user_id") or order.get("user_id") or 0
    uid_txt = html.escape(str(uid_raw)) if str(uid_raw).strip() else ""

    details_block = ""
    if delivery_choice or delivery_details:
        details_block = (
            "\n\n📦 <b>التسليم</b>: "
            + (f"<b>{html.escape(delivery_choice)}</b>" if delivery_choice else "—")
            + "\n<b>عنوان التسليم (كامل)</b>:\n"
            + f"<pre>{html.escape(delivery_details or '—')}</pre>"
            + (f"📞 <b>رقم العميل</b>: <code>{uid_txt}</code>" if uid_txt else "")
        )

    summary = (
        "🆓 <b>طلب مجاني (رسوم المنصة = 0)</b>\n"
        + (f"👤 العميل: <b>{cname}</b>\n" if cname else "")
        + f"{preview_html}"
        + f"{details_block}"
    )

    kb = admin_free_order_kb(order_id, int(client_id or ud.get("user_id") or 0))

    for aid in (ADMIN_IDS or []):
        try:
            await context.bot.send_message(
                chat_id=int(aid),
                text=summary,
                parse_mode="HTML",
                reply_markup=kb,
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)

    # (اختياري) فاتورة منصة داخلية للإدارة فقط برسوم 0
    try:
        await send_invoice_pdf(
            context,
            order_id,
            kind="preliminary",
            admin_only=True,
            invoice_for="platform",
        )
    except Exception as e:
        _swallow(e)

async def admin_forward_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    actor_id = q.from_user.id

    # صلاحية الادمن
    if actor_id not in ADMIN_IDS:
        await _alert(q, "غير مصرح")
        return

    data = q.data or ""
    try:
        _, order_id = data.split("|", 1)
    except Exception:
        await _alert(q, "بيانات غير صحيحة")
        return

    order_id = (order_id or "").strip()
    if not order_id:
        await _alert(q, "رقم طلب غير صحيح")
        return

    if not TEAM_CHAT_ID:
        await _alert(q, "لم يتم ضبط مجموعة التاجر")
        return

    # تحميل الطلب من الاكسل
    try:
        bundle = get_order_bundle(order_id)
        order = bundle.get("order", {}) or {}
        items = bundle.get("items", []) or []
    except Exception:
        await _alert(q, "تعذر قراءة بيانات الطلب")
        return

    if not order:
        await _alert(q, "لم يتم العثور على الطلب")
        return

    # ✅ اشعار العميل انه تم التحقق وتم اسناد طلبه للمنصة
    client_id = 0
    try:
        client_id = int(order.get("user_id") or 0)
    except Exception:
        client_id = 0

    # ✅ (1) إذا كان forwarded_to_team_at_utc موجود: Alert + قفل بصري للزر (بدون إخفاء الرسالة)
    if str(order.get("forwarded_to_team_at_utc") or "").strip():
        await _alert(q, "✅ تم إرسال الطلب للتاجر مسبقًا", force=True)
        try:
            await q.message.edit_reply_markup(
                reply_markup=admin_forward_kb_locked(order_id, client_id)
            )
        except Exception as e:
            _swallow(e)
        return

    ud_payload = {
        "order_id": str(order_id),
        "user_id": int(order.get("user_id") or 0),
        "user_name": str(order.get("user_name") or ""),
        "car_name": str(order.get("car_name") or ""),
        "car_model": str(order.get("car_model") or ""),
        "vin": str(order.get("vin") or ""),
        "notes": str(order.get("notes") or ""),
        "payment_method": str(order.get("payment_method") or ""),
        "price_sar": float(order.get("price_sar") or 0),
        "items": items,
    }

    # ارسال للمجموعة
    await notify_team(context, ud_payload)

    if client_id:
        try:
            await context.bot.send_message(
                chat_id=client_id,
                text=(
                    "✅ تم التحقق من الدفع بنجاح\n"
                    f"🧾 رقم الطلب: {_order_id_link_html(order_id)}\n\n"
                    "📤 تم اسناد طلبك للمنصة وارساله لمجموعة التجار\n"
                    "ستصلك عروض الأسعار فور توفرها\n\n"
                    "🔎 يمكنك المتابعة مع المنصة عند تاخر وصول العروض  "
                ),
                reply_markup=track_kb(order_id),
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)

    # ✅ إرسال فاتورة المنصة (PDF) للعميل فقط بعد التحقق (منع ارسالها للإدارة لتفادي التكدس)
    try:
        await send_invoice_pdf(
            context=context,
            order_id=order_id,
            kind="preliminary",
            tracking_number="",
            admin_only=False,
            invoice_for="platform",
            include_admins=False,  # ✅ العميل فقط
        )
    except Exception as e:
        _swallow(e)

    # ✅ رسالة واحدة للإدارة (بدون PDF) + تفاصيل كاملة + أزرار مراسلة
    try:
        client_name = str(order.get("user_name") or "").strip() or "—"

        trader_name = (order.get("accepted_trader_name") or order.get("quoted_trader_name") or "").strip()
        if not trader_name:
            try:
                tid = int(order.get("accepted_trader_id") or 0) if str(order.get("accepted_trader_id") or "").isdigit() else 0
                if tid:
                    tp = get_trader_profile(int(tid)) or {}
                    trader_name = (tp.get("display_name") or "").strip() or (tp.get("company_name") or "").strip()
            except Exception as e:
                _swallow(e)
        trader_name = trader_name or "—"

        amt = ""
        try:
            amt = str(order.get("total_amount_sar") or "").strip()
            if not amt:
                amt = str(order.get("price_sar") or "").strip()
        except Exception:
            amt = ""
        amt = amt or "—"

        st = ""
        try:
            st = str(order.get("order_status") or "").strip()
        except Exception:
            st = ""
        st = st or "تم ارسال الطلب لمجموعة التجار"

        brief = (
            "📌 إشعار إداري\n"
            f"🧾 رقم الطلب: {order_id}\n"
            f"👤 العميل: {client_name or '—'}\n"
            f"📞 الجوال: {phone or '—'}\n"
            f"🏙 المدينة: {city or '—'}\n"
            f"🧑‍💼 التاجر: {trader_name or '—'}\n"
            f"💰 إجمالي المبلغ: {amt}\n"
            f"📦 الحالة: {status_ar}\n\n"
            "اختر جهة المراسلة:"
        )

        for aid in ADMIN_IDS:
            try:
                await context.bot.send_message(
                    chat_id=int(aid),
                    text=brief,
                    reply_markup=admin_contact_kb(order_id),
                    disable_web_page_preview=True,
                )
            except Exception as e:
                _swallow(e)
    except Exception as e:
        _swallow(e)

    # تمييز الطلب انه تم تمريره بواسطة الادمن
    try:
        mark_order_forwarded(
            order_id,
            admin_id=actor_id,
            admin_name=_user_name(q),
            at_utc=utc_now_iso(),
        )
    except Exception as e:
        _swallow(e)

    # ✅ (2) بعد نجاح الإرسال وتمييزه forwarded: قفل بصري + Alert نجاح (بدون إخفاء الرسالة)
    try:
        await q.message.edit_reply_markup(
            reply_markup=admin_forward_kb_locked(order_id, client_id)
        )
    except Exception as e:
        _swallow(e)

    await _alert(q, "تم الإرسال ✅", force=True)
    return

async def receipt_photo_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    ud = get_ud(context, user_id)

    if ud.get(STAGE_KEY) != STAGE_AWAIT_RECEIPT:
        return

    order_id = (ud.get("order_id") or "").strip()
    if not order_id:
        await update.message.reply_text(f"{_user_name(update)}\nلا يوجد طلب مرتبط بالايصال حاليا")
        set_stage(context, user_id, STAGE_NONE)
        return

    photos = update.message.photo or []
    if not photos:
        await update.message.reply_text(f"{_user_name(update)}\nالايصال الزامي ارسل صورة ايصال الدفع فقط")
        return

    file_id = photos[-1].file_id

    try:
        update_order_fields(order_id, {
            "receipt_file_id": file_id,
            "payment_status": "awaiting_confirm",
        })
    except Exception as e:
        _swallow(e)

    # ✅ محاولة الإشعار بالطريقة الرئيسية
    sent_to_admin = False
    try:
        await notify_admins_receipt(
            context,
            ud,
            receipt_file_id=file_id,
            client_id=user_id,
            receipt_is_photo=True
        )
        sent_to_admin = True
    except Exception:
        sent_to_admin = False

    # ✅ Fallback مضمون: إذا notify_admins_receipt فشل لأي سبب (مثل اختلاف توقيع admin_forward_kb)
    if (not sent_to_admin) and ADMIN_IDS:
        try:
            preview_html = build_order_preview(ud)
        except Exception:
            preview_html = f"<b>معاينة الطلب</b>\n🧾 <b>رقم الطلب</b>: {html.escape(order_id)}"

        msg_html = (
            "💳 <b>إيصال دفع جديد (Fallback)</b>\n"
            f"🧾 <b>رقم الطلب</b>: {html.escape(order_id)}\n\n"
            f"{preview_html}\n\n"
            "⬇️ <b>اعتماد الطلب:</b> اضغط (ارسال الطلب للتاجر) أو (الغاء الطلب)"
        )

        # قصّ للـ caption
        msg_html = (msg_html or "").strip()
        if len(msg_html) > 950:
            msg_html = msg_html[:949].rstrip() + "…"

        kb = admin_forward_kb(order_id)  # الكيبورد الحالي عندك (باراميتر واحد)
        for aid in ADMIN_IDS:
            try:
                await context.bot.send_photo(
                    chat_id=aid,
                    photo=file_id,
                    caption=msg_html,
                    parse_mode="HTML",
                    reply_markup=kb,
                )
            except Exception as e:
                _swallow(e)

    # ✅ معاينة موحّدة للعميل بعد الإيصال
    try:
        await _send_client_payment_preview(context, user_id, order_id, pay_scope="platform")
    except Exception as e:
        _swallow(e)

    set_stage(context, user_id, STAGE_DONE)
    return

async def receipt_document_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    ud = get_ud(context, user_id)

    if ud.get(STAGE_KEY) != STAGE_AWAIT_RECEIPT:
        return

    order_id = (ud.get("order_id") or "").strip()
    if not order_id:
        await update.message.reply_text(f"{_user_name(update)}\nلا يوجد طلب مرتبط بالايصال حاليا")
        set_stage(context, user_id, STAGE_NONE)
        return

    doc = update.message.document
    if not doc:
        await update.message.reply_text(f"{_user_name(update)}\nالايصال الزامي ارسل صورة او PDF فقط")
        return

    mime = (doc.mime_type or "").lower()
    fname = (doc.file_name or "").lower()
    is_ok = (
        mime.startswith("image/")
        or mime.startswith("application/pdf")
        or fname.endswith((".jpg", ".jpeg", ".png", ".webp", ".pdf"))
    )
    if not is_ok:
        await update.message.reply_text(f"{_user_name(update)}\nالايصال الزامي ارسل صورة او PDF فقط")
        return

    file_id = doc.file_id

    try:
        update_order_fields(order_id, {
            "receipt_file_id": file_id,
            "receipt_mime": mime,
            "payment_status": "awaiting_confirm",
        })
    except Exception as e:
        _swallow(e)

    # ✅ محاولة الإشعار بالطريقة الرئيسية
    sent_to_admin = False
    try:
        await notify_admins_receipt(
            context,
            ud,
            receipt_file_id=file_id,
            client_id=user_id,
            receipt_is_photo=False
        )
        sent_to_admin = True
    except Exception:
        sent_to_admin = False

    # ✅ Fallback مضمون: إرسال مباشر للادمن (PDF/صورة كـ Document)
    if (not sent_to_admin) and ADMIN_IDS:
        try:
            preview_html = build_order_preview(ud)
        except Exception:
            preview_html = f"<b>معاينة الطلب</b>\n🧾 <b>رقم الطلب</b>: {html.escape(order_id)}"

        msg_html = (
            "💳 <b>إيصال دفع جديد (Fallback)</b>\n"
            f"🧾 <b>رقم الطلب</b>: {html.escape(order_id)}\n\n"
            f"{preview_html}\n\n"
            "⬇️ <b>اعتماد الطلب:</b> اضغط (ارسال الطلب للتاجر) أو (الغاء الطلب)"
        )

        msg_html = (msg_html or "").strip()
        if len(msg_html) > 950:
            msg_html = msg_html[:949].rstrip() + "…"

        kb = admin_forward_kb(order_id)
        for aid in ADMIN_IDS:
            try:
                await context.bot.send_document(
                    chat_id=aid,
                    document=file_id,
                    caption=msg_html,
                    parse_mode="HTML",
                    reply_markup=kb,
                )
            except Exception as e:
                _swallow(e)

    # ✅ معاينة موحّدة للعميل بعد الإيصال
    try:
        await _send_client_payment_preview(context, user_id, order_id, pay_scope="platform")
    except Exception as e:
        _swallow(e)

    set_stage(context, user_id, STAGE_DONE)
    return

async def delivery_ship_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    user_id = q.from_user.id
    ud = get_ud(context, user_id)

    # ✅ Guard: لازم تكون جلسة طلب فعّالة
    order_id = (ud.get("order_id") or "").strip()
    items = ud.get("items", []) or []
    if (not order_id) or (not items):
        reset_flow(context, user_id)
        try:
            await q.message.reply_text("⚠️ انتهت جلسة الطلب أو تم فتح زر قديم.\nابدأ من جديد بكتابة: pp")
        except Exception as e:
            _swallow(e)
        return

    ud["ship"] = {}
    set_stage(context, user_id, STAGE_ASK_SHIP_CITY)

    await q.message.reply_text(
        f"{_user_name(q)}\nاكتب اسم المدينة",
        reply_markup=_flow_nav_kb("delivery"),
    )

async def delivery_pickup_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, f"تمام يا {_user_name(q)}")
    user_id = q.from_user.id
    ud = get_ud(context, user_id)

    # ✅ Guard: لازم تكون جلسة طلب فعّالة
    order_id = (ud.get("order_id") or "").strip()
    items = ud.get("items", []) or []
    if (not order_id) or (not items):
        reset_flow(context, user_id)
        try:
            await q.message.reply_text("⚠️ انتهت جلسة الطلب أو تم فتح زر قديم.\nابدأ من جديد بكتابة: pp")
        except Exception as e:
            _swallow(e)
        return

    ud["delivery_choice"] = "استلام من الموقع"
    ud.setdefault("pickup", {})

    set_stage(context, user_id, STAGE_ASK_PICKUP_CITY)
    await q.message.reply_text(f"{_user_name(q)}\nاكتب مدينة الاستلام")

def team_locked_kb(order_id: str, reason: str = "🔒 الطلب مقفول") -> InlineKeyboardMarkup:
    # زر واحد فقط داخل المجموعة يوضح أن الطلب مقفول (بدون فتح الخاص)
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(reason, callback_data=f"pp_team_locked|{order_id}")]
    ])

async def _lock_team_post_keyboard(context: ContextTypes.DEFAULT_TYPE, order_id: str, reason: str = "🔒 الطلب مقفول") -> None:
    """Lock the original TEAM group order post keyboard (remove quote deeplink) once accepted/locked."""
    if not TEAM_CHAT_ID:
        return
    try:
        b = get_order_bundle(order_id)
        o = b.get("order", {}) or {}
        tm = o.get("team_message_id")
    except Exception:
        tm = None

    if not (str(tm).isdigit()):
        return

    try:
        await context.bot.edit_message_reply_markup(
            chat_id=TEAM_CHAT_ID,
            message_id=int(tm),
            reply_markup=team_locked_kb(order_id, reason=reason),
        )
    except Exception:
        # ignore (message may be too old / missing rights)
        return

async def team_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    data = (q.data or "").strip()
    parts = data.split("|")
    action = parts[0].strip() if len(parts) >= 1 else ""
    order_id = parts[1].strip() if len(parts) >= 2 else ""
    if not action or not order_id:
        return

    actor_id = q.from_user.id
    actor_name = (q.from_user.full_name or "").strip()
    actor_first = (q.from_user.first_name or actor_name or "").strip()

    # ===== مكان التنفيذ =====
    in_team_group = bool(TEAM_CHAT_ID and q.message and q.message.chat_id == TEAM_CHAT_ID)
    in_private = bool(q.message and q.message.chat.type == ChatType.PRIVATE)
    if not (in_team_group or in_private):
        return

    # ===== داخل المجموعة: أزرار محددة فقط =====
    if in_team_group and action not in (
        "pp_team_quote",
        "pp_trader_open",
        "pp_team_locked",
        "pp_team_quote_locked",
    ):
        return

    # ===== اسم التاجر =====
    def _actor_label() -> str:
        try:
            tp = get_trader_profile(actor_id) or {}
        except Exception:
            tp = {}
        dn = (tp.get("display_name") or "").strip() or actor_first or actor_name or "التاجر"
        cn = (tp.get("company_name") or "").strip()
        return f"{dn} ({cn})" if cn else dn

    # ===== زر مقفول (تنبيه فقط) =====
    if action in ("pp_team_locked", "pp_team_quote_locked"):
        try:
            # لا نعتمد على parts[2] إطلاقًا
            reason_code = "locked"

            tname = (actor_first or actor_name or "").strip() or "عزيزي التاجر"

            # ✅ تحديد سبب القفل (إلغاء من الإدارة vs قفل عام/دفع)
            try:
                ob = get_order_bundle(order_id) or {}
                oo = (ob.get("order", {}) or {}) if isinstance(ob, dict) else {}
            except Exception:
                oo = {}

            ost = str(oo.get("order_status") or "").strip().lower()
            if ost in ("canceled", "cancelled", "ملغي"):
                msg = (
                    f"{tname}\n"
                    "⛔ هذا الطلب معلق/ملغي من قبل الإدارة ولا يستقبل عروض حالياً.\n"
                    "نشكر لك اهتمامك وتعاونك."
                )
            else:
                msg = (
                    f"{tname}\n"
                    "🔒 هذا الطلب مقفول ولا يستقبل عروض جديدة حالياً.\n"
                    "نشكر لك اهتمامك وتعاونك."
                )

            # ✅ Popup واضح داخل المجموعة + رسالة خاصة مع زر (بدون تلويث المجموعة)
            await _alert(q, msg, force=True)
            try:
                await context.bot.send_message(
                    chat_id=actor_id,
                    text=msg,
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("حسناً", callback_data="pp_ui_close")]]),
                )
            except Exception:
                pass
            # ✅ Popup واضح
            await _alert(q, msg, force=True)
        except Exception:
            # fallback آمن
            try:
                await q.answer("🔒 هذا الطلب مقفول حالياً", show_alert=True)
            except Exception as e:
                _swallow(e)
        return

    # ==========================================================
    # 💰 تقديم عرض سعر (من المجموعة فقط)
    # ==========================================================
    if action == "pp_team_quote":
        if _is_maintenance_mode() and actor_id not in ADMIN_IDS:
            await _alert(q, "🟧 المنصة في وضع الصيانة حاليا\nتم ايقاف تقديم عروض السعر مؤقتا", force=True)
            return

        try:
            ob = get_order_bundle(order_id)
            oo = ob.get("order", {}) or {}
        except Exception:
            oo = {}

        order_status = str(oo.get("order_status") or "").lower()
        quote_locked = str(oo.get("quote_locked") or "").lower() == "yes"
        goods_pay_status = str(oo.get("goods_payment_status") or "").lower()

        accepted_tid = int(oo.get("accepted_trader_id") or 0)
        accepted_name = (oo.get("accepted_trader_name") or "").strip()

        is_final_locked = (
            order_status in ("closed", "delivered", "canceled", "cancelled")
            or quote_locked
            or goods_pay_status in ("awaiting_confirm", "confirmed")
        )

        if is_final_locked and actor_id not in ADMIN_IDS:
            who = accepted_name or "تاجر آخر"
            await _alert(q, f"🔒 الطلب منتهي/مغلق حاليا ومعلق لدى: {who}", force=True)
            return

        # ✅ شرط: لا يبدأ عرض سعر إلا بعد اكتمال ملف التاجر
        try:
            tp = get_trader_profile(int(actor_id or 0)) or {}
        except Exception:
            tp = {}

        required_fields = ["display_name","company_name","shop_phone","cr_no","vat_no","bank_name","iban","stc_pay"]
        if not all((tp.get(f) or "").strip() for f in required_fields):
            await _need_complete_trader_profile_notice(context, actor_id, _user_name(q), order_id)
            await _alert(q, "تم إرسال طريقة التفعيل بالخاص", force=True)
            return

        # تهيئة إدخال العرض
        ad = context.user_data.setdefault(actor_id, {})
        ad["quote_order_id"] = order_id
        set_stage(context, actor_id, STAGE_TRADER_SET_QUOTE)

        # ملخص الطلب
        order_snapshot = f"رقم الطلب: {order_id}"
        try:
            b = get_order_bundle(order_id)
            order = b.get("order", {}) or {}
            items = b.get("items", []) or []


            parts_txt = "\n".join(
                f"{i}- {it.get('name','')}"
                for i, it in enumerate(items, start=1)
                if it.get("name")
            ) or "لا يوجد"

            order_snapshot = (
                "📌 ملخص الطلب\n"
                f"رقم الطلب: {order_id}\n"
                f"السيارة: {order.get('car_name','')}\n"
                f"الموديل: {order.get('car_model','')}\n"
                f"VIN: {order.get('vin','')}\n\n"
                f"القطع:\n{parts_txt}"
            )
        except Exception as e:
            _swallow(e)

        try:
            bot_username = getattr(context.bot, "username", "") or ""
            quote_url = f"https://t.me/{bot_username}?start=ppq_{order_id}"
            open_url = f"https://t.me/{bot_username}?start=ppopen_{order_id}"

            await context.bot.send_message(
                chat_id=actor_id,
                text=(
                    f"{_user_name(q)}\n"
                    f"👤 {_actor_label()}\n"
                    "💰 تقديم عرض سعر\n\n"
                    f"{order_snapshot}\n\n"
                    "✍️ اتبع الخطوات داخل المنصة لإرسال عرض منسق."
                ),
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("💰 فتح شاشة العرض بالخاص", url=quote_url)],
                    [InlineKeyboardButton("↗️ فتح لوحة الطلب", url=open_url)],
                ]),
                disable_web_page_preview=True,
            )
            await _alert(q, "تم إرسال التعليمات بالخاص")  # عادي toast
        except Exception:
            await _alert(q, "افتح المنصة بالخاص ثم أعد المحاولة", force=True)
        return

    # ==========================================================
    # 🧰 فتح لوحة الطلب (للتاجر المقبول فقط)
    # ==========================================================
    if action == "pp_trader_open":
        try:
            b = get_order_bundle(order_id)
            order = b.get("order", {}) or {}
        except Exception:
            order = {}

        acc = int(order.get("accepted_trader_id") or 0)
        if not acc:
            await _alert(q, "🔒 لم يتم إسناد الطلب لتاجر بعد", force=True)
            return

        accepted_name = (order.get("accepted_trader_name") or "").strip() or "التاجر المستلم"

        if acc != actor_id and actor_id not in ADMIN_IDS:
            await _alert(q, f"🔒 الطلب مخصص للتاجر: {accepted_name}", force=True)
            return

        try:
            await context.bot.send_message(
                chat_id=actor_id,
                text=f"🧰 لوحة التحكم\n🧾 رقم الطلب: {order_id}\n👤 التاجر: {accepted_name}",
                reply_markup=trader_status_kb(order_id),
                disable_web_page_preview=True,
            )
            await _alert(q, "تم إرسال لوحة الطلب بالخاص")
        except Exception:
            await _alert(q, "تعذر إرسال اللوحة", force=True)
        return

    # ==========================================================
    # 🔐 باقي الأوامر: خاص فقط
    # ==========================================================
    if not in_private:
        return

    # ===== تأكيد استلام قيمة القطع =====
    if action == "pp_team_goods_confirm":
        assigned = _assigned_trader_id(order_id)
        if assigned and actor_id not in (assigned, *ADMIN_IDS):
            await _alert(q, "غير مصرح", force=True)
            return

        b = get_order_bundle(order_id)
        order = b.get("order", {}) or {}
        if not order.get("goods_amount_sar"):
            await q.message.reply_text("لا يوجد مبلغ مسجل لهذا الطلب")
            return

        # ✅✅✅ FIX: منع التكرار إذا كان تم التأكيد مسبقاً (Idempotent)
        try:
            gps_now = str(order.get("goods_payment_status") or "").strip().lower()
        except Exception:
            gps_now = ""
        if gps_now == "confirmed":
            try:
                await _alert(q, "تم تأكيد السداد مسبقا", force=False)
            except Exception:
                _swallow(Exception("toast"))
            return
        # ✅✅✅ END FIX

        # ✅ تحديد الحالة الصحيحة بعد التأكيد (بدون in_progress نهائيًا)
        try:
            ost_now = str(order.get("order_status") or "").strip().lower()
        except Exception:
            ost_now = ""

        inv_file_now = (str(order.get("seller_invoice_file_id") or order.get("shop_invoice_file_id") or "")).strip()

        # ✅ وضع الرابط: اعتبر وجود رابط الدفع = بديل للفواتير (حتى لا نرجّع الطلب للخلف)
        try:
            pay_method_now = str(order.get("goods_payment_method") or "").strip().lower()
        except Exception:
            pay_method_now = ""
        pay_link_now = (str(order.get("goods_payment_link") or "")).strip()
        inv_ok = bool(inv_file_now) or (bool(pay_link_now) and pay_method_now in ("pay_link", "link", "payment_link"))

        # لو الطلب وصل مراحل متقدمة لا نرجعه للخلف
        if ost_now in ("shipped", "delivered", "closed"):
            next_ost = ost_now
        else:
            if inv_ok:
                next_ost = "ready_to_ship"
            else:
                next_ost = ost_now if ost_now else "preparing"

        update_order_fields(order_id, {
            "goods_payment_status": "confirmed",
            "goods_payment_confirmed_at_utc": utc_now_iso(),
            "quote_locked": "yes",
            "order_status": next_ost,
        })

        # ✅ نسخة احتياطية ذكية بعد التأكيد
        try:
            app = getattr(context, "application", None)
            if app:
                if not app.bot_data.get("_backup_touch_goods_confirm"):
                    app.bot_data["_backup_touch_goods_confirm"] = True

                    async def _bk_job():
                        try:
                            await asyncio.sleep(5)
                            await _send_backup_excel(app, reason="goods_confirmed")
                        finally:
                            try:
                                app.bot_data["_backup_touch_goods_confirm"] = False
                            except Exception as e:
                                _swallow(e)

                asyncio.create_task(_bk_job())
        except Exception as e:
            _swallow(e)

        # 🔒 قفل زر المجموعة بصريًا
        try:
            await _lock_team_post_keyboard(
                context,
                order_id,
                reason="🔒 تم إيقاف العروض الطلب منتهي"
            )
        except Exception as e:
            _swallow(e)

        # ✅ إرسال فاتورة التاجر للعميل
        try:
            await send_trader_invoice_pdf(
                context=context,
                order_id=order_id,
                kind="preliminary",
                tracking_number="",
                admin_only=False,
            )
        except Exception as e:
            _swallow(e)

        # ✅ إرسال عنوان الشحن للتاجر + لوحة الطلب
        try:
            b3 = get_order_bundle(order_id) or {}
            o3 = b3.get("order", {}) or {}
            tid3 = int(o3.get("accepted_trader_id") or 0)

            ship_city = (o3.get("ship_city") or o3.get("pickup_city") or "").strip()
            ship_dist = (o3.get("ship_district") or "").strip()
            ship_short = (o3.get("ship_short_address") or "").strip()
            ship_phone = (o3.get("ship_phone") or "").strip()
            delivery_details = (o3.get("delivery_details") or "").strip()

            addr_lines = []
            if ship_city:
                addr_lines.append(f"المدينة: {ship_city}")
            if ship_dist:
                addr_lines.append(f"الحي: {ship_dist}")
            if ship_short:
                addr_lines.append(f"العنوان المختصر: {ship_short}")
            if delivery_details:
                addr_lines.append(f"تفاصيل إضافية: {delivery_details}")
            if ship_phone:
                addr_lines.append(f"📞 رقم الجوال: {ship_phone}")

            addr_block = "\n".join(addr_lines) if addr_lines else "—"

            # ===== ✅ المطلوب: إظهار (قيمة القطع + الشحن + الإجمالي) بشكل صحيح =====
            # قيمة القطع كما أدخلها التاجر
            try:
                goods_num3 = float(o3.get("goods_amount_sar") or o3.get("goods_total_sar") or 0)
            except Exception:
                goods_num3 = 0.0

            # الشحن: إذا لم يُدخل = 0
            try:
                ship_num3 = float(o3.get("shipping_fee_sar") or o3.get("shipping_fee") or 0)
            except Exception:
                ship_num3 = 0.0

            # هل الشحن مشمول؟
            try:
                ship_included = str(o3.get("ship_included") or o3.get("shipping_included") or "").strip().lower() in (
                    "yes", "y", "true", "1", "مشمول", "included"
                )
            except Exception:
                ship_included = False

            # الإجمالي: القطع + الشحن (إذا غير مشمول)
            total_num3 = goods_num3 + (0.0 if ship_included else ship_num3)

            amt3 = _money(goods_num3)     # قيمة القطع
            ship3 = _money(ship_num3)     # قيمة الشحن (قد تكون 0)
            total3 = _money(total_num3)   # إجمالي القطع + الشحن (حسب المشمول)

            # سطر الشحن: يظهر إذا غير مشمول (حتى لو 0)
            ship_line = ""
            total_label = "إجمالي الفاتورة"
            if ship_included:
                ship_line = "🚚 <b>الشحن</b>: <b>مشمول</b>\n"
                total_label = "إجمالي القطع (الشحن مشمول)"
            else:
                ship_line = f"🚚 <b>الشحن</b>: <b>{html.escape(str(ship3))}</b>\n"
                total_label = "إجمالي القطع + الشحن"
            # ===== نهاية المطلوب =====

            # اسم العميل الحقيقي (يظهر للتاجر فقط)
            client_name3 = (o3.get("user_name") or "").strip() or "العميل"

            # يوزر العميل (للإدارة فقط)
            client_username3 = (o3.get("user_username") or o3.get("username") or o3.get("client_username") or "").strip()
            if client_username3 and not client_username3.startswith("@"):
                client_username3 = f"@{client_username3}"
            if not client_username3:
                client_username3 = "—"

            if tid3:
                await context.bot.send_message(
                    chat_id=tid3,
                    text=(
                        "✅💳 <b>تم تأكيد سداد قيمة القطع</b>\n"
                        + f"🧾 <b>رقم الطلب</b>: {html.escape(order_id)}\n"
                        + f"👤 <b>العميل</b>: <b>{html.escape(client_name3)}</b>\n"
                        + f"🚗 <b>السيارة</b>: <b>{html.escape(str(o3.get('car_name') or o3.get('vehicle_name') or o3.get('car_model') or o3.get('car') or '—'))}</b>\n"
                        + f"📌 <b>الحالة</b>: <b>{_pay_status_ar(next_ost)}</b>\n"
                        + (f"🛠 <b>مدة التجهيز</b>: <b>{html.escape(str(o3.get('availability_days') or o3.get('quote_availability') or o3.get('availability') or '').strip())}</b>\n"
                           if str(o3.get('availability_days') or o3.get('quote_availability') or o3.get('availability') or '').strip() else "")
                        + (f"⏱ <b>مدة الشحن</b>: <b>{html.escape(str(o3.get('ship_eta') or o3.get('shipping_eta') or o3.get('ship_days') or '').strip())}</b>\n"
                           if str(o3.get('ship_eta') or o3.get('shipping_eta') or o3.get('ship_days') or '').strip() else "")
                        + f"💰 <b>قيمة القطع</b>: <b>{html.escape(str(amt3))}</b>\n"
                        + f"{ship_line}"
                        + f"🧾 <b>{html.escape(str(total_label))}</b>: <b>{html.escape(str(total3))}</b>\n\n"
                        + "🚀 <b>يرجى البدء بتجهيز الطلب</b> الآن\n"
                        + "🚚 <b>وعند الشحن</b>: حدّث الحالة + أرسل رقم التتبع (اختياري)\n\n"
                        + "📍 <b>عنوان الشحن (تم فك السرية بعد السداد)</b>:\n"
                        + f"<pre>{html.escape(addr_block)}</pre>\n\n"
                        + "⬇️ <b>لوحة الطلب</b>:"
                    ),
                    parse_mode="HTML",
                    disable_web_page_preview=True,
                    reply_markup=trader_status_kb(order_id),
                )

                # ✅ المهم: تحديث لوحة التاجر القديمة (لو كان فاتح لوحة سابقة)
                try:
                    await _show_order_panel_private(context, int(tid3), order_id)
                except Exception as e:
                    _swallow(e)
        except Exception as e:
            _swallow(e)

        # ✅ نسخة للإدارة: PDF مرة واحدة
        try:
            await send_trader_invoice_pdf(
                context=context,
                order_id=order_id,
                kind="preliminary",
                tracking_number="",
                admin_only=True,
            )
        except Exception as e:
            _swallow(e)

        # ✅ إشعار الإدارة (نصي): الاسم الحقيقي + اليوزر + المبالغ
        try:
            # نعيد جلب بيانات الطلب للتأكد أن المتغيرات موجودة حتى لو فشل try السابق
            b4 = get_order_bundle(order_id) or {}
            o4 = b4.get("order", {}) or {}

            try:
                goods_num4 = float(o4.get("goods_amount_sar") or o4.get("goods_total_sar") or 0)
            except Exception:
                goods_num4 = 0.0
            try:
                ship_num4 = float(o4.get("shipping_fee_sar") or o4.get("shipping_fee") or 0)
            except Exception:
                ship_num4 = 0.0
            try:
                ship_included4 = str(o4.get("ship_included") or o4.get("shipping_included") or "").strip().lower() in (
                    "yes", "y", "true", "1", "مشمول", "included"
                )
            except Exception:
                ship_included4 = False

            total_num4 = goods_num4 + (0.0 if ship_included4 else ship_num4)

            amt4 = _money(goods_num4)
            ship4 = _money(ship_num4)
            total4 = _money(total_num4)

            ship_line_admin = ""
            total_label_admin = "إجمالي الفاتورة"
            if ship_included4:
                ship_line_admin = "🚚 <b>الشحن</b>: <b>مشمول</b>\n"
                total_label_admin = "إجمالي القطع (الشحن مشمول)"
            else:
                ship_line_admin = f"🚚 <b>الشحن</b>: <b>{html.escape(str(ship4))}</b>\n"
                total_label_admin = "إجمالي القطع + الشحن"

            client_name_admin = (o4.get("user_name") or "").strip() or "العميل"

            client_username_admin = (
                o4.get("user_username")
                or o4.get("username")
                or o4.get("client_username")
                or ""
            )
            client_username_admin = (str(client_username_admin) or "").strip()

            # ✅ تنظيف اليوزر: لا نعرض (-) ولا (—) ولا أقواس فارغة
            _u_raw = client_username_admin.strip()
            if _u_raw in ("-", "—"):
                _u_raw = ""
            _u_raw = _u_raw.lstrip("@").strip()
            u_part_admin = f" (@{html.escape(_u_raw)})" if _u_raw else ""

            trader_name_admin = (o4.get("accepted_trader_name") or o4.get("quoted_trader_name") or "").strip()
            if not trader_name_admin:
                try:
                    tid_admin = int(o4.get("accepted_trader_id") or 0)
                except Exception:
                    tid_admin = 0
                if tid_admin:
                    try:
                        tp = get_trader_profile(int(tid_admin)) or {}
                        trader_name_admin = (tp.get("display_name") or "").strip() or (tp.get("company_name") or "").strip()
                    except Exception as e:
                        _swallow(e)
            trader_name_admin = trader_name_admin or "—"

            for aid in ADMIN_IDS:
                try:
                    await context.bot.send_message(
                        chat_id=int(aid),
                        text=(
                            "📌 <b>تأكيد استلام قيمة القطع</b>\n"
                            f"🧾 <b>رقم الطلب</b>: {html.escape(order_id)}\n"
                            f"👤 <b>العميل</b>: <b>{html.escape(client_name_admin)}</b>{u_part_admin}\n"
                            f"🧑‍💼 <b>التاجر</b>: <b>{html.escape(trader_name_admin)}</b>\n\n"
                            f"🧩 <b>قيمة القطع</b>: <b>{html.escape(str(amt4))}</b>\n"
                            f"{ship_line_admin}"
                            f"🧾 <b>{html.escape(str(total_label_admin))}</b>: <b>{html.escape(str(total4))}</b>\n"
                        ),
                        parse_mode="HTML",
                        disable_web_page_preview=True,
                    )
                except Exception as e:
                    _swallow(e)
        except Exception as e:
            _swallow(e)

        # ✅ إشعار العميل
        uid = get_order_user_id(order_id)
        if uid:
            try:
                # نعيد جلب للتأكد
                b5 = get_order_bundle(order_id) or {}
                o5 = b5.get("order", {}) or {}

                try:
                    goods_num5 = float(o5.get("goods_amount_sar") or o5.get("goods_total_sar") or 0)
                except Exception:
                    goods_num5 = 0.0
                try:
                    ship_num5 = float(o5.get("shipping_fee_sar") or o5.get("shipping_fee") or 0)
                except Exception:
                    ship_num5 = 0.0
                try:
                    ship_included5 = str(o5.get("ship_included") or o5.get("shipping_included") or "").strip().lower() in (
                        "yes", "y", "true", "1", "مشمول", "included"
                    )
                except Exception:
                    ship_included5 = False

                total_num5 = goods_num5 + (0.0 if ship_included5 else ship_num5)

                amt5 = _money(goods_num5)
                ship5 = _money(ship_num5)
                total5 = _money(total_num5)

                ship_line_client = ""
                total_label_client = "إجمالي الفاتورة"
                if ship_included5:
                    ship_line_client = "🚚 <b>الشحن</b>: <b>مشمول</b>\n"
                    total_label_client = "إجمالي القطع (الشحن مشمول)"
                else:
                    ship_line_client = f"🚚 <b>الشحن</b>: <b>{html.escape(str(ship5))}</b>\n"
                    total_label_client = "إجمالي القطع + الشحن"

                # ✅ إضافة اسم التاجر/المتجر للعميل (المطلوب)
                trader_name_client = (o5.get("accepted_trader_name") or o5.get("quoted_trader_name") or "").strip()
                trader_store_client = (o5.get("accepted_store_name") or o5.get("shop_name") or o5.get("store_name") or "").strip()
                if (not trader_name_client or not trader_store_client):
                    try:
                        tid5 = int(o5.get("accepted_trader_id") or 0)
                    except Exception:
                        tid5 = 0
                    if tid5:
                        try:
                            tp5 = get_trader_profile(int(tid5)) or {}
                        except Exception:
                            tp5 = {}
                        if not trader_name_client:
                            trader_name_client = (tp5.get("display_name") or "").strip() or (tp5.get("company_name") or "").strip()
                        if not trader_store_client:
                            trader_store_client = (tp5.get("company_name") or "").strip() or (tp5.get("store_name") or "").strip() or (tp5.get("display_name") or "").strip()
                trader_name_client = trader_name_client or "التاجر"
                trader_store_client = trader_store_client or "المتجر"

                await context.bot.send_message(
                    chat_id=uid,
                    text=(
                        "✅ <b>تم تأكيد استلام قيمة القطع بنجاح</b>\n"
                        f"🧾 <b>رقم الطلب</b>: {html.escape(order_id)}\n"
                        f"🧑‍💼 <b>التاجر</b>: <b>{html.escape(trader_name_client)}</b>\n"
                        f"🏪 <b>المتجر</b>: <b>{html.escape(trader_store_client)}</b>\n"
                        f"{ship_line_client}"
                        f"🧾 <b>{html.escape(str(total_label_client))}</b>: <b>{html.escape(str(total5))}</b>\n\n"
                        "🧰 الطلب الآن قيد التجهيز\n"
                        "🚚 سيتم تحديثك عند الشحن."
                    ),
                    parse_mode="HTML",
                    reply_markup=chat_nav_kb_for(context, to_uid, order_id, "pp_chat_trader_done"),
                    disable_web_page_preview=True,
                )
            except Exception as e:
                _swallow(e)

        return

async def media_router(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    ud = get_ud(context, user_id)
    stage = ud.get(STAGE_KEY)

    # =========================
    # Helpers (أمان + أرقام)
    # =========================
    def _s(x: object) -> str:
        return ("" if x is None else str(x)).strip()

    def _safe_int(x) -> int:
        try:
            return int(x or 0)
        except Exception:
            return 0

    def _num(x: object) -> float:
        s = _s(x)
        if not s:
            return 0.0
        try:
            return float(s)
        except Exception:
            try:
                s2 = re.sub(r"[^\d.]", "", s)
                return float(s2) if s2 else 0.0
            except Exception:
                return 0.0

    def _fmt_amt(x: object) -> str:
        v = _num(x)
        if v <= 0:
            return ""
        if abs(v - int(v)) < 1e-9:
            return f"{int(v)} ر.س"
        return f"{v:.2f}".rstrip("0").rstrip(".") + " ر.س"

    def _calc_goods_ship_total(order: dict, ud_local: dict | None = None) -> dict:
        ud_local = ud_local or {}
        goods = (
            _s(order.get("goods_amount_sar"))
            or _s(order.get("quote_goods_amount"))
            or _s(ud_local.get("quote_goods_amount"))
            or ""
        )
        ship = (
            _s(order.get("shipping_fee_sar"))
            or _s(order.get("shipping_fee"))
            or _s(order.get("quote_shipping_fee"))
            or _s(ud_local.get("quote_shipping_fee"))
            or ""
        )

        goods_val = _num(goods)
        ship_val = _num(ship)

        total_saved = _num(order.get("total_amount_sar") or order.get("total_amount") or "")
        total_val = goods_val + ship_val if (goods_val > 0 or ship_val > 0) else total_saved
        if total_saved > 0 and total_saved >= total_val:
            total_val = total_saved

        return {
            "goods_val": goods_val,
            "ship_val": ship_val,
            "total_val": total_val,
            "goods_txt": _fmt_amt(goods_val),
            "ship_txt": _fmt_amt(ship_val),
            "total_txt": _fmt_amt(total_val),
        }

    def _pick_media(m):
        try:
            if getattr(m, "photo", None):
                ph = m.photo[-1]
                return ("photo", getattr(ph, "file_id", None))
        except Exception:
            pass
        try:
            if getattr(m, "video", None):
                return ("video", getattr(m.video, "file_id", None))
        except Exception:
            pass
        try:
            if getattr(m, "document", None):
                return ("document", getattr(m.document, "file_id", None))
        except Exception:
            pass
        try:
            if getattr(m, "voice", None):
                return ("voice", getattr(m.voice, "file_id", None))
        except Exception:
            pass
        try:
            if getattr(m, "audio", None):
                return ("audio", getattr(m.audio, "file_id", None))
        except Exception:
            pass
        try:
            if getattr(m, "video_note", None):
                return ("video_note", getattr(m.video_note, "file_id", None))
        except Exception:
            pass
        return (None, None)

    async def _send_media(kind: str, file_id: str, chat_id: int, caption: str = "", kb=None):
        try:
            if kind == "photo":
                await context.bot.send_photo(
                    chat_id=chat_id,
                    photo=file_id,
                    caption=caption or None,
                    reply_markup=kb,
                    disable_web_page_preview=True,
                )
            elif kind == "video":
                await context.bot.send_video(
                    chat_id=chat_id,
                    video=file_id,
                    caption=caption or None,
                    reply_markup=kb,
                    disable_web_page_preview=True,
                )
            elif kind == "voice":
                await context.bot.send_voice(
                    chat_id=chat_id,
                    voice=file_id,
                    caption=caption or None,
                    reply_markup=kb,
                    disable_web_page_preview=True,
                )
            elif kind == "audio":
                await context.bot.send_audio(
                    chat_id=chat_id,
                    audio=file_id,
                    caption=caption or None,
                    reply_markup=kb,
                    disable_web_page_preview=True,
                )
            elif kind == "video_note":
                await context.bot.send_video_note(
                    chat_id=chat_id,
                    video_note=file_id,
                    reply_markup=kb,
                )
            else:
                await context.bot.send_document(
                    chat_id=chat_id,
                    document=file_id,
                    caption=caption or None,
                    reply_markup=kb,
                    disable_web_page_preview=True,
                )
        except Exception as e:
            _swallow(e)

    # =========================================================
    # ✅ حماية مراحل المراسلة (Stages) من التعليق داخل الوسائط
    # =========================================================
    try:
        idle_secs = int(os.getenv("PP_CHAT_IDLE_SECS", "1800") or 1800)
    except Exception:
        idle_secs = 1800
    try:
        now_ts = int(time.time())
    except Exception:
        now_ts = 0

    CHAT_STAGES = {
        STAGE_CHAT_TRADER,
        STAGE_TRADER_REPLY,
        STAGE_ADMIN_REPLY,
        STAGE_ADMIN_CHAT,
        STAGE_TRADER_CHAT_ADMIN,
        STAGE_SUPPORT_ADMIN_REPLY,
        STAGE_APPLICANT_CHAT_ADMIN,
    }

    if stage in CHAT_STAGES:
        try:
            prev_stage = (_s(ud.get("chat_stage_name"))).strip()
            if prev_stage != stage:
                ud["chat_stage_name"] = stage
                ud["chat_stage_started_at"] = now_ts
                ud["chat_stage_last_touch"] = now_ts

            last_touch = _safe_int(ud.get("chat_stage_last_touch")) or now_ts
            if now_ts and idle_secs and (now_ts - last_touch) > idle_secs:
                try:
                    ud.pop("chat_trader_order_id", None)
                    ud.pop("trader_reply_order_id", None)
                    ud.pop("admin_reply_order_id", None)
                    ud.pop("admin_chat_order_id", None)
                    ud.pop("admin_chat_peer_id", None)
                    ud.pop("admin_chat_role", None)
                    ud.pop("trader_chat_admin_order_id", None)
                    ud.pop("trader_chat_admin_peer_id", None)
                    ud.pop("support_admin_peer_id", None)
                    ud.pop("support_admin_role", None)
                    ud.pop("applicant_chat_peer_id", None)
                except Exception:
                    pass
                ud.pop("chat_stage_name", None)
                ud.pop("chat_stage_started_at", None)
                ud.pop("chat_stage_last_touch", None)
                set_stage(context, user_id, STAGE_NONE)
                stage = STAGE_NONE
            else:
                ud["chat_stage_last_touch"] = now_ts
        except Exception:
            pass

    # =========================================================
    # ✅ Media forwarding داخل نظام المراسلات (عميل/تاجر/إدارة)
    # =========================================================
    msg = getattr(update, "message", None)
    if msg:
        kind, file_id = _pick_media(msg)
        cap_raw = (getattr(msg, "caption", None) or "").strip()

        if len(cap_raw) > 900:
            cap_raw = cap_raw[:900].rstrip() + "…"

        # =========================================================
        # ✅✅ إصلاح صارم + Fallback للإيصال (حتى لو STAGE ضائع)
        # - السبب الحقيقي لتوقف الإيصال غالبًا: STAGE يتصفّر أو لا يصل للعميل
        #   فيرجع stage = NONE فتروح الوسائط للمراسلة أو تُهمل.
        # - هنا: إذا المستخدم أرسل صورة/PDF في الخاص، ونجد له طلب "بانتظار إيصال قيمة القطع"
        #   نربطه فورًا ونحوّله لـ STAGE_AWAIT_GOODS_RECEIPT ثم نمرره للهاندلر.
        # =========================================================
        try:
            is_private = bool(getattr(msg, "chat", None) and getattr(msg.chat, "type", None) == ChatType.PRIVATE)
        except Exception:
            is_private = False

        if is_private and kind and file_id:
            # هل الملف مقبول كإيصال (صورة أو PDF/صورة داخل document)
            is_receipt_media = False
            try:
                if kind == "photo":
                    is_receipt_media = True
                elif kind == "document":
                    doc0 = getattr(msg, "document", None)
                    if doc0:
                        mime0 = (doc0.mime_type or "").lower()
                        fname0 = (doc0.file_name or "").lower()
                        if mime0.startswith("application/pdf") or fname0.endswith(".pdf") or mime0.startswith("image/") or fname0.endswith((".jpg", ".jpeg", ".png", ".webp")):
                            is_receipt_media = True
            except Exception:
                is_receipt_media = False

            # إذا stage ليس مرحلة مراسلة + ليس مراحل الإيصالات الحالية، نعمل Fallback
            if is_receipt_media and stage not in CHAT_STAGES and stage not in (STAGE_AWAIT_GOODS_RECEIPT, STAGE_AWAIT_RECEIPT, STAGE_TRADER_SUB_AWAIT_RECEIPT, STAGE_JOIN_CR, STAGE_JOIN_LICENSE):
                try:
                    cand_order_id = ""
                    try:
                        orders = list_orders() or []
                    except Exception:
                        orders = []

                    best_num = -1
                    for o in orders:
                        try:
                            uid = int(o.get("user_id") or 0)
                        except Exception:
                            uid = 0
                        if uid != int(user_id):
                            continue

                        gps = str(o.get("goods_payment_status") or "").strip().lower()
                        # حالات متوقعة قبل استلام الإيصال
                        if gps not in ("awaiting_receipt", "awaiting_receipt_goods", "awaiting_goods_receipt", "awaiting_receipt_only"):
                            continue

                        # لا نختار طلب لديه إيصال محفوظ
                        gr = (str(o.get("goods_receipt_file_id") or o.get("goods_payment_receipt_file_id") or "")).strip()
                        if gr:
                            continue

                        oid = str(o.get("order_id") or "").strip()
                        if not oid:
                            continue

                        # اختيار أحدث طلب: نعتمد على رقم الطلب كأمان (0063..)
                        try:
                            on = int(re.sub(r"[^\d]", "", oid) or 0)
                        except Exception:
                            on = 0
                        if on > best_num:
                            best_num = on
                            cand_order_id = oid

                    if cand_order_id:
                        try:
                            ud["goods_order_id"] = cand_order_id
                        except Exception:
                            pass
                        try:
                            set_stage(context, user_id, STAGE_AWAIT_GOODS_RECEIPT)
                            stage = STAGE_AWAIT_GOODS_RECEIPT
                        except Exception:
                            pass

                        # الآن مرّرها للهاندلر الصحيح فورًا
                        try:
                            if update.message and update.message.photo:
                                return await goods_receipt_photo_handler(update, context)
                        except Exception as e:
                            _swallow(e)

                        try:
                            doc1 = update.message.document if update.message else None
                            if doc1:
                                mime1 = (doc1.mime_type or "").lower()
                                fname1 = (doc1.file_name or "").lower()
                                is_pdf1 = mime1.startswith("application/pdf") or fname1.endswith(".pdf")
                                is_img1 = mime1.startswith("image/") or fname1.endswith((".jpg", ".jpeg", ".png", ".webp"))
                                if is_pdf1 or is_img1:
                                    return await goods_receipt_document_handler(update, context)
                        except Exception as e:
                            _swallow(e)
                except Exception:
                    pass

        # =========================================================
        # 0) تمرير وسائط المراسلة الداخلية pp_chat_sessions (قبل أي STAGE)
        # =========================================================
        try:
            sessions = context.bot_data.get("pp_chat_sessions") or {}
            sess = sessions.get(str(user_id))
        except Exception:
            sessions = {}
            sess = None

        # ✅✅ إصلاح صارم: لا تسمح للـ session بخطف وسائط مراحل الإيصالات/الفواتير/البوابة
        try:
            BLOCK_SESS_STAGES = {
                STAGE_AWAIT_RECEIPT,              # إيصال رسوم المنصة
                STAGE_AWAIT_GOODS_RECEIPT,        # إيصال قيمة القطع
                STAGE_TRADER_SUB_AWAIT_RECEIPT,   # إيصال اشتراك التاجر
                STAGE_JOIN_CR,                    # بوابة التجار (سجل)
                STAGE_JOIN_LICENSE,               # بوابة التجار (رخصة)
            }
            if stage in BLOCK_SESS_STAGES:
                sess = None
            # رفع فاتورة التاجر داخل تحديث الحالة: لا نخليها تروح مراسلة
            if stage == STAGE_TRADER_STATUS_UPDATE and (_s(ud.get("tsu_kind")) == "seller_invoice"):
                sess = None
        except Exception:
            pass

        if isinstance(sess, dict) and kind and file_id:
            peer_id = _safe_int(sess.get("peer_id"))
            order_id_sess = _s(sess.get("order_id"))
            role = _s(sess.get("role"))  # client / trader

            # ⏱️ Timeout
            try:
                idle2 = int(os.getenv("PP_CHAT_IDLE_SECS", "1800") or 1800)
            except Exception:
                idle2 = 1800
            try:
                max_secs = int(os.getenv("PP_CHAT_MAX_SECS", "21600") or 21600)
            except Exception:
                max_secs = 21600
            try:
                now2 = int(time.time())
            except Exception:
                now2 = 0

            started_at = _safe_int(sess.get("started_at")) or now2
            last_touch = _safe_int(sess.get("last_touch")) or started_at

            expired = False
            if now2 and idle2 and (now2 - last_touch) > idle2:
                expired = True
            if now2 and max_secs and (now2 - started_at) > max_secs:
                expired = True

            if expired:
                try:
                    sessions.pop(str(user_id), None)
                    if peer_id:
                        sessions.pop(str(peer_id), None)
                    context.bot_data["pp_chat_sessions"] = sessions
                except Exception:
                    pass
            else:
                if peer_id and order_id_sess:
                    # تحديث آخر تفاعل
                    try:
                        sess["last_touch"] = now2
                        sessions[str(user_id)] = sess
                        context.bot_data["pp_chat_sessions"] = sessions
                    except Exception:
                        pass

                    try:
                        cn, tn = _order_parties(order_id_sess)
                        sender = f"👤 العميل: {cn}" if role == "client" else f"👤 التاجر: {tn}"
                        receiver = f"⬅️ إلى: {tn}" if role == "client" else f"⬅️ إلى: {cn}"
                        header = f"{sender}\n{receiver}\n{_order_tag_plain(order_id_sess)}"
                        caption = f"{header}\n💬 {cap_raw}" if cap_raw else f"{header}\n📎 مرفق"
                        kb_end = InlineKeyboardMarkup([
                            [InlineKeyboardButton("✖️ إنهاء المراسلة", callback_data=f"pp_chat_end|{order_id_sess}")]
                        ])

                        await _send_media(kind, file_id, peer_id, caption=caption, kb=kb_end)
                    except Exception as e:
                        _swallow(e)

                    return

        # =========================================================
        # ✅ Join Portal (بوابة التجار)
        # =========================================================
        try:
            if getattr(msg, "chat", None) and getattr(msg.chat, "type", None) == ChatType.PRIVATE:
                if stage in (STAGE_JOIN_CR, STAGE_JOIN_LICENSE) and kind and file_id:
                    jd = _join_ud(context, user_id)

                    if stage == STAGE_JOIN_CR:
                        jd["cr_kind"] = kind
                        jd["cr_file_id"] = file_id
                        set_stage(context, user_id, STAGE_JOIN_LICENSE)
                        await msg.reply_text(
                            "✅ تم استلام السجل التجاري.\n\n"
                            "الخطوة التالية:\n"
                            "أرسل رخصة/إثبات المتجر (صورة أو PDF).",
                            parse_mode="HTML",
                            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("✖️ إنهاء", callback_data="pp_join_done")]]),
                            disable_web_page_preview=True,
                        )
                        return

                    if stage == STAGE_JOIN_LICENSE:
                        jd["license_kind"] = kind
                        jd["license_file_id"] = file_id
                        set_stage(context, user_id, STAGE_JOIN_VAT)
                        await msg.reply_text(
                            "✅ تم استلام رخصة/إثبات المتجر.\n\n"
                            "الخطوة التالية:\n"
                            "اكتب الرقم الضريبي (15 رقم).",
                            parse_mode="HTML",
                            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("✖️ إنهاء", callback_data="pp_join_done")]]),
                            disable_web_page_preview=True,
                        )
                        return
        except Exception:
            pass

        # ===== STAGE_APPLICANT_CHAT_ADMIN (متقدم → إدارة) =====
        if stage == STAGE_APPLICANT_CHAT_ADMIN and kind and file_id:
            admin_id = _safe_int(ud.get("applicant_chat_admin_id"))
            if not admin_id:
                return
            caption = f"{_join_portal_brand()}\n\n📩 رسالة من متقدم → الإدارة\n🆔 ID: {user_id}\n" + (f"\n{cap_raw}" if cap_raw else "")
            try:
                await _send_media(
                    "photo" if kind == "photo" else "document",
                    file_id,
                    admin_id,
                    caption=caption,
                    kb=_join_portal_admin_kb(user_id),
                )
            except Exception:
                pass
            try:
                await msg.reply_text("✅ تم إرسال الرسالة.")
            except Exception:
                pass
            return

        # 2) مراحل المراسلة المعروفة (توجيه وسائط)
        try:
            name_now = _user_name(update)
        except Exception:
            name_now = "—"

        # ===== STAGE_SUPPORT_ADMIN_REPLY (رد الإدارة من /منصة) =====
        if user_id in ADMIN_IDS and stage == STAGE_SUPPORT_ADMIN_REPLY and kind and file_id:
            target_uid = _safe_int(ud.get("support_reply_to_uid"))
            order_id_reply = _s(ud.get("support_reply_order_id"))
            if not target_uid:
                return
            caption = (
                f"{PP_SUPPORT_LABEL}\n"
                f"🧾 رقم الطلب: {order_id_reply or '—'}\n"
                "────────────────\n"
                + (cap_raw if cap_raw else "📎 مرفق")
            )
            await _send_media(kind, file_id, target_uid, caption=caption, kb=_support_kb())
            try:
                await msg.reply_text("✅ تم إرسال ردّك للمستخدم.")
            except Exception as e:
                _swallow(e)
            ud[STAGE_KEY] = STAGE_NONE
            ud.pop("support_reply_to_uid", None)
            ud.pop("support_reply_order_id", None)
            return

        # ===== STAGE_ADMIN_TRADER_MSG (رسالة الإدارة للتاجر من ملف التاجر) =====
        if user_id in ADMIN_IDS and stage == STAGE_ADMIN_TRADER_MSG and kind and file_id:
            tid = _safe_int(ud.get("admin_msg_to_trader_id"))
            if not tid:
                return
            caption = f"{PP_SUPPORT_LABEL}\n💬 رسالة من الإدارة:\n" + (cap_raw if cap_raw else "📎 مرفق")
            await _send_media(kind, file_id, tid, caption=caption, kb=_admin_to_trader_reply_kb(user_id))
            try:
                await msg.reply_text("✅ تم إرسال الرسالة للتاجر.")
            except Exception as e:
                _swallow(e)
            ud[STAGE_KEY] = STAGE_NONE
            ud.pop("admin_msg_to_trader_id", None)
            return

        # ===== STAGE_CHAT_TRADER (عميل → تاجر) =====
        if stage == STAGE_CHAT_TRADER and kind and file_id:
            order_id = _s(ud.get("chat_trader_order_id"))
            tid = _assigned_trader_id(order_id) if order_id else None
            if not order_id or not tid:
                ud[STAGE_KEY] = STAGE_NONE
                try:
                    await msg.reply_text("⚠️ لا يوجد تاجر محدد لهذا الطلب حالياً.")
                except Exception as e:
                    _swallow(e)
                return

            try:
                snap = _order_snapshot(order_id) or {}
            except Exception:
                snap = {}

            head = []
            try:
                head.append(f"💬 {snap.get('client_name','العميل')}")
            except Exception:
                head.append("💬 العميل")
            head.append(f"🧾 الطلب: {order_id}")
            caption = "\n".join(head) + "\n" + (cap_raw if cap_raw else "📎 مرفق")

            await _send_media(kind, file_id, tid, caption=caption, kb=trader_reply_kb(order_id, user_id))
            try:
                await msg.reply_text("✅ تم إرسال رسالتك للتاجر.", reply_markup=chat_nav_kb_for(context, user_id, order_id, "pp_chat_trader_done"))
            except Exception as e:
                _swallow(e)
            return

        # ===== STAGE_TRADER_REPLY (تاجر → عميل) =====
        if stage == STAGE_TRADER_REPLY and kind and file_id:
            td_local = context.user_data.setdefault(user_id, {})
            to_uid = _safe_int(td_local.get("trader_reply_user_id"))
            order_id = _s(td_local.get("trader_reply_order_id"))
            if not to_uid or not order_id:
                ud[STAGE_KEY] = STAGE_NONE
                try:
                    await msg.reply_text("⚠️ تعذر تحديد العميل المرتبط بهذه المراسلة.")
                except Exception as e:
                    _swallow(e)
                return

            try:
                tprof = get_trader_profile(user_id) or {}
            except Exception:
                tprof = {}

            tname = _s(tprof.get("display_name")) or _s(name_now) or "التاجر"
            tcompany = _s(tprof.get("company_name"))
            tlabel = tname + (f" ({tcompany})" if tcompany else "")

            caption = (
                f"💬 {tlabel}\n"
                f"🧾 الطلب: {order_id}\n"
                + (cap_raw if cap_raw else "📎 مرفق")
            )

            await _send_media(kind, file_id, to_uid, caption=caption, kb=client_trader_chat_kb(order_id))
            try:
                await msg.reply_text("✅ تم إرسال ردّك للعميل.", reply_markup=trader_reply_done_kb())
            except Exception as e:
                _swallow(e)
            return

        # ===== STAGE_ADMIN_CHAT (إدارة → عميل/تاجر/متقدم) =====
        if stage == STAGE_ADMIN_CHAT and user_id in ADMIN_IDS and kind and file_id:
            order_id = _s(ud.get("admin_chat_order_id"))
            peer_id = _safe_int(ud.get("admin_chat_peer_id"))
            role = _s(ud.get("admin_chat_role"))  # client / trader / applicant
            if not peer_id:
                return

            header = "📩 رسالة من الإدارة"
            if role == "trader":
                header = "📩 رسالة من الإدارة → التاجر"
            elif role == "applicant":
                header = "📩 رسالة من الإدارة → متقدم"

            caption = (
                f"{header}\n"
                f"🧾 الطلب: {order_id or '—'}\n"
                "────────────────\n"
                + (cap_raw if cap_raw else "📎 مرفق")
            )

            kb = None
            try:
                if role == "client" and order_id:
                    kb = track_kb(order_id)
                elif role == "trader" and order_id:
                    kb = trader_chat_admin_kb(order_id, int(user_id))
                elif role == "applicant":
                    kb = InlineKeyboardMarkup([[InlineKeyboardButton("✍️ رد", callback_data=f"pp_join_chat|{peer_id}")]])
            except Exception:
                kb = None

            await _send_media(kind, file_id, peer_id, caption=caption, kb=kb)
            try:
                await msg.reply_text("✅ تم إرسال الرسالة.", reply_markup=chat_nav_kb(order_id, "pp_admin_chat_done"))
            except Exception as e:
                _swallow(e)
            return

        # ===== STAGE_TRADER_CHAT_ADMIN (تاجر → إدارة) =====
        if stage == STAGE_TRADER_CHAT_ADMIN and kind and file_id:
            order_id = _s(ud.get("trader_chat_order_id"))
            admin_id = _safe_int(ud.get("trader_chat_admin_id"))
            if not order_id or not admin_id:
                return

            caption = (
                "📩 رسالة من التاجر → الإدارة\n"
                f"🧾 رقم الطلب: {order_id}\n"
                f"🧑‍💼 التاجر: {name_now} ({user_id})\n"
                "────────────────\n"
                + (cap_raw if cap_raw else "📎 مرفق")
            )

            await _send_media(kind, file_id, admin_id, caption=caption, kb=admin_contact_kb(order_id))
            try:
                await msg.reply_text("✅ تم إرسال ردّك للإدارة.")
            except Exception as e:
                _swallow(e)
            return

        # ===== STAGE_ADMIN_REPLY (إدارة → عميل) =====
        if stage == STAGE_ADMIN_REPLY and user_id in ADMIN_IDS and kind and file_id:
            ad = context.user_data.setdefault(user_id, {})
            to_uid = _safe_int(ad.get("reply_user_id"))
            order_id = _s(ad.get("reply_order_id"))
            if not to_uid or not order_id:
                ud[STAGE_KEY] = STAGE_NONE
                return

            caption = (
                "📩 رسالة من الإدارة\n"
                f"🧾 الطلب: {order_id}\n"
                "────────────────\n"
                + (cap_raw if cap_raw else "📎 مرفق")
            )

            await _send_media(kind, file_id, to_uid, caption=caption, kb=track_kb(order_id))
            try:
                await msg.reply_text("✅ تم إرسال رسالتك للعميل.", reply_markup=admin_reply_done_kb())
            except Exception as e:
                _swallow(e)
            return

        # ===== STAGE_TRACK_ORDER (عميل → إدارة) =====
        if stage == STAGE_TRACK_ORDER and kind and file_id:
            order_id = _s(ud.get("track_order_id"))
            if not order_id:
                ud[STAGE_KEY] = STAGE_NONE
                return

            caption = (
                "📩 رسالة من العميل → الإدارة\n"
                f"🧾 الطلب: {order_id}\n"
                f"👤 العميل: {name_now} ({user_id})\n"
                "────────────────\n"
                + (cap_raw if cap_raw else "📎 مرفق")
            )

            for aid in ADMIN_IDS:
                await _send_media(kind, file_id, int(aid), caption=caption, kb=admin_reply_kb(order_id, user_id))

            try:
                await msg.reply_text("✅ تم استلام رسالتك وسيتم الرد عليك قريباً.")
            except Exception as e:
                _swallow(e)
            return

    # =========================================================
    # باقي منطقك كما هو (مراحل الإيصالات + فاتورة التاجر + صورة القطعة...)
    # =========================================================

    # === تحديث حالة التاجر: رفع فاتورة التاجر (PDF/صورة) ===
    if stage == STAGE_TRADER_STATUS_UPDATE and (ud.get("tsu_kind") or "").strip() == "seller_invoice":
        order_id2 = (ud.get("tsu_order_id") or "").strip()
        if not order_id2:
            set_stage(context, user_id, STAGE_NONE)
            return

        file_id = ""
        mime = ""
        is_photo = False

        if update.message and update.message.photo:
            file_id = update.message.photo[-1].file_id
            mime = "image/jpeg"
            is_photo = True
        else:
            doc = update.message.document if update.message else None
            if doc:
                mime = (doc.mime_type or "").lower()
                fname = (doc.file_name or "").lower()
                if mime.startswith("application/pdf") or mime.startswith("image/") or fname.endswith(".pdf"):
                    file_id = doc.file_id

        if not file_id:
            name = _user_name(update)
            await update.message.reply_text(f"{name}\nارسل فاتورة التاجر كملف PDF او صورة فقط")
            return

        update_order_fields(order_id2, {
            "seller_invoice_file_id": file_id,
            "seller_invoice_mime": mime,
            "seller_invoice_at": utc_now_iso(),
            "shop_invoice_file_id": file_id,
            "shop_invoice_mime": mime,
            "shop_invoice_at": utc_now_iso(),
            "order_status": "ready_to_ship",
        })

        try:
            await send_invoice_pdf(context, order_id2, kind="preliminary", admin_only=True)
        except Exception as e:
            _swallow(e)

        try:
            b2 = get_order_bundle(order_id2)
            o2 = b2.get("order", {}) or {}
            client_id2 = int(o2.get("user_id") or 0) if str(o2.get("user_id") or "").isdigit() else 0
            client_name2 = (_s(o2.get("user_name") or o2.get("client_name") or o2.get("customer_name")) or "العميل")
            car_name2 = (_s(o2.get("car_name") or o2.get("vehicle_name") or o2.get("car") or o2.get("car_model")) or "—")
            availability2 = _s(o2.get("availability_days") or o2.get("quote_availability") or o2.get("availability"))
            ship_eta2 = _s(o2.get("ship_eta") or o2.get("shipping_eta") or o2.get("ship_days"))

            try:
                tprof = get_trader_profile(user_id) or {}
            except Exception:
                tprof = {}

            tname = (_s(tprof.get("display_name")) or _s(_user_name(update)) or "التاجر")
            tcompany = _s(tprof.get("company_name"))
            tlabel = tname + (f" ({tcompany})" if tcompany else "")

            am = _calc_goods_ship_total(o2, ud_local=ud)

            try:
                if am.get("total_val", 0) > 0:
                    update_order_fields(order_id2, {"total_amount_sar": am.get("total_val")})
            except Exception as e:
                _swallow(e)

            money_lines = []
            if am.get("goods_txt"):
                money_lines.append(f"💰 قيمة القطع: {am['goods_txt']}")
            if am.get("ship_val", 0) > 0 and am.get("ship_txt"):
                money_lines.append(f"🚚 رسوم الشحن: {am['ship_txt']}")
            if am.get("total_txt"):
                money_lines.append(f"🧾 الإجمالي: {am['total_txt']}")
            money_block = ("\n".join(money_lines)).strip()
            if money_block:
                money_block = "\n" + money_block

            # 1) للعميل
            if client_id2:
                try:
                    cap_client = (
                        f"🧾 فاتورة المتجر الرسمية\n"
                        f"🧾 رقم الطلب: {order_id2}\n"
                        f"👤 العميل: {client_name2}\n"
                        f"🧑‍💼 التاجر: {tlabel}\n"
                        f"🚗 السيارة: {car_name2}\n"
                        f"📌 الحالة: {_pay_status_ar('ready_to_ship')}\n"
                        + (f"🛠 مدة التجهيز: {availability2}\n" if availability2 else "")
                        + (f"⏱ مدة الشحن: {ship_eta2}\n" if ship_eta2 else "")
                        + f"{money_block}\n"
                        "⬇️ الخطوة التالية: اختر طريقة الدفع ثم أرسل إيصال السداد لإكمال الشحن"
                    )
                    if is_photo:
                        await context.bot.send_photo(chat_id=client_id2, photo=file_id, caption=cap_client)
                    else:
                        await context.bot.send_document(chat_id=client_id2, document=file_id, caption=cap_client)
                except Exception as e:
                    _swallow(e)

            # 2) للإدارة
            for aid in ADMIN_IDS:
                try:
                    cap_admin = (
                        f"🧾 فاتورة تاجر (نسخة للمنصة)\n"
                        f"🧾 رقم الطلب: {order_id2}\n"
                        f"👤 العميل: {client_name2}\n"
                        f"🧑‍💼 التاجر: {tlabel}\n"
                        f"🚗 السيارة: {car_name2}\n"
                        f"📌 الحالة: {_pay_status_ar('ready_to_ship')}\n"
                        + (f"🛠 مدة التجهيز: {availability2}\n" if availability2 else "")
                        + (f"⏱ مدة الشحن: {ship_eta2}\n" if ship_eta2 else "")
                        + f"{money_block}"
                    )
                    if is_photo:
                        await context.bot.send_photo(chat_id=int(aid), photo=file_id, caption=cap_admin)
                    else:
                        await context.bot.send_document(chat_id=int(aid), document=file_id, caption=cap_admin)
                except Exception as e:
                    _swallow(e)

            # 3) نقل العميل لمرحلة اختيار/استلام إيصال قيمة القطع
            if client_id2:
                ud2 = get_ud(context, client_id2)
                ud2["goods_order_id"] = order_id2
                set_stage(context, client_id2, STAGE_AWAIT_GOODS_PAY_METHOD)
                update_order_fields(order_id2, {"goods_payment_status": "awaiting_method"})

                client_lines = [
                    "📌 <b>إشعار: تم إرسال فاتورة المتجر</b>",
                    "",
                    f"🧾 رقم الطلب: <b>{order_id2}</b>",
                    f"👤 العميل: <b>{html.escape(client_name2)}</b>",
                    f"🧑‍💼 التاجر: <b>{html.escape(tlabel)}</b>",
                    f"🚗 السيارة: <b>{html.escape(car_name2)}</b>",
                    f"📌 الحالة: <b>{_pay_status_ar('ready_to_ship')}</b>",
                ]
                if availability2:
                    client_lines.append(f"🛠 مدة التجهيز: <b>{html.escape(availability2)}</b>")
                if ship_eta2:
                    client_lines.append(f"⏱ مدة الشحن: <b>{html.escape(ship_eta2)}</b>")
                client_lines.append("")
                if am.get("goods_txt"):
                    client_lines.append(f"💰 مبلغ القطع: <b>{html.escape(am['goods_txt'])}</b>")
                if am.get("ship_val", 0) > 0 and am.get("ship_txt"):
                    client_lines.append(f"🚚 مبلغ الشحن: <b>{html.escape(am['ship_txt'])}</b>")
                elif _s(o2.get("ship_included")).lower() in ('yes', 'true', '1', 'included', 'مشمولة', 'مشمول'):
                    client_lines.append("🚚 مبلغ الشحن: <b>مشمول</b>")
                if am.get("total_txt"):
                    client_lines.append(f"🧾 المبلغ الإجمالي: <b>{html.escape(am['total_txt'])}</b>")
                client_lines.append("")
                client_lines.append("⬇️ <b>الخطوة القادمة</b>: اختر طريقة دفع قيمة القطع ثم أرسل إيصال السداد لإكمال الشحن")

                trader_pay_mode = "manual"
                try:
                    tid2 = int(o2.get("accepted_trader_id") or 0)
                    tp2 = get_trader_profile(tid2) if tid2 else {}
                    trader_pay_mode = (_s((tp2 or {}).get("payment_mode")).lower() or "manual")
                except Exception:
                    trader_pay_mode = "manual"

                if trader_pay_mode in ("link", "pay_link", "payment_link"):
                    link2 = _s(o2.get("goods_payment_link"))
                    if link2.startswith("http://") or link2.startswith("https://"):
                        try:
                            udc = get_ud(context, client_id2)
                            udc["goods_order_id"] = order_id2
                            set_stage(context, client_id2, STAGE_AWAIT_GOODS_RECEIPT)
                        except Exception as e:
                            _swallow(e)

                        await context.bot.send_message(
                            chat_id=client_id2,
                            text="\n".join(client_lines),
                            parse_mode="HTML",
                            reply_markup=InlineKeyboardMarkup([
                                [InlineKeyboardButton("💳 دفع الآن", url=link2)],
                                [InlineKeyboardButton("💬 مراسلة التاجر", callback_data=f"pp_chat_trader|{order_id2}")],
                            ]),
                            disable_web_page_preview=True,
                        )
                    else:
                        await context.bot.send_message(
                            chat_id=client_id2,
                            text="\n".join(client_lines[:-1] + [
                                "⏳ <b>بانتظار رابط الدفع من المتجر</b>\n"
                                "سيصلك رابط الدفع هنا داخل المنصة، وبعد الدفع أرسل الإيصال لإكمال الإجراء."
                            ]),
                            parse_mode="HTML",
                            reply_markup=InlineKeyboardMarkup([
                                [InlineKeyboardButton("💬 مراسلة التاجر", callback_data=f"pp_chat_trader|{order_id2}")],
                            ]),
                            disable_web_page_preview=True,
                        )
                else:
                    await context.bot.send_message(
                        chat_id=client_id2,
                        text="\n".join(client_lines),
                        parse_mode="HTML",
                        reply_markup=pay_goods_method_kb(order_id2, pay_mode="manual", has_link=bool(PP_PAY_LINK_URL)),
                        disable_web_page_preview=True,
                    )

        except Exception as e:
            _swallow(e)

        ud.pop("tsu_kind", None)
        ud.pop("tsu_order_id", None)
        set_stage(context, user_id, STAGE_NONE)

        name = _user_name(update)
        await update.message.reply_text(
            f"{name}\n✅ تم حفظ فاتورة المتجر وتحديث الحالة.\n⏳ بانتظار إكمال دفع قيمة القطع من العميل",
            disable_web_page_preview=True,
        )
        return

    # === مرحلة استلام ايصال قيمة القطع ===
    if stage == STAGE_AWAIT_GOODS_RECEIPT:
        if update.message and update.message.photo:
            return await goods_receipt_photo_handler(update, context)

        doc = update.message.document if update.message else None
        if doc:
            mime = (doc.mime_type or "").lower()
            fname = (doc.file_name or "").lower()
            is_pdf = mime.startswith("application/pdf") or fname.endswith(".pdf")
            is_img = mime.startswith("image/") or fname.endswith((".jpg", ".jpeg", ".png", ".webp"))
            if is_pdf or is_img:
                return await goods_receipt_document_handler(update, context)

        name = _user_name(update)
        await update.message.reply_text(f"{name}\nالايصال الزامي ارسل صورة او PDF فقط")
        return

    # === اشتراك التاجر: استلام إيصال رسوم الاشتراك ===
    if stage == STAGE_TRADER_SUB_AWAIT_RECEIPT:
        file_id = ""
        mime = ""
        is_photo = False

        if update.message and update.message.photo:
            file_id = update.message.photo[-1].file_id
            mime = "image/jpeg"
            is_photo = True
        else:
            doc = update.message.document if update.message else None
            if doc:
                mime = (doc.mime_type or "").lower()
                fname = (doc.file_name or "").lower()
                if mime.startswith("application/pdf") or mime.startswith("image/") or fname.endswith(".pdf"):
                    file_id = doc.file_id

        if not file_id:
            name = _user_name(update)
            await update.message.reply_text(f"{name}\nالايصال الزامي ارسل صورة او PDF فقط")
            return

        month = _s(ud.get("sub_month") or month_key_utc())
        amount = int(float(ud.get("sub_amount_sar") or 99))
        pm = _s(ud.get("sub_payment_method") or ud.get("payment_method") or "") or "—"

        try:
            upsert_trader_subscription(user_id, month, {
                "amount_sar": amount,
                "payment_method": pm,
                "payment_status": "pending",
                "receipt_file_id": file_id,
            })
        except Exception as e:
            _swallow(e)

        kb = InlineKeyboardMarkup([[
            InlineKeyboardButton("✅ تأكيد الاشتراك", callback_data=f"pp_admin_sub|confirm|{user_id}|{month}"),
            InlineKeyboardButton("❌ رفض", callback_data=f"pp_admin_sub|reject|{user_id}|{month}"),
        ]])

        cap = (
            "💳 <b>إيصال اشتراك تاجر</b>\n"
            f"👤 التاجر: <b>{html.escape(str(update.effective_user.full_name or update.effective_user.first_name or ''))}</b>\n"
            f"🆔 trader_id: <code>{user_id}</code>\n"
            f"📅 الشهر: <b>{html.escape(month)}</b>\n"
            f"💰 المبلغ: <b>{amount}</b> ريال\n"
            f"💳 الطريقة: <b>{html.escape(pm)}</b>\n\n"
            "⬇️ راجع الإيصال ثم أكد/ارفض:"
        )

        for aid in ADMIN_IDS:
            try:
                if is_photo:
                    await context.bot.send_photo(chat_id=int(aid), photo=file_id, caption=cap, parse_mode="HTML", reply_markup=kb)
                else:
                    await context.bot.send_document(chat_id=int(aid), document=file_id, caption=cap, parse_mode="HTML", reply_markup=kb)
            except Exception:
                try:
                    await context.bot.send_message(chat_id=int(aid), text=cap, parse_mode="HTML", reply_markup=kb)
                except Exception as e:
                    _swallow(e)

        set_stage(context, user_id, STAGE_NONE)
        await update.message.reply_text(
            f"{_user_name(update)}\n✅ تم استلام الإيصال وسيتم التحقق من الإدارة قريبًا",
            disable_web_page_preview=True,
        )
        return

    # === مرحلة استلام إيصال رسوم المنصة ===
    if stage == STAGE_AWAIT_RECEIPT:
        if update.message and update.message.photo:
            return await receipt_photo_handler(update, context)

        doc = update.message.document if update.message else None
        if doc:
            mime = (doc.mime_type or "").lower()
            fname = (doc.file_name or "").lower()
            is_pdf = mime.startswith("application/pdf") or fname.endswith(".pdf")
            is_img = mime.startswith("image/") or fname.endswith((".jpg", ".jpeg", ".png", ".webp"))
            if is_pdf or is_img:
                return await receipt_document_handler(update, context)

        name = _user_name(update)
        await update.message.reply_text(f"{name}\nالايصال الزامي ارسل صورة او PDF فقط")
        return

    # === مرحلة وسائط القطعة (اختيارية) ===
    if stage == STAGE_ASK_ITEM_PHOTO:
        items = ud.get("items", []) or []
        idx = ud.get("pending_item_idx", None)

        if idx is None or not isinstance(idx, int) or idx < 0 or idx >= len(items):
            set_stage(context, user_id, STAGE_CONFIRM_MORE)
            await update.message.reply_text(
                f"{_user_name(update)}\nلا يوجد قطعة مرتبطة بالصورة حاليا",
                reply_markup=more_kb()
            )
            return

        media_type = None
        file_id2 = None

        if update.message and update.message.photo:
            media_type = "photo"
            file_id2 = update.message.photo[-1].file_id
        elif update.message and update.message.document:
            media_type = "document"
            file_id2 = update.message.document.file_id
        elif update.message and update.message.video:
            media_type = "video"
            file_id2 = update.message.video.file_id
        elif update.message and update.message.video_note:
            media_type = "video_note"
            file_id2 = update.message.video_note.file_id
        elif update.message and update.message.voice:
            media_type = "voice"
            file_id2 = update.message.voice.file_id
        elif update.message and update.message.audio:
            media_type = "audio"
            file_id2 = update.message.audio.file_id

        if not file_id2:
            await update.message.reply_text(
                f"{_user_name(update)}\nارسل صورة الان (اختياري) او اكتب اسم القطعة التالية مباشرة",
                reply_markup=photo_prompt_kb(),
            )
            return

        it = items[idx]
        it["media_type"] = media_type
        it["file_id"] = file_id2
        it["photo_file_id"] = file_id2
        it.setdefault("created_at_utc", utc_now_iso())

        ud.pop("pending_item_idx", None)
        ud.pop("pending_item_name", None)

        set_stage(context, user_id, STAGE_CONFIRM_MORE)
        await update.message.reply_text(
            f"{_user_name(update)}\n"
            f"تم حفظ صورة القطعة رقم {idx + 1}\n"
            f"عدد القطع الحالي: {len(items)}\n\n"
            "يمكنك الان كتابة اسم قطعة جديدة مباشرة\n"
            "او اختيار انهاء وارسال الطلب",
            reply_markup=more_kb(),
        )
        return

    return
    
def _admin_to_trader_reply_kb(admin_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("💬 رد للإدارة", callback_data=f"pp_trader_reply_admin|{admin_id}")],
        [InlineKeyboardButton("🔒 إغلاق", callback_data="pp_ui_close")],
    ])


# ==============================
# Order Search + Legal Summary (Unified)
# ==============================

def _parse_order_search_input(s: str) -> dict:
    """
    يقبل:
      - pp0012 / PP0012 / pp10000 / pp250000
      - 0012 / 000012 / 10000 / 250000
      - 240217-0012 / 240217-10000
      - PP-240217-0012 / PP-240217-10000

    ويرجع:
      {"kind": "tail", "tail": "..."} أو {"kind":"full","order_id":"..."} أو {"kind":"none"}

    ملاحظة أمان:
      - رقم (4 أرقام) بدون pp لا يُعتبر بحث إلا إذا بدأ بصفر (مثل 0012) لتجنب التداخل مع إدخال الأسعار (مثل 1200).
      - 5 أرقام فأكثر تُقبل كبحث (لأن العداد قد يتجاوز 9999).
    """
    t = (s or "").strip()
    if not t:
        return {"kind": "none"}

    u = t.upper().replace("—", "-").replace("–", "-").strip()

    # pp + digits (4..9)
    m = re.match(r"^PP(\d{4,9})$", u, flags=re.I)
    if m:
        return {"kind": "tail", "tail": m.group(1)}

    # digits only:
    # - 4 digits => must start with 0 (0012) to avoid price confusion
    # - 5..9 digits => allowed (10000)
    m = re.match(r"^(\d{4,9})$", u)
    if m:
        d = m.group(1)
        if len(d) == 4 and not d.startswith("0"):
            return {"kind": "none"}
        return {"kind": "tail", "tail": d}

    # date-full: 240217-0012 / PP-240217-0012 (suffix 4..9)
    m = re.match(r"^(?:PP-)?(\d{6})-(\d{4,9})$", u, flags=re.I)
    if m:
        return {"kind": "full", "order_id": f"{m.group(1)}-{m.group(2)}"}

    return {"kind": "none"}


def _build_order_parts_details(order_id: str, items: list[dict], limit: int = 40) -> str:
    # يلخص القطع: مسعّر / غير مسعّر / غير متوفر (إن وجد)
    priced = []
    unpriced = []
    unavailable = []

    for it in (items or []):
        nm = (it.get("name") or it.get("item_name") or "").strip()
        if not nm:
            continue

        # مؤشرات محتملة
        st = str(it.get("status") or it.get("quote_status") or it.get("availability") or "").strip().lower()
        price = (it.get("price_sar") or it.get("unit_price_sar") or it.get("quote_price_sar") or it.get("price") or "")
        price = str(price or "").strip()

        if st in ("na", "not_available", "unavailable", "none", "0"):
            unavailable.append(nm)
        elif price:
            priced.append(f"{nm} — {price}")
        else:
            unpriced.append(nm)

        if (len(priced) + len(unpriced) + len(unavailable)) >= limit:
            break

    lines = []
    if priced:
        lines.append("تفاصيل القطع المسعّرة:")
        for x in priced[:20]:
            lines.append(f"• {x}")
    if unpriced:
        if lines:
            lines.append("")
        lines.append("تفاصيل القطع غير المسعّرة:")
        for x in unpriced[:20]:
            lines.append(f"• {x}")
    if unavailable:
        if lines:
            lines.append("")
        lines.append("تفاصيل القطع غير المتوفرة:")
        for x in unavailable[:20]:
            lines.append(f"• {x}")

    return "\n".join(lines).strip()

def _fmt_utc_ts_for_humans(s: str) -> str:
    """
    يحول ISO UTC timestamp إلى نص مفهوم (بتوقيت السعودية) مثل:
    2026-02-17 15:16
    يقبل:
    - 2026-02-17T12:16:15Z
    - 2026-02-17T12:16:15+00:00
    - 2026-02-17T12:16:15 (يُعتبر UTC)
    """
    v = ("" if s is None else str(s)).strip()
    if not v:
        return "—"

    try:
        # توحيد Z
        if v.endswith("Z"):
            v = v[:-1] + "+00:00"

        dt = datetime.fromisoformat(v)

        # إذا بدون tz نعتبره UTC
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        else:
            dt = dt.astimezone(timezone.utc)

        # تحويل لتوقيت السعودية
        try:
            from zoneinfo import ZoneInfo
            dt_ksa = dt.astimezone(ZoneInfo("Asia/Riyadh"))
        except Exception:
            dt_ksa = dt + timedelta(hours=3)

        return dt_ksa.strftime("%Y-%m-%d %H:%M")
    except Exception:
        # fallback: رجّع النص كما هو بدون كراش
        return ("" if s is None else str(s)).strip() or "—"


def _is_admin(uid: int) -> bool:
    try:
        s = str(int(uid or 0))
    except Exception:
        s = str(uid or "").strip()
    admin_set = set(str(x).strip() for x in (ADMIN_IDS or []))
    return bool(s and (s in admin_set))


def _viewer_role_for_order(viewer_id: int, order: dict) -> str:
    """
    يحدد دور المشاهد لهذا الطلب:
    - admin: إذا كان ضمن ADMIN_IDS
    - client: إذا كان viewer_id == user_id داخل الطلب
    - trader: إذا كان viewer_id هو التاجر المقبول (accepted_trader_id) أو التاجر الذي قدم عرض (quoted_trader_id)
    - otherwise: يرجع "" (غير مصرح)
    """
    vid = 0
    try:
        vid = int(viewer_id or 0)
    except Exception:
        vid = 0

    if not vid:
        return ""

    # ✅ الأدمن يشوف كل شيء حتى لو بيانات الطلب ناقصة
    if _is_admin(vid):
        return "admin"

    o = order or {}

    # client
    try:
        uid = int(o.get("user_id") or 0)
    except Exception:
        uid = 0
    if uid and uid == vid:
        return "client"

    # trader
    try:
        tid = int(o.get("accepted_trader_id") or 0)
    except Exception:
        tid = 0
    if tid and tid == vid:
        return "trader"

    try:
        qtid = int(o.get("quoted_trader_id") or 0)
    except Exception:
        qtid = 0
    if qtid and qtid == vid:
        return "trader"

    return ""
# ==================================================
# Helpers for order legal view
# ==================================================
async def _show_order_panel_private(context: ContextTypes.DEFAULT_TYPE, viewer_id: int, order_id: str):
    order_id = (order_id or "").strip()
    if not order_id:
        return

    uid = int(viewer_id or 0)
    if not uid:
        return

    # جلب الطلب
    try:
        b = get_order_bundle(order_id) or {}
        order = (b.get("order") or {}) if isinstance(b, dict) else {}
        items = (b.get("items") or []) if isinstance(b, dict) else []
    except Exception:
        order = {}
        items = []

    # تحديد الدور + الصلاحية
    try:
        role = _viewer_role_for_order(uid, order or {})
    except Exception:
        role = ""

    if not role:
        try:
            await context.bot.send_message(
                chat_id=uid,
                text=f"🔒 غير مصرح لك بعرض هذا الطلب.\n🧾 رقم الطلب: {order_id}",
                disable_web_page_preview=True,
            )
        except Exception:
            pass
        return

    # IDs
    def _to_int(v):
        try:
            return int(v)
        except Exception:
            try:
                return int(float(str(v).strip()))
            except Exception:
                return 0

    client_id = _to_int((order or {}).get("user_id"))
    trader_id = _to_int((order or {}).get("accepted_trader_id"))

    # ✅ حذف اللوحة السابقة إن وُجدت (لمنع التكدس فقط)
    td = getattr(context, "user_data", None) or {}
    last_mid = td.get("last_order_panel_mid")
    if last_mid:
        try:
            await context.bot.delete_message(chat_id=uid, message_id=int(last_mid))
        except Exception as e:
            _swallow(e)

    # ---------------- Helpers ----------------
    def _s(v):
        return ("" if v is None else str(v)).strip()

    def _load_item_prices_map() -> dict:
        pm = {}
        try:
            import ast
        except Exception:
            ast = None

        for k in ("quote_item_prices", "item_prices", "goods_item_prices", "quote_items_prices", "quote_item_price"):
            raw = (order or {}).get(k)
            if raw is None or raw == "":
                continue

            src = None
            try:
                if isinstance(raw, dict):
                    src = raw
                else:
                    sraw = str(raw).strip()
                    if sraw.startswith("{") and sraw.endswith("}"):
                        try:
                            src = json.loads(sraw)
                        except Exception:
                            src = None
                        if src is None and ast:
                            try:
                                src = ast.literal_eval(sraw)
                            except Exception:
                                src = None
            except Exception:
                src = None

            if isinstance(src, dict):
                for kk, vv in src.items():
                    ks = str(kk).strip()
                    vs = str(vv).strip()
                    if ks.isdigit() and vs not in ("", "0", "0.0", "0.00"):
                        pm[ks] = vs

        return pm

    def _pick_item_price(i: int, it: dict, pm: dict) -> str:
        cand_keys = (
            "price_sar",
            "unit_price_sar",
            "quote_price_sar",
            "item_price",
            "price",
            "unit_price",
            "amount_sar",
            "cost_sar",
            "cost",
            "sar",
        )
        for ck in cand_keys:
            try:
                vs = _s(it.get(ck))
            except Exception:
                vs = ""
            if vs and vs not in ("0", "0.0", "0.00"):
                return vs

        vs = _s(pm.get(str(i)))
        if vs and vs not in ("0", "0.0", "0.00"):
            return vs
        return ""

    def _is_unavailable(it: dict) -> bool:
        st = _s(it.get("status") or it.get("quote_status") or it.get("availability") or "").lower()
        return st in ("na", "not_available", "unavailable", "none", "0", "غير متوفر", "غير_متوفر")

    # ---------------- عرض بيانات ----------------
    client_name = (str(order.get("user_name") or order.get("client_name") or "—") or "—").strip() or "—"

    trader_name = "—"
    trader_store = ""
    if trader_id:
        trader_name = _trader_label(int(trader_id), "—")
        try:
            tp = get_trader_profile(int(trader_id)) or {}
        except Exception:
            tp = {}
        trader_store = (tp.get("company_name") or tp.get("shop_name") or "").strip()

    car_name = (str(order.get("car_name") or "—") or "—").strip() or "—"
    car_model = (str(order.get("car_model") or "—") or "—").strip() or "—"
    vin = (str(order.get("vin") or "—") or "—").strip() or "—"

    city = (str(order.get("shipping_city") or order.get("city") or "—") or "—").strip() or "—"
    delivery_type = (str(order.get("delivery_type") or order.get("shipping_method") or "—") or "—").strip() or "—"
    parts_type = (str(order.get("parts_type") or order.get("parts_condition") or "—") or "—").strip() or "—"

    # بيانات شحن إضافية (لو كانت موجودة بالطلب)
    shipping_phone = _s(order.get("user_phone") or order.get("client_phone") or order.get("phone") or order.get("mobile"))
    shipping_address = _s(order.get("shipping_address") or order.get("ship_address") or order.get("delivery_details") or order.get("address"))
    tracking_no = _s(order.get("shipping_tracking") or order.get("tracking_number") or order.get("tracking_no") or order.get("tracking"))

    ost = str(order.get("order_status") or "").strip().lower()
    status_ar = _order_status_display(order)

    # ✅ توضيح جهة الإلغاء داخل لوحة الطلب
    if ost == "cancelled":
        if str(order.get("cancelled_by_admin_id") or "").strip():
            status_ar = "ملغي من قبل الإدارة"
        elif str(order.get("cancelled_by_client_id") or "").strip():
            status_ar = "ملغي من قبل العميل"
        else:
            status_ar = "ملغي"

    ga = _money(order.get("goods_amount_sar") or order.get("quote_goods_amount") or "")
    sf = _money(order.get("shipping_fee_sar") or order.get("shipping_fee") or "")
    tot = _money(order.get("total_amount_sar") or order.get("total_amount") or order.get("quote_total_amount") or "")

    # ---------------- قطع + تسعير ----------------
    pm = _load_item_prices_map()
    priced_cnt = 0
    unpriced_cnt = 0
    unavailable_cnt = 0

    parts_lines = []
    for idx, it in enumerate(items or [], start=1):
        nm = (it.get("name") or it.get("item_name") or it.get("part_name") or "").strip()
        if not nm:
            continue

        partno = (it.get("part_no") or it.get("partno") or it.get("part_number") or "").strip()
        price = _pick_item_price(idx, it, pm)

        if _is_unavailable(it):
            unavailable_cnt += 1
            parts_lines.append(
                f"  {idx}) ❌ <b>{html.escape(nm)}</b>"
                + (f" — <code>{html.escape(partno)}</code>" if partno else "")
                + " — <i>غير متوفر</i>"
            )
        elif price:
            priced_cnt += 1
            parts_lines.append(
                f"  {idx}) ✅ <b>{html.escape(nm)}</b>"
                + (f" — <code>{html.escape(partno)}</code>" if partno else "")
                + f" — <b>{html.escape(str(price))}</b>"
            )
        else:
            unpriced_cnt += 1
            parts_lines.append(
                f"  {idx}) 🟡 <b>{html.escape(nm)}</b>"
                + (f" — <code>{html.escape(partno)}</code>" if partno else "")
                + " — <i>غير مسعر</i>"
            )

        if len(parts_lines) >= 60:
            break

    if priced_cnt > 0:
        pricing_status = "🟢 تم تسعير بعض/كل القطع"
    else:
        pricing_status = "🟢 تم التسعير" if (order.get("goods_amount_sar") or order.get("quote_goods_amount")) else "🟡 لم يتم التسعير"

    # ---------------- رسالة اللوحة (بدون خطوط طويلة) ----------------
    lines = []
    lines.append("📦 <b>لوحة الطلب</b>")
    lines.append("")
    lines.append(f"🧾 رقم الطلب: <b>{html.escape(order_id)}</b>")
    lines.append(f"📌 الحالة: <b>{html.escape(status_ar)}</b>")
    lines.append(f"💰 حالة التسعير: <b>{html.escape(pricing_status)}</b>")

    # الأطراف
    lines.append("")
    lines.append("👤 <b>الأطراف</b>")
    lines.append(f"• العميل: <b>{html.escape(client_name)}</b>")
    if shipping_phone:
        lines.append(f"• جوال العميل: <b>{html.escape(shipping_phone)}</b>")
    if client_id:
        lines.append(f"• معرف العميل: <code>{client_id}</code>")

    if trader_id:
        lines.append(f"• التاجر: <b>{html.escape(trader_name)}</b>")
        if trader_store:
            lines.append(f"• المتجر: <b>{html.escape(trader_store)}</b>")
    else:
        lines.append("• التاجر: <b>—</b>")

    # السيارة
    lines.append("")
    lines.append("🚗 <b>بيانات السيارة</b>")
    lines.append(f"• السيارة: <b>{html.escape(car_name)}</b>")
    lines.append(f"• سنة الموديل: <b>{html.escape(car_model)}</b>")
    lines.append(f"• رقم الهيكل VIN: <code>{html.escape(vin)}</code>")

    # الشحن
    lines.append("")
    lines.append("🚚 <b>الشحن</b>")
    lines.append(f"• المدينة: <b>{html.escape(city)}</b>")
    lines.append(f"• الطريقة: <b>{html.escape(delivery_type)}</b>")
    if shipping_address:
        lines.append(f"• العنوان: <b>{html.escape(shipping_address)}</b>")
    if tracking_no:
        lines.append(f"• رقم التتبع: <b>{html.escape(tracking_no)}</b>")

    # القطع
    lines.append("")
    lines.append("🧩 <b>القطع</b>")
    lines.append(f"• نوع القطع: <b>{html.escape(parts_type)}</b>")
    if parts_lines:
        lines.append(f"• الملخص: ✅ {priced_cnt} | 🟡 {unpriced_cnt} | ❌ {unavailable_cnt}")
        lines.append("")
        lines.extend(parts_lines)
    else:
        lines.append("• لا توجد قطع مسجلة لهذا الطلب حاليا")

    # المالي
    if ga or sf or tot:
        lines.append("")
        lines.append("💵 <b>الملخص المالي</b>")
        if ga:
            lines.append(f"• قيمة القطع: <b>{html.escape(ga)}</b>")
        if sf:
            lines.append(f"• رسوم الشحن: <b>{html.escape(sf)}</b>")
        if tot:
            lines.append(f"• الإجمالي: <b>{html.escape(tot)}</b>")

    msg = "\n".join([x for x in lines if str(x).strip()]).strip()

    # ---------------- كيبورد حسب الدور ----------------
    rows = []

    if role == "admin":
        rows.append([
            InlineKeyboardButton("💬 مراسلة العميل", callback_data=f"pp_admin_chat_client|{order_id}"),
            InlineKeyboardButton("🧑‍🔧 مراسلة التاجر", callback_data=f"pp_admin_chat_trader|{order_id}"),
        ])
        rows.append([InlineKeyboardButton("📜 سجل الطلب / آخر تحديث", callback_data=f"pp_order_legal|{order_id}")])
        rows.append([InlineKeyboardButton("🔒 إنهاء / إقفال الطلب (منجز)", callback_data=f"pp_order_finish|{order_id}")])
        rows.append([InlineKeyboardButton("✖️ إغلاق", callback_data="pp_ui_close")])

    elif role == "trader":
        rows.append([InlineKeyboardButton("💬 مراسلة العميل", callback_data=f"pp_chat_open|{order_id}")])
        rows.append([InlineKeyboardButton("🛎️ مراسلة المنصة", callback_data=f"pp_trader_chat_admin|{order_id}")])
        rows.append([InlineKeyboardButton("📜 سجل الطلب / آخر تحديث", callback_data=f"pp_order_legal|{order_id}")])
        try:
            tkb = trader_status_kb(order_id)
            for r in (tkb.inline_keyboard or []):
                rows.append(list(r))
        except Exception:
            pass
        rows.append([InlineKeyboardButton("✖️ إغلاق", callback_data="pp_ui_close")])

    else:  # client
        if trader_id:
            rows.append([InlineKeyboardButton("💬 مراسلة التاجر", callback_data=f"pp_chat_open|{order_id}")])

        # ✅ مراسلة المنصة مرتبطة بنفس الطلب مباشرة (بدون طلب رقم طلب لاحقًا)
        rows.append([InlineKeyboardButton("🛎️ مراسلة المنصة", callback_data=f"pp_track|{order_id}")])

        rows.append([InlineKeyboardButton("📜 سجل الطلب / آخر تحديث", callback_data=f"pp_order_legal|{order_id}")])
        rows.append([InlineKeyboardButton("✖️ إغلاق", callback_data="pp_ui_close")])

    kb = InlineKeyboardMarkup(rows)

    try:
        sent = await context.bot.send_message(
            chat_id=uid,
            text=msg,
            parse_mode="HTML",
            reply_markup=kb,
            disable_web_page_preview=True,
        )
        try:
            td["last_order_panel_mid"] = int(sent.message_id)
        except Exception:
            pass
    except Exception as e:
        _swallow(e)
        try:
            await context.bot.send_message(chat_id=uid, text="تعذر فتح لوحة الطلب حاليا")
        except Exception:
            pass


def _fmt_utc_ts_for_humans(ts: str) -> str:
    """Format ISO/UTC timestamp to Riyadh time in a compact human form."""
    s = (ts or "").strip()
    if not s:
        return ""

    from datetime import datetime, timezone
    try:
        from zoneinfo import ZoneInfo
        riyadh = ZoneInfo("Asia/Riyadh")
    except Exception:
        riyadh = None

    dt = None

    # epoch seconds?
    try:
        if re.fullmatch(r"\d{10}(?:\.\d+)?", s):
            dt = datetime.fromtimestamp(float(s), tz=timezone.utc)
    except Exception:
        dt = None

    if dt is None:
        try:
            ss = s.replace("Z", "+00:00")
            dt = datetime.fromisoformat(ss)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
        except Exception:
            dt = None

    if dt is None:
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                dt = datetime.strptime(s, fmt).replace(tzinfo=timezone.utc)
                break
            except Exception:
                dt = None

    if dt is None:
        return ""

    try:
        dt = dt.astimezone(riyadh) if riyadh else dt.astimezone(timezone.utc)
    except Exception:
        pass

    try:
        return dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return ""


def _fmt_utc_ts_for_humans(ts: str) -> str:
    """Format ISO/UTC timestamp to Riyadh time in a compact human form."""
    s = (ts or "").strip()
    if not s:
        return ""

    from datetime import datetime, timezone
    try:
        from zoneinfo import ZoneInfo
        riyadh = ZoneInfo("Asia/Riyadh")
    except Exception:
        riyadh = None

    dt = None

    # epoch seconds?
    try:
        if re.fullmatch(r"\d{10}(?:\.\d+)?", s):
            dt = datetime.fromtimestamp(float(s), tz=timezone.utc)
    except Exception:
        dt = None

    if dt is None:
        try:
            ss = s.replace("Z", "+00:00")
            dt = datetime.fromisoformat(ss)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
        except Exception:
            dt = None

    if dt is None:
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                dt = datetime.strptime(s, fmt).replace(tzinfo=timezone.utc)
                break
            except Exception:
                dt = None

    if dt is None:
        return ""

    try:
        dt = dt.astimezone(riyadh) if riyadh else dt.astimezone(timezone.utc)
    except Exception:
        pass

    try:
        return dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return ""


def build_order_legal_message(order_id: str, viewer_id: int) -> tuple[str, InlineKeyboardMarkup | None]:
    """رسالة  كاملة (ملخص حالة الطلب + خط زمني) + أزرار حسب الصلاحية."""
    try:
        b = get_order_bundle(order_id) or {}
        o = (b.get("order") or {}) if isinstance(b, dict) else {}
        items = (b.get("items") or []) if isinstance(b, dict) else []
    except Exception:
        o = {}
        items = []

    role = _viewer_role_for_order(viewer_id, o)
    if not role:
        return ("🔒 هذا الطلب ليس لديك.", InlineKeyboardMarkup([[InlineKeyboardButton("✖️ إغلاق", callback_data="pp_ui_close")]]))

    # أسماء
    client_name = (str(o.get("user_name") or o.get("client_name") or "—")).strip() or "—"
    car_name = (str(o.get("car_name") or o.get("vehicle_name") or o.get("car") or o.get("car_model") or "—")).strip() or "—"

    tid = 0
    try:
        tid = int(o.get("accepted_trader_id") or 0)
    except Exception:
        tid = 0

    trader_disp = _trader_label(tid, "—") if tid else "—"
    try:
        tp = get_trader_profile(tid) or {}
    except Exception:
        tp = {}
    trader_store = (tp.get("company_name") or tp.get("shop_name") or "").strip()
    if not trader_store:
        trader_store = (str(o.get("trader_store") or o.get("company_name") or "")).strip()

    # حالة الطلب
    ost = str(o.get("order_status") or "").strip().lower()
    display_status = _order_status_display(o)

    # ✅ توضيح جهة الإلغاء داخل سجل الطلب
    if ost == "cancelled":
        if str(o.get("cancelled_by_admin_id") or "").strip():
            display_status = "ملغي من قبل الإدارة"
        elif str(o.get("cancelled_by_client_id") or "").strip():
            display_status = "ملغي من قبل العميل"
        else:
            display_status = "ملغي"

    # مبالغ
    goods_line = ""
    ship_line = ""
    total_line = ""
    try:
        ga = _money(o.get("goods_amount_sar") or o.get("quote_goods_amount") or "")
        sf = _money(o.get("shipping_fee_sar") or o.get("shipping_fee") or "")
        # إجمالي (أولوية total_amount_sar)
        tot = _money(o.get("total_amount_sar") or o.get("total_amount") or "")
        if not tot:
            # نحاول حسابه إذا أمكن
            try:
                gnum = float(re.sub(r"[^0-9.]+", "", str(o.get("goods_amount_sar") or "")) or 0)
            except Exception:
                gnum = 0.0
            try:
                snum = float(re.sub(r"[^0-9.]+", "", str(o.get("shipping_fee_sar") or "")) or 0)
            except Exception:
                snum = 0.0
            if gnum or snum:
                tot = _money(gnum + snum)
        if ga:
            goods_line = f"قيمة القطع: <b>{html.escape(ga)}</b>"
        if sf:
            ship_line = f"قيمة الشحن: <b>{html.escape(sf)}</b>"
        if tot:
            total_line = f"الإجمالي: <b>{html.escape(tot)}</b>"
    except Exception:
        pass

    # المدد
    availability = (str(o.get("availability_days") or o.get("quote_availability") or o.get("availability") or "")).strip()
    ship_eta = (str(o.get("ship_eta") or o.get("shipping_eta") or o.get("ship_days") or "")).strip()

    # تفاصيل القطع
    parts_block = _build_order_parts_details(order_id, items)

    # منصة: مدفوع/مجاني
    fee_sar = str(o.get("price_sar") or "").strip()
    pm = str(o.get("payment_method") or "").strip().lower()
    ps = str(o.get("payment_status") or "").strip().lower()
    platform_fee_mode = "مجاني" if (fee_sar in ("", "0", "0.0") or pm == "free") else "مدفوع"
    platform_fee_status = _pay_status_ar(ps) if ps else "—"

    # خط زمني (best effort)
    timeline = []
    ts_created = _fmt_utc_ts_for_humans(str(o.get("created_at_utc") or o.get("created_at") or ""))
    if ts_created:
        timeline.append(f"• إنشاء الطلب: {ts_created} ({platform_fee_mode})")

    ts_pf = _fmt_utc_ts_for_humans(str(o.get("payment_confirmed_at_utc") or o.get("platform_payment_confirmed_at_utc") or ""))
    if ts_pf:
        timeline.append(f"• تأكيد رسوم المنصة: {ts_pf} ({platform_fee_status})")

    ts_quote = _fmt_utc_ts_for_humans(str(o.get("quote_sent_at_utc") or o.get("quote_sent_at") or ""))
    if ts_quote:
        timeline.append(f"• إرسال عرض السعر للعميل: {ts_quote}")

    ts_accept = _fmt_utc_ts_for_humans(str(o.get("accepted_at_utc") or o.get("quote_accepted_at_utc") or ""))
    if ts_accept:
        timeline.append(f"• قبول العرض: {ts_accept}")

    ts_inv = _fmt_utc_ts_for_humans(str(o.get("seller_invoice_at") or o.get("shop_invoice_at") or ""))
    if ts_inv:
        timeline.append(f"• رفع فاتورة المتجر: {ts_inv}")

    ts_goods = _fmt_utc_ts_for_humans(str(o.get("goods_payment_confirmed_at_utc") or o.get("goods_payment_at_utc") or ""))
    gps = str(o.get("goods_payment_status") or "").strip().lower()
    if ts_goods:
        timeline.append(f"• تأكيد سداد قيمة القطع: {ts_goods} ({_pay_status_ar(gps) if gps else '—'})")

    ts_ship = _fmt_utc_ts_for_humans(str(o.get("shipped_at_utc") or ""))
    if ts_ship:
        timeline.append(f"• تم شحن الطلب: {ts_ship}")

    ts_del = _fmt_utc_ts_for_humans(str(o.get("delivered_at_utc") or ""))
    if ts_del:
        timeline.append(f"• تم تسليم الطلب: {ts_del}")

    ts_close = _fmt_utc_ts_for_humans(str(o.get("closed_at_utc") or ""))
    if ts_close:
        timeline.append(f"• إغلاق الطلب: {ts_close}")

    tracking = (str(o.get("tracking_no") or o.get("tracking") or "")).strip()
    if tracking:
        timeline.append(f"• رقم التتبع: {html.escape(tracking)}")

    # موقع توقف الطلب (تشخيص بسيط)
    stop_hint = ""
    if ost in ("", "new"):
        stop_hint = "بانتظار تقديم العروض"
    elif ost in ("accepted", "quoted"):
        stop_hint = "بانتظار اختيار/قبول العرض"
    elif ost in ("preparing", "prep"):
        stop_hint = "بانتظار تجهيز الطلب"
    elif ost in ("ready_to_ship", "ready"):
        inv_file = (str(o.get("seller_invoice_file_id") or o.get("shop_invoice_file_id") or "")).strip()
        if not inv_file:
            stop_hint = "بانتظار رفع فاتورة المتجر"
        elif gps not in ("confirmed", "paid", "success", "successful", "done", "ok"):
            stop_hint = "بانتظار سداد العميل لقيمة القطع"
        else:
            stop_hint = "بانتظار شحن الطلب"
    elif ost == "shipped":
        stop_hint = "بانتظار تأكيد التسليم"
    elif ost in ("delivered", "closed"):
        stop_hint = "الطلب مكتمل"
    else:
        stop_hint = "قيد المتابعة"

    # بناء الرسالة (قانونية / مكتملة)
    lines = []
    # العنوان العام (موحد للجميع)
    lines.append("سجل الطلب ")
    lines.append("")
    lines.append(f"رقم الطلب: <b>{html.escape(order_id)}</b>")
    lines.append(f"العميل: <b>{html.escape(client_name)}</b>")
    lines.append(f"السيارة: <b>{html.escape(car_name)}</b>")
    lines.append(f"التاجر: <b>{html.escape(trader_disp)}</b>")
    if trader_store:
        lines.append(f"المتجر: <b>{html.escape(trader_store)}</b>")
    lines.append(f"الحالة الحالية: <b>{html.escape(display_status)}</b>")
    lines.append(f"موقع توقف الطلب: <b>{html.escape(stop_hint)}</b>")

    if availability:
        lines.append(f"مدة التجهيز: <b>{html.escape(availability)}</b>")
    if ship_eta:
        lines.append(f"مدة الشحن: <b>{html.escape(ship_eta)}</b>")

    if goods_line or ship_line or total_line:
        lines.append("")
        if goods_line: lines.append(goods_line)
        if ship_line: lines.append(ship_line)
        if total_line: lines.append(total_line)

    if parts_block:
        lines.append("")
        lines.append(parts_block)

    if timeline:
        lines.append("")
        lines.append("الخط الزمني:")
        lines.append("\n".join(timeline))

    msg = "\n".join(lines).strip()

    # أزرار
    rows = []
    # زر فتح لوحة الطلب (يبقى كما هو - قد يفشل للعميل حسب منطقك الحالي، لكنه مفيد للإدارة/التاجر)
    rows.append([InlineKeyboardButton("📦 فتح لوحة الطلب", callback_data=f"pp_open_order|{order_id}")])
    rows.append([InlineKeyboardButton("✖️ إغلاق", callback_data="pp_ui_close")])
    kb = InlineKeyboardMarkup(rows)

    return (msg, kb)


async def order_legal_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    data = (q.data or "").strip()
    try:
        _, oid = data.split("|", 1)
    except Exception:
        return
    oid = (oid or "").strip()
    if not oid:
        return

    msg, kb = build_order_legal_message(oid, int(q.from_user.id or 0))
    try:
        await q.message.reply_text(msg, parse_mode="HTML", reply_markup=kb, disable_web_page_preview=True)
    except Exception:
        try:
            await context.bot.send_message(chat_id=int(q.from_user.id), text=msg, parse_mode="HTML", reply_markup=kb, disable_web_page_preview=True)
        except Exception:
            pass


async def _resolve_and_show_order(context: ContextTypes.DEFAULT_TYPE, user_id: int, tail4: str):
    # ملاحظة: الاسم tail4 تاريخياً، لكنه الآن يمثل "الرقم التسلسلي العالمي" (قد يكون 4 أرقام أو أكثر)
    seq_in = str(tail4).strip()
    if not seq_in.isdigit():
        return

    uid = int(user_id or 0)
    if not uid:
        return

    try:
        seq_target = int(seq_in)
    except Exception:
        return

    def _extract_seq(oid: str) -> int | None:
        s = (oid or "").strip().upper()
        # يقبل:
        # PP-240217-0012  أو  240217-0012
        m = re.search(r"^(?:PP-)?\d{6}-(\d+)$", s)
        if not m:
            return None
        try:
            return int(m.group(1))
        except Exception:
            return None

    # ===== تحديد الأدمن بأمان =====
    try:
        _admins = set(int(x) for x in (ADMIN_IDS or []))
    except Exception:
        _admins = set()

    # نجمع الطلبات حسب الدور (تاجر/عميل/أدمن)
    orders = []
    try:
        # طلبات التاجر (إن وجدت)
        orders.extend(list_orders_for_trader(uid) or [])
    except Exception as e:
        _swallow(e)

    # طلبات العميل / الأدمن نأخذها من القائمة العامة ثم نفلتر
    try:
        all_orders = list_orders() or []
    except Exception:
        all_orders = []

    try:
        if uid in _admins:
            # الأدمن: يشوف كل الطلبات
            orders.extend(all_orders)
        else:
            # العميل: طلباته فقط (حسب user_id داخل الطلب)
            for o in all_orders:
                try:
                    if int(o.get("user_id") or 0) == uid:
                        orders.append(o)
                except Exception:
                    continue
    except Exception as e:
        _swallow(e)

    # إزالة التكرارات (حسب order_id)
    seen = set()
    uniq = []
    for o in orders:
        oid = str((o or {}).get("order_id") or "").strip()
        if not oid or oid in seen:
            continue
        seen.add(oid)
        uniq.append(o)

    # مطابقة "الرقم التسلسلي" بشكل صارم (لا يعتمد على endswith حتى لا يحدث تداخل بعد 9999)
    matches_info = []
    for o in uniq:
        oid = str(o.get("order_id") or "").strip()
        if not oid:
            continue

        seq = _extract_seq(oid)
        if seq is None:
            continue
        if int(seq) != int(seq_target):
            continue

        # معلومات مختصرة (بدون كشف سري خارج صلاحية المستخدم)
        try:
            cn = (str(o.get("user_name") or o.get("client_name") or "").strip() or "—")
        except Exception:
            cn = "—"
        try:
            st = _pay_status_ar(str(o.get("order_status") or "").strip() or "—")
        except Exception:
            st = "—"
        try:
            ct = _fmt_utc_ts_for_humans(str(o.get("created_at_utc") or o.get("created_at") or ""))
        except Exception:
            ct = ""

        matches_info.append({
            "order_id": oid,
            "client": cn,
            "status": st,
            "created": ct,
        })

    if not matches_info:
        # اسم الشخص الذي قام بالبحث (بدون أي معلومات عن صاحب الطلب الحقيقي)
        try:
            ch = await context.bot.get_chat(uid)
            intruder_name = (getattr(ch, "first_name", "") or getattr(ch, "full_name", "") or "").strip() or "عزيزي"
        except Exception:
            intruder_name = "عزيزي"

        # هل الرقم موجود بالنظام أصلاً؟ (بدون إظهار أي تفاصيل)
        exists_globally = False
        try:
            for o in (all_orders or []):
                oid = str((o or {}).get("order_id") or "").strip()
                if not oid:
                    continue
                seq = _extract_seq(oid)
                if seq is None:
                    continue
                if int(seq) == int(seq_target):
                    exists_globally = True
                    break
        except Exception as e:
            _swallow(e)

        if exists_globally:
            await context.bot.send_message(
                chat_id=uid,
                text=(
                    f"👋 {intruder_name}\n"
                    f"⚠️ الطلب رقم {seq_in} ليس ضمن سجل طلبات حسابك.\n"
                    "🔒 لا يمكن عرض طلبات الآخرين."
                ),
            )
            return

        await context.bot.send_message(
            chat_id=uid,
            text=f"🔍 لا يوجد طلب يحمل الرقم {seq_in}",
        )
        return

    # لو طلب واحد فقط → اعرض رسالة قانونية كاملة (ومنه زر فتح اللوحة)
    if len(matches_info) == 1:
        try:
            msg, kb = build_order_legal_message(matches_info[0]["order_id"], uid)
            await context.bot.send_message(
                chat_id=uid,
                text=msg,
                parse_mode="HTML",
                reply_markup=kb,
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)
        return

    # لو تعددت (نادر جداً: بيانات قديمة قبل توحيد التسلسل) نعرض قائمة اختيار
    lines = [
        f"⚠️ يوجد أكثر من طلب بنفس الرقم {seq_in} (بيانات قديمة قبل توحيد التسلسل).",
        "اختر الطلب الصحيح:",
        "",
    ]
    for i, it in enumerate(matches_info[:20], start=1):
        lines.append(
            f"{i}) {it['order_id']}  |  {it.get('status','—')}  |  {it.get('created','')}"
        )

    kb_rows = []
    for it in matches_info[:12]:
        # ✅ إصلاح مهم: كان ord_open (غير مستقبَل) → الآن pp_open_order (مستقبَل عندك)
        kb_rows.append([InlineKeyboardButton(f"فتح {it['order_id']}", callback_data=f"pp_open_order|{it['order_id']}")])
    kb = InlineKeyboardMarkup(kb_rows) if kb_rows else None

    try:
        await context.bot.send_message(
            chat_id=uid,
            text="\n".join(lines).strip(),
            reply_markup=kb,
            disable_web_page_preview=True,
        )
    except Exception as e:
        _swallow(e)
    return


async def open_order_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    # لا نزعج المستخدم بتنبيه — يكفي رد صامت
    try:
        await _alert(q, "")
    except Exception:
        try:
            await q.answer()
        except Exception:
            pass

    data = (q.data or "").strip()
    try:
        _, order_id = data.split("|", 1)
    except Exception:
        return

    order_id = (order_id or "").strip()
    if not order_id:
        return

    uid = int(getattr(q.from_user, "id", 0) or 0)
    if not uid:
        return

    # ✅ جلب الطلب لتحديد الصلاحية
    try:
        b = get_order_bundle(order_id) or {}
        order = (b.get("order") or {}) if isinstance(b, dict) else {}
    except Exception:
        order = {}

    try:
        client_id = int(order.get("user_id") or 0)
    except Exception:
        client_id = 0

    try:
        accepted_tid = int(order.get("accepted_trader_id") or 0)
    except Exception:
        accepted_tid = 0

    # ✅ الإدارة: تفتح أي طلب
    if uid in (ADMIN_IDS or []):
        try:
            await _show_order_panel_private(context, uid, order_id)
        except Exception as e:
            _swallow(e)
        return

    # ✅ التاجر المقبول: يفتح لوحة الطلب الخاصة بالتاجر
    if accepted_tid and uid == accepted_tid:
        try:
            await _show_order_panel_private(context, uid, order_id)
        except Exception as e:
            _swallow(e)
        return

    # ✅ العميل صاحب الطلب: يفتح لوحة الطلب العالمية الخاصة به
    if client_id and uid == client_id:
        try:
            await _show_order_panel_private(context, uid, order_id)
        except Exception as e:
            _swallow(e)
        return


    # ❌ غير مصرح
    try:
        await q.message.reply_text("🔒 هذا الطلب ليس ضمن طلباتك.")
    except Exception:
        pass


async def text_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    chat = update.effective_chat

    # ✅ Guard: لا نتعامل مع الرسائل النصية داخل المجموعات
    if not chat or chat.type != ChatType.PRIVATE:
        return

    # ✅ حماية: أحياناً يصل تحديث بدون message/text
    if not update.message or update.message.text is None:
        return

    raw_text = update.message.text
    text = _norm(raw_text)
    name = _user_name(update)

    # =========================
    # Helpers (تنسيق موحّد)
    # =========================
    def _clean(s: str) -> str:
        s = "" if s is None else str(s)
        s = s.replace("—", "-").replace("–", "-").replace("ـ", "-")
        return s.strip()

    async def _reply(msg: str, kb=None, parse_mode: str = None, web_preview: bool = False):
        try:
            await update.message.reply_text(
                msg,
                reply_markup=kb,
                parse_mode=parse_mode,
                disable_web_page_preview=(not web_preview),
            )
        except Exception as e:
            _swallow(e)

    async def _reply_html(title: str, lines: list[str], kb=None):
        body = "\n".join([l for l in (lines or []) if str(l).strip()])
        msg = f"{html.escape(name)}\n<b>{html.escape(title)}</b>"
        if body:
            msg += "\n" + body
        await _reply(msg, kb=kb, parse_mode="HTML", web_preview=False)

    def _safe_int(x) -> int:
        try:
            return int(x or 0)
        except Exception:
            return 0

    def _order_snapshot(order_id: str) -> dict:
        """
        Snapshot خفيف لعرض (العميل/المبلغ/الشحن/التاجر/الحالة).
        لا يرمي أخطاء.
        """
        out = {
            "client_name": "—",
            "client_id": "",
            "trader_id": 0,
            "trader_name": "—",
            "goods_amount": "",
            "shipping_fee": "",
            "total_amount": "",
            "status": "—",
            "ship_included": "",
            "ship_method": "",
            "ship_city": "",
        }
        oid = (order_id or "").strip()
        if not oid:
            return out

        try:
            b = get_order_bundle(oid) or {}
            o = (b.get("order") or {}) if isinstance(b, dict) else {}
        except Exception:
            o = {}

        try:
            out["client_name"] = (str(o.get("user_name") or "").strip() or "—")
        except Exception as e:
            _swallow(e)

        try:
            uid = _safe_int(o.get("user_id"))
            out["client_id"] = str(uid) if uid else ""
        except Exception as e:
            _swallow(e)

        try:
            out["trader_id"] = _safe_int(o.get("accepted_trader_id"))
        except Exception:
            out["trader_id"] = 0

        try:
            out["trader_name"] = _trader_label(out["trader_id"], "—") if out["trader_id"] else "—"
        except Exception:
            out["trader_name"] = "—"

        try:
            ga = str(o.get("goods_amount_sar") or o.get("quote_goods_amount") or "").strip()
            out["goods_amount"] = ga
        except Exception as e:
            _swallow(e)

        # الشحن (قد يكون موجود أو مشمول)
        try:
            sf = str(o.get("shipping_fee_sar") or o.get("shipping_fee") or "").strip()
            out["shipping_fee"] = sf
        except Exception as e:
            _swallow(e)

        try:
            out["ship_included"] = str(o.get("shipping_included") or "").strip()
        except Exception as e:
            _swallow(e)

        try:
            out["ship_method"] = str(o.get("ship_method") or "").strip()
            out["ship_city"] = str(o.get("ship_city") or "").strip()
        except Exception as e:
            _swallow(e)

        # الإجمالي
        try:
            ta = str(o.get("total_amount_sar") or "").strip()
            if not ta:
                ta = str(o.get("price_sar") or "").strip()
            out["total_amount"] = ta
        except Exception as e:
            _swallow(e)

        # الحالة (نص عربي إن أمكن)
        try:
            st = str(o.get("order_status") or "").strip()
            gps = str(o.get("goods_payment_status") or "").strip()
            out["status"] = _pay_status_ar(st or gps or "—")
        except Exception:
            out["status"] = "—"

        return out

    def _fmt_money(v: str) -> str:
        s = str(v or "").strip()
        if not s or s in ("0", "0.0"):
            return ""
        try:
            return _money(s)
        except Exception:
            return f"{s} ر.س"

    # =========================================================
    # 1) تمرير رسائل المراسلة الداخلية pp_chat_sessions (قبل أي STAGE)
    # ✅ محكم: لا يعمل إلا بعد زر فتح المراسلة + زر إنهاء + تايم آوت
    # =========================================================
    try:
        sessions = context.bot_data.get("pp_chat_sessions") or {}
        sess = sessions.get(str(user_id))
    except Exception:
        sessions = {}
        sess = None

    if isinstance(sess, dict):
        peer_id = _safe_int(sess.get("peer_id"))
        order_id_sess = (sess.get("order_id") or "").strip()
        role = (sess.get("role") or "").strip()  # client / trader

        # ⏱️ Timeout (افتراضي 30 دقيقة خمول / 6 ساعات كحد أقصى)
        try:
            idle_secs = int(os.getenv("PP_CHAT_IDLE_SECS", "1800") or 1800)
        except Exception:
            idle_secs = 1800
        try:
            max_secs = int(os.getenv("PP_CHAT_MAX_SECS", "21600") or 21600)
        except Exception:
            max_secs = 21600

        try:
            now_ts = int(time.time())
        except Exception:
            now_ts = 0

        started_at = _safe_int(sess.get("started_at")) or now_ts
        last_touch = _safe_int(sess.get("last_touch")) or started_at

        expired = False
        if now_ts and idle_secs and (now_ts - last_touch) > idle_secs:
            expired = True
        if now_ts and max_secs and (now_ts - started_at) > max_secs:
            expired = True

        if expired:
            # اغلاق الجلسة للطرفين بدون كسر التدفق
            try:
                sessions.pop(str(user_id), None)
                if peer_id:
                    sessions.pop(str(peer_id), None)
                context.bot_data["pp_chat_sessions"] = sessions
            except Exception:
                pass
        else:
            msg_body = _clean(text)
            if peer_id and order_id_sess and msg_body:
                # تحديث آخر تفاعل
                try:
                    sess["last_touch"] = now_ts
                    sessions[str(user_id)] = sess
                    context.bot_data["pp_chat_sessions"] = sessions
                except Exception:
                    pass

                try:
                    cn, tn = _order_parties(order_id_sess)
                    sender = f"👤 العميل: {cn}" if role == "client" else f"👤 التاجر: {tn}"
                    receiver = f"⬅️ إلى: {tn}" if role == "client" else f"⬅️ إلى: {cn}"
                    kb_end = InlineKeyboardMarkup([[InlineKeyboardButton("✖️ إنهاء المراسلة", callback_data=f"pp_chat_end|{order_id_sess}")]])
                    await context.bot.send_message(
                        chat_id=peer_id,
                        text=(
                            f"{sender}\n"
                            f"{receiver}\n"
                            f"{_order_tag_plain(order_id_sess)}\n"
                            f"💬 {msg_body}"
                        ),
                        reply_markup=kb_end,
                        disable_web_page_preview=True,
                    )
                except Exception as e:
                    _swallow(e)

                return
    # ✅ اجلب UD مرة واحدة فقط
    ud = get_ud(context, user_id)
    stage = ud.get(STAGE_KEY, STAGE_NONE)
    # =========================================================
    # ✅ حماية مراحل المراسلة (Stages) من التعليق:
    # - لا تعمل إلا داخل مرحلة مراسلة معروفة
    # - تنتهي تلقائيا بعد خمول (PP_CHAT_IDLE_SECS)
    # =========================================================
    try:
        idle_secs = int(os.getenv("PP_CHAT_IDLE_SECS", "1800") or 1800)
    except Exception:
        idle_secs = 1800
    try:
        now_ts = int(time.time())
    except Exception:
        now_ts = 0

    CHAT_STAGES = {
        STAGE_CHAT_TRADER,
        STAGE_TRADER_REPLY,
        STAGE_ADMIN_REPLY,
        STAGE_ADMIN_CHAT,
        STAGE_TRADER_CHAT_ADMIN,
        STAGE_SUPPORT_ADMIN_REPLY,
        STAGE_APPLICANT_CHAT_ADMIN,
    }

    if stage in CHAT_STAGES:
        try:
            prev_stage = (ud.get("chat_stage_name") or "").strip()
            if prev_stage != stage:
                ud["chat_stage_name"] = stage
                ud["chat_stage_started_at"] = now_ts
                ud["chat_stage_last_touch"] = now_ts
            last_touch = _safe_int(ud.get("chat_stage_last_touch")) or now_ts
            if now_ts and idle_secs and (now_ts - last_touch) > idle_secs:
                # انتهت تلقائيا بسبب الخمول: نفصل المرحلة حتى لا تتداخل مع أي إدخال عادي
                try:
                    ud.pop("chat_trader_order_id", None)
                    ud.pop("trader_reply_order_id", None)
                    ud.pop("admin_reply_order_id", None)
                    ud.pop("admin_chat_order_id", None)
                    ud.pop("admin_chat_peer_id", None)
                    ud.pop("admin_chat_role", None)
                    ud.pop("trader_chat_admin_order_id", None)
                    ud.pop("trader_chat_admin_peer_id", None)
                    ud.pop("support_admin_peer_id", None)
                    ud.pop("support_admin_role", None)
                    ud.pop("applicant_chat_peer_id", None)
                except Exception:
                    pass
                ud.pop("chat_stage_name", None)
                ud.pop("chat_stage_started_at", None)
                ud.pop("chat_stage_last_touch", None)
                set_stage(context, user_id, STAGE_NONE)
                stage = STAGE_NONE
        except Exception:
            pass
    if stage in CHAT_STAGES:
        try:
            ud["chat_stage_last_touch"] = now_ts
        except Exception:
            pass

    # =========================================================
    # ✅ Join Portal (بوابة التجار) — المرحلة 3/3 (الرقم الضريبي نص)
    # =========================================================
    if stage == STAGE_JOIN_VAT:
        vat_raw = (update.message.text or "").strip()
        vat = re.sub(r"[^0-9]+", "", vat_raw)

        # ✅ الرقم الضريبي السعودي = 15 رقم بالضبط
        if (not vat) or (len(vat) != 15):
            await update.message.reply_text(
                "⚠️ الرقم الضريبي غير صحيح.\n\n"
                "اكتب <b>15 رقم</b> بالضبط (أرقام فقط).\n"
                "مثال: <code>123456789012345</code>\n\n"
                "📌 اكتب الرقم الآن مرة أخرى:",
                parse_mode="HTML",
                disable_web_page_preview=True,
            )
            return

        jd = _join_ud(context, user_id)
        jd["vat_no"] = vat
        jd["join_pending"] = "1"
        set_stage(context, user_id, STAGE_NONE)

        # ✅ هوية المتقدم (اسم/يوزر) للإدارة
        who, username = _who_html(update.effective_user)

        async def _send_file_to_admins(kind: str, fid: str, cap: str):
            for aid in (ADMIN_IDS or []):
                try:
                    if kind == "photo":
                        await context.bot.send_photo(
                            chat_id=int(aid),
                            photo=fid,
                            caption=cap,
                            parse_mode="HTML",
                        )
                    else:
                        await context.bot.send_document(
                            chat_id=int(aid),
                            document=fid,
                            caption=cap,
                            parse_mode="HTML",
                        )
                except Exception:
                    pass

        # 1) إرسال المرفقات أولاً (بدون لوحة) حتى تكون واضحة
        cr_fid = str(jd.get("cr_file_id") or "").strip()
        cr_kind = str(jd.get("cr_kind") or "document").strip().lower()
        lic_fid = str(jd.get("license_file_id") or "").strip()
        lic_kind = str(jd.get("license_kind") or "document").strip().lower()

        if cr_fid:
            await _send_file_to_admins(
                cr_kind,
                cr_fid,
                f"{_join_portal_brand()}\n📎 <b>السجل التجاري</b>\n👤 <b>{who}</b>\n🆔 <code>{user_id}</code>",
            )
        if lic_fid:
            await _send_file_to_admins(
                lic_kind,
                lic_fid,
                f"{_join_portal_brand()}\n📎 <b>رخصة/إثبات المتجر</b>\n👤 <b>{who}</b>\n🆔 <code>{user_id}</code>",
            )

        # 2) ثم رسالة ملخص + 3 أزرار فقط (في آخر الرسالة)
        admin_text = (
            f"{_join_portal_brand()}\n\n"
            "📥 <b>طلب عضوية مكتمل — جاهز للمراجعة</b>\n\n"
            f"👤 الاسم: <b>{who}</b>\n"
            f"🆔 ID: <code>{user_id}</code>\n"
            + (f"🔗 المستخدم: @{html.escape(username, quote=False)}\n" if username else "")
            + f"🧾 الرقم الضريبي: <code>{vat}</code>\n\n"
            "✅ المرفقات أُرسلت أعلى هذه الرسالة."
        )
        await _notify_admins_private(context, admin_text, kb=_join_portal_admin_kb(user_id))

        # 3) تأكيد للمتقدم + زر مراسلة المنصة
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("💬 مراسلة المنصة", callback_data="pp_support_open")],
        ])
        await update.message.reply_text(
            f"{_join_portal_brand()}\n\n"
            "✅ تم إرسال مرفقاتك للإدارة للمراجعة.\n"
            "⏳ سيتم إشعارك بالقبول أو الرفض بعد المراجعة.\n\n"
            "إذا احتجت تواصل سريع اضغط الزر بالأسفل:",
            parse_mode="HTML",
            reply_markup=kb,
            disable_web_page_preview=True,
        )
        return
    # =========================================================
    # ✅ مراسلة المتقدم → الإدارة (داخل البوت)
    # =========================================================
    if stage == STAGE_APPLICANT_CHAT_ADMIN:
        admin_id = int(ud.get("applicant_chat_admin_id") or 0)
        txt = (update.message.text or "").strip()
        if not admin_id or not txt:
            return
        await context.bot.send_message(
            chat_id=admin_id,
            text=f"{_join_portal_brand()}\n\n💬 رسالة من متقدم\n🆔 ID: <code>{user_id}</code>\n\n{html.escape(txt, quote=False)}",
            parse_mode="HTML",
            reply_markup=_join_portal_admin_kb(user_id),
            disable_web_page_preview=True,
        )
        await update.message.reply_text("✅ تم إرسال رسالتك للإدارة.")
        return

    # ==================================================
    # 2) استدعاء لوحة طلب بكتابة رقم الطلب مباشرة
    # يقبل: pp0012 / 0012 / 240217-0012 / PP-240217-0012
    # ✅ للأدمن: يعمل حتى لو المرحلة ليست STAGE_NONE
    # (مع منع البحث داخل مراحل المراسلة حتى لا نكسر الشات)
    # ==================================================
    pr = _parse_order_search_input(raw_text or "")

    BLOCK_SEARCH_STAGES = {
        STAGE_ADMIN_CHAT,
        STAGE_SUPPORT_ADMIN_REPLY,
        STAGE_ADMIN_REPLY,
        STAGE_TRADER_REPLY,
        STAGE_CHAT_TRADER,
        STAGE_TRADER_CHAT_ADMIN,
        STAGE_APPLICANT_CHAT_ADMIN,
    }

    is_admin = _is_admin(user_id)
    allow_search = (stage == STAGE_NONE) or is_admin

    if allow_search and (stage not in BLOCK_SEARCH_STAGES or is_admin):
        if pr.get("kind") == "tail":
            await _resolve_and_show_order(context, user_id, pr.get("tail") or "")
            return

        if pr.get("kind") == "full":
            oid2 = (pr.get("order_id") or "").strip()
            if oid2:
                try:
                    msg, kb = build_order_legal_message(oid2, int(user_id or 0))
                    await context.bot.send_message(
                        chat_id=int(user_id),
                        text=msg,
                        parse_mode="HTML",
                        reply_markup=kb,
                        disable_web_page_preview=True,
                    )
                except Exception as e:
                    _swallow(e)
                return


    # ==================================================
    # 3) رد الإدارة داخل /منصة (بعد ضغط زر "رد")
    # ==================================================
    if user_id in ADMIN_IDS and stage == STAGE_SUPPORT_ADMIN_REPLY:
        target_uid = _safe_int(ud.get("support_reply_to_uid"))
        msg = _clean(raw_text)

        if not target_uid or not msg:
            await _reply_html("تنبيه", ["⚠️ اكتب ردّ واضح ثم أرسله."])
            return

        order_id_reply = str(ud.get("support_reply_order_id") or "").strip()

        try:
            await context.bot.send_message(
                chat_id=target_uid,
                text=(
                    f"{PP_SUPPORT_LABEL}\n"
                    f"🧾 رقم الطلب: {order_id_reply or '—'}\n"
                    "────────────────\n"
                    f"{msg}"
                ),
                reply_markup=_support_kb(),
                disable_web_page_preview=True,
            )
            await _reply_html(
                "تم الإرسال",
                [
                    "✅ تم إرسال الرد للمستخدم.",
                    f"🧾 رقم الطلب: {html.escape(order_id_reply or '—')}",
                ],
            )
        except Exception:
            await _reply_html("تعذر الإرسال", ["⚠️ قد لا يكون المستخدم بدأ البوت أو قام بحظر البوت."])
            return

        ud[STAGE_KEY] = STAGE_NONE
        ud.pop("support_reply_to_uid", None)
        ud.pop("support_reply_order_id", None)
        return

    # ==================================================
    # 4) مراسلة الأدمن للتاجر من ملف التاجر
    # ==================================================
    if user_id in ADMIN_IDS and stage == STAGE_ADMIN_TRADER_MSG:
        tid = _safe_int(ud.get("admin_msg_to_trader_id"))
        msg = _clean(raw_text)

        if not tid or not msg:
            await _reply_html("تنبيه", ["⚠️ اكتب رسالة صحيحة."])
            return

        try:
            await context.bot.send_message(
                chat_id=tid,
                text=f"{PP_SUPPORT_LABEL}\n💬 رسالة من الإدارة:\n{msg}",
                reply_markup=_admin_to_trader_reply_kb(user_id),
                disable_web_page_preview=True,
            )
            await _reply_html("تم الإرسال", ["✅ تم إرسال الرسالة للتاجر."])
        except Exception:
            await _reply_html("تعذر الإرسال", ["⚠️ قد لا يكون التاجر بدأ البوت أو قام بحظر البوت."])
            return

        ud[STAGE_KEY] = STAGE_NONE
        ud.pop("admin_msg_to_trader_id", None)
        return

    # ==================================================
    # 5) قناة /منصة للمستخدم (توجيه الرسائل للإدارة فقط)
    # ==================================================
    if _support_is_open(ud):
        if _support_should_close_by_time(ud):
            await _support_close(update, context, user_id, reason="ℹ️ تم الإغلاق تلقائياً لانتهاء مدة الجلسة.")
        else:
            cur_stage = ud.get(STAGE_KEY, STAGE_NONE)
            if cur_stage and cur_stage != STAGE_NONE:
                await _support_close(update, context, user_id, reason="ℹ️ تم الإغلاق تلقائياً لأنك بدأت عملية أخرى.")
            else:
                msg = _clean(raw_text)
                if msg:
                    _support_touch(ud)

                    msg_norm = _clean(msg)
                    order_id_support = str(
                        ud.get("support_order_id")
                        or ud.get("last_order_id")
                        or ud.get("order_id")
                        or ""
                    ).strip()

                    try:
                        m = re.search(r"(PP-\d{6}-\d{4,})", msg_norm, flags=re.IGNORECASE)
                        if m:
                            order_id_support = (m.group(1) or "").strip()
                    except Exception as e:
                        _swallow(e)

                    is_trader_sender = False
                    try:
                        tp0 = get_trader_profile(int(user_id or 0)) or {}
                    except Exception:
                        tp0 = {}

                    try:
                        if any((tp0.get(k) or "").strip() for k in ("display_name", "company_name", "bank_name", "iban", "stc_pay")):
                            is_trader_sender = True
                    except Exception as e:
                        _swallow(e)

                    if not order_id_support:
                        try:
                            m4 = re.search(r"\b(\d{4})\b", msg_norm)
                            tail4 = (m4.group(1) or "").strip() if m4 else ""
                        except Exception:
                            tail4 = ""

                        if tail4:
                            orders = []
                            uid = _safe_int(user_id)

                            try:
                                orders.extend(list_orders_for_trader(uid) or [])
                            except Exception as e:
                                _swallow(e)

                            try:
                                all_orders = list_orders() or []
                            except Exception:
                                all_orders = []

                            try:
                                if _is_admin(uid):
                                    orders.extend(all_orders)
                                else:
                                    for o in all_orders:
                                        try:
                                            if _safe_int(o.get("user_id")) == uid:
                                                orders.append(o)
                                        except Exception:
                                            continue
                            except Exception as e:
                                _swallow(e)

                            seen = set()
                            uniq = []
                            for o in orders:
                                oid = str((o or {}).get("order_id") or "").strip()
                                if not oid or oid in seen:
                                    continue
                                seen.add(oid)
                                uniq.append(o)

                            matches = []
                            for o in uniq:
                                oid = str(o.get("order_id") or "").strip()
                                if oid.endswith(tail4):
                                    matches.append(oid)

                            if len(matches) == 1:
                                order_id_support = matches[0]
                            elif len(matches) > 1:
                                kb = InlineKeyboardMarkup([
                                    [InlineKeyboardButton(f"🧾 {oid}", callback_data=f"pp_open_order|{oid}")]
                                    for oid in matches[:10]
                                ])
                                await _reply_html("اختر الطلب", ["🔎 وُجد أكثر من طلب بنفس الرقم، اختر الطلب من القائمة:"], kb=kb)
                                return

                    if not order_id_support and not is_trader_sender:
                        await _reply_html(
                            "مطلوب رقم الطلب",
                            [
                                "اكتب رقم الطلب كامل مثل:",
                                "<code>PP-260208-0003</code>",
                                "أو اكتب آخر 4 أرقام فقط مثل: <code>0003</code>",
                                "",
                                "ثم اكتب رسالتك للإدارة.",
                            ],
                            kb=_support_kb(),
                        )
                        return

                    try:
                        if order_id_support:
                            ud["support_order_id"] = order_id_support
                    except Exception as e:
                        _swallow(e)

                    reply_cb = (
                        f"pp_support_reply|{order_id_support}|{user_id}"
                        if order_id_support
                        else f"pp_support_reply|{user_id}"
                    )

                    header = "📩 رسالة عبر /منصة"
                    if is_trader_sender and not order_id_support:
                        header = "📩 رسالة تاجر عبر /منصة"

                    for aid in (ADMIN_IDS or []):
                        try:
                            await context.bot.send_message(
                                chat_id=int(aid),
                                text=(
                                    f"{header}\n"
                                    + (f"🧾 رقم الطلب: {order_id_support}\n" if order_id_support else "")
                                    + f"👤 {name}\n"
                                    f"🆔 {user_id}\n"
                                    "────────────────\n"
                                    f"{msg}"
                                ),
                                reply_markup=InlineKeyboardMarkup([[
                                    InlineKeyboardButton("✉️ رد", callback_data=reply_cb)
                                ]]),
                                disable_web_page_preview=True,
                            )
                        except Exception as e:
                            _swallow(e)

                    if order_id_support:
                        await _reply_html(
                            "تم الإرسال",
                            [
                                "✅ تم إرسال رسالتك للإدارة.",
                                f"🧾 رقم الطلب: {html.escape(order_id_support)}",
                            ],
                            kb=_support_kb(),
                        )
                    else:
                        await _reply_html("تم الإرسال", ["✅ تم إرسال رسالتك للإدارة."], kb=_support_kb())

                return

    # تشغيل بكلمة pp بدون سلاش (في الخاص فقط)
    if (text or "").lower() == "pp":
        await begin_flow(update, context)
        return
    # ==================================================
    # 6) إدخال رابط الدفع (يدوي) من الإدارة
    # ==================================================
    if user_id in ADMIN_IDS and stage == STAGE_ADMIN_SEND_PAYLINK:
        link = (raw_text or "").strip()
        if not (link.startswith("http://") or link.startswith("https://")):
            await _reply_html("رابط غير صحيح", ["⚠️ ارسل رابط يبدأ بـ <code>https://</code>"])
            return

        order_id = (ud.get("paylink_order_id") or "").strip()
        client_id = _safe_int(ud.get("paylink_client_id"))

        if not order_id or not client_id:
            await _reply_html("تعذر الإرسال", ["⚠️ تعذر تحديد الطلب/العميل، أعد المحاولة."])
            set_stage(context, user_id, STAGE_NONE)
            ud.pop("paylink_order_id", None)
            ud.pop("paylink_client_id", None)
            return

        try:
            update_order_fields(order_id, {
                "payment_method": "pay_link",
                "payment_status": "awaiting_receipt",
                "payment_link": link,
                "payment_link_sent_at_utc": utc_now_iso(),
            })
        except Exception as e:
            _swallow(e)

        try:
            b = get_order_bundle(order_id)
            order = b.get("order", {}) or {}
            fee = order.get("price_sar") or ud.get("price_sar") or ""
            fee_txt = f"{fee} ريال" if str(fee).strip() not in ("", "0", "0.0") else "—"

            await context.bot.send_message(
                chat_id=client_id,
                text=(
                    f"{_user_name(update)}\n"
                    "🔗 رابط دفع رسوم المنصة\n"
                    f"🧾 رقم الطلب: {order_id}\n"
                    f"💰 الرسوم: {fee_txt}\n\n"
                    "افتح الرابط وأكمل الدفع.\n"
                    "بعد الدفع أرسل الإيصال هنا داخل المنصة لإكمال الإجراء."
                ),
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔗 فتح رابط الدفع", url=link)]]),
                disable_web_page_preview=True,
            )
        except Exception:
            await _reply_html("تعذر الإرسال", ["⚠️ تعذر إرسال الرابط للعميل (قد لا يكون بدأ البوت)."])
            return

        await _reply_html("تم الإرسال", [f"✅ تم إرسال رابط الدفع للعميل.", f"🧾 رقم الطلب: {html.escape(order_id)}"])
        set_stage(context, user_id, STAGE_NONE)
        ud.pop("paylink_order_id", None)
        ud.pop("paylink_client_id", None)
        return
    # ==================================================
    # 7) لوحة التاجر (تاجر) - تعمل بالخاص فقط
    # ==================================================
    if (text or "").strip() == "تاجر":
        stage_now = ud.get(STAGE_KEY, STAGE_NONE)

        is_disabled = False
        try:
            is_disabled = _trader_is_disabled(int(user_id or 0))
        except Exception:
            is_disabled = False

        if stage_now != STAGE_NONE and not is_disabled:
            set_stage(context, user_id, STAGE_NONE)

        is_admin = user_id in ADMIN_IDS

        is_member = False
        try:
            is_member = await _is_trader_group_member(context, user_id)
        except Exception:
            is_member = False

        if not is_member and not is_admin:
            await _reply_html("غير مصرح", ["⛔ هذه الخدمة مخصصة للتجار المسجلين فقط."])
            return

        tp = {}
        try:
            tp = get_trader_profile(int(user_id or 0)) or {}
        except Exception:
            tp = {}

        is_registered_trader = False
        if not is_member and not is_admin and not tp:
            try:
                uid_s = str(int(user_id or 0))
                for t in (list_traders() or []):
                    if str(t.get("trader_id") or "").strip() == uid_s:
                        is_registered_trader = True
                        break
            except Exception:
                is_registered_trader = False

        if not is_member and not is_admin and not tp and not is_registered_trader:
            if not TRADERS_GROUP_ID:
                await _reply_html("تنبيه إعدادات", ["⚠️ متغير <code>PP_TRADERS_GROUP_ID</code> غير موجود في .env"])
            else:
                await _reply_html("غير مصرح", ["⛔ تأكد أنك منضم لمجموعة التجار، وأن البوت مشرف داخلها."])
            return

        set_stage(context, user_id, STAGE_NONE)
        try:
            await show_trader_panel(update, context, user_id)
        except Exception:
            await _reply_html("تعذر الفتح", ["⚠️ تعذر فتح لوحة التاجر حالياً. حاول لاحقاً."])
        return
    # ==================================================
    # 8) لوحة الادارة (pp25s) - ادمن فقط بالخاص
    # ==================================================
    if (text or "").strip().lower() == "pp25s":
        if user_id not in ADMIN_IDS:
            await _reply_html("غير مصرح", ["⛔ هذه الخدمة خاصة بالإدارة."])
            return
        set_stage(context, user_id, STAGE_NONE)
        try:
            await show_admin_panel(update, context, user_id)
        except Exception:
            await _reply_html("تعذر الفتح", ["⚠️ تعذر فتح لوحة الإدارة حالياً."])
        return
    # ==================================================
    # 9) بحث طلب من لوحة الإدارة
    # ==================================================
    if user_id in ADMIN_IDS and stage == STAGE_ADMIN_FIND_ORDER:
        raw_in = (text or "").strip()
        pr = _parse_order_search_input(raw_in)

        set_stage(context, user_id, STAGE_NONE)

        # ✅ pp0012 / 0012
        if pr.get("kind") == "tail":
            await _resolve_and_show_order(context, user_id, pr.get("tail") or "")
            return

        # ✅ 240217-0012 / PP-240217-0012
        if pr.get("kind") == "full":
            oid = (pr.get("order_id") or "").strip()
            if not oid:
                await _reply_html("غير موجود", ["⚠️ لم يتم العثور على الطلب."])
                return
            msg, kb = build_order_legal_message(oid, int(user_id))
            await _reply(msg, kb=kb, parse_mode="HTML")
            return

        # ✅ fallback: محاولة بالمعرف كما هو
        oid = raw_in
        try:
            ob = get_order_bundle(oid)
        except Exception:
            ob = None

        if not ob:
            await _reply_html("غير موجود", ["⚠️ لم يتم العثور على الطلب."])
            return

        try:
            msg, kb = build_order_legal_message(oid, int(user_id))
            await _reply(msg, kb=kb, parse_mode="HTML")
        except Exception:
            # fallback بسيط
            o = ob.get("order", {})
            msg2 = (
                f"📦 <b>الطلب</b> {html.escape(oid)}\n"
                f"👤 <b>العميل</b>: {html.escape(str(o.get('user_name','—')))}\n"
                f"🧑‍💼 <b>التاجر</b>: {html.escape(_trader_label(int(o.get('accepted_trader_id') or 0),'—'))}\n"
                f"💰 <b>قيمة القطع</b>: {html.escape(_money(o.get('goods_amount_sar')) or '—')}\n"
                f"📌 <b>الحالة</b>: {html.escape(_pay_status_ar(o.get('order_status','—')))}"
            )
            await _reply(msg2, kb=InlineKeyboardMarkup([[InlineKeyboardButton("↩️ رجوع", callback_data="pp_admin|home")]]), parse_mode="HTML")
        return

# ==================================================
    # 10) إدخال بيانات لوحة التاجر
    # ==================================================
    if stage == STAGE_TRADER_PROFILE_EDIT:
        field = (ud.get("tprof_field") or "").strip()
        val = (text or "").strip()

        if field not in ("display_name", "company_name", "shop_phone", "cr_no", "vat_no", "bank_name", "iban", "stc_pay"):
            set_stage(context, user_id, STAGE_NONE)
            await _reply_html("تعذر التعديل", ["⚠️ تعذر تحديد الحقل المراد تعديله."])
            return

        if field in ("display_name", "company_name", "bank_name") and len(val) < 2:
            await _reply_html("قيمة غير واضحة", ["⚠️ اكتب النص بشكل أوضح."])
            return


        if field == "shop_phone":
            v = re.sub(r"\s+", "", val)
            # allow KSA-style mobile 10 digits starting with 05
            if not re.fullmatch(r"05\d{8}", v):
                await _reply_html(
                    "رقم اتصال المتجر غير صحيح",
                    [
                        "⚠️ الصيغة المطلوبة: 10 أرقام ويبدأ بـ 05 (بدون مسافات).",
                        "مثال: <code>05XXXXXXXX</code>",
                    ],
                )
                return
            val = v

        if field == "cr_no":
            v = re.sub(r"\s+", "", val)
            # commercial register: digits only (flexible length to avoid false rejects)
            if not re.fullmatch(r"\d{6,15}", v):
                await _reply_html(
                    "رقم السجل التجاري غير صحيح",
                    [
                        "⚠️ الصيغة المطلوبة: أرقام فقط (من 6 إلى 15 رقم).",
                        "مثال: <code>1010XXXXXX</code>",
                    ],
                )
                return
            val = v

        if field == "vat_no":
            v = re.sub(r"\s+", "", val)
            if not re.fullmatch(r"\d{15}", v):
                await _reply_html(
                    "الرقم الضريبي غير صحيح",
                    [
                        "⚠️ الصيغة المطلوبة: 15 رقم (أرقام فقط).",
                        "مثال: <code>300XXXXXXXXXXXX</code>",
                    ],
                )
                return
            val = v

        if field == "iban":
            v = re.sub(r"\s+", "", val).upper()
            if not re.fullmatch(r"SA\d{22}", v):
                await _reply_html(
                    "IBAN غير صحيح",
                    [
                        "⚠️ الصيغة المطلوبة: يبدأ بـ SA ثم 22 رقم (بدون مسافات).",
                        "مثال: <code>SA1234567890123456789012</code>",
                    ],
                )
                return
            val = v

        if field == "stc_pay":
            v = re.sub(r"\s+", "", val)
            if not re.fullmatch(r"05\d{8}", v):
                await _reply_html(
                    "STC Pay غير صحيح",
                    [
                        "⚠️ الصيغة المطلوبة: 10 أرقام ويبدأ بـ 05.",
                        "مثال: <code>05XXXXXXXX</code>",
                    ],
                )
                return
            val = v

        try:
            upsert_trader_profile(int(user_id), {field: val})
        except Exception:
            await _reply_html("تعذر الحفظ", ["⚠️ تعذر حفظ البيانات حالياً. حاول لاحقاً."])
            return

        # ✅ جدولة نسخة احتياطية بعد حفظ ملف التاجر
        try:
            app = getattr(context, "application", None)
            if app:
                if not app.bot_data.get("_backup_touch_trader_profile"):
                    app.bot_data["_backup_touch_trader_profile"] = True

                    async def _bk_job():
                        try:
                            await asyncio.sleep(5)
                            await _send_backup_excel(app, reason="trader_profile_edit")
                        finally:
                            try:
                                app.bot_data["_backup_touch_trader_profile"] = False
                            except Exception as e:
                                _swallow(e)

                    asyncio.create_task(_bk_job())
        except Exception as e:
            _swallow(e)

        ud.pop("tprof_field", None)
        set_stage(context, user_id, STAGE_NONE)
        await _reply_html("تم الحفظ", ["✅ تم حفظ بياناتك بنجاح."])
        await show_trader_panel(update, context, user_id)
        return
    # ==================================================
    # 11) تحديث حالة التاجر (مدخلات إلزامية)
    # ==================================================
    if stage == STAGE_TRADER_STATUS_UPDATE:
        kind = (ud.get("tsu_kind") or "").strip()
        order_id2 = (ud.get("tsu_order_id") or "").strip()
        if not order_id2:
            set_stage(context, user_id, STAGE_NONE)
            return

        try:
            b2 = get_order_bundle(order_id2)
            o2 = b2.get("order", {}) or {}
        except Exception:
            o2 = {}

        client_id2 = _safe_int(o2.get("user_id"))

        tprof = get_trader_profile(user_id) or {}
        tname = (tprof.get("display_name") or "").strip() or (name or "").strip() or "التاجر"

        goods_amt = str(o2.get("goods_amount_sar") or o2.get("quote_goods_amount") or "").strip()
        ship_method = str(o2.get("ship_method") or "").strip()
        ship_city = str(o2.get("ship_city") or "").strip()

        # ✅ منع تكرار إشعار (رسالة التاجر مع 3 أزرار) إذا الطلب كان مشحون مسبقًا
        try:
            ost_before = str(o2.get("order_status") or "").strip().lower()
        except Exception:
            ost_before = ""
        already_shipped = ost_before in ("shipped", "delivered", "closed")

        # ===== Helper: إرسال "رسالة التاجر مع 3 أزرار" بشكل مفصل =====
        def _num_float(x) -> float:
            try:
                s = str(x or "").strip().replace(",", "")
                return float(s) if s not in ("", "—", "-") else 0.0
            except Exception:
                return 0.0

        def _is_yes(x) -> bool:
            try:
                v = str(x or "").strip().lower()
                return v in ("yes", "y", "true", "1", "مشمول", "included")
            except Exception:
                return False

        async def _send_trader_notice(tracking_value: str):
            # ✅ لا ترسل إشعار التاجر إذا already_shipped=True
            if already_shipped:
                return

            try:
                client_name2 = (str(o2.get("user_name") or "").strip() or "—")
            except Exception:
                client_name2 = "—"

            ship_included = _is_yes(o2.get("ship_included") or o2.get("shipping_included") or "")
            ship_fee_s = str(o2.get("shipping_fee_sar") or o2.get("shipping_fee") or "").strip()
            ship_fee_n = _num_float(ship_fee_s)

            goods_n = _num_float(goods_amt)
            total_n = goods_n + (0.0 if ship_included else ship_fee_n)

            def _fmt_num(n: float) -> str:
                try:
                    if abs(n - int(n)) < 1e-9:
                        return f"{int(n)} ر.س"
                    return (f"{n:.2f}".rstrip("0").rstrip(".") + " ر.س")
                except Exception:
                    return "—"

            goods_txt = _fmt_money(goods_amt) or (_fmt_num(goods_n) if goods_n > 0 else "—")
            ship_txt = "مشمول" if ship_included else (_fmt_money(ship_fee_s) or (_fmt_num(ship_fee_n) if ship_fee_n > 0 else "—"))
            total_txt = _fmt_num(total_n) if (goods_n > 0 or ship_fee_n > 0 or ship_included) else "—"

            trk = (tracking_value or "").strip() or "غير متوفر"

            try:
                await context.bot.send_message(
                    chat_id=user_id,
                    text=(
                        "📦 <b>تم تحديث الحالة إلى: تم الشحن</b>\n"
                        f"👤 العميل: <b>{html.escape(client_name2)}</b>\n"
                        f"🧾 رقم الطلب: <b>{html.escape(order_id2)}</b>\n"
                        f"🧩 مبالغ القطع: <b>{html.escape(str(goods_txt))}</b>\n"
                        f"🚚 الشحن: <b>{html.escape(str(ship_txt))}</b>\n"
                        f"💰 الإجمالي: <b>{html.escape(str(total_txt))}</b>\n"
                        f"📦 رقم التتبع: <b>{html.escape(trk)}</b>\n\n"
                        "اختر الإجراء التالي من الأزرار بالأسفل:"
                    ),
                    parse_mode="HTML",
                    reply_markup=trader_received_notice_kb(order_id2),
                    disable_web_page_preview=True,
                )
            except Exception as e:
                _swallow(e)

        
        if kind == "goods_paylink":
            link = (text or "").strip()
            if not (link.startswith("http://") or link.startswith("https://")):
                await update.message.reply_text(f"{name}\nارسل رابط يبدأ بـ https://")
                return

            try:
                update_order_fields(order_id2, {
                    "order_status": "ready_to_ship",
                    "goods_payment_method": "pay_link",
                    "goods_payment_status": "awaiting_receipt",
                    "goods_payment_link": link,
                })
            except Exception as e:
                _swallow(e)

            # إرسال زر الدفع للعميل + نقل العميل لمرحلة إرسال الإيصال
            if client_id2:
                try:
                    udc = get_ud(context, client_id2)
                    udc["goods_order_id"] = order_id2
                    set_stage(context, client_id2, STAGE_AWAIT_GOODS_RECEIPT)
                except Exception as e:
                    _swallow(e)

                try:
                    await context.bot.send_message(
                        chat_id=client_id2,
                        text=(
                            f"{_user_name(update)}\n"
                            "🔗 رابط دفع قيمة البضاعة\n"
                            f"🧾 رقم الطلب: {order_id2}\n\n"
                            "افتح الرابط وأكمل الدفع.\n"
                            "بعد الدفع أرسل الإيصال هنا داخل المنصة لإكمال الإجراء."
                        ),
                        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("💳 دفع الآن", url=link)]]),
                        disable_web_page_preview=True,
                    )
                except Exception:
                    await update.message.reply_text(f"{name}\nتعذر إرسال الرابط للعميل")
                    return

            # إشعار بسيط للإدارة
            try:
                for aid in (ADMIN_IDS or []):
                    try:
                        await context.bot.send_message(
                            chat_id=int(aid),
                            text=(f"🔔 تم إرسال رابط دفع قيمة البضاعة للعميل\n🧾 رقم الطلب: {order_id2}"),
                            disable_web_page_preview=True,
                        )
                    except Exception as e:
                        _swallow(e)
            except Exception as e:
                _swallow(e)

            ud.pop("tsu_kind", None)
            ud.pop("tsu_order_id", None)
            set_stage(context, user_id, STAGE_NONE)

            await update.message.reply_text(f"{name}\nتم إرسال الرابط للعميل وننتظر الإيصال")
            return

        if kind == "tracking":
            tracking = (text or "").strip()

            skip_words = ("تخطي", "تخطى", "تجاوز", "بدون تتبع", "بدون رقم", "skip", "no", "none")
            if tracking and tracking.lower() in [w.lower() for w in skip_words]:
                tracking = ""

                update_order_fields(order_id2, {
                    "order_status": "shipped",
                    "shipping_tracking": "",
                    "shipping_at": utc_now_iso(),
                })

                try:
                    await send_invoice_pdf(
                        context,
                        order_id2,
                        kind="shipping",
                        invoice_for="trader",
                        include_admins=False,
                        tracking_number="",
                    )
                except Exception as e:
                    await _notify_invoice_error(context, order_id2, "فاتورة الشحن", e)

                if client_id2:
                    try:
                        details_lines = []
                        if ship_method:
                            details_lines.append(f"🚚 طريقة التسليم: {ship_method}")
                        if ship_city:
                            details_lines.append(f"📍 المدينة: {ship_city}")
                        if goods_amt:
                            details_lines.append(f"💰 قيمة القطع: {_fmt_money(goods_amt) or (goods_amt + ' ر.س')}")

                        extra = ("\n".join(details_lines)).strip()
                        if extra:
                            extra = "\n\n" + extra

                        await context.bot.send_message(
                            chat_id=client_id2,
                            text=(
                                "✅ <b>تم شحن طلبك بنجاح</b>\n"
                                f"🧾 رقم الطلب: {html.escape(order_id2)}\n"
                                "📦 رقم التتبع: <b>غير متوفر</b>\n"
                                f"🧑‍💼 التاجر: <b>{html.escape(tname)}</b>"
                                f"{extra}\n\n"
                                "يمكنك مراسلة التاجر أو المتابعة من الزر بالأسفل."
                            ),
                            parse_mode="HTML",
                            reply_markup=client_trader_chat_kb(order_id2),
                            disable_web_page_preview=True,
                        )
                    except Exception as e:
                        _swallow(e)

                for aid in ADMIN_IDS:
                    try:
                        admin_lines = [
                            "📣 <b>تحديث حالة: تم الشحن</b>",
                            f"🧾 الطلب: {html.escape(order_id2)}",
                            f"🧑‍💼 التاجر: <b>{html.escape(tname)}</b> (<code>{user_id}</code>)",
                            "📦 التتبع: <b>غير متوفر</b>",
                        ]
                        if goods_amt:
                            admin_lines.append(f"💰 قيمة القطع: <b>{html.escape(_fmt_money(goods_amt) or goods_amt + ' ر.س')}</b>")
                        if ship_method or ship_city:
                            admin_lines.append(f"🚚 التسليم: {html.escape((ship_method + ' - ' + ship_city).strip(' -'))}")

                        await context.bot.send_message(
                            chat_id=aid,
                            text="\n".join(admin_lines),
                            parse_mode="HTML",
                            disable_web_page_preview=True,
                        )
                    except Exception as e:
                        _swallow(e)

                ud.pop("tsu_kind", None)
                ud.pop("tsu_order_id", None)
                set_stage(context, user_id, STAGE_NONE)

                await _reply_html(
                    "تم التحديث",
                    [
                        "✅ تم تحديث الحالة إلى: <b>تم الشحن</b>",
                        f"🧾 رقم الطلب: {html.escape(order_id2)}",
                        "📦 رقم التتبع: غير متوفر",
                    ],
                )

                # ✅ المطلوب: إرسال "رسالة التاجر مع 3 أزرار" بشكل مفصل + منع التكرار إذا already_shipped=True
                await _send_trader_notice("")

                return

            if len(tracking) < 4:
                await _reply_html("رقم تتبع غير واضح", ["⚠️ اكتب رقم التتبع بشكل صحيح (مثال: <code>7845123690</code>)."])
                return

            update_order_fields(order_id2, {
                "order_status": "shipped",
                "shipping_tracking": tracking,
                "shipping_at": utc_now_iso(),
            })

            try:
                await send_invoice_pdf(
                    context,
                    order_id2,
                    kind="shipping",
                    invoice_for="trader",
                    include_admins=False,
                    tracking_number=tracking,
                )
            except Exception as e:
                await _notify_invoice_error(context, order_id2, "فاتورة الشحن", e)

            if client_id2:
                try:
                    details_lines = []
                    if ship_method:
                        details_lines.append(f"🚚 طريقة التسليم: {ship_method}")
                    if ship_city:
                        details_lines.append(f"📍 المدينة: {ship_city}")
                    if goods_amt:
                        details_lines.append(f"💰 قيمة القطع: {_fmt_money(goods_amt) or (goods_amt + ' ر.س')}")

                    extra = ("\n".join(details_lines)).strip()
                    if extra:
                        extra = "\n\n" + extra

                    await context.bot.send_message(
                        chat_id=client_id2,
                        text=(
                            "✅ <b>تم شحن طلبك بنجاح</b>\n"
                            f"🧾 رقم الطلب: {html.escape(order_id2)}\n"
                            f"📦 رقم التتبع: <b>{html.escape(tracking)}</b>\n"
                            f"🧑‍💼 التاجر: <b>{html.escape(tname)}</b>"
                            f"{extra}\n\n"
                            "يمكنك مراسلة التاجر أو المتابعة من الزر بالأسفل."
                        ),
                        parse_mode="HTML",
                        reply_markup=client_trader_chat_kb(order_id2),
                        disable_web_page_preview=True,
                    )
                except Exception as e:
                    _swallow(e)

            for aid in ADMIN_IDS:
                try:
                    admin_lines = [
                        "📣 <b>تحديث حالة: تم الشحن</b>",
                        f"🧾 الطلب: {html.escape(order_id2)}",
                        f"🧑‍💼 التاجر: <b>{html.escape(tname)}</b> (<code>{user_id}</code>)",
                        f"📦 التتبع: <b>{html.escape(tracking)}</b>",
                    ]
                    if goods_amt:
                        admin_lines.append(f"💰 قيمة القطع: <b>{html.escape(_fmt_money(goods_amt) or goods_amt + ' ر.س')}</b>")
                    if ship_method or ship_city:
                        admin_lines.append(f"🚚 التسليم: {html.escape((ship_method + ' - ' + ship_city).strip(' -'))}")

                    await context.bot.send_message(
                        chat_id=aid,
                        text="\n".join(admin_lines),
                        parse_mode="HTML",
                        disable_web_page_preview=True,
                    )
                except Exception as e:
                    _swallow(e)

            ud.pop("tsu_kind", None)
            ud.pop("tsu_order_id", None)
            set_stage(context, user_id, STAGE_NONE)

            await _reply_html(
                "تم التحديث",
                [
                    "✅ تم تحديث الحالة إلى: <b>تم الشحن</b>",
                    f"🧾 رقم الطلب: {html.escape(order_id2)}",
                    f"📦 رقم التتبع: <b>{html.escape(tracking)}</b>",
                ],
            )

            # ✅ المطلوب: إرسال "رسالة التاجر مع 3 أزرار" بشكل مفصل + منع التكرار إذا already_shipped=True
            await _send_trader_notice(tracking)

            return

        await _reply_html(
            "فاتورة التاجر مطلوبة",
            [
                f"🧾 رقم الطلب: {html.escape(order_id2)}",
                "ارسل الفاتورة كـ PDF أو صورة واضحة.",
                "بدون الفاتورة لن يتم اعتماد التحديث.",
            ],
        )
        return
    # ==================================================
    # 12) إدخال عرض السعر من التاجر
    # ==================================================
    if stage == STAGE_TRADER_SET_QUOTE:
        if _trader_is_disabled(user_id):
            set_stage(context, user_id, STAGE_NONE)
            td0 = context.user_data.setdefault(user_id, {})
            td0.pop("quote_order_id", None)
            td0.pop("quote_step", None)
            await _reply_html("الحساب موقوف", [html.escape(_trader_disabled_msg())])
            return

        txt = (text or "").strip().lower()
        if txt in ("الغاء", "إلغاء", "cancel", "خروج", "رجوع", "انهاء", "إنهاء"):
            td = context.user_data.setdefault(user_id, {})
            old_oid = str(td.get("quote_order_id") or "").strip()

            td.pop("quote_order_id", None)
            td.pop("quote_step", None)
            td.pop("quote_goods_amount", None)
            td.pop("quote_parts_type", None)
            td.pop("quote_ship_method", None)
            td.pop("quote_ship_included", None)
            td.pop("quote_shipping_fee", None)
            td.pop("quote_ship_eta", None)
            td.pop("quote_availability", None)
            td.pop("quote_started_at_utc", None)

            try:
                _qreset(td, old_oid or "")
            except Exception as e:
                _swallow(e)

            td.pop("quote_item_prices", None)
            td.pop("quote_pending_item_idx", None)
            td.pop("quote_pending_item_name", None)

            set_stage(context, user_id, STAGE_NONE)
            await _reply_html("تم الإلغاء", ["✅ تم إنهاء وضع عرض السعر."])
            return

        td = context.user_data.setdefault(user_id, {})
        order_id = str(td.get("quote_order_id") or "").strip()
        if not order_id:
            set_stage(context, user_id, STAGE_NONE)
            await _reply_html(
                "لا يوجد طلب مرتبط",
                [
                    "ℹ️ لا يوجد طلب مرتبط بعرض السعر حالياً.",
                    "ارجع لنفس الطلب واضغط زر <b>(تقديم عرض سعر)</b> ثم حاول مرة أخرى.",
                ],
            )
            return

        # ===== هيدر موحّد لرسائل عرض السعر =====
        def _fmt_money_local(v: object) -> str:
            s = str(v or "").strip()
            if not s or s in ("0", "0.0"):
                return ""
            try:
                return _fmt_money(s)
            except Exception:
                try:
                    return _money(s)
                except Exception:
                    return f"{s} ر.س"

        def _quote_hdr(oid: str, goods_total: str = "") -> list:
            snap = _order_snapshot(oid) if "_order_snapshot" in globals() or "_order_snapshot" in locals() else {}
            client_name = str((snap or {}).get("client_name") or (snap or {}).get("user_name") or "—").strip() or "—"

            goods_now = str(goods_total or "").strip()
            if not goods_now:
                goods_now = str(td.get("quote_goods_amount") or "").strip()

            ship_fee = str(td.get("quote_shipping_fee") or "").strip()

            lines = [
                f"🧾 رقم الطلب: {_order_id_link_html(oid, context)}",
                f"👤 العميل: <b>{html.escape(client_name)}</b>",
            ]

            if goods_now:
                lines.append(f"💰 إجمالي القطع: <b>{html.escape(goods_now)}</b> ر.س")

            ship_txt = _fmt_money_local(ship_fee)
            ship_val = None
            try:
                ship_val = float(str(ship_fee).strip()) if str(ship_fee).strip() not in ("", "—") else None
            except Exception:
                ship_val = None

            if ship_txt and ship_txt not in ("0 ر.س", "0.0 ر.س") and ship_val is not None and ship_val > 0:
                lines.append(f"🚚 الشحن: <b>{html.escape(ship_txt)}</b>")

            try:
                g = float(str(goods_now).strip()) if str(goods_now).strip() not in ("", "—") else None
            except Exception:
                g = None

            if g is not None and ship_val is not None:
                total_calc = g + ship_val
                if abs(total_calc - int(total_calc)) < 1e-9:
                    total_txt = f"{int(total_calc)} ر.س"
                else:
                    total_txt = f"{total_calc:.2f}".rstrip("0").rstrip(".") + " ر.س"
                lines.append(f"🧾 الإجمالي: <b>{html.escape(total_txt)}</b>")

            return lines

        try:
            tp = get_trader_profile(int(user_id or 0)) or {}
        except Exception:
            tp = {}

        required_fields = ["display_name","company_name","shop_phone","cr_no","vat_no","bank_name","iban","stc_pay"]
        if not all((tp.get(f) or "").strip() for f in required_fields):
            await _reply_html(
                "أكمل ملف التاجر",
                [
                    "ℹ️ لا يمكنك تقديم عرض قبل إكمال بيانات ملف التاجر:",
                    "• اسم المتجر",
                    "• IBAN",
                    "• STC Pay",
                ],
            )
            return

        try:
            started = _parse_utc_iso(td.get("quote_started_at_utc") or "")
            if started:
                now = _dt_utc_now()
                if (now - started).total_seconds() > 20 * 60:
                    try:
                        _qreset(td, order_id)
                    except Exception as e:
                        _swallow(e)

                    td.pop("quote_order_id", None)
                    td.pop("quote_step", None)
                    td.pop("quote_started_at_utc", None)
                    td.pop("quote_item_prices", None)
                    td.pop("quote_pending_item_idx", None)
                    td.pop("quote_pending_item_name", None)
                    set_stage(context, user_id, STAGE_NONE)

                    await _reply_html(
                        "انتهت الجلسة",
                        [
                            "ℹ️ انتهت جلسة عرض السعر بسبب عدم النشاط.",
                            "ابدأ من جديد من الطلب.",
                        ],
                    )
                    return
        except Exception as e:
            _swallow(e)

        def _get_items_for_quote(oid: str):
            try:
                ensure_workbook()
            except Exception as e:
                _swallow(e)

            try:
                obx = get_order_bundle(oid) or {}
                its = (obx.get("items") or []) if isinstance(obx, dict) else []
            except Exception:
                its = []

            if not isinstance(its, list):
                its = []

            out = []
            for it in its:
                if isinstance(it, dict):
                    nm = (it.get("name") or it.get("item_name") or "").strip()
                    pn = (it.get("part_no") or it.get("partno") or it.get("part_number") or "").strip()
                    out.append({"name": nm, "part_no": pn})
                else:
                    s = str(it).strip()
                    out.append({"name": s, "part_no": ""})
            return out

        def _price_map():
            try:
                qs = _qget(td, order_id, create=True)
            except Exception:
                qs = {}
            pm = qs.get("item_prices") if isinstance(qs, dict) else None
            if not isinstance(pm, dict):
                pm = {}
            out = {}
            for k, v in pm.items():
                ks = str(k).strip()
                vs = str(v).strip()
                if ks.isdigit() and vs:
                    out[ks] = vs
            try:
                if isinstance(qs, dict):
                    qs["item_prices"] = out
                    qs["last_touch"] = int(_dt_utc_now().timestamp())
            except Exception as e:
                _swallow(e)
            td["quote_item_prices"] = dict(out)
            return out

        def _calc_total(items: list, price_map: dict):
            total = 0.0
            for i in range(1, len(items) + 1):
                v = str(price_map.get(str(i), "")).strip()
                if not v:
                    continue
                try:
                    total += float(v)
                except Exception as e:
                    _swallow(e)
            if abs(total - int(total)) < 1e-9:
                return str(int(total))
            return f"{total:.2f}".rstrip("0").rstrip(".")

        def _items_kb_local(oid: str, items: list, price_map: dict):
            rows = []
            for i, it in enumerate(items, start=1):
                nm = (it.get("name") or "").strip() or f"قطعة {i}"
                pn = (it.get("part_no") or "").strip()
                price = str(price_map.get(str(i), "")).strip()
                tail = f" — {price} ر.س" if price else " — اضف سعر"
                label = f"🧩 {i}) {nm}"
                if pn:
                    label += f" ({pn})"
                label += tail
                rows.append([InlineKeyboardButton(label, callback_data=f"ppq_it|{oid}|{i}")])

            rows.append([InlineKeyboardButton("✅ اكمال خطوات العرض", callback_data=f"ppq_it_done|{oid}")])
            return InlineKeyboardMarkup(rows)

        step = str(td.get("quote_step") or "start")

        if step == "start":
            await _reply_html(
                "بناء عرض السعر",
                _quote_hdr(order_id) + ["👇 اضغط زر <b>(بدء بناء عرض السعر)</b> ثم اتبع الخطوات بالترتيب."],
                kb=trader_quote_start_kb(order_id),
            )
            return

        if step == "it_price":
            m_amt = re.search(r"(\d+(?:\.\d+)?)", text)
            if not m_amt:
                await _reply_html(
                    "سعر غير صحيح",
                    _quote_hdr(order_id) + ["ℹ️ اكتب السعر بالأرقام فقط.", "مثال: <code>120</code> أو <code>120.50</code>"],
                )
                return

            price = m_amt.group(1)

            idx = _safe_int(td.get("quote_pending_item_idx"))
            items = _get_items_for_quote(order_id)
            if not items or idx < 1 or idx > len(items):
                td["quote_step"] = "it_pick"
                td.pop("quote_pending_item_idx", None)
                td.pop("quote_pending_item_name", None)
                await _reply_html("تعذر تحديد القطعة", _quote_hdr(order_id) + ["ℹ️ ارجع للكيبورد واختر القطعة مرة أخرى."])
                return

            pm = _price_map()
            pm[str(idx)] = price

            try:
                qs = _qget(td, order_id, create=True)
                qs["item_prices"] = dict(pm)
                qs["last_touch"] = int(_dt_utc_now().timestamp())
                try:
                    qs.pop("pending_item_idx", None)
                    qs.pop("pending_item_name", None)
                except Exception as e:
                    _swallow(e)
            except Exception as e:
                _swallow(e)
            td["quote_item_prices"] = dict(pm)

            td["quote_step"] = "it_pick"
            td.pop("quote_pending_item_idx", None)
            td.pop("quote_pending_item_name", None)

            total_now = _calc_total(items, pm)
            await _reply_html(
                "تم حفظ السعر",
                _quote_hdr(order_id, goods_total=total_now) + ["✅ تم حفظ سعر القطعة.", "👇 اختر قطعة أخرى أو أكمل الخطوات:"],
                kb=_items_kb_local(order_id, items, pm),
            )
            return

        if step == "it_all_price":
            m_amt = re.search(r"(\d+(?:\.\d+)?)", text)
            if not m_amt:
                await _reply_html(
                    "سعر غير صحيح",
                    _quote_hdr(order_id) + ["ℹ️ اكتب السعر بالأرقام فقط.", "مثال: <code>50</code> أو <code>75.5</code>"],
                )
                return

            price = m_amt.group(1)
            items = _get_items_for_quote(order_id)
            if not items:
                td["quote_step"] = "it_pick"
                await _reply_html("لا توجد بنود", _quote_hdr(order_id) + ["ℹ️ لا توجد قطع لتسعيرها حالياً."])
                return

            pm = _price_map()
            for i in range(1, len(items) + 1):
                pm[str(i)] = price

            try:
                qs = _qget(td, order_id, create=True)
                qs["item_prices"] = dict(pm)
                qs["last_touch"] = int(_dt_utc_now().timestamp())
            except Exception as e:
                _swallow(e)
            td["quote_item_prices"] = dict(pm)

            td["quote_step"] = "it_pick"
            total_now = _calc_total(items, pm)

            await _reply_html(
                "تم تطبيق السعر",
                _quote_hdr(order_id, goods_total=total_now)
                + ["✅ تم تطبيق السعر على جميع القطع.", "👇 اضغط (اكمل خطوات العرض) للمتابعة:"],
                kb=_items_kb_local(order_id, items, pm),
            )
            return

        if step == "amount":
            m_amt = re.search(r"(\d+(?:\.\d+)?)", text)
            if not m_amt:
                await _reply_html(
                    "مبلغ غير صحيح",
                    _quote_hdr(order_id) + ["ℹ️ اكتب مبلغ القطع بالأرقام فقط.", "مثال: <code>850</code> أو <code>850.50</code>"],
                )
                return
            amount = m_amt.group(1)
            td["quote_goods_amount"] = amount
            td["quote_step"] = "type"
            await _reply_html(
                "نوع القطع",
                _quote_hdr(order_id, goods_total=amount) + ["👇 اختر نوع القطع من الأزرار:"],
                kb=trader_quote_type_kb(order_id),
            )
            return

        if step == "shipping_fee":
            m_fee = re.search(r"(\d+(?:\.\d+)?)", text)
            if not m_fee:
                await _reply_html(
                    "قيمة شحن غير صحيحة",
                    _quote_hdr(order_id) + ["ℹ️ اكتب قيمة الشحن بالأرقام فقط.", "مثال: <code>25</code> أو <code>40.5</code>"],
                )
                return

            fee = m_fee.group(1)
            td["quote_shipping_fee"] = fee
            td["quote_step"] = "availability"

            # ✅ تثبيت الشحن داخل الطلب فوراً (حتى تظهر في معاينة التاجر)
            try:
                update_order_fields(order_id, {
                    "ship_included": "no",
                    "shipping_fee_sar": str(fee).strip(),
                })
            except Exception as e:
                _swallow(e)

            await _reply_html(
                "مدة التجهيز",
                _quote_hdr(order_id) + [f"🚚 الشحن: <b>{html.escape(fee)}</b> ر.س", "👇 حدد مدة التجهيز من الأزرار:"],
                kb=trader_quote_availability_kb(order_id),
            )
            return

        if step == "eta_custom":
            v = (text or "").strip()
            if len(v) < 2:
                await _reply_html(
                    "مدة غير واضحة",
                    _quote_hdr(order_id) + ["ℹ️ اكتبها بصيغة مفهومة.", "مثال: <code>2-3 أيام</code>"]
                )
                return

            td["quote_ship_eta"] = v

            # ✅ بدل الإرسال المباشر: اعرض المعاينة مع الأزرار
            td["quote_step"] = "preview"
            try:
                await show_quote_preview(context, user_id, update.message, order_id)
            except Exception:
                # fallback آمن: لو فشل العرض لأي سبب، لا نرسل تلقائياً
                await _reply_html(
                    "تنبيه",
                    _quote_hdr(order_id) + ["ℹ️ تعذر عرض المعاينة الآن، حاول مرة أخرى من الأزرار."]
                )
            return

        if step == "avail_custom":
            v = (text or "").strip()
            if len(v) < 2:
                await _reply_html(
                    "مدة غير واضحة",
                    _quote_hdr(order_id) + ["ℹ️ اكتبها بصيغة مفهومة.", "مثال: <code>5 أيام</code>"]
                )
                return

            td["quote_availability"] = v
            td["quote_step"] = "eta"
            await _reply_html(
                "مدة الشحن",
                _quote_hdr(order_id) + [f"⏳ مدة التجهيز: <b>{html.escape(v)}</b>", "👇 حدد مدة الشحن من الأزرار:"],
                kb=trader_quote_eta_kb(order_id),
            )
            return

        await _reply_html(
            "تنبيه",
            _quote_hdr(order_id) + ["ℹ️ أكمل الخطوات باستخدام الأزرار حتى لا تتداخل المراحل."]
        )
        return

    # ==================================================
    # 13) مراحل دفع قيمة القطع
    # ==================================================
    if stage == STAGE_AWAIT_GOODS_PAY_METHOD:
        await _reply_html("اختيار طريقة الدفع", ["👇 اختر طريقة الدفع من الأزرار بالأسفل."])
        return

    if stage == STAGE_AWAIT_GOODS_RECEIPT:
        await _reply_html("الإيصال مطلوب", ["⚠️ الرجاء إرسال صورة إيصال الدفع فقط."])
        return
    # ==================================================
    # 14) مراسلة التاجر (بدون كشف الهوية)
    # ==================================================
    if stage == STAGE_CHAT_TRADER:
        order_id = ud.get("chat_trader_order_id", "")
        tid = _assigned_trader_id(order_id) if order_id else None
        if not order_id or not tid:
            set_stage(context, user_id, STAGE_NONE)
            await _reply_html("لا يوجد تاجر", ["⚠️ لا يوجد تاجر محدد لهذا الطلب حالياً."])
            return

        try:
            tprof = get_trader_profile(tid) or {}
            tname = (tprof.get("display_name") or "").strip() or "التاجر"
        except Exception:
            tname = "التاجر"

        # ===== احسب (القطع + الشحن) فقط — بدون رسوم منصة =====
        snap = _order_snapshot(order_id)

        def _s(x: object) -> str:
            return ("" if x is None else str(x)).strip()

        def _is_yes(x: object) -> bool:
            v = _s(x).lower()
            return v in ("yes", "y", "true", "1", "مشمول", "included")

        goods_str = ""
        ship_str = ""
        ship_included = False

        try:
            b = get_order_bundle(order_id) or {}
            o = b.get("order", {}) or {}
            goods_str = _s(o.get("goods_amount_sar") or snap.get("goods_amount") or "")
            ship_str = _s(o.get("shipping_fee_sar") or snap.get("shipping_fee") or "")
            ship_included = _is_yes(o.get("ship_included") or o.get("shipping_included") or snap.get("ship_included") or "")
        except Exception:
            goods_str = _s(snap.get("goods_amount") or "")
            ship_str = _s(snap.get("shipping_fee") or "")
            ship_included = _is_yes(snap.get("ship_included") or "")

        goods_txt = _fmt_money(goods_str or "")
        ship_for_total = "0" if ship_included else (ship_str or "0")

        total_txt = ""
        try:
            # _calc_totals ترجع أرقام كنصوص (g,s,t)
            g_num, s_num, t_num = _calc_totals(goods_str or "0", ship_for_total or "0")
            total_txt = _fmt_money(t_num or "")
        except Exception:
            total_txt = ""

        head_lines = [f"💬 {snap.get('client_name','العميل')}"]
        head_lines.append(f"🧾 الطلب: {order_id}")

        # نظهر فقط منطق (قطع + شحن) إذا كان عندنا قيمة قطع فعلية
        if goods_txt:
            head_lines.append(f"🧩 قيمة القطع: {goods_txt}")
            if ship_included:
                head_lines.append("🚚 الشحن: مشمول")
            else:
                ship_txt = _fmt_money(ship_str or "")
                if ship_txt:
                    head_lines.append(f"🚚 الشحن: {ship_txt}")
            if total_txt:
                head_lines.append(f"💰 إجمالي القطع + الشحن: {total_txt}")

        msg = "\n".join(head_lines) + "\n" + (text or "")
        try:
            await context.bot.send_message(
                chat_id=tid,
                text=msg,
                parse_mode="Markdown",
                reply_markup=trader_reply_kb(order_id, user_id),
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)

        await _reply_html("تم الإرسال", ["✅ تم إرسال رسالتك للتاجر."], kb=chat_nav_kb_for(context, user_id, order_id, "pp_chat_trader_done"))
        return
    # ==================================================
    # 15) رد التاجر (يصل للعميل باسم التاجر)
    # ==================================================
    if stage == STAGE_TRADER_REPLY:
        td = context.user_data.setdefault(user_id, {})

        to_uid = _safe_int(td.get("trader_reply_user_id"))
        order_id = str(td.get("trader_reply_order_id") or "").strip()

        if not to_uid or not order_id:
            await _reply_html("تعذر الإرسال", ["⚠️ تعذر تحديد العميل المرتبط بهذه المراسلة."])
            set_stage(context, user_id, STAGE_NONE)
            return

        tprof = get_trader_profile(user_id) or {}
        tname = (tprof.get("display_name") or "").strip() or (_user_name(update) or "").strip() or "التاجر"
        tcompany = (tprof.get("company_name") or "").strip()
        tlabel = tname + (f" ({tcompany})" if tcompany else "")

        body = (text or "").strip()
        if not body:
            await _reply_html("مطلوب نص", ["ℹ️ اكتب رسالتك ثم أرسلها."])
            return

        # ===== احسب (القطع + الشحن) فقط — بدون رسوم منصة =====
        snap = _order_snapshot(order_id)

        def _s(x: object) -> str:
            return ("" if x is None else str(x)).strip()

        def _is_yes(x: object) -> bool:
            v = _s(x).lower()
            return v in ("yes", "y", "true", "1", "مشمول", "included")

        goods_str = ""
        ship_str = ""
        ship_included = False

        try:
            b = get_order_bundle(order_id) or {}
            o = b.get("order", {}) or {}
            goods_str = _s(o.get("goods_amount_sar") or snap.get("goods_amount") or "")
            ship_str = _s(o.get("shipping_fee_sar") or snap.get("shipping_fee") or "")
            ship_included = _is_yes(o.get("ship_included") or o.get("shipping_included") or snap.get("ship_included") or "")
        except Exception:
            goods_str = _s(snap.get("goods_amount") or "")
            ship_str = _s(snap.get("shipping_fee") or "")
            ship_included = _is_yes(snap.get("ship_included") or "")

        goods_txt = _fmt_money(goods_str or "")
        ship_for_total = "0" if ship_included else (ship_str or "0")

        ship_txt = ""
        total_txt = ""
        try:
            g_num, s_num, t_num = _calc_totals(goods_str or "0", ship_for_total or "0")
            total_txt = _fmt_money(t_num or "")
            if ship_included:
                ship_txt = "مشمول"
            else:
                ship_txt = _fmt_money(s_num or "") or _fmt_money(ship_str or "")
        except Exception:
            if ship_included:
                ship_txt = "مشمول"
            else:
                ship_txt = _fmt_money(ship_str or "")
            total_txt = ""

        head_lines = [
            f"💬 {html.escape(tlabel)}",
            f"🧾 الطلب: {html.escape(order_id)}",
        ]

        if goods_txt:
            head_lines.append(f"🧩 قيمة القطع: <b>{html.escape(goods_txt)}</b>")
            if ship_txt:
                head_lines.append(f"🚚 الشحن: <b>{html.escape(ship_txt)}</b>")
            if total_txt:
                head_lines.append(f"💰 إجمالي القطع + الشحن: <b>{html.escape(total_txt)}</b>")

        msg_to_client = "\n".join(head_lines) + "\n" + html.escape(body)

        try:
            await context.bot.send_message(
                chat_id=to_uid,
                text=msg_to_client,
                parse_mode="HTML",
                reply_markup=chat_nav_kb_for(context, to_uid, order_id, "pp_chat_trader_done"),
                disable_web_page_preview=True,
            )
            await _reply_html("تم الإرسال", ["✅ تم إرسال ردّك للعميل."], kb=chat_nav_kb(order_id, "pp_trader_reply_done"))
        except Exception:
            await _reply_html("تعذر الإرسال", ["⚠️ قد لا يكون العميل بدأ البوت أو قام بحظر البوت."])
        return
    # ==================================================
    # 16) مراسلة الإدارة (إلى عميل/تاجر) — مختصر
    # ==================================================
    if stage == STAGE_ADMIN_CHAT:
        if user_id not in ADMIN_IDS:
            set_stage(context, user_id, STAGE_NONE)
            await _reply_html("غير مصرح", ["⛔ هذه الخدمة خاصة بالإدارة."])
            return

        order_id = str(ud.get("admin_chat_order_id") or "").strip()
        peer_id = _safe_int(ud.get("admin_chat_peer_id"))
        role = str(ud.get("admin_chat_role") or "").strip()  # client / trader
        body = (text or "").strip()

        if not order_id or not peer_id or not body:
            await _reply_html("بيانات ناقصة", ["⚠️ اكتب رسالة صحيحة."])
            return

        snap = _order_snapshot(order_id)
        body_esc = html.escape(body)

        # ===== احسب (القطع + الشحن) فقط — بدون رسوم منصة =====
        def _s(x: object) -> str:
            return ("" if x is None else str(x)).strip()

        def _is_yes(x: object) -> bool:
            v = _s(x).lower()
            return v in ("yes", "y", "true", "1", "مشمول", "included")

        goods_str = ""
        ship_str = ""
        ship_included = False

        try:
            b = get_order_bundle(order_id) or {}
            o = b.get("order", {}) or {}
            goods_str = _s(o.get("goods_amount_sar") or snap.get("goods_amount") or "")
            ship_str = _s(o.get("shipping_fee_sar") or snap.get("shipping_fee") or "")
            ship_included = _is_yes(o.get("ship_included") or o.get("shipping_included") or snap.get("ship_included") or "")
        except Exception:
            goods_str = _s(snap.get("goods_amount") or "")
            ship_str = _s(snap.get("shipping_fee") or "")
            ship_included = _is_yes(snap.get("ship_included") or "")

        goods_txt = _fmt_money(goods_str or "")
        ship_for_total = "0" if ship_included else (ship_str or "0")

        ship_txt = ""
        total_txt = ""
        try:
            g_num, s_num, t_num = _calc_totals(goods_str or "0", ship_for_total or "0")
            total_txt = _fmt_money(t_num or "")
            if ship_included:
                ship_txt = "مشمول"
            else:
                ship_txt = _fmt_money(s_num or "") or _fmt_money(ship_str or "")
        except Exception:
            if ship_included:
                ship_txt = "مشمول"
            else:
                ship_txt = _fmt_money(ship_str or "")
            total_txt = ""

        amount_lines = ""
        if goods_txt:
            amount_lines = (
                f"🧩 قيمة القطع: <b>{html.escape(goods_txt)}</b>\n"
                f"🚚 الشحن: <b>{html.escape(ship_txt or '—')}</b>\n"
                + (f"💰 إجمالي القطع + الشحن: <b>{html.escape(total_txt)}</b>\n" if total_txt else "")
            )

        try:
            if role == "client":
                msg = (
                    "📩 <b>رسالة من الإدارة</b>\n"
                    f"🧾 الطلب: {html.escape(order_id)}\n"
                    f"👤 العميل: <b>{html.escape(snap.get('client_name','—'))}</b>\n"
                    f"{amount_lines}"
                    f"🧑‍💼 التاجر: <b>{html.escape(snap.get('trader_name','—'))}</b>\n"
                    f"📌 الحالة: <b>{html.escape(snap.get('status','—'))}</b>\n"
                    "────────────────\n"
                    f"{body_esc}"
                )
                await context.bot.send_message(
                    chat_id=peer_id,
                    text=msg,
                    parse_mode="HTML",
                    reply_markup=track_kb(order_id),
                    disable_web_page_preview=True,
                )
            else:
                msg = (
                    "📩 <b>رسالة من الإدارة → التاجر</b>\n"
                    f"🧾 الطلب: {html.escape(order_id)}\n"
                    f"👤 العميل: <b>{html.escape(snap.get('client_name','—'))}</b>\n"
                    f"{amount_lines}"
                    f"📌 الحالة: <b>{html.escape(snap.get('status','—'))}</b>\n"
                    "────────────────\n"
                    f"{body_esc}"
                )
                try:
                    context.bot_data.setdefault("pp_admin_trader_sessions", {})[str(peer_id)] = {
                        "order_id": order_id,
                        "peer_admin_id": int(user_id),
                    }
                except Exception as e:
                    _swallow(e)

                await context.bot.send_message(
                    chat_id=peer_id,
                    text=msg,
                    parse_mode="HTML",
                    reply_markup=trader_chat_admin_kb(order_id, int(user_id)),
                    disable_web_page_preview=True,
                )

            await _reply_html("تم الإرسال", ["✅ تم إرسال الرسالة."])
        except Exception:
            await _reply_html("تعذر الإرسال", ["⚠️ تعذر إرسال الرسالة حالياً."])
        return
    # ==================================================
    # 17) رد التاجر للإدارة (قناة مستقلة)
    # ==================================================
    if stage == STAGE_TRADER_CHAT_ADMIN:
        order_id = str(ud.get("trader_chat_order_id") or "").strip()
        admin_id = _safe_int(ud.get("trader_chat_admin_id"))
        body = (text or "").strip()
        if not order_id or not admin_id or not body:
            await _reply_html("بيانات ناقصة", ["⚠️ اكتب رسالة صحيحة."])
            return

        try:
            tprof = get_trader_profile(user_id) or {}
            tname = (tprof.get("display_name") or "").strip() or (_user_name(update) or "").strip() or "التاجر"
        except Exception:
            tname = _user_name(update) or "التاجر"

        snap = _order_snapshot(order_id)

        # ===== احسب (القطع + الشحن) فقط — بدون رسوم منصة =====
        def _s(x: object) -> str:
            return ("" if x is None else str(x)).strip()

        def _is_yes(x: object) -> bool:
            v = _s(x).lower()
            return v in ("yes", "y", "true", "1", "مشمول", "included")

        goods_str = ""
        ship_str = ""
        ship_included = False

        try:
            b = get_order_bundle(order_id) or {}
            o = b.get("order", {}) or {}
            goods_str = _s(o.get("goods_amount_sar") or snap.get("goods_amount") or "")
            ship_str = _s(o.get("shipping_fee_sar") or snap.get("shipping_fee") or "")
            ship_included = _is_yes(o.get("ship_included") or o.get("shipping_included") or snap.get("ship_included") or "")
        except Exception:
            goods_str = _s(snap.get("goods_amount") or "")
            ship_str = _s(snap.get("shipping_fee") or "")
            ship_included = _is_yes(snap.get("ship_included") or "")

        goods_txt = _fmt_money(goods_str or "")
        ship_for_total = "0" if ship_included else (ship_str or "0")

        ship_txt = ""
        total_txt = ""
        try:
            g_num, s_num, t_num = _calc_totals(goods_str or "0", ship_for_total or "0")
            total_txt = _fmt_money(t_num or "")
            if ship_included:
                ship_txt = "مشمول"
            else:
                ship_txt = _fmt_money(s_num or "") or _fmt_money(ship_str or "")
        except Exception:
            if ship_included:
                ship_txt = "مشمول"
            else:
                ship_txt = _fmt_money(ship_str or "")
            total_txt = ""

        client_id_part = ""
        try:
            if snap.get("client_id"):
                client_id_part = f" (<code>{html.escape(str(snap.get('client_id')))}</code>)"
        except Exception:
            client_id_part = ""

        lines = [
            "📩 <b>رسالة من التاجر → الإدارة</b>",
            f"🧾 رقم الطلب: {html.escape(order_id)}",
            f"🧑‍💼 التاجر: <b>{html.escape(tname)}</b> (<code>{user_id}</code>)",
            f"👤 العميل: <b>{html.escape(snap.get('client_name','—'))}</b>{client_id_part}",
        ]

        if goods_txt:
            lines.append(f"🧩 قيمة القطع: <b>{html.escape(goods_txt)}</b>")
            lines.append(f"🚚 قيمة الشحن: <b>{html.escape(ship_txt or '—')}</b>")
            if total_txt:
                lines.append(f"💰 إجمالي القطع + الشحن: <b>{html.escape(total_txt)}</b>")

        lines.append(f"📌 الحالة: <b>{html.escape(snap.get('status','—'))}</b>")
        lines.append("────────────────")
        lines.append(html.escape(body))

        msg = "\n".join(lines)

        try:
            await context.bot.send_message(
                chat_id=admin_id,
                text=msg,
                parse_mode="HTML",
                disable_web_page_preview=True,
                reply_markup=admin_contact_kb(order_id),
            )
            await _reply_html("تم الإرسال", ["✅ تم إرسال ردّك للإدارة."])
        except Exception:
            await _reply_html("تعذر الإرسال", ["⚠️ تعذر إرسال ردّك حالياً."])
        return
    # ==================================================
    # 18) رد الإدارة (يصل للعميل باسم الإدارة)
    # ==================================================
    if stage == STAGE_ADMIN_REPLY:
        if user_id not in ADMIN_IDS:
            set_stage(context, user_id, STAGE_NONE)
            await _reply_html("غير مصرح", ["⛔ هذه الخدمة خاصة بالإدارة."])
            return

        ad = context.user_data.setdefault(user_id, {})
        to_uid = _safe_int(ad.get("reply_user_id"))
        order_id = str(ad.get("reply_order_id") or "").strip()

        if not to_uid or not order_id:
            await _reply_html("تعذر الإرسال", ["⚠️ تعذر تحديد العميل المرتبط بهذه المراسلة."])
            set_stage(context, user_id, STAGE_NONE)
            return

        body = (text or "").strip()
        if not body:
            await _reply_html("مطلوب نص", ["ℹ️ اكتب رسالتك ثم أرسلها."])
            return

        snap = _order_snapshot(order_id)

        # ===== احسب (القطع + الشحن) فقط — بدون رسوم منصة =====
        def _s(x: object) -> str:
            return ("" if x is None else str(x)).strip()

        def _is_yes(x: object) -> bool:
            v = _s(x).lower()
            return v in ("yes", "y", "true", "1", "مشمول", "included")

        goods_str = ""
        ship_str = ""
        ship_included = False

        try:
            b = get_order_bundle(order_id) or {}
            o = b.get("order", {}) or {}
            goods_str = _s(o.get("goods_amount_sar") or snap.get("goods_amount") or "")
            ship_str = _s(o.get("shipping_fee_sar") or snap.get("shipping_fee") or "")
            ship_included = _is_yes(o.get("ship_included") or o.get("shipping_included") or snap.get("ship_included") or "")
        except Exception:
            goods_str = _s(snap.get("goods_amount") or "")
            ship_str = _s(snap.get("shipping_fee") or "")
            ship_included = _is_yes(snap.get("ship_included") or "")

        goods_txt = _fmt_money(goods_str or "")
        ship_for_total = "0" if ship_included else (ship_str or "0")

        ship_txt = ""
        total_txt = ""
        try:
            g_num, s_num, t_num = _calc_totals(goods_str or "0", ship_for_total or "0")
            total_txt = _fmt_money(t_num or "")
            if ship_included:
                ship_txt = "مشمول"
            else:
                ship_txt = _fmt_money(s_num or "") or _fmt_money(ship_str or "")
        except Exception:
            if ship_included:
                ship_txt = "مشمول"
            else:
                ship_txt = _fmt_money(ship_str or "")
            total_txt = ""

        amount_lines = ""
        if goods_txt:
            amount_lines = (
                f"🧩 قيمة القطع: <b>{html.escape(goods_txt)}</b>\n"
                f"🚚 الشحن: <b>{html.escape(ship_txt or '—')}</b>\n"
                + (f"💰 إجمالي القطع + الشحن: <b>{html.escape(total_txt)}</b>\n" if total_txt else "")
            )

        msg_to_client = (
            "📩 <b>رسالة من الإدارة</b>\n"
            f"🧾 الطلب: {html.escape(order_id)}\n"
            f"{amount_lines}"
            f"📌 الحالة: <b>{html.escape(snap.get('status','—'))}</b>\n"
            "────────────────\n"
            f"{html.escape(body)}"
        )

        try:
            await context.bot.send_message(
                chat_id=to_uid,
                text=msg_to_client,
                parse_mode="HTML",
                reply_markup=track_kb(order_id),
                disable_web_page_preview=True,
            )
            await _reply_html(
                "تم الإرسال",
                [f"✅ تم إرسال رسالتك للعميل باسم <b>{html.escape(PP_SUPPORT_LABEL)}</b>"],
                kb=admin_reply_done_kb(),
            )
        except Exception:
            await _reply_html("تعذر الإرسال", ["⚠️ قد لا يكون العميل بدأ البوت أو قام بحظر البوت."])
        return
    # ==================================================
    # 19) متابعة الطلب (قناة تواصل بدون كشف الهوية)
    # ==================================================
    if stage == STAGE_TRACK_ORDER:
        order_id = str(ud.get("track_order_id", "") or "").strip()

        real_name = ""
        try:
            b = get_order_bundle(order_id)
            o = b.get("order", {}) or {}
            real_name = str(o.get("user_name") or "").strip()
        except Exception:
            real_name = ""

        if not real_name:
            try:
                real_name = (
                    update.effective_user.full_name
                    or update.effective_user.first_name
                    or ""
                ).strip()
            except Exception:
                real_name = ""

        uname = ""
        try:
            uname = (update.effective_user.username or "").strip()
        except Exception:
            uname = ""

        name_line = real_name or "—"
        if uname:
            name_line = f"{name_line} @{uname}"

        msg = (
            "📩 <b>رسالة من العميل → الإدارة</b>\n"
            f"🧾 الطلب: {html.escape(order_id)}\n"
            f"👤 العميل: <b>{html.escape(name_line)}</b>\n"
            "────────────────\n"
            f"{html.escape(text)}"
        )

        for aid in ADMIN_IDS:
            try:
                await context.bot.send_message(
                    chat_id=aid,
                    text=msg,
                    parse_mode="HTML",
                    reply_markup=admin_reply_kb(order_id, user_id),
                    disable_web_page_preview=True,
                )
            except Exception as e:
                _swallow(e)

        await _reply_html("تم الاستلام", [f"✅ تم استلام رسالتك وسيتم الرد عليك من <b>{html.escape(PP_SUPPORT_LABEL)}</b>."])
        return
    # ==================================================
    # 20) ملاحظة قبل الدفع (اختيارية)
    # ==================================================
    if stage == STAGE_PREPAY_NOTES:
        ud["notes"] = text
        try:
            update_order_fields(ud.get("order_id", ""), {"notes": text})
        except Exception as e:
            _swallow(e)

        await _reply(build_order_preview(ud), parse_mode="HTML")
        await _reply_html(
            "تم الحفظ",
            ["✅ تم حفظ الملاحظة.", "يمكنك إرسال ملاحظة جديدة للتعديل أو اختيار (تخطي) للمتابعة."],
            kb=prepay_notes_kb(),
        )
        return

    if stage == STAGE_PREPAY_NOTES_TEXT:
        ud["notes"] = text
        try:
            update_order_fields(ud.get("order_id", ""), {"notes": text})
        except Exception as e:
            _swallow(e)

        await _reply(build_order_preview(ud), parse_mode="HTML")
        set_stage(context, user_id, STAGE_PREPAY_NOTES)
        await _reply_html(
            "تم الحفظ",
            ["✅ تم حفظ الملاحظة.", "يمكنك إرسال ملاحظة جديدة للتعديل أو اختيار (تخطي) للمتابعة."],
            kb=prepay_notes_kb(),
        )
        return
    # ==================================================
    # 21) الإيصال إلزامي (رسائل نصية ترفض)
    # ==================================================
    if stage == STAGE_AWAIT_RECEIPT:
        await _reply_html("الإيصال مطلوب", ["⚠️ الرجاء إرسال صورة إيصال الدفع فقط."])
        return
    # ==================================================
    # 22) استلام من الموقع (مدينة + جوال)
    # ==================================================
    if stage == STAGE_ASK_PICKUP_CITY:
        if len(text) < 2:
            await _reply_html("اسم مدينة غير واضح", ["⚠️ اكتب اسم المدينة بشكل أوضح."])
            return
        ud.setdefault("pickup", {})["city"] = text.strip()
        set_stage(context, user_id, STAGE_ASK_PICKUP_PHONE)
        await _reply_html("رقم الجوال", ["📱 اكتب رقم الجوال للاستلام. مثال: <code>05xxxxxxxx</code>"])
        return

    if stage == STAGE_ASK_PICKUP_PHONE:
        phone = re.sub(r"\D+", "", text or "")
        if not (phone.startswith("05") and len(phone) == 10):
            await _reply_html(
                "رقم غير صحيح",
                ["⚠️ اكتب الرقم أرقام فقط ويبدأ بـ 05 ويكون 10 أرقام.", "مثال: <code>05xxxxxxxx</code>"],
            )
            return

        pick = ud.setdefault("pickup", {})
        pick["phone"] = phone

        order_id = (ud.get("order_id") or "").strip()
        if not order_id:
            await _reply_html("تعذر الربط", ["⚠️ تعذر ربط بيانات الاستلام بالطلب. اعد المحاولة من البداية."])
            set_stage(context, user_id, STAGE_NONE)
            return

        details = (
            f"المدينة: {pick.get('city','')}\n"
            f"رقم الجوال: {pick.get('phone','')}\n"
            "سيتم تحديد موقع الاستلام من التاجر عند جاهزية الطلب"
        )

        try:
            update_delivery(order_id, "pickup", details)
        except Exception as e:
            _swallow(e)

        try:
            update_order_fields(order_id, {
                "ship_method": "استلام من الموقع",
                "ship_city": pick.get("city", ""),
                "delivery_details": details,
                "delivery_choice": "استلام من الموقع",
            })
        except Exception as e:
            _swallow(e)

        ud["delivery_choice"] = "استلام من الموقع"
        ud["delivery_details"] = details
        ud["ship_method"] = "استلام من الموقع"
        ud["ship_city"] = pick.get("city", "")

        fee = 0
        try:
            fee = int(float(ud.get("price_sar") or 0))
        except Exception:
            fee = 0
        non_cnt = _safe_int(ud.get("non_consumable_count"))
        cons_cnt = _safe_int(ud.get("consumable_count"))

        # ✅ إذا وضع المنصة مجاني: نثبت الرسوم = 0 قبل المعاينة ليظهر صحيح
        if _is_platform_fee_free_mode():
            ud["price_sar"] = 0
            fee = 0

        # ✅ معاينة إلزامية قبل الدفع/الإرسال
        ud["preview_details"] = details
        ud["preview_mode"] = ("pay_pickup" if fee > 0 else "free_pickup")
        ud["preview_non_cnt"] = non_cnt
        ud["preview_cons_cnt"] = cons_cnt
        set_stage(context, user_id, STAGE_CONFIRM_CLIENT_PREVIEW)

        try:
            safe_details = html.escape(details)
            await _reply(
                build_order_preview(ud)
                + "\n\n<b>📍 تفاصيل الاستلام</b>:\n<pre>"
                + safe_details
                + "</pre>\n\n"
                "<b>🔎 راجع طلبك قبل المتابعة</b>",
                parse_mode="HTML",
                kb=client_preview_kb(),
            )
        except Exception:
            await _reply_html("معاينة الطلب", ["🔎 راجع طلبك قبل المتابعة"], kb=client_preview_kb())

        return
    # ==================================================
    # 23) بيانات الشحن (مدينة -> عنوان مختصر -> جوال)
    # ==================================================
    if stage == STAGE_ASK_SHIP_CITY:
        if len(text) < 2:
            await _reply_html(
                "اسم مدينة غير واضح",
                ["⚠️ اكتب اسم المدينة بشكل أوضح."],
                kb=_flow_nav_kb("ship_city"),
            )
            return
        ud.setdefault("ship", {})["city"] = text.strip()
        set_stage(context, user_id, STAGE_ASK_SHIP_STREET)
        await _reply_html(
            "العنوان المختصر",
            ["🏠 اكتب العنوان الوطني المختصر."],
            kb=_flow_nav_kb("ship_city"),
        )
        return

    if stage == STAGE_ASK_SHIP_STREET:
        if len(text) < 3:
            await _reply_html(
                "عنوان غير واضح",
                ["⚠️ اكتب العنوان المختصر بشكل أوضح."],
                kb=_flow_nav_kb("ship_street"),
            )
            return
        ud.setdefault("ship", {})["short"] = text.strip()
        set_stage(context, user_id, STAGE_ASK_SHIP_PHONE)
        await _reply_html(
            "رقم الاتصال",
            ["📱 اكتب رقم الاتصال. مثال: <code>05xxxxxxxx</code>"],
            kb=_flow_nav_kb("ship_street"),
        )
        return

    if stage == STAGE_ASK_SHIP_PHONE:
        phone = re.sub(r"\D+", "", text or "")
        if not (phone.startswith("05") and len(phone) == 10):
            await _reply_html(
                "رقم غير صحيح",
                ["⚠️ اكتب الرقم أرقام فقط ويبدأ بـ 05 ويكون 10 أرقام.", "مثال: <code>05xxxxxxxx</code>"],
                kb=_flow_nav_kb("ship_phone"),
            )
            return

        ship = ud.setdefault("ship", {})
        ship["phone"] = phone

        order_id = (ud.get("order_id") or "").strip()
        if not order_id:
            await _reply_html(
                "تعذر الربط",
                ["⚠️ تعذر ربط عنوان الشحن بالطلب. اعد المحاولة من البداية."],
                kb=_flow_nav_kb("ship_phone"),
            )
            set_stage(context, user_id, STAGE_NONE)
            return

        details = (
            f"المدينة: {ship.get('city','')}\n"
            f"العنوان الوطني المختصر: {ship.get('short','')}\n"
            f"رقم الاتصال: {ship.get('phone','')}"
        )

        try:
            update_delivery(order_id, "ship", details)
        except Exception as e:
            _swallow(e)

        try:
            update_order_fields(order_id, {
                "ship_method": "شحن",
                "ship_city": ship.get("city", ""),
                "delivery_details": details,
                "delivery_choice": "شحن",
            })
        except Exception as e:
            _swallow(e)

        ud["delivery_choice"] = "شحن"
        ud["delivery_details"] = details
        ud["ship_method"] = "شحن"
        ud["ship_city"] = ship.get("city", "")

        fee = 0
        try:
            fee = int(float(ud.get("price_sar") or 0))
        except Exception:
            fee = 0
        non_cnt = _safe_int(ud.get("non_consumable_count"))
        cons_cnt = _safe_int(ud.get("consumable_count"))

        # ✅ إذا وضع المنصة مجاني: نثبت الرسوم = 0 قبل المعاينة ليظهر صحيح
        if _is_platform_fee_free_mode():
            ud["price_sar"] = 0
            fee = 0

        # ✅ معاينة إلزامية قبل الدفع/الإرسال
        ud["preview_details"] = details
        ud["preview_mode"] = ("pay_ship" if fee > 0 else "free_ship")
        ud["preview_non_cnt"] = non_cnt
        ud["preview_cons_cnt"] = cons_cnt
        set_stage(context, user_id, STAGE_CONFIRM_CLIENT_PREVIEW)

        try:
            safe_details = html.escape(details)
            await _reply(
                build_order_preview(ud)
                + "\n\n<b>📦 تفاصيل الشحن</b>:\n<pre>"
                + safe_details
                + "</pre>\n\n"
                "<b>🔎 راجع طلبك قبل المتابعة</b>",
                parse_mode="HTML",
                kb=client_preview_kb(),
            )
        except Exception:
            await _reply_html("معاينة الطلب", ["🔎 راجع طلبك قبل المتابعة"], kb=client_preview_kb())

        return
    # ==================================================
    # 24) بيانات السيارة
    # ==================================================
    if stage == STAGE_ASK_CAR:
        if len(text) < 3:
            await _reply_html("بيانات غير واضحة", ["⚠️ اكتب اسم السيارة بشكل أوضح."], kb=cancel_only_kb())
            return
        ud["car_name"] = text
        set_stage(context, user_id, STAGE_ASK_MODEL)
        await _reply_html("سنة الموديل", ["📌 اكتب سنة الموديل فقط (4 أرقام). مثال: <code>2023</code>"], kb=cancel_only_kb())
        return

    if stage == STAGE_ASK_MODEL:
        s = (text or "").strip()
        if not re.fullmatch(r"(19|20)\d{2}", s):
            await _reply_html("صيغة غير صحيحة", ["⚠️ اكتب سنة الموديل 4 أرقام فقط. مثال: <code>2023</code>"], kb=cancel_only_kb())
            return

        ud["car_model"] = s
        set_stage(context, user_id, STAGE_ASK_VIN)
        await _reply_html("رقم الهيكل VIN", ["🔎 اكتب رقم الهيكل (17 خانة). مثال: <code>LVVDC12B4RD012345</code>"], kb=cancel_only_kb())
        return

    if stage == STAGE_ASK_VIN:
        vin = (text or "").replace(" ", "").upper()
        if not _looks_like_vin(vin):
            await _reply_html("VIN غير صحيح", ["⚠️ رقم الهيكل يجب أن يكون 17 خانة. مثال: <code>LVVDC12B4RD012345</code>"], kb=cancel_only_kb())
            return

        ud["vin"] = vin
        set_stage(context, user_id, STAGE_ASK_ITEM_NAME)
        await _reply_html("اسم القطعة", ["🧩 اكتب اسم القطعة رقم 1"], kb=cancel_only_kb())
        return
    # ==================================================
    # 25) لو المستخدم كتب اسم قطعة أثناء شاشة التأكيد
    # ==================================================
    if stage == STAGE_CONFIRM_MORE and text:
        items = ud.get("items", []) or []
        if len(items) >= MAX_ITEMS:
            await _reply_html("تم الوصول للحد", [f"⚠️ وصلت للحد الأقصى من القطع ({MAX_ITEMS}).", "اختر (إنهاء وإرسال الطلب)."], kb=more_kb())
            return

        if len(text) < 2:
            await _reply_html("اسم غير واضح", ["⚠️ اكتب اسم القطعة بشكل أوضح."], kb=cancel_only_kb())
            return

        ud.pop("pending_item_idx", None)
        ud.pop("pending_item_name", None)

        ud["pending_item_name"] = text
        set_stage(context, user_id, STAGE_ASK_ITEM_PARTNO)
        await _reply_html("رقم القطعة", ["اكتب رقم القطعة (اختياري) أو اختر (تخطي)."], kb=partno_kb())
        return
    # ==================================================
    # 26) إدخال اسم القطعة
    # ==================================================
    if stage == STAGE_ASK_ITEM_NAME:
        if len(text) < 2:
            await _reply_html("اسم غير واضح", ["⚠️ اكتب اسم القطعة بشكل أوضح."], kb=cancel_only_kb())
            return

        ud["pending_item_name"] = text
        set_stage(context, user_id, STAGE_ASK_ITEM_PARTNO)
        await _reply_html("رقم القطعة", ["اكتب رقم القطعة (اختياري) أو اختر (تخطي)."], kb=partno_kb())
        return
    # ==================================================
    # 27) إدخال رقم القطعة (اختياري)
    # ==================================================
    if stage == STAGE_ASK_ITEM_PARTNO:
        pending_name = _norm(ud.get("pending_item_name", ""))
        if not pending_name:
            set_stage(context, user_id, STAGE_ASK_ITEM_NAME)
            await _reply_html("مطلوب اسم القطعة", ["⚠️ اكتب اسم القطعة أولاً."])
            return

        part_no = (text or "").strip()

        ud.setdefault("items", []).append({
            "name": pending_name,
            "part_no": part_no,
            "photo_file_id": "",
            "created_at_utc": utc_now_iso(),
        })

        ud.pop("pending_item_name", None)
        ud["pending_item_idx"] = len(ud["items"]) - 1

        set_stage(context, user_id, STAGE_ASK_ITEM_PHOTO)
        item_no = len(ud["items"])
        await _reply_html(
            "تمت الإضافة",
            [f"✅ تمت إضافة القطعة رقم <b>{item_no}</b>.", "📷 ارسل صورة (اختياري) أو اكتب اسم القطعة التالية مباشرة."],
            kb=photo_prompt_kb(),
        )
        return
    # ==================================================
    # 28) لو المستخدم كتب نص أثناء مرحلة الصورة: اعتبره اسم قطعة جديدة مباشرة
    # ==================================================
    if stage == STAGE_ASK_ITEM_PHOTO and text:
        ud.pop("pending_item_idx", None)
        ud.pop("pending_item_name", None)

        if len(text) < 2:
            await _reply_html("اسم غير واضح", ["⚠️ اكتب اسم القطعة بشكل أوضح."], kb=cancel_only_kb())
            return

        ud["pending_item_name"] = text
        set_stage(context, user_id, STAGE_ASK_ITEM_PARTNO)
        await _reply_html("رقم القطعة", ["اكتب رقم القطعة (اختياري) أو اختر (تخطي)."], kb=partno_kb())
        return
    # ==================================================
    # 29) رد التاجر للإدارة (رسائل ملف التاجر)
    # ==================================================
    if stage == "trader_reply_admin_msg":
        admin_id = _safe_int(ud.get("reply_to_admin_id"))
        msg = _clean(raw_text)
        if not admin_id or not msg:
            await _reply_html("تنبيه", ["⚠️ اكتب رسالة صحيحة."])
            return

        try:
            await context.bot.send_message(
                chat_id=admin_id,
                text=(
                    "💬 <b>رد من التاجر</b>\n"
                    f"🆔 التاجر: <b>{user_id}</b>\n"
                    f"👤 الاسم: <b>{html.escape(name)}</b>\n"
                    "────────────────\n"
                    f"{html.escape(msg)}"
                ),
                parse_mode="HTML",
                disable_web_page_preview=True,
            )
            await _reply_html("تم الإرسال", ["✅ تم إرسال الرد للإدارة."])
        except Exception:
            await _reply_html("تعذر الإرسال", ["⚠️ تعذر إرسال الرد للإدارة حالياً."])
            return

        ud[STAGE_KEY] = STAGE_NONE
        ud.pop("reply_to_admin_id", None)
        return
    # ==================================================
    # 30) إفتراضي: ما فيه مسار مطابق
    # ==================================================
    await _reply_html(
        "تنبيه",
        [
            "ℹ️ لم أفهم الرسالة.",
            "اكتب <code>PP</code> لبدء طلب جديد، أو اكتب رقم الطلب <code>PP0001</code> لفتح لوحة الطلب، أو اكتب <code>/منصة</code> للتواصل مع فريق الدعم."
        ],
        kb=_support_kb(),
    )
    return
        
async def admin_cancel_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    actor_id = q.from_user.id

    if actor_id not in ADMIN_IDS:
        await _alert(q, "غير مصرح")
        return

    data = q.data or ""
    try:
        _, order_id = data.split("|", 1)
    except Exception:
        await _alert(q, "بيانات غير صحيحة")
        return

    order_id = (order_id or "").strip()
    if not order_id:
        await _alert(q, "رقم طلب غير صحيح")
        return

    update_order_status(order_id, "cancelled")
    update_order_fields(order_id, {
        "cancelled_by_admin_id": actor_id,
        "cancelled_by_admin_name": _user_name(q),
        "cancelled_at_utc": utc_now_iso(),
        # ✅ قفل صارم: يمنع أي عروض جديدة + يوقف إعادة النشر
        "quote_locked": "yes",
        "rebroadcast_disabled": "1",
        "rebroadcast_disabled_at_utc": utc_now_iso(),
        "rebroadcast_disabled_by_id": str(actor_id),
    })

    # ✅ قفل كيبورد رسالة مجموعة التجار (زر عرض السعر يتحول لمقفول)
    try:
        await _lock_team_post_keyboard(context, order_id, reason="🔒 الطلب معلق من الإدارة")
    except Exception as e:
        _swallow(e)
    # اشعار العميل / التاجر / الإدارة (بنفس كيبورد التنقّل الموحد)
    uid = get_order_user_id(order_id)

    # هل توجد مبالغ/دفعات مؤكدة؟ (لتفعيل زر مراسلة التاجر للعميل)
    has_paid = False
    accepted_tid = 0
    try:
        b = get_order_bundle(order_id) or {}
        o = (b.get("order", {}) or {}) if isinstance(b, dict) else {}
        gps = str(o.get("goods_payment_status") or o.get("payment_status") or "").strip().lower()
        has_paid = gps in ("paid", "confirmed")
        accepted_tid = int(o.get("accepted_trader_id") or 0) if str(o.get("accepted_trader_id") or "").strip().isdigit() else 0
    except Exception:
        has_paid = False
        accepted_tid = 0

    # ✅ العميل
    if uid:
        try:
            await context.bot.send_message(
                chat_id=uid,
                text=(
                    "⛔ تم تعليق/إلغاء الطلب من قبل الإدارة\n" f"🧾 رقم الطلب: {order_id}"
                ),
                reply_markup=notice_kb_for(context, uid, order_id, include_chat_trader=bool(has_paid), include_support=True),
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)

    # ✅ التاجر المقبول (إن وجد)
    if accepted_tid:
        try:
            await context.bot.send_message(
                chat_id=int(accepted_tid),
                text=(
                    "⛔ <b>تم تعليق/إلغاء الطلب من قبل الإدارة</b>\n"
                    f"🧾 الطلب: <b>{html.escape(order_id)}</b>\n"
                    "لن يتم قبول أي عروض/تحديثات على هذا الطلب حتى يتم إعادة فتحه."
                ),
                parse_mode="HTML",
                reply_markup=notice_kb_for(context, int(accepted_tid), order_id, include_chat_trader=False, include_support=True),
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)

    # ✅ إشعار الإدارة (كل الأدمن)
    for aid in (ADMIN_IDS or []):
        try:
            await context.bot.send_message(
                chat_id=int(aid),
                text=(
                    "⛔ <b>تم تعليق/إلغاء الطلب</b>\n"
                    f"🧾 الطلب: <b>{html.escape(order_id)}</b>\n"
                    f"👤 بواسطة: <b>{html.escape(_user_name(q))}</b> (<code>{actor_id}</code>)"
                ),
                parse_mode="HTML",
                reply_markup=notice_kb_for(context, int(aid), order_id, include_chat_trader=False, include_support=True),
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)

    # اشعار الفريق
    if TEAM_CHAT_ID:
        try:
            await context.bot.send_message(
                chat_id=TEAM_CHAT_ID,
                text="⛔ تم الغاء الطلب من قبل الادارة\n"
                     f"رقم الطلب: {order_id}"
            )
        except Exception as e:
            _swallow(e)

    await _alert(q, "تم الغاء الطلب")
    try:
        await q.message.reply_text(f"{_user_name(q)}\nتم الغاء الطلب #{order_id}", reply_markup=notice_kb_for(context, actor_id, order_id, include_chat_trader=False, include_support=True))
    except Exception as e:
        _swallow(e)


async def admin_republish_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    actor_id = q.from_user.id

    if actor_id not in ADMIN_IDS:
        await _alert(q, "غير مصرح")
        return

    data = q.data or ""
    try:
        _, order_id = data.split("|", 1)
    except Exception:
        await _alert(q, "بيانات غير صحيحة")
        return

    order_id = (order_id or "").strip()
    if not order_id:
        await _alert(q, "رقم طلب غير صحيح")
        return

    # ✅ إعادة فتح الطلب (فك القفل + تفعيل إعادة النشر) ثم إعادة نشره للمجموعة
    now_iso = utc_now_iso()
    try:
        update_order_fields(order_id, {
            "order_status": "",
            "quote_locked": "no",
            "rebroadcast_disabled": "",
            "rebroadcast_disabled_at_utc": "",
            "rebroadcast_disabled_by_id": "",
            "forwarded_to_team_at_utc": now_iso,
            "last_group_broadcast_at_utc": now_iso,
            "rebroadcast_count": "0",
            "last_noquote_user_ping_at_utc": "",
            "admin_noquote_24h_sent_at_utc": "",
            "republished_by_admin_id": str(actor_id),
            "republished_by_admin_name": _user_name(q),
            "republished_at_utc": now_iso,
        })
    except Exception as e:
        _swallow(e)

    # إعادة نشر للمجموعة
    try:
        await notify_team(context, {"order_id": order_id})
    except Exception as e:
        _swallow(e)
    # إشعار العميل / التاجر / الإدارة (بنفس كيبورد التنقّل الموحد)
    uid = get_order_user_id(order_id)

    has_paid = False
    accepted_tid = 0
    try:
        b = get_order_bundle(order_id) or {}
        o = (b.get("order", {}) or {}) if isinstance(b, dict) else {}
        gps = str(o.get("goods_payment_status") or o.get("payment_status") or "").strip().lower()
        has_paid = gps in ("paid", "confirmed")
        accepted_tid = int(o.get("accepted_trader_id") or 0) if str(o.get("accepted_trader_id") or "").strip().isdigit() else 0
    except Exception:
        has_paid = False
        accepted_tid = 0

    # ✅ العميل
    if uid:
        try:
            await context.bot.send_message(
                chat_id=uid,
                text=(
                    "✅ تم إعادة فتح الطلب وإعادة نشره للتجار\n" f"🧾 رقم الطلب: {order_id}"
                ),
                reply_markup=notice_kb_for(context, uid, order_id, include_chat_trader=bool(has_paid), include_support=True),
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)

    # ✅ التاجر المقبول (إن وجد)
    if accepted_tid:
        try:
            await context.bot.send_message(
                chat_id=int(accepted_tid),
                text=(
                    "✅ <b>تم إعادة فتح الطلب</b>\n"
                    f"🧾 الطلب: <b>{html.escape(order_id)}</b>\n"
                    "الطلب عاد نشط وسيظهر للتجار مجددًا."
                ),
                parse_mode="HTML",
                reply_markup=notice_kb_for(context, int(accepted_tid), order_id, include_chat_trader=False, include_support=True),
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)

    # ✅ إشعار الإدارة (كل الأدمن)
    for aid in (ADMIN_IDS or []):
        try:
            await context.bot.send_message(
                chat_id=int(aid),
                text=(
                    "🔁 <b>إعادة نشر الطلب</b>\n"
                    f"🧾 الطلب: <b>{html.escape(order_id)}</b>\n"
                    f"👤 بواسطة: <b>{html.escape(_user_name(q))}</b> (<code>{actor_id}</code>)"
                ),
                parse_mode="HTML",
                reply_markup=notice_kb_for(context, int(aid), order_id, include_chat_trader=False, include_support=True),
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)


    await _alert(q, "تم إعادة نشر الطلب")
    try:
        await q.message.reply_text(f"{_user_name(q)}\nتم إعادة نشر الطلب #{order_id}", reply_markup=notice_kb_for(context, actor_id, order_id, include_chat_trader=False, include_support=True))
    except Exception as e:
        _swallow(e)

async def track_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    user_id = q.from_user.id

    data = q.data or ""
    try:
        _, order_id = data.split("|", 1)
    except Exception:
        return

    order_id = (order_id or "").strip()
    if not order_id:
        return

    ud = get_ud(context, user_id)
    ud["track_order_id"] = order_id
    set_stage(context, user_id, STAGE_TRACK_ORDER)

    await q.message.reply_text(
        f"{_user_name(q)}\nاكتب رسالتك بخصوص الطلب {order_id}\nسيتم الرد عليك من {PP_SUPPORT_LABEL}",
    )

async def admin_reply_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    actor_id = q.from_user.id
    if actor_id not in ADMIN_IDS:
        await _alert(q, "غير مصرح")
        return

    data = q.data or ""
    # pp_admin_reply|order_id|user_id
    parts = data.split("|")
    if len(parts) != 3:
        return
    _, order_id, uid = parts
    try:
        uid_int = int(uid)
    except Exception:
        return

    ad = context.user_data.setdefault(actor_id, {})
    ad["reply_order_id"] = order_id
    ad["reply_user_id"] = uid_int
    set_stage(context, actor_id, STAGE_ADMIN_REPLY)

    await q.message.reply_text(
        f"👤 {_admin_public_name()}\n✍️ رد كالإدارة للعميل\n{_order_tag_plain(order_id)}\n\nاكتب ردك الآن وسيصل للعميل باسم {_admin_public_name()}",
        reply_markup=admin_reply_done_kb(),
        disable_web_page_preview=True,
    )

async def admin_reply_done_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    actor_id = q.from_user.id
    if actor_id not in ADMIN_IDS:
        return
    ad = context.user_data.setdefault(actor_id, {})
    ad.pop("reply_order_id", None)
    ad.pop("reply_user_id", None)
    set_stage(context, actor_id, STAGE_NONE)
    await q.message.reply_text("تم انهاء وضع الرد")

# === شات مباشر بين العميل والتاجر (Relay) ===
def _assigned_trader_id(order_id: str) -> int:
    try:
        b = get_order_bundle(order_id)
        o = b.get("order", {}) or {}
    except Exception:
        o = {}

    # ✅ الأهم: إذا فيه تاجر مقبول (accepted_trader_id) اعتبره هو المعني دائمًا
    try:
        acc = int(o.get("accepted_trader_id") or 0)
    except Exception:
        acc = 0
    if acc:
        return acc

    # fallback: آخر تاجر قدّم عرض
    try:
        qt = int(o.get("quoted_trader_id") or 0)
    except Exception:
        qt = 0
    return qt

async def chat_trader_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    user_id = q.from_user.id

    data = q.data or ""
    try:
        _, order_id = data.split("|", 1)
    except Exception:
        return
    order_id = (order_id or "").strip()
    if not order_id:
        return

    tid = _assigned_trader_id(order_id)
    if not tid:
        await q.message.reply_text(f"{_user_name(q)}\nلم يتم تحديد تاجر لهذا الطلب بعد")
        return

    ud = get_ud(context, user_id)
    ud["chat_trader_order_id"] = order_id
    set_stage(context, user_id, STAGE_CHAT_TRADER)

    await q.message.reply_text(
        f"{_user_name(q)}\nاكتب رسالتك للتاجر بخصوص الطلب {order_id}",
        reply_markup=chat_nav_kb_for(context, user_id, order_id, "pp_chat_trader_done"),
    )

async def chat_trader_done_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    user_id = q.from_user.id
    ud = get_ud(context, user_id)
    ud.pop("chat_trader_order_id", None)
    set_stage(context, user_id, STAGE_NONE)
    await q.message.reply_text(f"{_user_name(q)}\nتم انهاء المراسلة")
    
async def confirm_received_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    user_id = q.from_user.id

    # ✅ استخراج يوزر العميل لاستخدامه بالإشعار الإداري
    username = (q.from_user.username or "").strip()
    username = f"@{username}" if username else "—"

    data = (q.data or "").strip()
    try:
        _, order_id = data.split("|", 1)
    except Exception:
        return

    order_id = (order_id or "").strip()
    if not order_id:
        return

    b = get_order_bundle(order_id)
    order = b.get("order", {}) or {}

    # ✅ السماح للعميل صاحب الطلب فقط
    client_id = 0
    try:
        client_id = int(order.get("user_id") or 0)
    except Exception:
        client_id = 0

    if not client_id or user_id != client_id:
        await _alert(q, "⛔ غير مصرح")
        return

    # ✅ منع التكرار لو الطلب أصلاً مُستلم/مغلق
    ost = str(order.get("order_status") or "").strip().lower()
    if ost in ("delivered", "closed"):
        await _alert(q, "✅ تم تأكيد الاستلام مسبقًا")
        return

    # ============================
    # ⏳ مؤقت 7 أيام
    # ============================
    if user_id not in ADMIN_IDS:
        now_utc = datetime.now(timezone.utc)

        expires_raw = (order.get("chat_expires_at_utc") or "").strip()
        expires_dt = None
        if expires_raw:
            try:
                expires_dt = datetime.fromisoformat(expires_raw.replace("Z", "+00:00")).astimezone(timezone.utc)
            except Exception:
                expires_dt = None

        if not expires_dt:
            st = str(order.get("order_status") or "").strip().lower()
            gps = str(order.get("goods_payment_status") or "").strip().lower()

            base_raw = (
                (order.get("shipped_at_utc") or "").strip()
                or (order.get("closed_at_utc") or "").strip()
                or (order.get("goods_payment_confirmed_at_utc") or "").strip()
                or (order.get("shipped_at") or "").strip()
                or (order.get("closed_at") or "").strip()
            )
            base_dt = None
            if base_raw:
                try:
                    base_dt = datetime.fromisoformat(base_raw.replace("Z", "+00:00")).astimezone(timezone.utc)
                except Exception:
                    base_dt = None

            if not base_dt and (gps == "confirmed" or st in ("shipped", "delivered", "closed")):
                base_dt = now_utc

            if base_dt:
                expires_dt = base_dt + timedelta(days=7)
                try:
                    update_order_fields(order_id, {"chat_expires_at_utc": expires_dt.isoformat()})
                except Exception as e:
                    _swallow(e)

        if expires_dt and now_utc > expires_dt:
            await _alert(q, "🔒 انتهت مدة المتابعة/التأكيد (7 أيام) لهذا الطلب")
            return

    # ✅ تحديث الحالة إلى (تم الاستلام)
    fields = {
        "order_status": "delivered",
        "closed_at_utc": utc_now_iso(),
        "delivered_confirmed_at_utc": utc_now_iso(),
        "delivered_confirmed_by": str(user_id),
    }
    try:
        update_order_fields(order_id, fields)
    except Exception as e:
        _swallow(e)

    # ✅ اسم التاجر
    accepted_tid = 0
    try:
        accepted_tid = int(order.get("accepted_trader_id") or 0)
    except Exception:
        accepted_tid = 0

    tname = (order.get("accepted_trader_name") or order.get("quoted_trader_name") or "").strip() or "التاجر"
    if accepted_tid and not (order.get("accepted_trader_name") or "").strip():
        try:
            tp = get_trader_profile(int(accepted_tid)) or {}
            tname = (tp.get("display_name") or "").strip() or (tp.get("company_name") or "").strip() or tname
        except Exception as e:
            _swallow(e)

    # ===== حساب المبالغ =====
    goods_total = 0
    shipping_fee = 0
    ship_included = False

    try:
        ship_included = str(order.get("ship_included") or order.get("shipping_included") or "").strip().lower() in (
            "yes", "y", "true", "1", "مشمول", "included"
        )
    except Exception:
        ship_included = False

    try:
        goods_total = int(float(order.get("goods_amount_sar") or order.get("goods_total_sar") or 0))
    except Exception:
        goods_total = 0

    try:
        shipping_fee = int(float(order.get("shipping_fee_sar") or order.get("shipping_fee") or 0))
    except Exception:
        shipping_fee = 0

    grand_total = goods_total + (0 if ship_included else shipping_fee)
    ship_txt = "مشمول" if ship_included else f"{shipping_fee} ر.س"

    # ✅ إشعار العميل
    try:
        await q.message.reply_text(
            f"{_user_name(q)}\n"
            "✅ تم تأكيد استلام القطع بنجاح\n"
            f"🧾 رقم الطلب: {order_id}\n"
            f"👤 التاجر: {tname}\n\n"
            f"🧩 قيمة القطع: {goods_total} ر.س\n"
            f"🚚 الشحن: {ship_txt}\n"
            f"💰 الإجمالي (قطع + شحن): {grand_total} ر.س\n\n"
            "تم إشعار التاجر والإدارة.",
            disable_web_page_preview=True,
        )
    except Exception as e:
        _swallow(e)

    # ✅ إشعار التاجر
    if accepted_tid:
        try:
            await context.bot.send_message(
                chat_id=accepted_tid,
                text=(
                    "✅ تأكيد استلام من العميل\n"
                    f"🧾 رقم الطلب: {order_id}\n"
                    f"👤 العميل: {_user_name(q)}\n\n"
                    f"🧩 قيمة القطع: {goods_total} ر.س\n"
                    f"🚚 الشحن: {ship_txt}\n"
                    f"💰 الإجمالي (قطع + شحن): {grand_total} ر.س\n"
                ),
                reply_markup=trader_received_notice_kb(order_id),
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)

    # ✅ إشعار الإدارة (تم التعديل هنا فقط)
    for aid in ADMIN_IDS:
        try:
            await context.bot.send_message(
                chat_id=int(aid),
                text=(
                    "📌 تأكيد استلام من العميل\n"
                    f"🧾 رقم الطلب: {order_id}\n"
                    f"👤 العميل: {_user_name(q)} ({username})\n"
                    f"🧑‍💼 التاجر: {tname} ({accepted_tid})\n\n"
                    f"🧩 قيمة القطع: {goods_total} ر.س\n"
                    f"🚚 الشحن: {ship_txt}\n"
                    f"💰 الإجمالي (قطع + شحن): {grand_total} ر.س\n"
                    "✅ تم تحديث الحالة إلى: تم الاستلام"
                ),
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)

    await _alert(q, "تم التأكيد ✅")

async def trader_reply_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    actor_id = q.from_user.id

    data = (q.data or "").strip()
    parts = data.split("|")
    if len(parts) != 3:
        return

    _, order_id, uid = parts
    order_id = (order_id or "").strip()

    try:
        uid_int = int(uid)
    except Exception:
        return

    # يسمح فقط للتاجر المسند له الطلب (او الادمن)
    assigned = _assigned_trader_id(order_id)
    if assigned and actor_id not in (assigned, *ADMIN_IDS):
        await _alert(q, "⛔ غير مصرح")
        return

    # تجهيز وضع الرد
    td = context.user_data.setdefault(actor_id, {})
    td["trader_reply_order_id"] = order_id
    td["trader_reply_user_id"] = uid_int
    set_stage(context, actor_id, STAGE_TRADER_REPLY)

    # اسم التاجر (اختياري) من لوحة التاجر
    tp = get_trader_profile(actor_id) or {}
    tname = (tp.get("display_name") or "").strip() or (q.from_user.first_name or q.from_user.full_name or "").strip() or "التاجر"
    tco = (tp.get("company_name") or "").strip()
    tline = f"👤 <b>{html.escape(tname)}</b>" + (f"  •  🏢 <b>{html.escape(tco)}</b>" if tco else "")

    # ملخص سريع
    try:
        b = get_order_bundle(order_id)
        o = b.get("order", {}) or {}

        # قيمة القطع
        amt = _money(o.get("goods_amount_sar") or "")

        # بيانات السيارة
        car = (o.get("car_name") or "").strip()
        model = (o.get("car_model") or "").strip()
    except Exception:
        amt = ""
        car = ""
        model = ""

    # سطر السيارة (مستقل)
    car_line = ""
    if car or model:
        car_line = f"🚗 السيارة: <b>{html.escape((car + ' ' + model).strip())}</b>\n"

    # سطر قيمة القطع (مستقل)
    amt_line = ""
    if amt:
        amt_line = f"💰 قيمة القطع: <b>{html.escape(amt)}</b>\n"

    # ملخص إضافي (اختياري)
    summary = []

    msg = (
        "🟦 <b>مراسلة العميل</b>\n"
        f"{tline}\n"
        f"🧾 رقم الطلب: <b>{html.escape(order_id)}</b>\n"
        f"{car_line}"
        f"{amt_line}"
        + (("—\n" + " • ".join(summary) + "\n") if summary else "")
        + "\n"
        "✍️ اكتب ردّك الآن وسيصل للعميل داخل المنصة.\n"
        "⚠️ لا تكتب بيانات حساسة خارج سياق الطلب."
    )

    await q.message.reply_text(
        msg,
        parse_mode="HTML",
        reply_markup=trader_reply_done_kb(),
        disable_web_page_preview=True,
    )

    data = q.data or ""
    parts = data.split("|")
    if len(parts) != 3:
        await _alert(q, "🟥 بيانات الزر غير مكتملة")
        return

    _, order_id, uid = parts
    order_id = (order_id or "").strip()

    if not order_id:
        await _alert(q, "🟥 رقم الطلب غير صحيح")
        return

    try:
        uid_int = int(uid)
    except Exception:
        await _alert(q, "🟥 تعذر تحديد العميل لهذا الطلب")
        return

    # يسمح فقط للتاجر المسند له الطلب (او الادمن)
    assigned = _assigned_trader_id(order_id)
    if assigned and actor_id not in (assigned, *ADMIN_IDS):
        intruder_name = (q.from_user.first_name or q.from_user.full_name or "").strip() or "هذا التاجر"
        # اسم التاجر المخصص (إن وجد)
        accepted_name = ""
        try:
            b0 = get_order_bundle(order_id)
            o0 = b0.get("order", {}) or {}
            accepted_name = (o0.get("accepted_trader_name") or "").strip()
            if not accepted_name and assigned:
                tp0 = get_trader_profile(int(assigned)) or {}
                accepted_name = (tp0.get("display_name") or "").strip()
        except Exception:
            accepted_name = ""

        who = accepted_name or "تاجر آخر"
        await _alert(q, f"🔒 الطلب معلق\n👤 {intruder_name}\nهذا الطلب مخصص لـ: {who}")
        return

    # اسم التاجر الذي سيظهر للعميل (اختياري)
    tprof = get_trader_profile(actor_id) or {}
    tname = (tprof.get("display_name") or "").strip() or (q.from_user.first_name or q.from_user.full_name or "").strip() or "التاجر"
    tcompany = (tprof.get("company_name") or "").strip()

    # ملخص الطلب للتاجر أثناء الرد
    snap = ""
    try:
        b = get_order_bundle(order_id)
        o = b.get("order", {}) or {}
        items = b.get("items", []) or []

        parts_lines = []
        for i, it in enumerate(items, start=1):
            nm = (it.get("name") or "").strip()
            pn = (it.get("part_no") or it.get("item_part_no") or "").strip()
            if nm and pn:
                parts_lines.append(f"{i}- {nm} (رقم: {pn})")
            elif nm:
                parts_lines.append(f"{i}- {nm}")
        parts_txt = "\n".join(parts_lines) if parts_lines else "لا يوجد"

        amt = (o.get("goods_amount_sar") or "").strip()
        amt_line = f"\n💰 مبلغ العرض: {amt} ريال" if amt else ""

        car_name = (o.get("car_name") or "").strip()
        car_model = (o.get("car_model") or "").strip()
        vin = (o.get("vin") or "").strip()
        notes = (o.get("notes") or "").strip()

        snap = (
            "📦 <b>ملخص الطلب</b>\n"
            f"🧾 <b>رقم الطلب</b>: {order_id}\n"
            f"🚗 <b>السيارة</b>: {car_name or '—'}\n"
            f"📌 <b>الموديل/الفئة</b>: {car_model or '—'}\n"
            f"🔎 <b>VIN</b>: {vin or '—'}\n"
            f"📝 <b>ملاحظات</b>: {notes or 'لا يوجد'}"
            f"{amt_line}\n\n"
            f"🧩 <b>القطع</b>:\n{parts_txt}\n"
        )
    except Exception:
        snap = ""

    ad = context.user_data.setdefault(actor_id, {})
    ad["trader_reply_order_id"] = order_id
    ad["trader_reply_user_id"] = uid_int
    set_stage(context, actor_id, STAGE_TRADER_REPLY)

    trader_line = f"{tname}" + (f" ({tcompany})" if tcompany else "")

async def trader_reply_done_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    actor_id = q.from_user.id
    ad = context.user_data.setdefault(actor_id, {})
    ad.pop("trader_reply_order_id", None)
    ad.pop("trader_reply_user_id", None)
    set_stage(context, actor_id, STAGE_NONE)
    await q.message.reply_text("تم انهاء وضع الرد")

# ===== Trader/Admin panel callbacks =====
async def trader_panel_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    ensure_workbook()  # ✅ مهم جداً: يضمن قراءة/كتابة بيانات التاجر والطلبات من الإكسل
    q = update.callback_query
    await _alert(q, "")
    parts = (q.data or "").split("|")

    if len(parts) < 2:
        return
    # pp_tprof|edit|field  OR pp_tprof|orders|pending
    action = parts[1].strip() if len(parts) >= 2 else ""
    sub = parts[2].strip() if len(parts) >= 3 else ""

    uid = q.from_user.id
    ud = get_ud(context, uid)

    # ==========================================================
    # ✅ توحيد التنبيهات المهمة (Popup) + نداء لطيف باسم التاجر
    # ==========================================================
    def _trader_hi() -> str:
        try:
            tp = get_trader_profile(int(uid or 0)) or {}
        except Exception:
            tp = {}
        nm = (tp.get("display_name") or "").strip()
        if not nm:
            try:
                nm = (q.from_user.first_name or q.from_user.full_name or "").strip()
            except Exception:
                nm = ""
        return nm if nm else "عزيزي التاجر"

    async def _pop(msg: str):
        await _alert(q, f"{_trader_hi()}\n{msg}", force=True)

    async def _toast(msg: str):
        await _alert(q, msg, force=False)

    # ✅ التاجر الموقوف: يسمح بفتح اللوحة فقط، ويمنع الأفعال التنفيذية
    if uid not in ADMIN_IDS and _trader_is_disabled(uid) and action in ("edit", "orders", "sub"):
        # Popup واضح
        await _pop("⛔ لا يمكنك استخدام هذه الخاصية لأن حسابك موقوف")
        try:
            await show_trader_panel(q.message, context, uid)
        except Exception as e:
            _swallow(e)
        return

    
    if action == "paymode":
        # pp_tprof|paymode|choose  OR  pp_tprof|paymode|bank|link
        if sub == "choose":
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("🏦 تحويل بنكي / STC Pay", callback_data="pp_tprof|paymode|bank")],
                [InlineKeyboardButton("🔗 رابط دفع إلكتروني فقط", callback_data="pp_tprof|paymode|link")],
            ])
            await q.message.reply_text(f"{_user_name(q)}\nاختر طريقة الدفع:", reply_markup=kb)
            return

        new_mode = sub.strip().lower()
        if new_mode not in ("bank", "link"):
            await _toast("خيار غير صحيح")
            return

        try:
            upsert_trader_profile(int(uid or 0), {"payment_mode": new_mode})
        except Exception as e:
            _swallow(e)

        try:
            await show_trader_panel(q.message, context, uid)
        except Exception as e:
            _swallow(e)
        return

    if action == "edit":
        field = sub
        labels = {
            "display_name": "اسم التاجر المعروض",
            "company_name": "اسم المتجر",
            "shop_phone": "رقم اتصال المتجر",
            "cr_no": "رقم السجل التجاري",
            "vat_no": "الرقم الضريبي",
            "bank_name": "اسم البنك",
            "iban": "رقم الايبان",
            "stc_pay": "رقم STC Pay",
        }
        title = labels.get(field, "البيان")
        ud["tprof_field"] = field
        set_stage(context, uid, STAGE_TRADER_PROFILE_EDIT)
        await q.message.reply_text(
            f"{_user_name(q)}\n🟦 <b>تعديل {html.escape(title)}</b>\nاكتب القيمة الان وسيتم حفظها مباشرة",
            parse_mode="HTML",
        )
        return

    if action == "orders":
        mode = sub or "pending"
        orders = list_orders_for_trader(uid)
        rows = []
        for o in orders:
            oid = str(o.get("order_id") or "").strip()
            if not oid:
                continue
            gps = str(o.get("goods_payment_status") or "").strip().lower()
            ost = str(o.get("order_status") or "").strip().lower()
            amt = _money(o.get("goods_amount_sar") or "")
            show = False
            if mode == "done":
                show = (gps == "confirmed") or (ost in ("closed", "delivered"))
            else:
                show = not ((gps == "confirmed") or (ost in ("closed", "delivered")))
            if show:
                rows.append(f"• {oid} — {amt or '—'} — {ost or gps or 'pending'}")

        if not rows:
            # ✅ مهم؟ هنا لا تعتبر منع/خطأ، نخليها Toast لطيف
            await _toast("لا توجد طلبات")
            return

        header = "📦 طلباتك المعلقة" if mode != "done" else "✅ طلباتك المنجزة"
        msg = "🟩 <b>%s</b>\n\n%s" % (html.escape(header), html.escape("\n".join(rows)))
        await q.message.reply_text(msg, parse_mode="HTML", disable_web_page_preview=True)
        return

    if action == "sub":
        # 💳 اشتراك شهري للتاجر (99 ر.س)
        month = month_key_utc()
        amount = 99
        ud["sub_month"] = month
        ud["sub_amount_sar"] = amount
        ud["sub_kind"] = "trader_subscription"
        set_stage(context, uid, STAGE_TRADER_SUB_AWAIT_PAY_METHOD)

        try:
            upsert_trader_subscription(uid, month, {
                "amount_sar": amount,
                "payment_status": "awaiting",
            })
        except Exception as e:
            _swallow(e)

        msg = (
            "💳 <b>اشتراك المنصة للتاجر</b>\n"
            f"📅 الشهر: <b>{html.escape(month)}</b>\n"
            f"💰 قيمة الاشتراك: <b>{amount}</b> ريال\n\n"
            "⬇️ اختر طريقة الدفع ثم ارسل إيصال السداد هنا."
        )
        await q.message.reply_text(msg, parse_mode="HTML", reply_markup=pay_method_kb(), disable_web_page_preview=True)
        return

    # default: refresh
    try:
        await show_trader_panel(q.message, context, uid)
    except Exception as e:
        _swallow(e)
    
async def trader_reply_admin_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    tid = q.from_user.id

    parts = (q.data or "").split("|")
    admin_id = 0
    if len(parts) >= 2:
        try:
            admin_id = int(parts[1] or 0)
        except Exception:
            admin_id = 0

    if not admin_id:
        await _alert(q, "بيانات غير مكتملة")
        return

    ud = get_ud(context, tid)
    ud["reply_to_admin_id"] = int(admin_id)
    ud[STAGE_KEY] = "trader_reply_admin_msg"

    msg = (
        "💬 <b>رد للإدارة</b>\n\n"
        "اكتب رسالتك الآن وسيتم إرسالها للإدارة مباشرة."
    )
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("🔒 إغلاق", callback_data="pp_ui_close")],
    ])
    await _admin_edit_or_send(q, msg, kb)  # نفس دالة edit لتفادي التكدس
    
async def _admin_edit_or_send(q, text: str, kb: InlineKeyboardMarkup = None):
    """تحديث نفس رسالة اللوحة قدر الإمكان لتفادي التشوه البصري + عدم الصمت."""
    # ✅ Guard: لو النص نفسه، لا نحاول edit (يتجنب 400 message is not modified)
    try:
        if q and getattr(q, "message", None):
            old_text = (q.message.text or q.message.caption or "")
            if (old_text or "").strip() == (text or "").strip():
                return
    except Exception as e:
        _swallow(e)

    try:
        await q.edit_message_text(
            text=text,
            parse_mode="HTML",
            reply_markup=kb,
            disable_web_page_preview=True,
        )
        return
    except Exception as e:
        # ✅ إذا تيليجرام قال "Message is not modified" لا نسوي fallback برسالة جديدة
        try:
            msg = str(e).lower()
            if "message is not modified" in msg:
                return
        except Exception as e:
            _swallow(e)

    # fallback: رسالة جديدة إذا تعذر التعديل
    try:
        await q.message.reply_text(
            text,
            parse_mode="HTML",
            reply_markup=kb,
            disable_web_page_preview=True,
        )
    except Exception:
        # آخر حل: تنبيه فقط
        try:
            await _alert(q, "تعذر عرض الصفحة")
        except Exception as e:
            _swallow(e)

async def _notify_admins(context: ContextTypes.DEFAULT_TYPE, text: str, exclude_id: int = 0):
    for aid in ADMIN_IDS:
        if exclude_id and aid == exclude_id:
            continue
        try:
            await context.bot.send_message(
                chat_id=aid,
                text=text,
                parse_mode="HTML",
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)

async def admin_panel_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    ensure_workbook()
    q = update.callback_query
    await _alert(q, "")
    uid = q.from_user.id

    # ==========================================================
    # ✅ Popup للرسائل المهمة + نداء لطيف باسم الأدمن (داخل الدالة فقط)
    # ==========================================================
    def _admin_hi() -> str:
        try:
            n = (q.from_user.first_name or q.from_user.full_name or "").strip()
        except Exception:
            n = ""
        # ✅ تعديل بسيط: لو مو أدمن نخليها تحية مناسبة للتاجر
        if uid not in ADMIN_IDS:
            return n if n else "عزيزي التاجر"
        return n if n else "عزيزي الأدمن"

    async def _pop(msg: str):
        # Popup مع تحية لطيفة
        await _alert(q, f"{_admin_hi()}\n{msg}", force=True)

    async def _toast(msg: str):
        # Toast عادي
        await _alert(q, msg, force=False)

    # ✅ نحدد action قبل شرط الأدمن (عشان نستثني tledgerpdf فقط)
    parts = (q.data or "").split("|")
    action = parts[1].strip() if len(parts) >= 2 else "home"

    # ✅ شرط الصلاحية (كما هو) مع استثناء tledgerpdf للتاجر لنفسه فقط
    if uid not in ADMIN_IDS:
        if action != "tledgerpdf":
            await _pop("⛔ غير مصرح")
            return

        # action == tledgerpdf -> مسموح للتاجر فقط إذا tid == uid
        tid_chk = 0
        if len(parts) >= 3:
            try:
                tid_chk = int(parts[2] or 0)
            except Exception:
                tid_chk = 0

        if not tid_chk or int(tid_chk) != int(uid):
            await _pop("⛔ غير مصرح")
            return

    # ==================================================================
    # 🔒 زر إغلاق الطلب (بعد 7 أيام من shipped_at_utc فقط) — بدون حذف الكيبورد
    # ==================================================================
    if action == "pp_order_finish":
        order_id = parts[2].strip() if len(parts) >= 3 else ""
        if not order_id:
            return

        b = get_order_bundle(order_id) or {}
        order = b.get("order", {}) or {}

        shipped_raw = str(order.get("shipped_at_utc") or "").strip()
        if not shipped_raw:
            try:
                await q.answer(
                    "⏳ لا يمكن إغلاق الطلب الآن.\n\n"
                    "🔒 سيتم تفعيل زر الإغلاق بعد الشحن وانتهاء مدة المتابعة (7 أيام).",
                    show_alert=True
                )
            except Exception as e:
                _swallow(e)
            return

        try:
            shipped_dt = datetime.fromisoformat(shipped_raw.replace("Z", "+00:00"))
        except Exception:
            shipped_dt = None

        if not shipped_dt:
            try:
                await q.answer(
                    "⏳ زر الإغلاق غير متاح حاليًا.\n\n"
                    "🔒 سيتم تفعيله تلقائيًا بعد انتهاء مدة المتابعة (7 أيام من الشحن).",
                    show_alert=True
                )
            except Exception as e:
                _swallow(e)
            return

        now_utc = datetime.now(timezone.utc)
        expires_dt = shipped_dt + timedelta(days=7)

        # قبل انتهاء 7 أيام → Alert فقط (لا تعديل ولا حذف للكيبورد)
        if now_utc < expires_dt:
            remaining = expires_dt - now_utc
            days_left = max(remaining.days, 0)
            try:
                await q.answer(
                    "⏳ زر الإغلاق غير متاح الآن.\n\n"
                    "🔒 يتم تفعيل زر الإغلاق بعد انتهاء مدة المتابعة (7 أيام من الشحن).\n"
                    f"🕒 المتبقي تقريبًا: {days_left} يوم.",
                    show_alert=True
                )
            except Exception as e:
                _swallow(e)
            return

        # ✅ بعد انتهاء 7 أيام → يغلق الطلب
        update_order_fields(order_id, {
            "order_status": "closed",
            "closed_at_utc": utc_now_iso(),
        })

        try:
            await q.answer("✅ تم إغلاق الطلب بنجاح.", show_alert=True)
        except Exception as e:
            _swallow(e)

        # ✅ مهم: لا نحذف الكيبورد — فقط نحدثه ليظل ظاهر ويتحوّل حسب trader_status_kb
        try:
            if q.message:
                await q.message.edit_reply_markup(reply_markup=trader_status_kb(order_id))
        except Exception as e:
            _swallow(e)

        return

    async def _go_home():
        # ===== احصائيات القطع (للتجار) =====
        try:
            st0 = compute_admin_financials()
            total_amt0 = _money(st0.get("total_confirmed_amount", 0))
            total_cnt0 = int(st0.get("total_confirmed_count", 0) or 0)
        except Exception:
            total_amt0, total_cnt0 = "", 0

        # ===== رسوم المنصة =====
        try:
            st1 = compute_revenue_breakdown()
            platform_confirmed = _money(st1.get("platform_fees_confirmed", 0))
        except Exception:
            platform_confirmed = ""

        body0 = (
            "🟥 <b>لوحة الادارة</b>\n\n"
            f"🧾 عدد طلبات القطع المؤكدة: <b>{total_cnt0}</b>\n"
            f"💰 إجمالي قيمة القطع المؤكدة (للتجار): <b>{html.escape(total_amt0)}</b>\n"
            f"🏦 رسوم المنصة المؤكدة: <b>{html.escape(platform_confirmed)}</b>\n\n"
            "اختر من الازرار لعرض التفاصيل."
        )

        await _admin_edit_or_send(q, body0, admin_panel_kb())

    async def _admin_show_traders_manage():
        # قائمة التجار -> فتح ملف التاجر + تفعيل/تعطيل مباشر
        try:
            trs = list_traders() or []
        except Exception:
            trs = []

        if not trs:
            msg = "🧑‍💼 <b>إدارة التجار</b>\nلا يوجد تجار مسجلين بعد"
            kb = InlineKeyboardMarkup([[InlineKeyboardButton("↩️ رجوع", callback_data="pp_admin|home")]])
            await _admin_edit_or_send(q, msg, kb)
            return

        # ترتيب: المفعل أولاً ثم الموقوف
        def _en(t):
            try:
                tid0 = int(t.get("trader_id") or 0)
            except Exception:
                tid0 = 0
            if not tid0:
                return 9
            try:
                return 0 if is_trader_enabled(tid0) else 1
            except Exception:
                return 0

        trs = sorted(trs, key=_en)[:40]

        rows = []
        for t in trs:
            try:
                tid = int(t.get("trader_id") or 0)
            except Exception:
                tid = 0
            if not tid:
                continue

            tlabel = _trader_label(tid, "")
            try:
                en_now = is_trader_enabled(tid)
            except Exception:
                en_now = True

            # زر ملف التاجر
            rows.append([InlineKeyboardButton(f"👤 ملف — {tlabel}", callback_data=f"pp_admin|tview|{tid}")])

            # زر تفعيل/تعطيل مباشر
            rows.append([InlineKeyboardButton(
                f"{'⛔ تعطيل' if en_now else '✅ تفعيل'} — {tlabel}",
                callback_data=f"pp_admin|tset|{tid}|{'off' if en_now else 'on'}"
            )])

        msg = "🧑‍💼 <b>إدارة التجار</b>\nاختر تاجر لفتح ملفه أو تفعيل/تعطيل:"
        kb = InlineKeyboardMarkup(rows + [[InlineKeyboardButton("↩️ رجوع", callback_data="pp_admin|home")]])
        await _admin_edit_or_send(q, msg, kb)

    # ===== FINANCE =====
    if action == "finance":
        try:
            st = compute_revenue_breakdown()
        except Exception:
            st = {
                "platform_fees_confirmed": 0,
                "platform_fees_pending": 0,
                "traders_goods_confirmed": 0,
                "shipping_confirmed": 0,
            }

        msg = (
            "💼 <b>التقارير المالية</b>\n\n"
            f"🏦 دخل المنصة (مؤكد): <b>{_money(st.get('platform_fees_confirmed', 0))}</b>\n"
            f"⌛ دخل المنصة (غير مؤكد): <b>{_money(st.get('platform_fees_pending', 0))}</b>\n\n"
            f"🧾 قيمة قطع التجار (مؤكد): <b>{_money(st.get('traders_goods_confirmed', 0))}</b>\n"
            f"🚚 رسوم الشحن (مؤكد): <b>{_money(st.get('shipping_confirmed', 0))}</b>"
        )
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("↩️ رجوع", callback_data="pp_admin|home")]])
        await _admin_edit_or_send(q, msg, kb)
        return

    # ===== FIND ORDER =====
    if action == "find_order":
        try:
            set_stage(context, uid, STAGE_ADMIN_FIND_ORDER)
        except Exception as e:
            _swallow(e)
        msg = "🔎 <b>بحث عن طلب</b>\n\nاكتب رقم الطلب الآن:"
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("↩️ رجوع", callback_data="pp_admin|home")]])
        await _admin_edit_or_send(q, msg, kb)
        return

    # ===== HOME =====
    if action in ("home", ""):
        await _go_home()
        return

    # ===== STATS =====
    if action == "stats":
        try:
            st = compute_admin_financials()
            total_amt = _money(st.get("total_confirmed_amount", 0))
            total_cnt = int(st.get("total_confirmed_count", 0) or 0)
            msg = (
                "📊 <b>احصائيات المنصة</b>\n"
                f"✅ عدد الطلبات المؤكدة: <b>{total_cnt}</b>\n"
                f"💰 اجمالي المبالغ المؤكدة: <b>{html.escape(total_amt)}</b>\n"
            )
        except Exception:
            msg = "🟥 <b>احصائيات المنصة</b>\nتعذر قراءة الاحصائيات"

        kb = InlineKeyboardMarkup([[InlineKeyboardButton("↩️ رجوع", callback_data="pp_admin|home")]])
        await _admin_edit_or_send(q, msg, kb)
        return

    # ===== Toggle Platform Fee Free Mode =====
    if action == "fee_free":
        enabled = _is_platform_fee_free_mode()
        status = "✅ مفعل (رسوم المنصة = 0)" if enabled else "⛔ غير مفعل (الرسوم طبيعية)"
        msg = f"🎁 <b>العرض المجاني لرسوم المنصة</b>\nالحالة: {status}"

        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ تفعيل المجاني", callback_data="pp_admin|fee_free_on")],
            [InlineKeyboardButton("⛔ إلغاء المجاني", callback_data="pp_admin|fee_free_off")],
            [InlineKeyboardButton("🏠 الرئيسية", callback_data="pp_admin|home")],
        ])
        await _admin_edit_or_send(q, msg, kb)
        return

    if action == "fee_free_on":
        _set_platform_fee_free_mode(True)
        await _toast("تم تفعيل العرض المجاني")
        await _admin_edit_or_send(
            q,
            "✅ تم تفعيل العرض المجاني لرسوم المنصة (رسوم المنصة = 0)",
            InlineKeyboardMarkup([[InlineKeyboardButton("↩️ رجوع", callback_data="pp_admin|fee_free")]])
        )
        return

    if action == "fee_free_off":
        _set_platform_fee_free_mode(False)
        await _toast("تم إلغاء العرض المجاني")
        await _admin_edit_or_send(
            q,
            "⛔ تم إلغاء العرض المجاني (رجعت رسوم المنصة كما كانت)",
            InlineKeyboardMarkup([[InlineKeyboardButton("↩️ رجوع", callback_data="pp_admin|fee_free")]])
        )
        return

    # ===== BACKUP / RESTORE (زرّين فقط) =====
    if action == "backup_now":
        await _toast("جاري النسخ...")
        try:
            sent = await _send_backup_excel(context.application, reason="manual_admin")

            # ✅ لا تطبع "فشل" إذا السبب مجرد حدّ أدنى بين النسخ
            if not sent:
                # إذا كانت هناك نسخة قريبة جدًا، نعرض تنبيه مناسب بدل "فشل"
                try:
                    last = str(get_setting("last_backup_at_utc", "") or "").strip()
                except Exception:
                    last = ""
                try:
                    if last:
                        last_dt = datetime.fromisoformat(last.replace("Z", "+00:00"))
                        delta = (datetime.now(timezone.utc) - last_dt).total_seconds()
                        if delta < PP_BACKUP_MIN_SECONDS:
                            mins = max(1, int(delta // 60))
                            await q.message.reply_text(f"ℹ️ تم أخذ نسخة احتياطية قبل {mins} دقيقة. حاول بعد قليل.")
                            return
                except Exception as e:
                    _swallow(e)

                try:
                    await q.message.reply_text("❌ تعذر إرسال النسخة الاحتياطية. تحقق من أن PP_BACKUP_CHAT_ID صحيح وأن البوت لديه صلاحية الإرسال في مجموعة النسخ.")
                except Exception as e:
                    _swallow(e)
                return

            # ✅ ثبّت آخر نسخة
            try:
                if PP_BACKUP_CHAT_ID:
                    await context.bot.pin_chat_message(
                        chat_id=int(str(PP_BACKUP_CHAT_ID).strip()),
                        message_id=sent.message_id,
                        disable_notification=True,
                    )
            except Exception as e:
                _swallow(e)

            try:
                await q.message.reply_text("✅ تم إرسال نسخة احتياطية الآن إلى مجموعة النسخ وتم تثبيتها.")
            except Exception as e:
                _swallow(e)

        except Exception as e:
            try:
                await q.message.reply_text(f"❌ فشل إرسال النسخة الاحتياطية.\n{e}")
            except Exception as e:
                _swallow(e)
        return

    if action == "restore_last_pinned":
        # فقط في الخاص للأدمن
        if q.message and q.message.chat and q.message.chat.type != ChatType.PRIVATE:
            await _pop("⚠️ هذا الخيار يعمل في الخاص فقط")
            return

        await _toast("جاري الاسترجاع...")
        try:
            if not PP_BACKUP_CHAT_ID:
                await _pop("⚠️ PP_BACKUP_CHAT_ID غير مضبوط")
                return

            chat_id = int(str(PP_BACKUP_CHAT_ID).strip())

            chat_obj = await context.bot.get_chat(chat_id)
            pm = getattr(chat_obj, "pinned_message", None)
            doc = getattr(pm, "document", None) if pm else None

            if not doc or not (doc.file_name or "").lower().endswith(".xlsx"):
                try:
                    await q.message.reply_text(
                        "⚠️ لا يوجد ملف إكسل مثبت في مجموعة النسخ.\n"
                        "⚠️ تأكد أن التثبيت تم في (المحادثة العامة) وليس داخل موضوع (Topic).\n"
                        "اضغط (🗂 نسخ احتياطي الآن) أولاً ثم أعد الاسترجاع."
                    )
                except Exception as e:
                    _swallow(e)
                return

            path = _excel_path()
            f = await context.bot.get_file(doc.file_id)
            await f.download_to_drive(custom_path=path)

            try:
                await q.message.reply_text("✅ تم استرجاع آخر نسخة مثبتة وتشغيلها فورًا.")
            except Exception as e:
                _swallow(e)

        except Exception as e:
            try:
                await q.message.reply_text(f"❌ فشل الاسترجاع.\n{e}")
            except Exception as e:
                _swallow(e)
        return

    # ===== MAINT =====
    if action == "maint":
        on = _is_maintenance_mode()
        state = "🟧 مفعّل" if on else "🟩 غير مفعّل"
        msg = (
            "⚙️ <b>وضع الصيانة</b>\n"
            f"الحالة الحالية: <b>{state}</b>\n\n"
            "عند التفعيل سيتم منع استقبال الطلبات الجديدة وتقديم عروض السعر (لغير الادمن)."
        )
        kb = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("🟧 تفعيل الصيانة", callback_data="pp_admin|maint_on"),
                InlineKeyboardButton("🟩 إيقاف الصيانة", callback_data="pp_admin|maint_off"),
            ],
            [InlineKeyboardButton("↩️ رجوع", callback_data="pp_admin|home")],
        ])
        await _admin_edit_or_send(q, msg, kb)
        return

    if action in ("maint_on", "maint_off"):
        on = (action == "maint_on")
        try:
            set_setting("maintenance_mode", "on" if on else "off", actor_id=uid, actor_name=(q.from_user.full_name or ""))
            try:
                append_legal_log(uid, (q.from_user.full_name or ""), "maintenance_mode", f"{'on' if on else 'off'}")
            except Exception as e:
                _swallow(e)

            try:
                await _notify_admins(
                    context,
                    f"⚙️ <b>تحديث وضع الصيانة</b>\n"
                    f"👤 بواسطة: <b>{html.escape(q.from_user.full_name or str(uid))}</b>\n"
                    f"🔁 الحالة: <b>{'مفعّل' if on else 'متوقف'}</b>",
                    exclude_id=uid
                )
            except Exception as e:
                _swallow(e)

            await _toast("تم التحديث ✅")
        except Exception:
            await _pop("❌ فشل التحديث")

        await _go_home()
        return

    # ===== TRADERS STATS =====
    if action == "traders":
        try:
            st = compute_admin_financials()
            per_amt = st.get("per_trader_amount", {}) or {}
            per_cnt = st.get("per_trader_count", {}) or {}
        except Exception:
            per_amt, per_cnt = {}, {}

        if not per_amt:
            msg = "👥 <b>احصائيات التجار</b>\nلا توجد بيانات مؤكدة بعد"
        else:
            lines = []
            for tid, amt in sorted(per_amt.items(), key=lambda x: float(x[1] or 0), reverse=True)[:30]:
                tlabel = _trader_label(int(tid), "")
                lines.append(f"• {tlabel} — {_money(amt)} — {int(per_cnt.get(tid, 0) or 0)} طلب")
            msg = "👥 <b>احصائيات التجار</b>\n\n" + html.escape("\n".join(lines))

        kb = InlineKeyboardMarkup([[InlineKeyboardButton("↩️ رجوع", callback_data="pp_admin|home")]])
        await _admin_edit_or_send(q, msg, kb)
        return

    # ===== ORDERS =====
    if action == "orders":
        try:
            orders = list_orders() or []
        except Exception:
            orders = []

        def _dt(o):
            v = str(o.get("created_at_utc") or "")
            try:
                return datetime.fromisoformat(v.replace("Z", "+00:00"))
            except Exception:
                return datetime.min.replace(tzinfo=timezone.utc)

        orders_sorted = sorted(orders, key=_dt, reverse=True)[:20]

        lines = []
        for o in orders_sorted:
            oid = str(o.get("order_id") or "").strip()
            if not oid:
                continue
            uname = str(o.get("user_name") or "").strip() or "عميل"
            ost = str(o.get("order_status") or o.get("status") or "").strip() or "—"
            amt = _money(o.get("goods_amount_sar") or o.get("quote_amount_sar") or "") or "—"
            lines.append(f"• {oid} — {uname} — {amt} — {ost}")

        msg = "📦 <b>أحدث الطلبات</b>\n\n" + html.escape("\n".join(lines) or "لا يوجد")
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("↩️ رجوع", callback_data="pp_admin|home")]])
        await _admin_edit_or_send(q, msg, kb)
        return

    # ===== SUBS =====
    if action == "subs":
        month = month_key_utc()
        try:
            subs = list_trader_subscriptions(month) or []
        except Exception:
            subs = []

        confirmed = set()
        pending = set()
        for s in subs:
            try:
                tid = int(s.get("trader_id") or 0)
            except Exception:
                tid = 0
            stv = str(s.get("payment_status") or "").strip().lower()
            if stv == "confirmed":
                confirmed.add(tid)
            elif stv in ("pending", "awaiting"):
                pending.add(tid)

        try:
            traders = list_traders() or []
        except Exception:
            traders = []

        overdue_lines = []
        paid_lines = []
        for t in traders:
            try:
                tid = int(t.get("trader_id") or 0)
            except Exception:
                tid = 0
            name = (t.get("display_name") or t.get("company_name") or "").strip() or str(tid)
            if tid in confirmed:
                paid_lines.append(f"🟩 {name} — مدفوع")
            elif tid in pending:
                overdue_lines.append(f"🟨 {name} — قيد التحقق")
            else:
                overdue_lines.append(f"🟥 {name} — متأخر")

        text = (
            f"💳 <b>اشتراكات التجار</b>\n"
            f"📅 الشهر: <b>{html.escape(month)}</b>\n\n"
            f"✅ المدفوع: <b>{len(paid_lines)}</b>\n"
            f"⏳/❌ المتأخر/قيد التحقق: <b>{len(overdue_lines)}</b>\n\n"
            "<b>🟩 المدفوع</b>\n" + (html.escape("\n".join(paid_lines)) if paid_lines else "—") + "\n\n"
            "<b>🟥/🟨 المتأخر / قيد التحقق</b>\n" + (html.escape("\n".join(overdue_lines[:40])) if overdue_lines else "—")
        )

        kb = InlineKeyboardMarkup([[InlineKeyboardButton("↩️ رجوع", callback_data="pp_admin|home")]])
        await _admin_edit_or_send(q, text, kb)
        return

    # ===== TRADERS MANAGE (list -> profiles) =====
    if action == "traders_manage":
        await _admin_show_traders_manage()
        return

    # ===== TRADER PROFILE (tview) =====
    if action == "tview":
        tid = 0
        if len(parts) >= 3:
            try:
                tid = int(parts[2] or 0)
            except Exception:
                tid = 0

        if not tid:
            await _pop("⚠️ بيانات غير مكتملة")
            return

        try:
            prof = get_trader_profile(tid) or {}
        except Exception:
            prof = {}

        tname = (prof.get("display_name") or "").strip()
        tcompany = (prof.get("company_name") or "").strip()
        bank = (prof.get("bank_name") or "").strip()
        iban = (prof.get("iban") or "").strip()
        stc = (prof.get("stc_pay") or "").strip()
        upd = (prof.get("updated_at_utc") or "").strip()

        label = (tname or "التاجر") + (f" ({tcompany})" if tcompany else "")

        try:
            enabled = is_trader_enabled(tid)
        except Exception:
            enabled = True

        enabled_txt = "🟩 مفعل" if enabled else "🟥 موقوف"

        month = month_key_utc()
        sub_status = "—"
        try:
            subs = list_trader_subscriptions(month) or []
            st_map = {}
            for s in subs:
                try:
                    x = int(s.get("trader_id") or 0)
                except Exception:
                    x = 0
                if not x:
                    continue
                st_map[x] = str(s.get("payment_status") or "").strip().lower()
            stv = st_map.get(int(tid), "")
            if stv == "confirmed":
                sub_status = "🟩 مدفوع"
            elif stv in ("pending", "awaiting"):
                sub_status = "🟨 قيد التحقق"
            else:
                sub_status = "🟥 متأخر"
        except Exception as e:
            _swallow(e)

        total_orders = 0
        confirmed_orders = 0
        confirmed_amt = 0.0
        last_order_id = ""
        last_order_ts = ""

        try:
            orders = list_orders_for_trader(tid) or []
        except Exception:
            orders = []

        def _dt(o):
            v = str(o.get("created_at_utc") or "")
            try:
                return datetime.fromisoformat(v.replace("Z", "+00:00"))
            except Exception:
                return datetime.min.replace(tzinfo=timezone.utc)

        if orders:
            orders_sorted = sorted(orders, key=_dt, reverse=True)
            total_orders = len(orders_sorted)

            lo = orders_sorted[0]
            last_order_id = str(lo.get("order_id") or "").strip()
            last_order_ts = str(lo.get("created_at_utc") or "").strip()

            for o in orders_sorted:
                gps = str(o.get("goods_payment_status") or "").strip().lower()
                ost = str(o.get("order_status") or "").strip().lower()
                if gps == "confirmed" or ost in ("closed", "delivered"):
                    confirmed_orders += 1
                    raw = str(o.get("goods_amount_sar") or "").strip()
                    try:
                        confirmed_amt += float(re.sub(r"[^0-9.]+", "", raw) or 0)
                    except Exception as e:
                        _swallow(e)

        msg = (
            "👤 <b>ملف التاجر</b>\n\n"
            f"🆔 ID: <b>{tid}</b>\n"
            f"👤 الاسم: <b>{html.escape(label)}</b>\n"
            f"🔘 الحالة: <b>{enabled_txt}</b>\n"
            f"💳 الاشتراك ({html.escape(month)}): <b>{sub_status}</b>\n\n"
            f"📦 عدد الطلبات: <b>{total_orders}</b>\n"
            f"✅ طلبات مؤكدة (قيمة القطع): <b>{confirmed_orders}</b>\n"
            f"💰 إجمالي مؤكد للتاجر: <b>{html.escape(_money(confirmed_amt))}</b>\n\n"
            f"🏦 البنك: <b>{html.escape(bank or '—')}</b>\n"
            f"🏷️ IBAN: <b>{html.escape(iban or '—')}</b>\n"
            f"📱 STC Pay: <b>{html.escape(stc or '—')}</b>\n"
            f"🕓 آخر تحديث: <b>{html.escape(upd or '—')}</b>\n\n"
            f"🧾 آخر طلب: <b>{html.escape(last_order_id or '—')}</b>\n"
            f"🗓️ وقت آخر طلب: <b>{html.escape(last_order_ts or '—')}</b>"
        )

        kb_rows = [
            [InlineKeyboardButton("💬 مراسلة التاجر", callback_data=f"pp_admin|tmsg|{tid}")],
            [InlineKeyboardButton("📤 كشف معاملات (CSV)", callback_data=f"pp_admin|texport|{tid}")],
            [InlineKeyboardButton("🧾 سجل التاجر (PDF)", callback_data=f"pp_admin|tledgerpdf|{tid}")],
            [InlineKeyboardButton("📦 آخر طلبات التاجر", callback_data=f"pp_admin|torders|{tid}")],
            [InlineKeyboardButton(
                "⛔ تعطيل التاجر" if enabled else "✅ تفعيل التاجر",
                callback_data=f"pp_admin|tset|{tid}|{'off' if enabled else 'on'}"
            )],
            [InlineKeyboardButton("↩️ رجوع لقائمة التجار", callback_data="pp_admin|traders_manage")],
            [InlineKeyboardButton("🏠 الرئيسية", callback_data="pp_admin|home")],
        ]

        await _admin_edit_or_send(q, msg, InlineKeyboardMarkup(kb_rows))
        return

    # ===== TRADER LEDGER PDF (tledgerpdf) =====
    if action == "tledgerpdf":
        tid = 0
        if len(parts) >= 3:
            try:
                tid = int(parts[2] or 0)
            except Exception:
                tid = 0

        if not tid:
            await _pop("⚠️ بيانات غير مكتملة")
            return

        # ✅ يولّد PDF "سجل التاجر" ويرسله لمن طلبه (أدمن أو التاجر نفسه حسب شرط الأعلى)
        try:
            await send_trader_ledger_pdf(context=context, trader_id=tid, admin_chat_id=uid)
            await _pop("✅ تم إرسال سجل التاجر (PDF) في الخاص")
        except Exception as e:
            _swallow(e)
            await _pop("⚠️ تعذر إنشاء سجل التاجر الآن")

        return

    # ===== TRADER ORDERS (torders) =====
    if action == "torders":
        tid = 0
        if len(parts) >= 3:
            try:
                tid = int(parts[2] or 0)
            except Exception:
                tid = 0
        if not tid:
            await _pop("⚠️ بيانات غير مكتملة")
            return

        try:
            orders = list_orders_for_trader(tid) or []
        except Exception:
            orders = []

        def _dt(o):
            v = str(o.get("created_at_utc") or "")
            try:
                return datetime.fromisoformat(v.replace("Z", "+00:00"))
            except Exception:
                return datetime.min.replace(tzinfo=timezone.utc)

        orders_sorted = sorted(orders, key=_dt, reverse=True)[:15]

        try:
            prof = get_trader_profile(tid) or {}
        except Exception:
            prof = {}
        nm = (prof.get("display_name") or "").strip() or str(tid)

        if not orders_sorted:
            msg = f"📦 <b>طلبات التاجر</b>\nالتاجر: <b>{html.escape(nm)}</b>\n\nلا يوجد طلبات بعد"
        else:
            lines = []
            for o in orders_sorted:
                oid = str(o.get("order_id") or "").strip()
                ost = str(o.get("order_status") or "").strip() or "—"
                amt = _money(o.get("goods_amount_sar") or o.get("quote_amount_sar") or "") or "—"
                ts = str(o.get("created_at_utc") or "").strip()
                lines.append(f"• {oid} — {amt} — {ost} — {ts}")
            msg = f"📦 <b>طلبات التاجر</b>\nالتاجر: <b>{html.escape(nm)}</b>\n\n" + html.escape("\n".join(lines))

        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("↩️ رجوع لملف التاجر", callback_data=f"pp_admin|tview|{tid}")],
            [InlineKeyboardButton("↩️ رجوع لقائمة التجار", callback_data="pp_admin|traders_manage")],
        ])
        await _admin_edit_or_send(q, msg, kb)
        return

    # ===== MESSAGE TRADER (tmsg) =====
    if action == "tmsg":
        tid = 0
        if len(parts) >= 3:
            try:
                tid = int(parts[2] or 0)
            except Exception:
                tid = 0
        if not tid:
            await _pop("⚠️ بيانات غير مكتملة")
            return

        ud = get_ud(context, uid)
        ud["admin_msg_to_trader_id"] = int(tid)
        ud[STAGE_KEY] = STAGE_ADMIN_TRADER_MSG

        msg = (
            "💬 <b>مراسلة التاجر</b>\n\n"
            f"🆔 التاجر: <b>{tid}</b>\n"
            "اكتب رسالتك الآن وسيتم إرسالها للتاجر مباشرة."
        )
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("↩️ رجوع لملف التاجر", callback_data=f"pp_admin|tview|{tid}")],
            [InlineKeyboardButton("🏠 الرئيسية", callback_data="pp_admin|home")],
        ])
        await _admin_edit_or_send(q, msg, kb)
        return

    # ===== EXPORT TRADER CSV (texport) =====
    if action == "texport":
        tid = 0
        if len(parts) >= 3:
            try:
                tid = int(parts[2] or 0)
            except Exception:
                tid = 0
        if not tid:
            await _pop("⚠️ بيانات غير مكتملة")
            return

        try:
            import io, csv
        except Exception:
            await _pop("❌ تعذر التصدير")
            return

        try:
            prof = get_trader_profile(tid) or {}
        except Exception:
            prof = {}
        nm = (prof.get("display_name") or "").strip() or str(tid)

        try:
            orders = list_orders_for_trader(tid) or []
        except Exception:
            orders = []

        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow([
            "order_id",
            "created_at_utc",
            "order_status",
            "user_name",
            "goods_amount_sar",
            "goods_payment_status",
            "shipping_fee_sar",
            "payment_status",
            "price_sar",
        ])
        for o in (orders or []):
            w.writerow([
                str(o.get("order_id") or ""),
                str(o.get("created_at_utc") or ""),
                str(o.get("order_status") or o.get("status") or ""),
                str(o.get("user_name") or ""),
                str(o.get("goods_amount_sar") or ""),
                str(o.get("goods_payment_status") or ""),
                str(o.get("shipping_fee_sar") or ""),
                str(o.get("payment_status") or ""),
                str(o.get("price_sar") or ""),
            ])

        data = buf.getvalue()
        b = io.BytesIO(data.encode("utf-8-sig"))
        b.name = f"trader_{tid}_orders_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.csv"

        try:
            await context.bot.send_document(
                chat_id=uid,
                document=b,
                caption=f"📤 كشف معاملات التاجر (CSV)\nالتاجر: {nm}\nID: {tid}",
            )
            try:
                append_legal_log(uid, (q.from_user.full_name or ""), "export_trader_csv", f"trader_id={tid}; rows={len(orders or [])}")
            except Exception as e:
                _swallow(e)
            await _toast("تم إرسال الملف ✅")
        except Exception:
            await _pop("❌ تعذر إرسال الملف")

        await _admin_edit_or_send(
            q,
            f"✅ تم تجهيز كشف التاجر: <b>{html.escape(nm)}</b>\nID: <b>{tid}</b>",
            InlineKeyboardMarkup([
                [InlineKeyboardButton("👤 فتح ملف التاجر", callback_data=f"pp_admin|tview|{tid}")],
                [InlineKeyboardButton("↩️ رجوع لقائمة التجار", callback_data="pp_admin|traders_manage")],
            ])
        )
        return

    # ===== TRADER ENABLE/DISABLE (tset) =====
    if action == "tset":
        tid = 0
        flag = "on"
        if len(parts) >= 3:
            try:
                tid = int(parts[2] or 0)
            except Exception:
                tid = 0
        if len(parts) >= 4:
            flag = (parts[3] or "on").strip().lower()

        if not tid:
            await _pop("⚠️ بيانات غير مكتملة")
            return

        enable = (flag == "on")
        try:
            set_trader_enabled(tid, enable)
            try:
                append_legal_log(uid, (q.from_user.full_name or ""), "trader_enable",
                                 f"trader_id={tid}; enabled={'yes' if enable else 'no'}")
            except Exception as e:
                _swallow(e)

            try:
                await _notify_admins(
                    context,
                    f"🧑‍💼 <b>تحديث حالة تاجر</b>\n"
                    f"👤 بواسطة: <b>{html.escape(q.from_user.full_name or str(uid))}</b>\n"
                    f"🆔 التاجر: <b>{tid}</b>\n"
                    f"🔁 الحالة: <b>{'مفعل' if enable else 'موقوف'}</b>",
                    exclude_id=uid
                )
            except Exception as e:
                _swallow(e)

            await _toast("تم تحديث حالة التاجر ✅")
        except Exception:
            await _pop("❌ فشل تحديث التاجر")

        await _admin_edit_or_send(
            q,
            "✅ تم تحديث حالة التاجر",
            InlineKeyboardMarkup([
                [InlineKeyboardButton("👤 فتح ملف التاجر", callback_data=f"pp_admin|tview|{tid}")],
                [InlineKeyboardButton("↩️ رجوع لقائمة التجار", callback_data="pp_admin|traders_manage")],
                [InlineKeyboardButton("🏠 الرئيسية", callback_data="pp_admin|home")],
            ])
        )
        return

    # ===== LOG =====
    if action == "log":
        try:
            logs = list_legal_log(limit=30) or []
        except Exception:
            logs = []

        if not logs:
            msg = "🧾 <b>سجل الإجراءات</b>\nلا يوجد سجل بعد"
        else:
            lines = []
            for e in logs:
                ts = str(e.get("ts_utc") or "")
                an = str(e.get("actor_name") or "") or str(e.get("actor_id") or "")
                ac = str(e.get("action") or "")
                det = str(e.get("details") or "")
                line = f"• {ts} — {an} — {ac}"
                if det:
                    line += f" — {det}"
                lines.append(line)
            msg = "🧾 <b>سجل الإجراءات (آخر 30)</b>\n\n" + html.escape("\n".join(lines))

        kb = InlineKeyboardMarkup([[InlineKeyboardButton("↩️ رجوع", callback_data="pp_admin|home")]])
        await _admin_edit_or_send(q, msg, kb)
        return

    await _pop("⚠️ أمر غير معروف")


def trader_panel_kb(tid: int) -> InlineKeyboardMarkup:
    # ✅ إخفاء/إظهار أزرار البنك حسب وضع الدفع المحفوظ للتاجر
    try:
        tp = get_trader_profile(int(tid or 0)) or {}
    except Exception:
        tp = {}
    pay_mode = (str(tp.get("payment_mode") or "").strip().lower())
    if pay_mode not in ("link", "bank"):
        pay_mode = "bank"

    rows = [
        [InlineKeyboardButton("🧑‍💼 تعديل اسم التاجر", callback_data="pp_tprof|edit|display_name")],
        [InlineKeyboardButton("🏢 تعديل اسم المتجر", callback_data="pp_tprof|edit|company_name")],
        [InlineKeyboardButton("📞 تعديل رقم اتصال المتجر", callback_data="pp_tprof|edit|shop_phone")],
        [InlineKeyboardButton("🏷️ تعديل رقم السجل التجاري", callback_data="pp_tprof|edit|cr_no")],
        [InlineKeyboardButton("🧾 تعديل الرقم الضريبي", callback_data="pp_tprof|edit|vat_no")],

        # ✅ جديد: اختيار وضع الدفع (تحويل / رابط)
        [InlineKeyboardButton("💳 طريقة الدفع", callback_data="pp_tprof|paymode|choose")],
    ]

    # ✅ وضع التحويل: نعرض أزرار البنك/IBAN/STC كما هي
    if pay_mode != "link":
        rows += [
            [InlineKeyboardButton("🏦 تعديل اسم البنك", callback_data="pp_tprof|edit|bank_name")],
            [InlineKeyboardButton("💳 تعديل رقم الايبان", callback_data="pp_tprof|edit|iban")],
            [InlineKeyboardButton("📱 تعديل رقم STC Pay", callback_data="pp_tprof|edit|stc_pay")],
        ]

    rows += [
        # [InlineKeyboardButton("💳 سداد اشتراك المنصة (99 ر.س)", callback_data="pp_tprof|sub|start")],
        [InlineKeyboardButton("📦 طلباتي المعلقة", callback_data="pp_tprof|orders|pending")],
        [InlineKeyboardButton("✅ طلباتي المنجزة", callback_data="pp_tprof|orders|done")],
        [InlineKeyboardButton("🧾 سجل التاجر (PDF)", callback_data=f"pp_admin|tledgerpdf|{int(tid)}")],
        [InlineKeyboardButton("📩 اتصل بالمنصة", callback_data="pp_support_open")],
        [InlineKeyboardButton("✖️ إغلاق", callback_data="pp_ui_close")],
    ]

    return InlineKeyboardMarkup(rows)

def admin_panel_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("👥 إدارة التجار", callback_data="pp_admin|traders_manage")],
        [InlineKeyboardButton("📊 التقارير المالية", callback_data="pp_admin|finance")],
        [InlineKeyboardButton("🔎 بحث عن طلب", callback_data="pp_admin|find_order")],
        [InlineKeyboardButton("🎁 عرض مجاني لرسوم المنصة", callback_data="pp_admin|fee_free")],

        # ✅ الزران المطلوبان فقط
        [InlineKeyboardButton("🗂 نسخ احتياطي الآن", callback_data="pp_admin|backup_now")],
        [InlineKeyboardButton("♻️ استرجاع آخر نسخة مثبتة", callback_data="pp_admin|restore_last_pinned")],

        [InlineKeyboardButton("⚙️ الصيانة", callback_data="pp_admin|maint")],
        [InlineKeyboardButton("✖️ إغلاق", callback_data="pp_ui_close")],
    ])

async def show_trader_panel(update_or_q, context: ContextTypes.DEFAULT_TYPE, trader_id: int):
    ensure_workbook()  # ✅ مهم: يضمن قراءة/كتابة بيانات التاجر من الاكسل بشكل سليم

    # ✅ سياسة صارمة: لوحة التاجر تعمل بالخاص فقط (لا تعمل بالمجموعة إطلاقًا)
    try:
        chat_type = None
        if hasattr(update_or_q, "message") and update_or_q.message and update_or_q.message.chat:
            chat_type = getattr(update_or_q.message.chat, "type", None)
        else:
            msg = getattr(update_or_q, "message", None)
            chat = getattr(msg, "chat", None) if msg else None
            chat_type = getattr(chat, "type", None) if chat else None

        if chat_type and str(chat_type).lower() != "private":
            # لا نرسل لوحة، ولا نعمل بالمجموعة
            try:
                if hasattr(update_or_q, "answer"):
                    await update_or_q.answer("افتح الخاص لعرض لوحة التاجر", show_alert=True)
                elif hasattr(update_or_q, "message") and update_or_q.message:
                    await update_or_q.message.reply_text("افتح الخاص لعرض لوحة التاجر")
            except Exception as e:
                _swallow(e)
            return
    except Exception as e:
        _swallow(e)

    # ✅ الشرط الأساسي: لازم يكون عضو في مجموعة التجار (عضو عادي يكفي)
    try:
        is_member = await _is_trader_group_member(context, int(trader_id or 0))
    except Exception:
        is_member = False

    # ✅ الأدمن فقط مستثنى من شرط العضوية
    is_admin = int(trader_id or 0) in (ADMIN_IDS or [])

    # ✅ إصلاح الثغرة: العضوية إلزامية للتاجر (حتى لو لديه ملف/سجل سابق)
    # - إذا انحذف من مجموعة التجار => يمنع من لوحة التاجر
    # - الاستثناء الوحيد: الأدمن
    if not is_member and not is_admin:
        try:
            if hasattr(update_or_q, "message") and update_or_q.message:
                await update_or_q.message.reply_text("غير مصرح")
            else:
                try:
                    await update_or_q.answer("غير مصرح", show_alert=True)
                except Exception:
                    try:
                        await update_or_q.edit_message_text("غير مصرح")
                    except Exception as e:
                        _swallow(e)
        except Exception as e:
            _swallow(e)
        return

    # ✅ نحضر ملف التاجر من الشيت
    tp = get_trader_profile(int(trader_id or 0)) or {}
    tp = tp or {}

    # ✅ مهم: ننشئ سجل تاجر جديد إذا كان عضو مجموعة أو أدمن
    if not tp and (is_member or is_admin):
        try:
            upsert_trader_profile(int(trader_id or 0), {"trader_id": int(trader_id or 0)})
            tp = get_trader_profile(int(trader_id or 0)) or {}
        except Exception:
            tp = tp or {}

    dn = (tp.get("display_name") or "").strip() or (
        getattr(update_or_q, "from_user", None).full_name if getattr(update_or_q, "from_user", None) else ""
    ) or "التاجر"
    cn = (tp.get("company_name") or "").strip() or "غير محدد"
    pay_block = _trade_payment_block(tp)

    # ✅ مصدر الحقيقة الوحيد للحالة
    try:
        enabled = is_trader_enabled(int(trader_id or 0))
    except Exception:
        enabled = False  # ✅ آمن: لا نُظهره "مفعل" إذا فشلنا نقرأ الحالة

    status_txt = "مفعل ✅" if enabled else "موقوف ⛔"

    # ✅ بانر واضح للموقوف
    banner = ""
    if not enabled:
        banner = (
            "⛔ <b>تنبيه:</b> حسابك موقوف حاليًا، يمكنك استعراض بياناتك فقط.\n"
            "للاستفسار تواصل مع الإدارة من الزر بالأسفل.\n\n"
        )

    txt = (
        f"{banner}"
        "🟩 <b>لوحة التاجر</b>\n"
        f"🔒 الحالة: <b>{status_txt}</b>\n"
        f"👤 الاسم المعروض: <b>{html.escape(dn)}</b>\n"
        f"🏢 المتجر: <b>{html.escape(cn)}</b>\n"
        f"🧾 بيانات التحويل:\n<pre>{html.escape(pay_block)}</pre>\n"
        "ℹ️ هذه البيانات تحفظ مباشرة داخل ملف المنصة وتبقى حتى بعد اعادة التشغيل.\n"
    )

    # ✅ كيبورد اللوحة:
    # - للتاجر المفعل: كما هو trader_panel_kb()
    # - للتاجر الموقوف: نفس الكيبورد + زر مراسلة الإدارة بالأسفل (كما عندك)
    try:
        kb = trader_panel_kb(int(trader_id or 0))
    except Exception:
        kb = None

    if kb and not enabled:
        try:
            rows = [row[:] for row in (kb.inline_keyboard or [])]
            kb = InlineKeyboardMarkup(rows)
        except Exception as e:
            _swallow(e)

    # ✅ منع تكدس اللوحات: نحدّث نفس رسالة اللوحة إن أمكن، وإلا نحذف القديمة ثم نرسل الجديدة
    try:
        ud = get_ud(context, int(trader_id or 0))
    except Exception:
        ud = {}

    old_chat_id = 0
    old_msg_id = 0
    try:
        old_chat_id = int(ud.get("trader_panel_chat_id") or 0)
    except Exception:
        old_chat_id = 0
    try:
        old_msg_id = int(ud.get("trader_panel_msg_id") or 0)
    except Exception:
        old_msg_id = 0

    # ✅ إذا كان الاستدعاء من Callback داخل الخاص: حدّث نفس الرسالة (الأكثر سلاسة)
    try:
        if not (hasattr(update_or_q, "message") and update_or_q.message):
            # callback_query غالبًا
            msg = getattr(update_or_q, "message", None)
            chat = getattr(msg, "chat", None) if msg else None
            ctype = getattr(chat, "type", None) if chat else None

            if msg and ctype and str(ctype).lower() == "private":
                try:
                    await update_or_q.edit_message_text(
                        txt,
                        parse_mode="HTML",
                        reply_markup=kb,
                        disable_web_page_preview=True,
                    )
                    try:
                        ud["trader_panel_chat_id"] = int(msg.chat_id)
                        ud["trader_panel_msg_id"] = int(msg.message_id)
                    except Exception as e:
                        _swallow(e)
                    return
                except Exception as e:
                    _swallow(e)
    except Exception as e:
        _swallow(e)

    # ✅ غير ذلك: احذف لوحة التاجر السابقة (إن وجدت) ثم أرسل لوحة جديدة بالخاص
    try:
        if old_msg_id:
            await context.bot.delete_message(chat_id=int(old_chat_id or trader_id or 0), message_id=int(old_msg_id))
    except Exception as e:
        _swallow(e)

    sent = None
    try:
        sent = await context.bot.send_message(
            chat_id=int(trader_id or 0),
            text=txt,
            parse_mode="HTML",
            reply_markup=kb,
            disable_web_page_preview=True,
        )
    except Exception:
        sent = None

    if sent:
        try:
            ud["trader_panel_chat_id"] = int(sent.chat_id)
            ud["trader_panel_msg_id"] = int(sent.message_id)
        except Exception as e:
            _swallow(e)
           
async def trader_welcome_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat = update.effective_chat
    if not chat:
        return

    # فقط مجموعة التجار
    if int(chat.id) != int(TRADERS_GROUP_ID or 0):
        return

    new = update.chat_member.new_chat_member
    old = update.chat_member.old_chat_member

    # فقط عند الانضمام الحقيقي
    if old.status in ("left", "kicked") and new.status in ("member", "restricted"):
        user = new.user
        name = user.full_name

        bot_username = context.bot.username
        deeplink = f"https://t.me/{bot_username}?start=trader_{user.id}"

        text = (
            f"👋 مرحبًا {name}\n\n"
            "أنت الآن ضمن *مجموعة التجار* ✅\n\n"
            "🔹 لتفعيل حسابك وبدء تقديم العروض:\n"
            "1️⃣ افتح الخاص مع البوت\n"
            "2️⃣ اضغط الزر بالأسفل\n"
            "3️⃣ أكمل بياناتك (اسم المتجر – IBAN – STC Pay)\n\n"
            "⚠️ لا يمكن تقديم أي عرض قبل إكمال ملف التاجر."
        )

        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("🧑‍💼 فتح لوحة التاجر", url=deeplink)]
        ])

        try:
            await context.bot.send_message(
                chat_id=chat.id,
                text=text,
                reply_markup=kb,
                parse_mode="Markdown",
            )
        except Exception as e:
            _swallow(e)

async def pp25s_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # ✅ فتح لوحة الإدارة بالأمر /pp25s (خاص فقط + ادمن فقط)
    try:
        chat = update.effective_chat
        user = update.effective_user
        if not chat or not user:
            return
        if chat.type != ChatType.PRIVATE:
            return
        user_id = int(user.id)
        if user_id not in ADMIN_IDS:
            try:
                await update.message.reply_text(f"{_user_name(update)}\nغير مصرح")
            except Exception as e:
                _swallow(e)
            return
        set_stage(context, user_id, STAGE_NONE)
        await show_admin_panel(update, context, user_id)
    except Exception:
        try:
            await update.message.reply_text("تعذر فتح لوحة الادارة حاليا")
        except Exception as e:
            _swallow(e)

async def show_admin_panel(update_or_q, context: ContextTypes.DEFAULT_TYPE, admin_id: int):
    """لوحة الادارة: تعديل نفس الرسالة قدر الإمكان لتفادي التشوه البصري + ضمان عمل الرجوع."""
    ensure_workbook()  # مهم لقراءة الاحصائيات والاعدادات

    # ===== احصائيات القطع (للتجار) =====
    try:
        st0 = compute_admin_financials()
        total_amt0 = _money(st0.get("total_confirmed_amount", 0))
        total_cnt0 = int(st0.get("total_confirmed_count", 0) or 0)
    except Exception:
        total_amt0, total_cnt0 = "", 0

    # ===== رسوم المنصة =====
    try:
        st1 = compute_revenue_breakdown()
        platform_confirmed = _money(st1.get("platform_fees_confirmed", 0))
    except Exception:
        platform_confirmed = ""

    body = (
        "🟥 <b>لوحة الادارة</b>\n\n"
        f"🧾 عدد طلبات القطع المؤكدة: <b>{total_cnt0}</b>\n"
        f"💰 إجمالي قيمة القطع المؤكدة (للتجار): <b>{html.escape(total_amt0)}</b>\n"
        f"🏦 رسوم المنصة المؤكدة: <b>{html.escape(platform_confirmed)}</b>\n\n"
        "اختر من الازرار لعرض التفاصيل."
    )

    kb = admin_panel_kb()

    # نحاول نحدد الرسالة التي سنعدلها
    msg = None
    try:
        # CallbackQuery
        if hasattr(update_or_q, "message") and getattr(update_or_q, "message", None):
            msg = update_or_q.message
        # Message
        elif hasattr(update_or_q, "edit_text"):
            msg = update_or_q
        # Update
        elif hasattr(update_or_q, "effective_message") and getattr(update_or_q, "effective_message", None):
            msg = update_or_q.effective_message
    except Exception:
        msg = None

    # edit-in-place اولاً
    if msg is not None:
        try:
            await msg.edit_text(body, parse_mode="HTML", reply_markup=kb, disable_web_page_preview=True)
            return
        except Exception as e:
            _swallow(e)

    # fallback: رسالة جديدة
    try:
        await context.bot.send_message(
            chat_id=admin_id,
            text=body,
            parse_mode="HTML",
            reply_markup=kb,
            disable_web_page_preview=True,
        )
    except Exception as e:
        _swallow(e)
  
async def admin_sub_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await _alert(q, "")
    actor_id = q.from_user.id
    if actor_id not in ADMIN_IDS:
        await _alert(q, "⛔ غير مصرح")
        return

    data = (q.data or "").strip()
    parts = data.split("|")
    if len(parts) < 4:
        return
    act = (parts[1] or "").strip()
    try:
        trader_id = int(parts[2] or 0)
    except Exception:
        trader_id = 0
    month = (parts[3] or "").strip()

    if not trader_id or not month:
        return

    if act == "confirm":
        try:
            upsert_trader_subscription(trader_id, month, {
                "payment_status": "confirmed",
                "paid_at_utc": utc_now_iso(),
            })
        except Exception as e:
            _swallow(e)

        # إشعار التاجر
        try:
            await context.bot.send_message(
                chat_id=trader_id,
                text=(
                    "✅ <b>تم تأكيد اشتراكك في المنصة</b>\n"
                    f"📅 الشهر: <b>{html.escape(month)}</b>\n"
                    "يمكنك الآن تقديم عروض السعر بشكل طبيعي."
                ),
                parse_mode="HTML",
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)

        # إرسال فاتورة اشتراك بسيطة للتاجر + نسخة للإدارة
        try:
            await send_trader_subscription_invoice_pdf(context, trader_id, month, 99)
        except Exception as e:
            _swallow(e)

        try:
            await q.message.reply_text("✅ تم تأكيد الاشتراك")
        except Exception as e:
            _swallow(e)
        return

    if act == "reject":
        try:
            upsert_trader_subscription(trader_id, month, {
                "payment_status": "rejected",
            })
        except Exception as e:
            _swallow(e)

        try:
            await context.bot.send_message(
                chat_id=trader_id,
                text=(
                    "❌ <b>تم رفض إيصال الاشتراك</b>\n"
                    f"📅 الشهر: <b>{html.escape(month)}</b>\n"
                    "يرجى إعادة إرسال إيصال واضح أو التواصل بكتابة: منصة"
                ),
                parse_mode="HTML",
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)

        try:
            await q.message.reply_text("تم الرفض")
        except Exception as e:
            _swallow(e)
        return

## ===== Backup helpers =====
def _excel_path() -> str:
    # pp_excel يعتمد على PP_EXCEL_PATH
    return (os.getenv("PP_EXCEL_PATH") or "pp_data.xlsx").strip() or "pp_data.xlsx"


async def _auto_restore_last_pinned_on_boot(application) -> bool:
    """
    Auto-restore the latest pinned XLSX from PP_BACKUP_CHAT_ID at boot (Render-safe).
    Returns True if restore succeeded.
    """
    try:
        on = (os.getenv("PP_AUTO_RESTORE_ON_BOOT", "1") or "1").strip().lower()
        if on in ("0", "false", "no", "off"):
            log.info("Auto-restore on boot disabled via PP_AUTO_RESTORE_ON_BOOT=%s", on)
            return False
    except Exception:
        pass

    try:
        if not PP_BACKUP_CHAT_ID:
            log.warning("PP_BACKUP_CHAT_ID not set; skip auto-restore")
            return False
        chat_id = int(str(PP_BACKUP_CHAT_ID).strip())
    except Exception as e:
        try:
            log.error("Invalid PP_BACKUP_CHAT_ID: %s", e)
        except Exception:
            pass
        return False

    path = _excel_path()

    try:
        chat_obj = await application.bot.get_chat(chat_id)
        pm = getattr(chat_obj, "pinned_message", None)
        doc = getattr(pm, "document", None) if pm else None

        if not doc or not (doc.file_name or "").lower().endswith(".xlsx"):
            log.warning("No pinned XLSX found in backup chat (or pinned inside Topic).")
            return False

        f = await application.bot.get_file(doc.file_id)
        await f.download_to_drive(custom_path=path)

        # ✅ upgrade headers/sheets without destroying data
        try:
            ensure_workbook(path)
        except Exception as e:
            try:
                log.error("ensure_workbook after restore failed: %s", e)
            except Exception:
                pass

        log.info("✅ Auto-restore OK from pinned message -> %s", path)
        return True

    except Exception as e:
        try:
            log.error("❌ Auto-restore failed: %s", e)
        except Exception:
            pass
        # IMPORTANT: do NOT overwrite an existing file on failure.
        # If file doesn't exist at all, create a fresh workbook so bot can still run.
        try:
            if not os.path.exists(path):
                ensure_workbook(path)
        except Exception:
            pass
        return False


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()

def _riyadh_tz():
    # منطقة السعودية (لا يوجد DST حالياً، لكن نستخدم ZoneInfo لضمان الدقة)
    try:
        from zoneinfo import ZoneInfo
        return ZoneInfo("Asia/Riyadh")
    except Exception:
        return timezone(timedelta(hours=3))

async def _notify_admins(app: Application, text: str) -> None:
    try:
        ids = list(ADMIN_IDS or [])
    except Exception:
        ids = []
    if not ids:
        return
    for aid in ids:
        try:
            await app.bot.send_message(chat_id=int(aid), text=text, disable_web_page_preview=True)
        except Exception as e:
            _swallow(e)

def _should_throttle_notice(key: str, min_seconds: int = 3600) -> bool:
    # True => اسمح بالإشعار الآن. False => اسكت (لمنع السبام).
    try:
        last = str(get_setting(key, "") or "").strip()
    except Exception:
        last = ""
    try:
        if last:
            last_dt = datetime.fromisoformat(last.replace("Z", "+00:00"))
            if (datetime.now(timezone.utc) - last_dt).total_seconds() < min_seconds:
                return False
    except Exception as e:
        _swallow(e)
    try:
        set_setting(key, _utc_now_iso())
    except Exception as e:
        _swallow(e)
    return True

# ===== Backup (send + daily schedule) =====
async def _send_backup_excel(app: Application, reason: str = "scheduled"):
    """
    يرسل ملف الإكسل لمجموعة النسخ.
    ✅ يرجّع رسالة الإرسال (sent Message) عند النجاح ليستفاد منها في pin
    ✅ يرجّع None عند الفشل
    """

    # ✅ 0) اقرأ chat_id من الإعدادات أولًا ثم من env
    try:
        backup_chat_id_raw = (get_setting("backup_chat_id", "") or "").strip()
    except Exception:
        backup_chat_id_raw = ""

    chat_id_raw = backup_chat_id_raw or (str(PP_BACKUP_CHAT_ID).strip() if PP_BACKUP_CHAT_ID else "")
    if not chat_id_raw:
        if _should_throttle_notice("last_backup_warn_no_chat_id_utc", 6 * 3600):
            await _notify_admins(app, "⚠️ النسخ الاحتياطي متوقف: PP_BACKUP_CHAT_ID غير مضبوط.")
        return None

    try:
        chat_id = int(chat_id_raw)
    except Exception:
        if _should_throttle_notice("last_backup_warn_bad_chat_id_utc", 6 * 3600):
            await _notify_admins(app, f"⚠️ chat_id غير صالح: {chat_id_raw}")
        return None

    # 1) تحقق من ملف الإكسل
    path = _excel_path()
    if not os.path.exists(path):
        if _should_throttle_notice("last_backup_warn_no_excel_utc", 6 * 3600):
            await _notify_admins(app, f"⚠️ ملف الإكسل غير موجود:\n{path}")
        return None

    # ✅ تحقق الحجم
    try:
        sz = os.path.getsize(path)
    except Exception:
        sz = -1
    if sz <= 0:
        if _should_throttle_notice("last_backup_warn_excel_empty_utc", 30 * 60):
            await _notify_admins(app, f"❌ ملف الإكسل فارغ/تالف.\nPATH: {path}\nSIZE: {sz}")
        return None

    # ✅ تحقق محتوى الملف: لا نرسل نسخة "فارغة" (مهمة لحماية بياناتك)
    try:
        from openpyxl import load_workbook as _lw
        wb_chk = _lw(path, read_only=True, data_only=True)
        empty_orders = True
        try:
            if "orders" in wb_chk.sheetnames:
                ws_o = wb_chk["orders"]
                empty_orders = (ws_o.max_row or 0) <= 1
        except Exception:
            empty_orders = False
        empty_traders = True
        try:
            if "traders" in wb_chk.sheetnames:
                ws_t = wb_chk["traders"]
                empty_traders = (ws_t.max_row or 0) <= 1
        except Exception:
            empty_traders = False
        try:
            wb_chk.close()
        except Exception:
            pass

        if empty_orders and empty_traders:
            await _notify_admins(app, "⛔ تم إيقاف النسخ الاحتياطي: ملف الإكسل الحالي يبدو فارغًا (لا طلبات ولا تجار).\nتحقق من الاسترجاع/التثبيت قبل أخذ نسخة.")
            return None
    except Exception:
        pass

    # ✅ تم إلغاء منع التكرار بالكامل: سيتم الإرسال بأي وقت

    caption = f"🗂 نسخة احتياطية (PP)\n🕑 UTC: {_utc_now_iso()}\n📌 السبب: {reason}"

    async def _try_send(target_chat_id: int):
        with open(path, "rb") as f:
            return await app.bot.send_document(
                chat_id=target_chat_id,
                document=InputFile(f, filename=os.path.basename(path)),
                caption=caption,
            )

    try:
        sent = await _try_send(chat_id)

        # ✅ حفظ معلومات آخر نسخة
        try:
            if sent and getattr(sent, "document", None):
                set_setting("last_backup_file_id", sent.document.file_id)
                set_setting("last_backup_file_name", sent.document.file_name or os.path.basename(path))
                set_setting("last_backup_at_utc", _utc_now_iso())
        except Exception as e:
            _swallow(e)

        return sent

    except BadRequest as e:
        msg = str(e)

        # ✅ تحصين الهجرة: Group -> Supergroup
        # بعض النسخ ترجع النص وفيه "New chat id: -100..."
        if "migrated to supergroup" in msg.lower() and "new chat id" in msg.lower():
            new_id = None
            try:
                m = re.search(r"new chat id:\s*(-?\d+)", msg, flags=re.IGNORECASE)
                if m:
                    new_id = int(m.group(1))
            except Exception:
                new_id = None

            if new_id:
                # خزّن الـ id الجديد داخليًا
                try:
                    set_setting("backup_chat_id", str(new_id))
                except Exception as e:
                    _swallow(e)

                # أعد الإرسال مرة واحدة بالـ id الجديد
                try:
                    sent2 = await _try_send(new_id)

                    try:
                        if sent2 and getattr(sent2, "document", None):
                            set_setting("last_backup_file_id", sent2.document.file_id)
                            set_setting("last_backup_file_name", sent2.document.file_name or os.path.basename(path))
                            set_setting("last_backup_at_utc", _utc_now_iso())
                    except Exception as e:
                        _swallow(e)

                    if _should_throttle_notice("last_backup_info_migrated_utc", 6 * 3600):
                        await _notify_admins(
                            app,
                            f"✅ تم تحديث مجموعة النسخ تلقائيًا بعد الهجرة.\nOLD: {chat_id}\nNEW: {new_id}"
                        )
                    return sent2
                except Exception as e2:
                    if _should_throttle_notice("last_backup_warn_send_error_utc", 30 * 60):
                        await _notify_admins(app, f"❌ فشل الإرسال حتى بعد تحديث chat_id.\n{e2}")
                    return None

        # خطأ BadRequest عادي
        if _should_throttle_notice("last_backup_warn_send_error_utc", 30 * 60):
            await _notify_admins(app, f"❌ BadRequest أثناء إرسال النسخة:\n{msg}\nCHAT_ID: {chat_id}\nPATH: {path}\nSIZE: {sz}")
        return None

    except Exception as e:
        if _should_throttle_notice("last_backup_warn_send_error_utc", 30 * 60):
            await _notify_admins(app, f"❌ خطأ أثناء إرسال النسخة:\n{e}")
        return None

def _seconds_until_next_riyadh_1am() -> int:
    tz = _riyadh_tz()
    now = datetime.now(tz)
    target = now.replace(hour=1, minute=0, second=0, microsecond=0)
    if now >= target:
        target = target + timedelta(days=1)
    delta = (target - now).total_seconds()
    return max(1, int(delta))

async def _backup_loop(app: Application) -> None:
    # ✅ جدولة يومية الساعة 1:00 صباحاً بتوقيت السعودية
    while True:
        try:
            await asyncio.sleep(_seconds_until_next_riyadh_1am())
            await _send_backup_excel(app, reason="daily_01:00_riyadh")
        except Exception as e:
            if _should_throttle_notice("last_backup_warn_loop_error_utc", 30 * 60):
                await _notify_admins(app, f"❌ خطأ داخل جدولة النسخ الاحتياطي:\n{e}")
            await asyncio.sleep(60)

def _start_backup_tasks(application: Application) -> None:
    # تشغيل واحد فقط
    try:
        if application.bot_data.get("_backup_loop_started"):
            return
        application.bot_data["_backup_loop_started"] = True
    except Exception as e:
        _swallow(e)
    try:
        asyncio.create_task(_backup_loop(application))
    except Exception as e:
        _swallow(e)

# ===== Restore helpers (Group + Private) =====
def _restore_is_admin(uid: int) -> bool:
    try:
        return int(uid) in (ADMIN_IDS or [])
    except Exception:
        return False

async def _restore_excel_from_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    استرجاع يدوي بإرسال ملف .xlsx:
    - في مجموعة النسخ: يُقبل فقط من الأدمن
    - في الخاص: يُقبل فقط من الأدمن (بدون كلمة مرور)
    """
    msg = update.message
    if not msg or not msg.document:
        return

    doc = msg.document
    if not (doc.file_name or "").lower().endswith(".xlsx"):
        return

    chat = msg.chat
    uid = (msg.from_user.id if msg.from_user else 0)

    # السماح فقط للأدمن
    if not _restore_is_admin(uid):
        return

    # 1) مجموعة النسخ المحددة
    if PP_BACKUP_CHAT_ID and chat.id == PP_BACKUP_CHAT_ID:
        # ✅ سجّل كآخر نسخة (اختياري)
        try:
            set_setting("last_backup_file_id", doc.file_id)
            set_setting("last_backup_file_name", doc.file_name or "pp_data.xlsx")
            set_setting("last_backup_at_utc", _utc_now_iso())
        except Exception as e:
            _swallow(e)

    # 2) الخاص مع البوت
    elif chat.type == ChatType.PRIVATE:
        pass

    else:
        return

    path = _excel_path()
    try:
        f = await doc.get_file()
        await f.download_to_drive(custom_path=path)
        await msg.reply_text("✅ تم استرجاع قاعدة البيانات بنجاح وتم تشغيلها فورًا.")
    except Exception:
        try:
            await msg.reply_text("❌ فشل استرجاع النسخة، حاول مرة أخرى.")
        except Exception as e:
            _swallow(e)

async def pprs_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # نفس restorepass لكن باسم أسهل
    # يسمح: /pprs أو /pprs T194525i
    return await restorepass_cmd(update, context)


async def pp_rb_stop_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    try:
        await q.answer()
    except Exception as e:
        _swallow(e)

    data = q.data or ""
    try:
        _, order_id = data.split("|", 1)
    except Exception:
        return

    order_id = (order_id or "").strip()
    if not order_id:
        return

    uid = 0
    try:
        uid = int(q.from_user.id)
    except Exception:
        uid = 0
    if not uid:
        return

    # ✅ الأدمن ما يستخدم هذا الزر نهائياً (هذا زر عميل فقط)
    is_admin = str(uid) in set([str(x) for x in (ADMIN_IDS or [])])
    if is_admin:
        try:
            await q.answer("هذا الزر خاص بصاحب الطلب فقط", show_alert=True)
        except Exception:
            pass
        return

    # ✅ السماح لصاحب الطلب فقط
    owner_id = 0
    try:
        owner_id = int(get_order_user_id(order_id) or 0)
    except Exception:
        owner_id = 0

    if owner_id != uid:
        try:
            await q.answer("غير مصرح", show_alert=True)
        except Exception:
            pass
        return

    # =========================================================
    # ✅ إلغاء فعلي من العميل + قفل عروض + إيقاف إعادة نشر
    # =========================================================
    try:
        update_order_status(order_id, "cancelled")
    except Exception as e:
        _swallow(e)

    try:
        update_order_fields(order_id, {
            "cancelled_by_client_id": str(uid),
            "cancelled_by_client_name": (q.from_user.full_name or q.from_user.first_name or "العميل").strip(),
            "cancelled_at_utc": utc_now_iso(),
            "quote_locked": "yes",
            "rebroadcast_disabled": "1",
            "rebroadcast_disabled_at_utc": utc_now_iso(),
            "rebroadcast_disabled_by_id": str(uid),
        })
    except Exception as e:
        _swallow(e)

    # ✅ قفل بوست مجموعة التجار (زر عرض السعر يصير مقفول بسبب العميل)
    try:
        await _lock_team_post_keyboard(context, order_id, reason="🔒 ملغي من قبل العميل")
    except Exception as e:
        _swallow(e)

    # تعطيل الزر في نفس الرسالة (إن أمكن)
    try:
        await q.edit_message_reply_markup(reply_markup=None)
    except Exception as e:
        _swallow(e)

    # نحاول نجيب التاجر المقبول (إن وجد) + هل فيه دفعات مؤكدة (لزر مراسلة التاجر)
    accepted_tid = 0
    has_paid = False
    try:
        b = get_order_bundle(order_id) or {}
        o = (b.get("order", {}) or {}) if isinstance(b, dict) else {}
        if str(o.get("accepted_trader_id") or "").strip().isdigit():
            accepted_tid = int(o.get("accepted_trader_id") or 0)
        gps = str(o.get("goods_payment_status") or o.get("payment_status") or "").strip().lower()
        has_paid = gps in ("paid", "confirmed")
    except Exception:
        accepted_tid = 0
        has_paid = False

    # ✅ إشعار العميل (مع الكيبورد الموحد)
    try:
        await context.bot.send_message(
            chat_id=uid,
            text=(
                "⛔ تم إلغاء الطلب بناءً على طلبك ولن يستقبل عروض جديدة\n"
                f"🧾 رقم الطلب: {_order_id_link_html(order_id)}"
            ),
            parse_mode="HTML",
            reply_markup=notice_kb_for(context, uid, order_id, include_chat_trader=bool(has_paid), include_support=True),
            disable_web_page_preview=True,
        )
    except Exception as e:
        _swallow(e)

    # ✅ إشعار الإدارة
    if ADMIN_IDS:
        for aid in (ADMIN_IDS or []):
            try:
                await context.bot.send_message(
                    chat_id=int(aid),
                    text=(
                        "⛔ <b>تم إلغاء الطلب من قبل العميل</b>\n"
                        f"🧾 رقم الطلب: <b>{html.escape(order_id)}</b>\n"
                        f"👤 العميل: <code>{uid}</code>"
                    ),
                    parse_mode="HTML",
                    reply_markup=notice_kb_for(context, int(aid), order_id, include_chat_trader=False, include_support=True),
                    disable_web_page_preview=True,
                )
            except Exception as e:
                _swallow(e)

    # ✅ إشعار التاجر المقبول (إن وجد)
    if accepted_tid:
        try:
            await context.bot.send_message(
                chat_id=int(accepted_tid),
                text=(
                    "⛔ <b>تم إلغاء الطلب من قبل العميل</b>\n"
                    f"🧾 رقم الطلب: <b>{html.escape(order_id)}</b>"
                ),
                parse_mode="HTML",
                reply_markup=notice_kb_for(context, int(accepted_tid), order_id, include_chat_trader=False, include_support=True),
                disable_web_page_preview=True,
            )
        except Exception as e:
            _swallow(e)

def build_app():
    if not BOT_TOKEN:
        raise SystemExit("PP_BOT_TOKEN غير موجود في .env")
    if not TEAM_CHAT_ID:
        raise SystemExit("PARTS_TEAM_CHAT_ID غير صحيح او غير موجود في .env")

    # تحقق اجباري للدفع اليدوي
    missing = []
    if not PP_IBAN:
        missing.append("PP_IBAN")
    if not PP_STC_PAY:
        missing.append("PP_STC_PAY")
    if not PP_BANK_NAME:
        missing.append("PP_BANK_NAME")
    if not PP_BENEFICIARY:
        missing.append("PP_BENEFICIARY")
    if missing:
        raise SystemExit("متغيرات ناقصة في .env: " + ", ".join(missing))

    # ✅ نتركها (تضمن وجود الشيتات والهيدرز)
    ensure_workbook()

    # ✅ تحسين اتصال تيليجرام لتفادي TimedOut تحت الضغط
    try:
        request = HTTPXRequest(
            connect_timeout=20.0,
            read_timeout=40.0,
            write_timeout=40.0,
            pool_timeout=20.0,
            connection_pool_size=64,
        )
        app = Application.builder().token(BOT_TOKEN).request(request).build()
    except Exception:
        app = Application.builder().token(BOT_TOKEN).build()

    # 🟢 [HANDLER] Error Handler
    app.add_error_handler(globals().get('on_error') or _on_error_fallback)

    # 🟢 [HANDLER] Commands
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("chatid", chatid))

    # 🟢 [HANDLER] Admin Panel (PP25S) بطريقتين
    app.add_handler(CommandHandler("pp25s", pp25s_cmd))
    app.add_handler(MessageHandler(filters.Regex(r"(?i)^pp25s$"), pp25s_cmd))  # بدون /

    # 🟢 [HANDLER] Support (/منصة)
    app.add_handler(MessageHandler(filters.Regex(r"^/منصة(?:@\w+)?(?:\s|$)"), support_cmd))
    # ✅ أوامر إنجليزية صالحة
    app.add_handler(CommandHandler(["h", "help"], support_cmd))

    app.add_handler(ChatMemberHandler(trader_welcome_cb, ChatMemberHandler.CHAT_MEMBER))
    app.add_handler(ChatJoinRequestHandler(traders_join_request_cb))

    # 🟢 [HANDLER] UI / Cancel / Close
    app.add_handler(CallbackQueryHandler(cancel_cb, pattern=r"^pp_cancel$"))
    app.add_handler(CallbackQueryHandler(pp_rb_stop_cb, pattern=r"^pp_rb_stop\|"))
    app.add_handler(CallbackQueryHandler(start_new_order_cb, pattern=r"^pp_start_new$"))

    # 🟢 [HANDLER] Join Portal
    app.add_handler(CallbackQueryHandler(pp_join_done_cb, pattern=r"^pp_join_done$"))
    app.add_handler(CallbackQueryHandler(pp_join_chat_cb, pattern=r"^pp_join_chat\|"))
    app.add_handler(CallbackQueryHandler(applicant_chat_admin_cb, pattern=r"^pp_applicant_chat_admin\|"))
    app.add_handler(CallbackQueryHandler(applicant_chat_admin_done_cb, pattern=r"^pp_applicant_chat_admin_done$"))
    app.add_handler(CallbackQueryHandler(pp_join_admin_action_cb, pattern=r"^pp_join_(appr|decl)\|"))
    app.add_handler(CallbackQueryHandler(back_cb, pattern=r"^pp_back\|"))
    app.add_handler(CallbackQueryHandler(ui_close_cb, pattern="^pp_ui_close$"))
    app.add_handler(CallbackQueryHandler(ui_locked_cb, pattern="^pp_ui_locked$"))
    app.add_handler(CallbackQueryHandler(client_confirm_preview_cb, pattern="^pp_client_confirm_preview$"))

    # 🟢 [HANDLER] Support Close / Admin Reply / Done + Open from Button
    app.add_handler(CallbackQueryHandler(support_close_cb, pattern="^pp_support_close$"))
    app.add_handler(CallbackQueryHandler(pp_support_reply_cb, pattern=r"^pp_support_reply\|"))
    app.add_handler(CallbackQueryHandler(support_admin_done_cb, pattern="^pp_support_admin_done$"))
    app.add_handler(CallbackQueryHandler(support_open_cb, pattern="^pp_support_open$"))

    app.add_handler(CallbackQueryHandler(more_yes_cb, pattern="^pp_more_yes$"))
    app.add_handler(CallbackQueryHandler(more_no_cb, pattern="^pp_more_no$"))

    app.add_handler(CallbackQueryHandler(skip_photo_cb, pattern="^pp_skip_photo$"))
    app.add_handler(CallbackQueryHandler(partno_skip_cb, pattern="^pp_partno_skip$"))
    app.add_handler(CallbackQueryHandler(skip_notes_cb, pattern="^pp_skip_notes$"))
    app.add_handler(CallbackQueryHandler(prepay_notes_skip_cb, pattern="^pp_prepay_notes_skip$"))

    app.add_handler(CallbackQueryHandler(ppq_cb, pattern=r"^ppq"))
    app.add_handler(CallbackQueryHandler(track_cb, pattern=r"^pp_track\|"))
    app.add_handler(CallbackQueryHandler(open_order_cb, pattern=r"^pp_open_order\|"))
    app.add_handler(CallbackQueryHandler(order_legal_cb, pattern=r"^pp_order_legal\|"))
    app.add_handler(CallbackQueryHandler(admin_reply_cb, pattern=r"^pp_admin_reply\|"))
    app.add_handler(CallbackQueryHandler(admin_reply_done_cb, pattern="^pp_admin_reply_done$"))

    app.add_handler(CallbackQueryHandler(chat_trader_cb, pattern=r"^pp_chat_trader\|"))
    app.add_handler(CallbackQueryHandler(chat_trader_done_cb, pattern="^pp_chat_trader_done$"))
    app.add_handler(CallbackQueryHandler(trader_reply_cb, pattern=r"^pp_trader_reply\|"))
    app.add_handler(CallbackQueryHandler(trader_reply_done_cb, pattern=r"^pp_trader_reply_done$"))
    app.add_handler(CallbackQueryHandler(trader_reply_admin_cb, pattern=r"^pp_trader_reply_admin\|"))

    app.add_handler(CallbackQueryHandler(copy_iban_cb, pattern="^pp_copy_iban$"))
    app.add_handler(CallbackQueryHandler(copy_beneficiary_cb, pattern="^pp_copy_beneficiary$"))
    app.add_handler(CallbackQueryHandler(copy_stc_cb, pattern="^pp_copy_stc$"))

    app.add_handler(CallbackQueryHandler(pay_bank_cb, pattern="^pp_pay_bank$"))
    app.add_handler(CallbackQueryHandler(pay_stc_cb, pattern="^pp_pay_stc$"))
    app.add_handler(CallbackQueryHandler(pay_link_cb, pattern="^pp_pay_link$"))
    app.add_handler(CallbackQueryHandler(quote_ok_cb, pattern=r"^pp_quote_ok\|"))
    app.add_handler(CallbackQueryHandler(quote_no_cb, pattern=r"^pp_quote_no\|"))

    app.add_handler(CallbackQueryHandler(admin_paylink_cb, pattern=r"^pp_admin_paylink\|"))
    app.add_handler(CallbackQueryHandler(admin_sub_cb, pattern=r"^pp_admin_sub\|"))

    app.add_handler(CallbackQueryHandler(goods_pay_bank_cb, pattern=r"^pp_goods_pay_bank\|"))
    app.add_handler(CallbackQueryHandler(goods_pay_stc_cb, pattern=r"^pp_goods_pay_stc\|"))
    app.add_handler(CallbackQueryHandler(trader_status_cb, pattern=r"^pp_trader_status\|"))
    app.add_handler(CallbackQueryHandler(tsu_skip_tracking_cb, pattern=r"^pp_tsu_skip_tracking\|"))
    app.add_handler(CallbackQueryHandler(pp_chat_end_cb, pattern=r"^pp_chat_end(\|.*)?$"))
    app.add_handler(CallbackQueryHandler(chat_open_cb, pattern=r"^pp_chat_open\|"))
    app.add_handler(CallbackQueryHandler(order_finish_cb, pattern=r"^pp_order_finish\|"))

    app.add_handler(CallbackQueryHandler(admin_chat_client_cb, pattern=r"^pp_admin_chat_client\|"))
    app.add_handler(CallbackQueryHandler(admin_chat_trader_cb, pattern=r"^pp_admin_chat_trader\|"))
    app.add_handler(CallbackQueryHandler(admin_chat_done_cb, pattern=r"^pp_admin_chat_done$"))
    app.add_handler(CallbackQueryHandler(trader_chat_admin_cb, pattern=r"^pp_trader_chat_admin\|"))
    app.add_handler(CallbackQueryHandler(trader_chat_admin_done_cb, pattern=r"^pp_trader_chat_admin_done$"))

    app.add_handler(CallbackQueryHandler(admin_panel_cb, pattern=r"^pp_admin\|"))
    app.add_handler(CallbackQueryHandler(trader_panel_cb, pattern=r"^pp_tprof\|"))

    app.add_handler(CallbackQueryHandler(goods_pay_link_cb, pattern=r"^pp_goods_pay_link\|"))

    app.add_handler(CallbackQueryHandler(delivery_ship_cb, pattern="^pp_delivery_ship$"))
    app.add_handler(CallbackQueryHandler(delivery_pickup_cb, pattern="^pp_delivery_pickup$"))

    app.add_handler(CallbackQueryHandler(admin_forward_cb, pattern=r"^pp_admin_forward\|"))
    app.add_handler(CallbackQueryHandler(admin_cancel_cb, pattern=r"^pp_admin_cancel\|"))
    app.add_handler(CallbackQueryHandler(admin_republish_cb, pattern=r"^pp_admin_republish\|"))

    app.add_handler(CallbackQueryHandler(team_cb, pattern=r"^(pp_team_|pp_trader_open\|)"))

    app.add_handler(CallbackQueryHandler(confirm_received_cb, pattern=r"^pp_confirm_received\|"))

    # 🟢 [HANDLER] Media Router
    app.add_handler(MessageHandler(
        filters.PHOTO | filters.VIDEO | filters.Document.ALL | filters.VOICE | filters.AUDIO | filters.VIDEO_NOTE,
        media_router
    ))

    # 🟢 [HANDLER] Text Router
    app.add_handler(MessageHandler(filters.ChatType.PRIVATE & filters.TEXT & ~filters.COMMAND, text_handler))

    # 🟢 [HANDLER] Jobs
    try:
        if app.job_queue:
            app.job_queue.run_repeating(
                _rebroadcast_noquote_orders_job,
                interval=86400,  # ✅ كل 24 ساعة
                first=600,       # ✅ أول فحص بعد 10 دقائق من الإقلاع
                name="rebroadcast_noquote_orders",
            )
    except Exception as e:
        try:
            log.warning(f"JobQueue warning: {e}")
        except Exception as e:
            _swallow(e)

    # 🟢 [HANDLER] Restore DB (Admin only) — (اختياري) قبول إرسال ملف xlsx يدويًا
    # ملاحظة: زر "استرجاع آخر نسخة مثبتة" هو الأساس في السيناريو الأخير
    try:
        app.add_handler(MessageHandler(filters.Document.ALL, _restore_excel_from_message), group=0)
    except Exception as e:
        try:
            log.error(f"Restore handlers error: {e}")
        except Exception as e:
            _swallow(e)

    # 🟢 [TASK] Backup (daily 01:00 Riyadh) — الباك اب اليدوي من لوحة الأدمن هو الأساس قبل أي Restart
    async def _post_init(application):
        try:
            _start_backup_tasks(application)
        except Exception as e:
            try:
                log.error(f"Backup tasks start error: {e}")
            except Exception as e:
                _swallow(e)

    try:
        app.post_init = _post_init
    except Exception as e:
        try:
            log.error(f"post_init attach error: {e}")
        except Exception as e:
            _swallow(e)

    return app

class _HealthHandler(BaseHTTPRequestHandler):
    def do_HEAD(self):
        self.send_response(200)
        self.send_header("Content-type", "text/plain; charset=utf-8")
        self.end_headers()

    def do_GET(self):
        # ✅ health check (Render/UptimeRobot) - يرد OK على أي مسار (/ أو /healthz ...)
        self.send_response(200)
        self.send_header("Content-type", "text/plain; charset=utf-8")
        self.end_headers()
        self.wfile.write(b"OK")

    def log_message(self, format, *args):
        # تقليل الضجيج في اللوق (Render / UptimeRobot)
        return

def _start_health_server():
    port = int(os.getenv("PORT", "10000"))
    HTTPServer(("0.0.0.0", port), _HealthHandler).serve_forever()

async def _run_webhook():
    """تشغيل PP كبوت Webhook داخل Web Service (Render) مع /healthz."""
    application = build_app()

    # ✅ Auto-restore آخر نسخة مثبتة قبل أي قراءة/كتابة للإكسل
    try:
        await _auto_restore_last_pinned_on_boot(application)
    except Exception as e:
        try:
            log.error(f"Auto-restore on boot error: {e}")
        except Exception:
            pass

    # ✅ تجهيز التطبيق (بدون run_polling)
    await application.initialize()
    await application.start()

    # ✅ تشغيل مهام النسخ (post_init لا يعمل هنا لأننا لا نستخدم run_webhook)
    try:
        _start_backup_tasks(application)
    except Exception as e:
        try:
            log.error(f"Backup tasks start error (webhook): {e}")
        except Exception as e:
            _swallow(e)

    # ✅ إعداد Webhook URL
    base_url = (os.getenv("WEBHOOK_BASE_URL") or os.getenv("RENDER_EXTERNAL_URL") or "").strip().rstrip("/")
    if not base_url:
        log.warning("⚠️ WEBHOOK_BASE_URL غير موجود. شغّل WEBHOOK_BASE_URL=https://<your-service>.onrender.com")
    webhook_path = (os.getenv("WEBHOOK_PATH") or "webhook").strip().lstrip("/")
    webhook_url = f"{base_url}/{webhook_path}" if base_url else ""

    # ✅ حماية اختيارية (تيليجرام يدعم Secret Token)
    secret = (os.getenv("WEBHOOK_SECRET_TOKEN") or "").strip() or None

    # ✅ فعّل الويبهوك في تيليجرام (إذا توفر base_url)
    if webhook_url:
        try:
            await application.bot.set_webhook(
                url=webhook_url,
                secret_token=secret,
                drop_pending_updates=True,
            )
            log.info("✅ Webhook set: %s", webhook_url)
        except Exception as e:
            log.exception("❌ فشل ضبط Webhook: %s", e)

    # ✅ Web server (aiohttp) على PORT
    port = int(os.getenv("PORT", "10000"))

    async def healthz(_request):
        return web.Response(text="OK")

    async def webhook_handler(request: web.Request):
        # ✅ تحقق secret_token (اختياري)
        if secret:
            got = request.headers.get("X-Telegram-Bot-Api-Secret-Token")
            if got != secret:
                return web.Response(status=403, text="forbidden")

        try:
            data = await request.json()
        except Exception:
            return web.Response(status=400, text="bad json")

        try:
            update = Update.de_json(data, application.bot)
            await application.process_update(update)
        except Exception as e:
            try:
                log.exception("WEBHOOK_PROCESS_UPDATE_FAILED: %s", e)
            except Exception as e:
                _swallow(e)

        return web.Response(text="OK")

    web_app = web.Application()
    web_app.router.add_get("/", healthz)
    web_app.router.add_get("/healthz", healthz)
    web_app.router.add_post(f"/{webhook_path}", webhook_handler)

    runner = web.AppRunner(web_app)
    await runner.setup()
    site = web.TCPSite(runner, "0.0.0.0", port)
    await site.start()

    log.info("🌐 Web server running on 0.0.0.0:%s | webhook_path=/%s", port, webhook_path)

    # ✅ ابقِ العملية شغالة
    try:
        while True:
            await asyncio.sleep(3600)
    finally:
        try:
            await runner.cleanup()
        except Exception as e:
            _swallow(e)
        try:
            await application.stop()
            await application.shutdown()
        except Exception as e:
            _swallow(e)

def main():
    # ✅ اختر الوضع عبر متغير البيئة:
    enabled = (os.getenv("PP_WEBHOOK_ENABLED") or "").strip().lower() in ("1", "true", "yes", "on")

    if enabled:
        log.info("PP Bot is running (webhook)...")
        asyncio.run(_run_webhook())
        return

    # Polling mode (قديم): نبدأ health server على PORT ثم polling
    try:
        threading.Thread(target=_start_health_server, daemon=True).start()
    except Exception as e:
        _swallow(e)

    app = build_app()
    log.info("PP Bot is running (polling)...")
    app.run_polling(
        allowed_updates=Update.ALL_TYPES,
        drop_pending_updates=True,
    )


if __name__ == "__main__":
    main()