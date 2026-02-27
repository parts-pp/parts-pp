import os
import uuid
import json
import re
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook, load_workbook

import threading
import tempfile
import time

_EXCEL_LOCK = threading.RLock()

def _atomic_save(wb, path: str) -> None:
    """
    Save XLSX safely:
    - write to temp file
    - os.replace() to swap atomically
    يقلل جدا احتمال تلف الملف عند انقطاع/تزامن.
    """
    d = os.path.dirname(path) or "."
    fd, tmp_path = tempfile.mkstemp(prefix="pp_", suffix=".xlsx", dir=d)
    os.close(fd)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, path)
    finally:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass
        


def _save_and_close(wb, path: str) -> None:
    """Atomic save then close workbook (thread-safe)."""
    with _EXCEL_LOCK:
        _atomic_save(wb, path)
        try:
            wb.close()
        except Exception:
            pass
DEFAULT_PATH = os.getenv("PP_EXCEL_PATH") or "pp_data.xlsx"

SHEET_ORDERS = "orders"
SHEET_ITEMS = "items"
SHEET_EVENTS = "events"
SHEET_MESSAGES = "messages"
SHEET_TRADERS = "traders"
SHEET_SETTINGS = "settings"
SHEET_LEGAL_LOG = "legal_log"

SHEET_TRADER_SUBS = "trader_subs"

ORDERS_HEADERS = [
    "order_id",
    "user_id",
    "user_name",
    "car_name",
    "car_model",
    "vin",
    "notes",
    "items_count",
    "price_sar",
    "status",
    "payment_method",
    "payment_status",
    "receipt_file_id",
    "payment_confirmed_at_utc",
    "delivery_choice",
    "delivery_details",
    "assigned_admin_id",
    "assigned_admin_name",
    "assigned_at_utc",
    "forwarded_to_team_at_utc",
    "forwarded_by_admin_id",
    "forwarded_by_admin_name",
    "created_at_utc",
    "order_status",
    "quote_status",
    "quote_locked",
    "goods_amount_sar",
    "quote_item_prices",
    "parts_type",
    "ship_method",
    "ship_carrier",
    "ship_eta",
    "ship_included",
    "availability_days",
    "quoted_trader_id",
    "quoted_trader_name",
    "accepted_trader_id",
    "accepted_trader_name",
    "accepted_at_utc",
    "accepted_trader_notified",
    "goods_payment_status",
    "goods_payment_method",
    "goods_payment_link",
    "goods_receipt_file_id",
    "goods_payment_confirmed_at_utc",
    "closed_at_utc",
    "team_message_id",
    "ship_phone",
    "shipping_fee_sar",
    "seller_invoice_kind",
    "shop_invoice_at",
    "shop_invoice_mime",
    "shop_invoice_file_id",
    "seller_invoice_at",
    "seller_invoice_mime",
    "seller_invoice_file_id",
    "trader_invoice_pre_no",
    "trader_invoice_pre_at",
    "trader_invoice_ship_no",
    "trader_invoice_ship_at",
    "ship_city",
    "pickup_city",
    "pickup_location",
    "invoice_pre_no",
    "invoice_pre_at",
    "invoice_ship_no",
    "invoice_ship_at",
    "invoice_platform_pre_pdf_sent",
    "invoice_platform_ship_pdf_sent",
    "invoice_trader_pre_pdf_sent",
    "invoice_trader_ship_pdf_sent",
    "shipped_at_utc",
    "delivered_at_utc",
    "chat_expires_at_utc",
    "last_group_broadcast_at_utc",
    "last_noquote_user_ping_at_utc",
    "last_unpaid_invoice_user_ping_at_utc",
    "last_paid_trader_ping_at_utc",
    "admin_noquote_24h_sent_at_utc",
    "shipping_tracking",
    "shipping_at",
    "rebroadcast_count",
    "rebroadcast_disabled",
    "rebroadcast_disabled_at_utc",
    "rebroadcast_disabled_by_id",
]

ITEMS_HEADERS = [
    "order_id",
    "idx",
    "item_name",
    "photo_file_id",
    "created_at_utc",

    "item_part_no",
]

EVENTS_HEADERS = [
    "event_id",
    "order_id",
    "event_type",
    "actor_role",
    "actor_id",
    "actor_name",
    "payload_json",
    "created_at_utc",
]

TRADERS_HEADERS = [
    "trader_id",
    "display_name",
    "company_name",
    "shop_phone",
    "cr_no",
    "vat_no",
    "payment_mode",
    "bank_name",
    "iban",
    "stc_pay",
    "joined_at_utc",
    "is_enabled",
    "updated_at_utc",
]

SETTINGS_HEADERS = [
    "key",
    "value",
    "updated_at_utc",
]

LEGAL_LOG_HEADERS = [
    "ts_utc",
    "actor_id",
    "actor_name",
    "action",
    "details",
]


TRADER_SUBS_HEADERS = [
    "sub_id",
    "trader_id",
    "month",  # YYYY-MM
    "amount_sar",
    "payment_method",
    "payment_status",  # awaiting/confirmed
    "receipt_file_id",
    "paid_at_utc",
    "created_at_utc",
]

MESSAGES_HEADERS = [
    "msg_id",
    "order_id",
    "sender_role",
    "sender_id",
    "sender_name",
    "receiver_role",
    "receiver_id",
    "text",
    "file_id",
    "created_at_utc",
]
def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()

# ✅ public alias (used by subscriptions / external callers)
def utc_now_iso() -> str:
    return _utc_now_iso()

def ensure_workbook(path: str = DEFAULT_PATH) -> None:
    """Create workbook + sheets if missing and append missing headers safely."""
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_ORDERS
        ws.append(ORDERS_HEADERS)

        ws2 = wb.create_sheet(SHEET_ITEMS)
        ws2.append(ITEMS_HEADERS)

        ws3 = wb.create_sheet(SHEET_EVENTS)
        ws3.append(EVENTS_HEADERS)

        ws4 = wb.create_sheet(SHEET_MESSAGES)
        ws4.append(MESSAGES_HEADERS)

        # ✅ مهم: إنشاء شيت التجار من البداية
        ws5 = wb.create_sheet(SHEET_TRADERS)
        ws5.append(TRADERS_HEADERS)

        ws6 = wb.create_sheet(SHEET_SETTINGS)
        ws6.append(SETTINGS_HEADERS)

        ws7 = wb.create_sheet(SHEET_LEGAL_LOG)
        ws7.append(LEGAL_LOG_HEADERS)

        ws8 = wb.create_sheet(SHEET_TRADER_SUBS)
        ws8.append(TRADER_SUBS_HEADERS)

        with _EXCEL_LOCK:
            _atomic_save(wb, path)
            wb.close()
        return

    wb = load_workbook(path)

    def _ensure_sheet(name: str, headers: list[str]):
        if name not in wb.sheetnames:
            wsx = wb.create_sheet(name)
            wsx.append(headers)
            return wsx
        wsx = wb[name]
        if wsx.max_row < 1:
            wsx.append(headers)

        # ensure headers exist (append missing at end)
        existing = []
        for col in range(1, wsx.max_column + 1):
            v = wsx.cell(row=1, column=col).value
            existing.append(str(v).strip() if v is not None else "")

        for hname in headers:
            if hname not in existing:
                wsx.cell(row=1, column=wsx.max_column + 1).value = hname
                existing.append(hname)

        return wsx

    _ensure_sheet(SHEET_ORDERS, ORDERS_HEADERS)
    _ensure_sheet(SHEET_ITEMS, ITEMS_HEADERS)
    _ensure_sheet(SHEET_EVENTS, EVENTS_HEADERS)
    _ensure_sheet(SHEET_MESSAGES, MESSAGES_HEADERS)
    _ensure_sheet(SHEET_TRADERS, TRADERS_HEADERS)
    _ensure_sheet(SHEET_SETTINGS, SETTINGS_HEADERS)
    _ensure_sheet(SHEET_LEGAL_LOG, LEGAL_LOG_HEADERS)
    _ensure_sheet(SHEET_TRADER_SUBS, TRADER_SUBS_HEADERS)
    _save_and_close(wb, path)
def _load(path: str = DEFAULT_PATH):
    """
    تحميل ملف الإكسل بثبات:
    - قفل (Lock) لمنع تداخل القراءة/الحفظ
    - Retry عند أخطاء التزامن
    """
    ensure_workbook(path)
    last_err = None

    for _ in range(10):
        try:
            with _EXCEL_LOCK:
                return load_workbook(path)
        except Exception as e:
            last_err = e
            try:
                time.sleep(0.25)
            except Exception:
                pass

    raise last_err

def _header_index(ws) -> dict:
    idx = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is not None:
            idx[str(v).strip()] = col
    return idx

def add_order(order: dict, path: str = DEFAULT_PATH) -> None:
    wb = _load(path)
    ws = wb[SHEET_ORDERS]
    h = _header_index(ws)
    row = ws.max_row + 1

    for k, v in order.items():
        if k in h:
            ws.cell(row=row, column=h[k]).value = v

    if "created_at_utc" not in order and "created_at_utc" in h:
        ws.cell(row=row, column=h["created_at_utc"]).value = _utc_now_iso()

    with _EXCEL_LOCK:
        _atomic_save(wb, path)
        wb.close()

def add_items(order_id: str, items: list[dict], path: str = DEFAULT_PATH) -> None:
    wb = _load(path)
    ws = wb[SHEET_ITEMS]
    for i, it in enumerate(items, start=1):
        ws.append([
            str(order_id),
            i,
            it.get("name", ""),
            it.get("photo_file_id", ""),
            it.get("created_at_utc", _utc_now_iso()),
            it.get("part_no", ""),
        ])
    with _EXCEL_LOCK:
        _atomic_save(wb, path)
        wb.close()

def update_order_fields(order_id: str, fields: dict, path: str = DEFAULT_PATH) -> None:
    wb = _load(path)
    ws = wb[SHEET_ORDERS]
    h = _header_index(ws)

    order_col = h.get("order_id", 1)

    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=order_col).value) == str(order_id):
            for k, v in fields.items():
                if k in h:
                    ws.cell(row=row, column=h[k]).value = v
            break

    with _EXCEL_LOCK:
        _atomic_save(wb, path)
        wb.close()
def update_order_status(order_id: str, status: str, path: str = DEFAULT_PATH) -> None:
    update_order_fields(order_id, {"status": status}, path=path)

def update_order_payment(order_id: str, payment_status: str, receipt_file_id: str = "", confirmed_at_utc: str = "", path: str = DEFAULT_PATH) -> None:
    fields = {"payment_status": payment_status}
    if receipt_file_id is not None:
        fields["receipt_file_id"] = receipt_file_id
    if confirmed_at_utc is not None:
        fields["payment_confirmed_at_utc"] = confirmed_at_utc
    update_order_fields(order_id, fields, path=path)

def update_delivery(order_id: str, choice: str, details: str, path: str = DEFAULT_PATH) -> None:
    update_order_fields(order_id, {"delivery_choice": choice, "delivery_details": details}, path=path)


def get_order_assignment(order_id: str, path: str = DEFAULT_PATH) -> dict:
    """Return assignment info: {'assigned_admin_id': int|None, 'assigned_admin_name': str, 'assigned_at_utc': str}."""
    wb = _load(path)
    ws = wb[SHEET_ORDERS]
    h = _header_index(ws)
    order_col = h.get("order_id", 1)
    out = {"assigned_admin_id": None, "assigned_admin_name": "", "assigned_at_utc": ""}
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=order_col).value) == str(order_id):
            aid = ws.cell(row=row, column=h.get("assigned_admin_id", 0)).value if h.get("assigned_admin_id") else None
            aname = ws.cell(row=row, column=h.get("assigned_admin_name", 0)).value if h.get("assigned_admin_name") else ""
            at = ws.cell(row=row, column=h.get("assigned_at_utc", 0)).value if h.get("assigned_at_utc") else ""
            try:
                out["assigned_admin_id"] = int(aid) if aid not in (None, "") else None
            except Exception:
                out["assigned_admin_id"] = None
            out["assigned_admin_name"] = str(aname or "")
            out["assigned_at_utc"] = str(at or "")
            break
    return out

def get_order_user_id(order_id: str, path: str = DEFAULT_PATH) -> int | None:
    try:
        wb = _load(path)
        ws = wb[SHEET_ORDERS]
        h = _header_index(ws)
        order_col = h.get("order_id", 1)
        uid_col = h.get("user_id", 2)
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row=row, column=order_col).value) == str(order_id):
                uid = ws.cell(row=row, column=uid_col).value
                try:
                    return int(uid)
                except Exception:
                    return None
    except Exception:
        return None
    return None


def get_order_bundle(order_id: str, path: str = DEFAULT_PATH) -> dict:
    """Return dict with order fields and items list for a given order_id."""
    wb = _load(path)
    try:

        # --- order ---
        ws = wb[SHEET_ORDERS]
        h = _header_index(ws)
        order_col = h.get("order_id", 1)

        order = None
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row=row, column=order_col).value) == str(order_id):
                order = {}
                for k, col in h.items():
                    order[k] = ws.cell(row=row, column=col).value
                break

        if order is None:
            return {"order": {}, "items": []}

        # --- items ---
        ws_items = wb[SHEET_ITEMS]
        hi = _header_index(ws_items)

        def _get(r: int, key: str, fallback_col: int):
            col = hi.get(key, fallback_col)
            try:
                return ws_items.cell(row=r, column=col).value
            except Exception:
                return None

        def _looks_like_file_id(v: str) -> bool:
            if not v:
                return False
            s = str(v)
            # common telegram file_id patterns often start with "Ag" or contain many underscores/dashes
            return s.startswith("Ag") or ("_" in s and len(s) > 20)

        def _looks_like_iso(v: str) -> bool:
            if not v:
                return False
            s = str(v)
            return bool(re.match(r"^\d{4}-\d{2}-\d{2}T", s))

        items = []
        for r in range(2, ws_items.max_row + 1):
            if str(ws_items.cell(row=r, column=hi.get("order_id", 1)).value) != str(order_id):
                continue

            item_no = _get(r, "idx", 2) or _get(r, "item_no", 2)
            name = _get(r, "item_name", 3) or _get(r, "name", 3) or ""
            photo = _get(r, "photo_file_id", 4) or ""
            created = _get(r, "created_at_utc", 5) or _utc_now_iso()
            part_no = _get(r, "item_part_no", 6) or _get(r, "part_no", 6) or ""

            # Legacy misalignment recovery (older rows written in wrong column order)
            # Old write order was: [order_id, idx, name, part_no, photo_file_id, created_at]
            if photo and not _looks_like_file_id(str(photo)) and (not part_no or _looks_like_iso(str(part_no))):
                # 'photo' is probably part_no
                part_no = str(photo)
                photo = str(created) if _looks_like_file_id(str(created)) else ""
                created = str(part_no) if _looks_like_iso(str(part_no)) else created

            if created and not _looks_like_iso(str(created)) and _looks_like_iso(str(part_no)):
                # created and part_no swapped
                created, part_no = part_no, created

            items.append({
                "order_id": str(order_id),
                "item_no": item_no,
                "name": str(name or ""),
                "part_no": str(part_no or "").strip(),
                "photo_file_id": str(photo or "").strip(),
                "created_at_utc": str(created or _utc_now_iso()),
            })

        return {"order": order, "items": items}


    finally:
        try:
            wb.close()
        except Exception:
            pass
def mark_order_forwarded(order_id: str, admin_id: int, admin_name: str, at_utc: str = "", path: str = DEFAULT_PATH) -> None:
    if not at_utc:
        at_utc = _utc_now_iso()
    update_order_fields(order_id, {
        "forwarded_to_team_at_utc": at_utc,
        "forwarded_by_admin_id": str(admin_id),
        "forwarded_by_admin_name": admin_name,
    }, path=path)


def generate_order_id(prefix: str = "PP", path: str = DEFAULT_PATH) -> str:
    """
    Generate human-friendly order id: PP-YYMMDD-####

    ✅ Global sequence (never resets daily):
    - date part changes (YYMMDD)
    - numeric part is global, monotonically increasing across days
    - minimum width is 4 digits, and it automatically expands (10000 => 5 digits, etc.)

    Storage:
    - uses settings sheet key: order_seq
    """
    ensure_workbook(path)

    # YYMMDD (UTC) for readability only (does not affect uniqueness)
    today = datetime.now(timezone.utc).strftime("%y%m%d")

    SEQ_KEY = "order_seq"

    with _EXCEL_LOCK:
        wb = _load(path)
        ws = wb[SHEET_SETTINGS]

        # ensure headers
        headers = _sheet_headers(ws)
        if "key" not in headers or "value" not in headers:
            try:
                ws.delete_rows(1, ws.max_row)
            except Exception:
                pass
            ws.append(["key", "value", "updated_at_utc"])

        # find seq row
        col_key = 1
        col_val = 2
        found_row = None
        for r in range(2, ws.max_row + 1):
            try:
                k = str(ws.cell(row=r, column=col_key).value or "").strip()
            except Exception:
                k = ""
            if k == SEQ_KEY:
                found_row = r
                break

        # read current value
        cur = 0
        if found_row:
            try:
                cur = int(str(ws.cell(row=found_row, column=col_val).value or "0").strip() or "0")
            except Exception:
                cur = 0

        # bump
        nxt = int(cur) + 1

        ts = _utc_now_iso()
        if found_row:
            ws.cell(row=found_row, column=col_val).value = str(nxt)
            try:
                ws.cell(row=found_row, column=3).value = ts
            except Exception:
                pass
        else:
            ws.append([SEQ_KEY, str(nxt), ts])

        _save_and_close(wb, path)

    seq_str = str(int(nxt)).zfill(4)  # expands automatically when seq > 9999
    return f"{prefix}-{today}-{seq_str}"


def log_event(
    order_id: str,
    event_type: str,
    actor_role: str = "",
    actor_id: int = 0,
    actor_name: str = "",
    payload: dict | None = None,
    path: str = DEFAULT_PATH,
) -> None:
    """Append an event row; payload stored as JSON string."""
    wb = _load(path)
    ws = wb[SHEET_EVENTS]
    ws.append([
        str(uuid.uuid4()),
        str(order_id),
        str(event_type),
        str(actor_role),
        int(actor_id or 0),
        str(actor_name or ""),
        json.dumps(payload or {}, ensure_ascii=False),
        _utc_now_iso(),
    ])
    _save_and_close(wb, path)
def log_message(
    order_id: str,
    sender_role: str,
    sender_id: int,
    sender_name: str,
    receiver_role: str,
    receiver_id: int,
    text: str = "",
    file_id: str = "",
    path: str = DEFAULT_PATH,
) -> None:
    wb = _load(path)
    ws = wb[SHEET_MESSAGES]
    ws.append([
        str(uuid.uuid4()),
        str(order_id),
        str(sender_role),
        int(sender_id or 0),
        str(sender_name or ""),
        str(receiver_role),
        int(receiver_id or 0),
        str(text or ""),
        str(file_id or ""),
        _utc_now_iso(),
    ])
    _save_and_close(wb, path)
# === Traders profile (auto) ===
def get_trader_profile(trader_id: int, path: str = DEFAULT_PATH) -> dict:
    """
    Return trader profile dict.

    ✅ Always returns keys compatible with TRADERS_HEADERS when present:
    trader_id, display_name, company_name, shop_phone, cr_no, vat_no, payment_mode, bank_name, iban, stc_pay,
    joined_at_utc, is_enabled, updated_at_utc.

    ✅ Backward compatibility:
    If the XLSX has legacy columns, we will READ from them as fallback:
    - cr_number -> cr_no
    - vat_number -> vat_no
    - created_at_utc -> joined_at_utc
    - phone/shop_phone -> shop_phone (best-effort)
    """
    trader_id = int(trader_id or 0)
    if trader_id <= 0:
        return {}

    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb[SHEET_TRADERS]

    headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}

    tid_col = idx.get("trader_id", 1)

    # aliases (legacy -> new)
    aliases = {
        "cr_no": ["cr_no", "cr_number", "cr", "commercial_register", "commercial_register_no"],
        "vat_no": ["vat_no", "vat_number", "vat", "tax_no", "tax_number"],
        "joined_at_utc": ["joined_at_utc", "created_at_utc", "joined_at", "created_at"],
        "shop_phone": ["shop_phone", "phone", "shop_mobile", "mobile", "contact_phone"],
    }

    def _read_col(r: int, key: str):
        col = idx.get(key)
        if not col:
            return ""
        try:
            v = ws.cell(row=r, column=col).value
            return "" if v is None else str(v).strip()
        except Exception:
            return ""

    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=tid_col).value
        try:
            if int(v or 0) != trader_id:
                continue
        except Exception:
            continue

        out = {}
        # read all known headers if present
        for k in TRADERS_HEADERS:
            if k in idx:
                out[k] = _read_col(r, k)

        # ensure trader_id always present
        out["trader_id"] = str(trader_id)

        # ✅ default payment_mode for backward compatibility
        pm = (out.get("payment_mode") or "").strip().lower()
        if pm not in ("bank","link"):
            out["payment_mode"] = "bank"

        # fill fallbacks for missing new keys
        for new_k, legacy_keys in aliases.items():
            cur = (out.get(new_k) or "").strip()
            if cur:
                continue
            for lk in legacy_keys:
                if lk == new_k:
                    continue
                vv = _read_col(r, lk)
                if vv:
                    out[new_k] = vv
                    break

        return out

    return {}
def upsert_trader_profile(trader_id: int, fields: dict, path: str = DEFAULT_PATH) -> None:
    """Create/update trader profile row."""
    trader_id = int(trader_id or 0)
    if trader_id <= 0:
        return
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb[SHEET_TRADERS]

    headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}

    tid_col = idx.get("trader_id", 1)

    target_row = None
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=tid_col).value
        try:
            if int(v or 0) == trader_id:
                target_row = r
                break
        except Exception:
            continue

    if target_row is None:
        target_row = ws.max_row + 1
        ws.cell(row=target_row, column=tid_col).value = trader_id


    # ✅ تثبيت تاريخ الانضمام أول مرة فقط
    try:
        if "joined_at_utc" in idx:
            cur = ws.cell(row=target_row, column=idx["joined_at_utc"]).value
            if cur in (None, "", 0):
                ws.cell(row=target_row, column=idx["joined_at_utc"]).value = datetime.now(timezone.utc).isoformat()
    except Exception:
        pass

    for k, val in (fields or {}).items():
        if k not in idx:
            continue
        ws.cell(row=target_row, column=idx[k]).value = val

        # ✅ backward-safe mirroring (if legacy columns exist in the same file)
        # This prevents "missing fields" when old bot versions read legacy headers.
        try:
            if k == "cr_no" and "cr_number" in idx:
                ws.cell(row=target_row, column=idx["cr_number"]).value = val
            if k == "vat_no" and "vat_number" in idx:
                ws.cell(row=target_row, column=idx["vat_number"]).value = val
            if k == "joined_at_utc" and "created_at_utc" in idx:
                # only if legacy empty
                cur = ws.cell(row=target_row, column=idx["created_at_utc"]).value
                if cur in (None, "", 0):
                    ws.cell(row=target_row, column=idx["created_at_utc"]).value = val
            if k == "shop_phone" and "phone" in idx:
                ws.cell(row=target_row, column=idx["phone"]).value = val
        except Exception:
            pass

    if "updated_at_utc" in idx:
        ws.cell(row=target_row, column=idx["updated_at_utc"]).value = datetime.now(timezone.utc).isoformat()
    _save_and_close(wb, path)
# === Admin/Trader dashboards helpers (lightweight) ===
def list_orders(path: str = DEFAULT_PATH) -> list[dict]:
    """Return all orders as list of dicts (best-effort)."""
    ensure_workbook(path)
    wb = _load(path)
    try:
        ws = wb[SHEET_ORDERS]
        headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
        idx = {h: i + 1 for i, h in enumerate(headers) if h}
        out: list[dict] = []
        for r in range(2, ws.max_row + 1):
            row = {}
            for h, c in idx.items():
                v = ws.cell(row=r, column=c).value
                row[h] = v if v is not None else ""
            if str(row.get("order_id") or "").strip():
                out.append(row)
        return out

    finally:
        try:
            wb.close()
        except Exception:
            pass
def list_orders_for_trader(trader_id: int, path: str = DEFAULT_PATH) -> list[dict]:
    """Orders where accepted_trader_id equals trader_id."""
    tid = int(trader_id or 0)
    if tid <= 0:
        return []
    orders = list_orders(path)
    out = []
    for o in orders:
        try:
            atid = int(o.get("accepted_trader_id") or 0)
        except Exception:
            atid = 0
        if atid == tid:
            out.append(o)
    return out

def compute_admin_financials(path: str = DEFAULT_PATH) -> dict:
    """Compute simple sums by trader for confirmed goods payments."""
    orders = list_orders(path)
    total_confirmed = 0.0
    per_trader: dict[int, float] = {}
    per_trader_count: dict[int, int] = {}
    total_count = 0
    for o in orders:
        gps = str(o.get("goods_payment_status") or "").strip().lower()
        ost = str(o.get("order_status") or "").strip().lower()
        if gps == "confirmed" or ost in ("closed", "delivered"):
            amt_raw = str(o.get("goods_amount_sar") or "").strip()
            try:
                amt = float(re.sub(r"[^0-9.]+", "", amt_raw) or 0)
            except Exception:
                amt = 0.0
            if amt <= 0:
                continue
            total_confirmed += amt
            total_count += 1
            try:
                tid = int(o.get("accepted_trader_id") or 0)
            except Exception:
                tid = 0
            if tid > 0:
                per_trader[tid] = per_trader.get(tid, 0.0) + amt
                per_trader_count[tid] = per_trader_count.get(tid, 0) + 1
    return {
        "total_confirmed_amount": total_confirmed,
        "total_confirmed_count": total_count,
        "per_trader_amount": per_trader,
        "per_trader_count": per_trader_count,
    }


def compute_revenue_breakdown(path: str = DEFAULT_PATH) -> dict:
    """
    Platform & traders revenue breakdown:
    - platform_fees: price_sar (confirmed / pending)
    - traders_goods: goods_amount_sar (confirmed)
    - shipping_fees: shipping_fee_sar (confirmed)
    """
    orders = list_orders(path)

    platform_confirmed = 0.0
    platform_pending = 0.0
    traders_goods_confirmed = 0.0
    shipping_confirmed = 0.0

    for o in orders:
        # platform fee
        ps = str(o.get("payment_status") or "").strip().lower()
        price_raw = str(o.get("price_sar") or "").strip()
        try:
            price = float(re.sub(r"[^0-9.]+", "", price_raw) or 0)
        except Exception:
            price = 0.0

        if price > 0:
            if ps == "confirmed":
                platform_confirmed += price
            else:
                platform_pending += price

        # trader goods
        gps = str(o.get("goods_payment_status") or "").strip().lower()
        if gps == "confirmed":
            amt_raw = str(o.get("goods_amount_sar") or "").strip()
            try:
                traders_goods_confirmed += float(re.sub(r"[^0-9.]+", "", amt_raw) or 0)
            except Exception:
                pass

            ship_raw = str(o.get("shipping_fee_sar") or "").strip()
            try:
                shipping_confirmed += float(re.sub(r"[^0-9.]+", "", ship_raw) or 0)
            except Exception:
                pass

    return {
        "platform_fees_confirmed": platform_confirmed,
        "platform_fees_pending": platform_pending,
        "traders_goods_confirmed": traders_goods_confirmed,
        "shipping_confirmed": shipping_confirmed,
    }


def _sheet_headers(ws):
    return [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]

def get_setting(key: str, default: str = "", path: str = DEFAULT_PATH) -> str:
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb[SHEET_SETTINGS]
    headers = _sheet_headers(ws)
    if "key" not in headers or "value" not in headers:
        wb.close()
        return default
    col_key = headers.index("key") + 1
    col_val = headers.index("value") + 1
    for r in range(2, ws.max_row + 1):
        k = str(ws.cell(row=r, column=col_key).value or "").strip()
        if k == key:
            v = str(ws.cell(row=r, column=col_val).value or "")
            wb.close()
            return v
    wb.close()
    return default

def set_setting(key: str, value: str, actor_id: int = 0, actor_name: str = "", path: str = DEFAULT_PATH) -> None:
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb[SHEET_SETTINGS]
    headers = _sheet_headers(ws)
    col_key = headers.index("key") + 1
    col_val = headers.index("value") + 1
    col_upd = headers.index("updated_at_utc") + 1 if "updated_at_utc" in headers else None

    target_row = None
    for r in range(2, ws.max_row + 1):
        k = str(ws.cell(row=r, column=col_key).value or "").strip()
        if k == key:
            target_row = r
            break
    if target_row is None:
        target_row = ws.max_row + 1
        ws.cell(row=target_row, column=col_key).value = key

    ws.cell(row=target_row, column=col_val).value = str(value)
    if col_upd:
        ws.cell(row=target_row, column=col_upd).value = datetime.now(timezone.utc).isoformat()
    _save_and_close(wb, path)

def append_legal_log(actor_id: int, actor_name: str, action: str, details: str = "", path: str = DEFAULT_PATH) -> None:
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb[SHEET_LEGAL_LOG]
    headers = _sheet_headers(ws)

    def _col(name: str):
        return headers.index(name) + 1 if name in headers else None

    r = ws.max_row + 1
    ws.cell(row=r, column=_col("ts_utc") or 1).value = datetime.now(timezone.utc).isoformat()
    ws.cell(row=r, column=_col("actor_id") or 2).value = int(actor_id or 0)
    ws.cell(row=r, column=_col("actor_name") or 3).value = str(actor_name or "")
    ws.cell(row=r, column=_col("action") or 4).value = str(action or "")
    ws.cell(row=r, column=_col("details") or 5).value = str(details or "")
    _save_and_close(wb, path)

def list_legal_log(limit: int = 50, path: str = DEFAULT_PATH) -> list[dict]:
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb[SHEET_LEGAL_LOG]
    headers = _sheet_headers(ws)
    out = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers, start=1):
            row[h] = ws.cell(row=r, column=c).value
        out.append(row)
    wb.close()
    # newest last; return newest first
    out = out[-limit:][::-1]
    return out

def list_traders(path: str = DEFAULT_PATH) -> list[dict]:
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb[SHEET_TRADERS]
    headers = _sheet_headers(ws)
    out = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers, start=1):
            row[h] = ws.cell(row=r, column=c).value
        tid = str(row.get("trader_id") or "").strip()
        if tid:
            out.append(row)
    wb.close()
    return out

def set_trader_enabled(trader_id: int, enabled: bool, path: str = DEFAULT_PATH) -> None:
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb[SHEET_TRADERS]
    headers = _sheet_headers(ws)
    if "trader_id" not in headers:
        wb.close()
        return
    col_tid = headers.index("trader_id") + 1
    col_en = headers.index("is_enabled") + 1 if "is_enabled" in headers else None
    col_upd = headers.index("updated_at_utc") + 1 if "updated_at_utc" in headers else None

    target_row = None
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col_tid).value
        try:
            if int(v or 0) == int(trader_id or 0):
                target_row = r
                break
        except Exception:
            continue
    if target_row is None:
        # create row
        target_row = ws.max_row + 1
        ws.cell(row=target_row, column=col_tid).value = int(trader_id or 0)

    if col_en:
        ws.cell(row=target_row, column=col_en).value = "yes" if enabled else "no"
    if col_upd:
        ws.cell(row=target_row, column=col_upd).value = datetime.now(timezone.utc).isoformat()
    _save_and_close(wb, path)

def is_trader_enabled(trader_id: int, path: str = DEFAULT_PATH) -> bool:
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb[SHEET_TRADERS]
    headers = _sheet_headers(ws)

    if "trader_id" not in headers:
        wb.close()
        return True

    col_tid = headers.index("trader_id") + 1
    col_en = headers.index("is_enabled") + 1 if "is_enabled" in headers else None

    if not col_en:
        wb.close()
        return True

    for r in range(2, ws.max_row + 1):
        try:
            tid = int(ws.cell(row=r, column=col_tid).value or 0)
        except Exception:
            continue
        if tid == int(trader_id):
            v = str(ws.cell(row=r, column=col_en).value or "yes").strip().lower()
            wb.close()
            return v != "no"

    wb.close()
    return True


# =========================
# Trader subscriptions (monthly)
# =========================

def month_key_utc(ts: str | None = None) -> str:
    """Return YYYY-MM for given ISO ts or now (UTC)."""
    if ts:
        try:
            d = datetime.fromisoformat(str(ts).replace("Z", "+00:00"))
        except Exception:
            d = datetime.now(timezone.utc)
    else:
        d = datetime.now(timezone.utc)
    d = d.astimezone(timezone.utc)
    return f"{d.year:04d}-{d.month:02d}"

def upsert_trader_subscription(trader_id: int, month: str, fields: dict, path: str = DEFAULT_PATH) -> dict:
    """Create or update a trader subscription row for (trader_id, month). Returns the row dict."""
    wb = _load(path)
    try:
        ws = wb[SHEET_TRADER_SUBS]
        h = _header_index(ws)

        trader_id = int(trader_id or 0)
        month = str(month or "").strip()
        if not trader_id or not month:
            return {}

        target_row = 0
        for r in range(2, ws.max_row + 1):
            try:
                tid = int(ws.cell(row=r, column=h.get("trader_id")).value or 0)
            except Exception:
                tid = 0
            m = str(ws.cell(row=r, column=h.get("month")).value or "").strip()
            if tid == trader_id and m == month:
                target_row = r
                break

        now = utc_now_iso()
        if not target_row:
            target_row = ws.max_row + 1
            base = {
                "sub_id": str(uuid.uuid4()),
                "trader_id": trader_id,
                "month": month,
                "amount_sar": fields.get("amount_sar", 99),
                "payment_method": fields.get("payment_method", ""),
                "payment_status": fields.get("payment_status", "awaiting"),
                "receipt_file_id": fields.get("receipt_file_id", ""),
                "paid_at_utc": fields.get("paid_at_utc", ""),
                "created_at_utc": now,
            }
            for k, v in base.items():
                if k in h:
                    ws.cell(row=target_row, column=h[k]).value = v

        for k, v in (fields or {}).items():
            if k in h:
                ws.cell(row=target_row, column=h[k]).value = v
        _save_and_close(wb, path)
        return get_trader_subscription(trader_id, month, path=path)

    finally:
        try:
            wb.close()
        except Exception:
            pass
def get_trader_subscription(trader_id: int, month: str, path: str = DEFAULT_PATH) -> dict:
    wb = _load(path)
    ws = wb[SHEET_TRADER_SUBS]
    h = _header_index(ws)
    trader_id = int(trader_id or 0)
    month = str(month or "").strip()
    for r in range(2, ws.max_row + 1):
        try:
            tid = int(ws.cell(row=r, column=h.get("trader_id")).value or 0)
        except Exception:
            tid = 0
        m = str(ws.cell(row=r, column=h.get("month")).value or "").strip()
        if tid == trader_id and m == month:
            row = {}
            for k, c in h.items():
                row[k] = ws.cell(row=r, column=c).value
            return row
    return {}

def list_trader_subscriptions(month: str = "", path: str = DEFAULT_PATH) -> list[dict]:
    wb = _load(path)
    ws = wb[SHEET_TRADER_SUBS]
    h = _header_index(ws)
    out = []
    month = str(month or "").strip()
    for r in range(2, ws.max_row + 1):
        m = str(ws.cell(row=r, column=h.get("month")).value or "").strip()
        if month and m != month:
            continue
        row = {}
        for k, c in h.items():
            row[k] = ws.cell(row=r, column=c).value
        out.append(row)
    return out